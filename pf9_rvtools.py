#!/usr/bin/env python3
# Platform9 RVTools-like Export
# + XLSX export
# + CSV export
# + Delta / trend tracking
# + Optional customer-safe masking
# Includes:
#   Domains, Projects
#   Servers, Volumes, Snapshots, Images
#   Hypervisors (HW nodes), Flavors
#   Networks, Subnets, Ports, Routers, Floating IPs
# Adds:
#   project_name + tenant_name/domain_name on all relevant sheets
#   Ports also get tenant info via network fallback
#   Servers sheet now includes attached_networks, attached_ips, attached_ports
# (NO EMAIL)

import os
import csv
import json
import argparse
import traceback
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Load .env file if running outside Docker
env_file = Path(__file__).parent / '.env'
if env_file.exists() and not os.getenv('PF9_USERNAME'):
    with open(env_file, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, value = line.split('=', 1)
                # Strip quotes from value if present
                value = value.strip().strip('"').strip("'")
                os.environ[key.strip()] = value

from p9_common import (
    CFG, ERRORS, now_utc_str,
    mask_value,
    get_session_best_scope,
    list_domains_all, list_projects_all,
    list_users_all_domains, get_all_users_multi_domain, list_roles_all, list_role_assignments_all, list_groups_all,
    nova_servers_all, nova_hypervisors_all, nova_flavors,
    nova_keypairs, nova_server_groups, nova_aggregates, nova_availability_zones, nova_quotas,
    cinder_volumes_all, cinder_snapshots_all, cinder_volume_types, cinder_quotas,
    glance_images,
    neutron_list, neutron_quotas,
    infer_user_roles_from_data,
)

from db_writer import (
    db_connect,
    start_inventory_run,
    finish_inventory_run,
    upsert_domains,
    upsert_projects,
    upsert_hypervisors,
    upsert_servers,
    upsert_volumes,
    upsert_networks,
    upsert_subnets,
    upsert_ports,
    upsert_routers,
    upsert_floating_ips,
    upsert_flavors,       # ✅ Flavors
    upsert_images,        # ✅ Images
    upsert_snapshots,     # ✅ Snapshots (NEW)
    write_users,          # ✅ Users (NEW)
    write_roles,          # ✅ Roles (NEW)
    write_role_assignments, # ✅ Role Assignments (NEW)
    write_groups,         # ✅ Groups (NEW)
    upsert_security_groups,      # ✅ Security Groups
    upsert_security_group_rules, # ✅ Security Group Rules
    upsert_keypairs,             # ✅ Keypairs
    upsert_server_groups,        # ✅ Server Groups
    upsert_host_aggregates,      # ✅ Host Aggregates
    upsert_volume_types,         # ✅ Volume Types
    upsert_project_quotas,       # ✅ Project Quotas
)

STATE_FILE = "p9_rvtools_state.json"

ENABLE_DB = os.getenv("PF9_ENABLE_DB", "1") == "1"


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------

def autosize(ws):
    """
    Auto-size Excel columns based on cell contents.
    """
    for column_cells in ws.columns:
        length = 0
        col = column_cells[0].column  # 1-based index
        for cell in column_cells:
            try:
                value = cell.value
                if value is None:
                    continue
                length = max(length, len(str(value)))
            except Exception:
                continue
        if length > 0:
            ws.column_dimensions[get_column_letter(col)].width = min(length + 2, 80)


def excel_safe(value):
    """
    Convert Python value to something safe for Excel.
    - convert booleans and numbers to themselves
    - convert dict / list to JSON strings
    - convert others to strings
    """
    if value is None:
        return None
    if isinstance(value, (bool, int, float)):
        return value
    if isinstance(value, (list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return value


def write_table(ws, rows):
    """Write list[dict] to worksheet with Excel-safe values."""
    if not rows:
        ws.append(["(no data)"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)

    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(bold=True)

    for r in rows:
        ws.append([excel_safe(r.get(h)) for h in headers])

    ws.freeze_panes = "A2"
    autosize(ws)


def export_csv(base_dir, name, rows):
    """Export list[dict] to CSV with union-of-keys header."""
    if not rows:
        return

    os.makedirs(base_dir, exist_ok=True)
    path = os.path.join(base_dir, f"{name}.csv")

    # Build union of keys to keep CSV header consistent
    keys = set()
    for r in rows:
        keys.update(r.keys())
    header = sorted(keys)

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=header)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in header})


def load_state():
    if not os.path.exists(STATE_FILE):
        return {}
    try:
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_state(data):
    try:
        with open(STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception:
        traceback.print_exc()


# ------------------------------------------------------------------
# Data munging / enrichment for project + tenant info
# ------------------------------------------------------------------

def build_project_domain_maps(projects, domains):
    """
    Build lookup dictionaries:
      - project_id -> project record (with domain_id)
      - domain_id -> domain record
    """
    project_by_id = {}
    for p in projects:
        pid = p.get("id")
        if pid:
            project_by_id[pid] = p

    domain_by_id = {}
    for d in domains:
        did = d.get("id")
        if did:
            domain_by_id[did] = d

    return project_by_id, domain_by_id


def cleanup_old_records(conn):
    """Clean up records that haven't been seen in recent scans.
    
    Deletes rows whose last_seen_at is older than the cutoff (15 min).
    Order respects foreign-key constraints (children before parents).
    Every deletion is logged to deletions_history for the audit trail.
    """
    from datetime import datetime, timedelta, timezone
    
    # Very aggressive cleanup - remove records older than 15 minutes
    # This ensures database always reflects current Platform9 state exactly
    cutoff = datetime.now(timezone.utc) - timedelta(minutes=15)
    
    with conn.cursor() as cur:
        # Order matters: child tables first, then parent tables
        # security_group_rules has ON DELETE CASCADE from security_groups so it's auto-handled,
        # but we include it explicitly for audit logging.
        tables = [
            'security_group_rules', 'security_groups',
            'snapshots', 'images', 'flavors',
            'servers', 'volumes',
            'floating_ips', 'ports', 'routers', 'subnets', 'networks',
            'hypervisors',
            'projects', 'domains',
        ]
        
        # Map table names to their resource types for deletions_history
        resource_type_map = {
            'snapshots': 'snapshot',
            'images': 'image',
            'flavors': 'flavor',
            'servers': 'server',
            'volumes': 'volume',
            'floating_ips': 'floating_ip',
            'ports': 'port',
            'routers': 'router',
            'subnets': 'subnet',
            'networks': 'network',
            'hypervisors': 'hypervisor',
            'domains': 'domain',
            'projects': 'project',
            'security_groups': 'security_group',
            'security_group_rules': 'security_group_rule',
        }
        
        total_removed = 0
        for table in tables:
            try:
                resource_type = resource_type_map.get(table, table)

                # Clear ALL foreign key references before deleting parent rows
                if table == 'projects':
                    stale_proj = "SELECT id FROM projects WHERE last_seen_at < %s"
                    # Null out user FK references
                    cur.execute(f"UPDATE users SET default_project_id = NULL WHERE default_project_id IN ({stale_proj})", (cutoff,))
                    # Delete child resources that reference stale projects
                    for child in ['security_group_rules', 'security_groups',
                                  'snapshots', 'servers', 'volumes',
                                  'floating_ips', 'ports', 'routers', 'subnets', 'networks']:
                        try:
                            if child == 'security_group_rules':
                                cur.execute(f"DELETE FROM security_group_rules WHERE security_group_id IN (SELECT id FROM security_groups WHERE project_id IN ({stale_proj}))", (cutoff,))
                            elif child == 'subnets':
                                cur.execute(f"DELETE FROM subnets WHERE network_id IN (SELECT id FROM networks WHERE project_id IN ({stale_proj}))", (cutoff,))
                            else:
                                cur.execute(f"DELETE FROM {child} WHERE project_id IN ({stale_proj})", (cutoff,))
                        except Exception:
                            conn.rollback()
                elif table == 'domains':
                    stale_dom = "SELECT id FROM domains WHERE last_seen_at < %s"
                    # Null out user FK references
                    cur.execute(f"UPDATE users SET domain_id = NULL WHERE domain_id IN ({stale_dom})", (cutoff,))
                    # Orphan projects referencing stale domains
                    cur.execute(f"UPDATE projects SET domain_id = NULL WHERE domain_id IN ({stale_dom})", (cutoff,))

                # Determine which columns exist in the table
                cur.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = %s AND table_schema = 'public'", (table,))
                cols = set(r[0] for r in cur.fetchall())
                # Use NULL for missing columns
                name_col = 'name as resource_name' if 'name' in cols else 'NULL as resource_name'
                project_name_col = 'project_name' if 'project_name' in cols else 'NULL as project_name'
                domain_name_col = 'domain_name' if 'domain_name' in cols else 'NULL as domain_name'
                # Build the SELECT statement dynamically
                select_stmt = f"SELECT %s::text as resource_type, id::text as resource_id, {name_col}, {project_name_col}, {domain_name_col}, last_seen_at as last_seen_before_deletion, 'Automatic cleanup - not seen in recent RVTools scan' as reason, row_to_json({table}.*)::jsonb as raw_json_snapshot FROM {table} WHERE last_seen_at < %s"
                cur.execute(f"""
                    INSERT INTO deletions_history 
                        (resource_type, resource_id, resource_name, project_name, 
                         domain_name, last_seen_before_deletion, reason, raw_json_snapshot)
                    {select_stmt}
                """, (resource_type, cutoff))
                logged = cur.rowcount
                cur.execute(f"DELETE FROM {table} WHERE last_seen_at < %s", (cutoff,))
                removed = cur.rowcount
                if removed > 0:
                    print(f"[DB] Cleaned {removed} old records from {table} (logged {logged} to deletions_history)")
                    total_removed += removed
                conn.commit()
            except Exception as e:
                print(f"[DB] Warning: Could not clean {table}: {e}")
                conn.rollback()
        
        # Also clean up metering_pricing entries for flavors that no longer exist
        try:
            cur.execute("""
                DELETE FROM metering_pricing
                WHERE category = 'flavor'
                  AND item_name NOT IN (SELECT name FROM flavors)
            """)
            pricing_removed = cur.rowcount
            if pricing_removed > 0:
                print(f"[DB] Cleaned {pricing_removed} stale flavor pricing entries")
                total_removed += pricing_removed
            conn.commit()
        except Exception as e:
            print(f"[DB] Warning: Could not clean stale pricing: {e}")
            conn.rollback()

        if total_removed > 0:
            print(f"[DB] Total cleaned: {total_removed} old records (all logged to audit trail)")
        else:
            print("[DB] No old records to clean")


def enrich_project_records(projects, domain_by_id):
    """Add tenant_name/domain_name to project records."""
    for p in projects:
        dom_id = p.get("domain_id")
        dom = domain_by_id.get(dom_id)
        if dom:
            name = dom.get("name")
            p["tenant_name"] = name
            p["domain_name"] = name


def attach_project_info(obj, project_by_id, domain_by_id):
    """
    Attach project_name + tenant/domain info based 
    on any project/tenant key we find.
    This is called for each resource row (servers, volumes, snapshots, ports, etc.).
    """
    if not isinstance(obj, dict):
        return

    pid = (
        obj.get("project_id")
        or obj.get("tenant_id")
        or obj.get("os-tenant-id")
        or obj.get("os-vol-tenant-attr:tenant_id")
        or obj.get("os-extended-snapshot-attributes:project_id")  # Cinder snapshots
        or obj.get("owner")
        or obj.get("owner_id")
    )

    if not pid:
        return

    proj = project_by_id.get(pid)
    if not proj:
        return

    # Basic project info
    obj["project_id"] = proj.get("id")
    obj["project_name"] = proj.get("name")

    dom_id = proj.get("domain_id")
    dom = domain_by_id.get(dom_id)
    if dom:
        obj["tenant_name"] = proj.get("name")  # ✅ Use project name as tenant name
        obj["domain_name"] = dom.get("name")
        obj["domain_id"] = dom.get("id")


def enrich_ports_from_networks(ports, networks):
    """
    For ports that still have no project or tenant, fall back to their network.
    """
    net_by_id = {n.get("id"): n for n in networks if n.get("id")}

    for p in ports:
        if not isinstance(p, dict):
            continue

        # If we already have project or tenant info, skip
        if p.get("project_id") or p.get("tenant_name") or p.get("domain_name"):
            continue

        net = net_by_id.get(p.get("network_id"))
        if not net:
            continue

        # Copy relevant fields
        for field in ("project_id", "project_name", "tenant_name",
                      "domain_name", "domain_id"):
            if field in net and p.get(field) is None:
                p[field] = net.get(field)


def enrich_servers_with_ports(servers, ports, networks):
    """
    For each server, add:
      - attached_networks: comma-separated Neutron network names
      - attached_ips:      comma-separated fixed IP addresses
      - attached_ports:    comma-separated Neutron port IDs

    This uses Neutron ports (device_id == server.id) and networks.
    """
    # Map network_id -> network object
    net_by_id = {n.get("id"): n for n in networks if n.get("id")}

    # Build a mapping from server_id -> list of ports
    ports_by_server = {}
    for p in ports:
        if not isinstance(p, dict):
            continue
        dev_id = p.get("device_id")
        if not dev_id:
            continue
        ports_by_server.setdefault(dev_id, []).append(p)

    for s in servers:
        if not isinstance(s, dict):
            continue
        sid = s.get("id")
        if not sid:
            continue

        attached = ports_by_server.get(sid, [])
        net_names = []
        ip_addrs = []
        port_ids = []

        for p in attached:
            port_id = p.get("id")
            if port_id:
                port_ids.append(port_id)

            # Collect fixed IP addresses
            for addr in p.get("fixed_ips", []):
                ip = addr.get("ip_address")
                if ip:
                    ip_addrs.append(ip)

            # Resolve network name
            net = net_by_id.get(p.get("network_id"))
            if net:
                nname = net.get("name")
                if nname:
                    net_names.append(nname)

        # Deduplicate and store as comma-separated strings
        s["attached_networks"] = ", ".join(sorted(set(net_names)))
        s["attached_ips"] = ", ".join(sorted(set(ip_addrs)))
        s["attached_ports"] = ", ".join(sorted(set(port_ids)))


def enrich_all_with_project_and_tenant(projects, domains,
                                       servers, volumes, snapshots,
                                       images, networks, subnets,
                                       ports, routers, floatingips,
                                       security_groups=None, security_group_rules=None):
    project_by_id, domain_by_id = build_project_domain_maps(projects, domains)

    # Projects get tenant/domain names directly
    enrich_project_records(projects, domain_by_id)

    # For all other resource lists, attach project_name + tenant/domain_name
    all_collections = [servers, volumes, snapshots, images,
                       networks, subnets, ports, routers, floatingips]
    if security_groups:
        all_collections.append(security_groups)
    for coll in all_collections:
        for obj in coll:
            attach_project_info(obj, project_by_id, domain_by_id)

    # Extra: for ports that still have no project/tenant, fall back to network
    enrich_ports_from_networks(ports, networks)

    # NEW: enrich servers with network/port/IP info
    enrich_servers_with_ports(servers, ports, networks)

    # Enrich security groups with VM + network associations
    if security_groups:
        enrich_security_groups(security_groups, security_group_rules or [], ports, servers, networks)


def enrich_security_groups(security_groups, rules, ports, servers, networks):
    """
    Enrich security groups with associated VMs, networks, and rule summaries.
    """
    sg_by_id = {sg.get("id"): sg for sg in security_groups if sg.get("id")}
    server_by_id = {s.get("id"): s for s in servers if s.get("id")}
    net_by_id = {n.get("id"): n for n in networks if n.get("id")}

    # Build mapping: security_group_id -> list of ports
    sg_ports = {}
    for p in ports:
        if not isinstance(p, dict):
            continue
        sgs = p.get("security_groups", [])
        if not isinstance(sgs, list):
            continue
        for sg_id in sgs:
            sg_ports.setdefault(sg_id, []).append(p)

    # Build mapping: security_group_id -> list of rules
    sg_rules = {}
    for r in rules:
        sg_id = r.get("security_group_id")
        if sg_id:
            sg_rules.setdefault(sg_id, []).append(r)

    for sg in security_groups:
        sg_id = sg.get("id")
        if not sg_id:
            continue

        associated_ports = sg_ports.get(sg_id, [])

        # Attached VMs
        vm_ids = set()
        vm_names = []
        net_ids = set()
        net_names = []

        for p in associated_ports:
            dev_id = p.get("device_id")
            dev_owner = p.get("device_owner", "")
            if dev_id and "compute" in dev_owner:
                vm_ids.add(dev_id)
                srv = server_by_id.get(dev_id)
                if srv:
                    vm_names.append(srv.get("name", dev_id))
                else:
                    vm_names.append(dev_id)

            net_id = p.get("network_id")
            if net_id and net_id not in net_ids:
                net_ids.add(net_id)
                net = net_by_id.get(net_id)
                if net:
                    net_names.append(net.get("name", net_id))
                else:
                    net_names.append(net_id)

        sg["attached_vm_ids"] = ", ".join(sorted(vm_ids))
        sg["attached_vm_names"] = ", ".join(sorted(set(vm_names)))
        sg["attached_vm_count"] = len(vm_ids)
        sg["attached_network_ids"] = ", ".join(sorted(net_ids))
        sg["attached_network_names"] = ", ".join(sorted(set(net_names)))
        sg["attached_network_count"] = len(net_ids)

        # Rule counts
        my_rules = sg_rules.get(sg_id, [])
        ingress = [r for r in my_rules if r.get("direction") == "ingress"]
        egress = [r for r in my_rules if r.get("direction") == "egress"]
        sg["ingress_rule_count"] = len(ingress)
        sg["egress_rule_count"] = len(egress)
        sg["total_rule_count"] = len(my_rules)


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--customer-view", action="store_true")
    args = parser.parse_args()

    conn = None
    run_id = None
    if ENABLE_DB:
        try:
            # Check if we can connect to the database
            conn = db_connect()
            run_id = start_inventory_run(conn, source="pf9_rvtools")
            print("[DB] Connected to database successfully")
        except Exception as e:
            print(f"[DB] Database connection failed: {e}")
            print("[DB] Continuing without database storage...")
            traceback.print_exc()
            conn = None
            run_id = None

    print("[1] Auth")
    session, token, body, scope_mode, _ = get_session_best_scope()
    project_id = body["token"]["project"]["id"]

    print("[2] Inventory")

    # Keystone - Core Resources
    domains  = list_domains_all(session)
    projects = list_projects_all(session)
    
    # Keystone - User Management (NEW)
    print("    [KEY] Collecting users and roles across all domains...")
    users           = get_all_users_multi_domain()
    roles           = list_roles_all(session)
    role_assignments = list_role_assignments_all(session)
    groups          = list_groups_all(session)
    
    # If no role assignments were collected, try to infer them
    if len(role_assignments) == 0:
        print("    [INFO] No role assignments collected, attempting to infer roles...")
        role_assignments = infer_user_roles_from_data(users, projects, roles)
    
    print(f"    [OK] Found {len(users)} users, {len(roles)} roles, {len(role_assignments)} role assignments, {len(groups)} groups")

    # Nova
    servers      = nova_servers_all(session)
    hypervisors  = nova_hypervisors_all(session)
    flavors      = nova_flavors(session)

    # Nova additional metadata
    print("    [NOVA] Collecting keypairs, server groups, aggregates, AZs...")
    keypairs = []
    server_groups = []
    aggregates = []
    availability_zones = []
    try:
        keypairs = nova_keypairs(session)
    except Exception as e:
        print(f"    [WARN] Keypairs collection failed (non-critical): {e}")
    try:
        server_groups = nova_server_groups(session)
    except Exception as e:
        print(f"    [WARN] Server groups collection failed (non-critical): {e}")
    try:
        aggregates = nova_aggregates(session)
    except Exception as e:
        print(f"    [WARN] Aggregates collection failed (non-critical): {e}")
    try:
        availability_zones = nova_availability_zones(session)
    except Exception as e:
        print(f"    [WARN] Availability zones collection failed (non-critical): {e}")
    print(f"    [OK] Found {len(keypairs)} keypairs, {len(server_groups)} server groups, {len(aggregates)} aggregates, {len(availability_zones)} AZs")

    # Cinder
    volumes   = cinder_volumes_all(session, project_id)
    snapshots = cinder_snapshots_all(session, project_id)

    # Cinder additional metadata
    print("    [CINDER] Collecting volume types...")
    volume_types = []
    try:
        volume_types = cinder_volume_types(session)
    except Exception as e:
        print(f"    [WARN] Volume types collection failed (non-critical): {e}")
    print(f"    [OK] Found {len(volume_types)} volume types")

    # Glance
    images = glance_images(session)

    # Neutron
    networks    = neutron_list(session, "networks")
    subnets     = neutron_list(session, "subnets")
    ports       = neutron_list(session, "ports")
    routers     = neutron_list(session, "routers")
    floatingips = neutron_list(session, "floatingips")

    # Security Groups (Neutron)
    print("    [NET] Collecting security groups...")
    security_groups = neutron_list(session, "security-groups")
    security_group_rules = neutron_list(session, "security-group-rules")
    print(f"    [OK] Found {len(security_groups)} security groups, {len(security_group_rules)} rules")

    # --------------------------------------------------------------
    # Enrich records with project_name + tenant_name/domain_name
    # + server network/port/IP info
    # --------------------------------------------------------------
    enrich_all_with_project_and_tenant(
        projects, domains,
        servers, volumes, snapshots,
        images, networks, subnets,
        ports, routers, floatingips,
        security_groups=security_groups,
        security_group_rules=security_group_rules,
    )

    # Enrich servers with OS info (from image metadata)
    image_os_lookup = {}
    for img in images:
        iid = img.get('id')
        if iid:
            image_os_lookup[iid] = {
                'os_distro': img.get('os_distro') or img.get('properties', {}).get('os_distro') or '',
                'os_version': img.get('os_version') or img.get('properties', {}).get('os_version') or '',
                'name': img.get('name', ''),
            }
    for s in servers:
        # Try server metadata first
        md = s.get('metadata') or {}
        os_distro = md.get('os_distro', '')
        os_version = md.get('os_version', '')
        # Fallback to image
        image_ref = s.get('image')
        image_id = None
        if isinstance(image_ref, dict):
            image_id = image_ref.get('id')
        elif isinstance(image_ref, str) and image_ref:
            image_id = image_ref
        if image_id and image_id in image_os_lookup:
            il = image_os_lookup[image_id]
            if not os_distro:
                os_distro = il['os_distro']
            if not os_version:
                os_version = il['os_version']
        s['os_distro'] = os_distro
        s['os_version'] = os_version

    # Collect quotas as flat rows for Excel export
    quota_export_rows = []
    project_name_map = {p.get('id'): p.get('name', p.get('id', '')) for p in projects}
    try:
        for p in projects:
            pid = p.get('id')
            pname = p.get('name', pid)
            if not pid:
                continue
            for svc_name, svc_fn in [('nova', nova_quotas), ('cinder', cinder_quotas), ('neutron', neutron_quotas)]:
                try:
                    qdata = svc_fn(session, pid)
                    if isinstance(qdata, dict):
                        for resource, val in qdata.items():
                            if resource in ('id',):
                                continue
                            if isinstance(val, dict):
                                quota_export_rows.append({
                                    'project_name': pname, 'project_id': pid, 'service': svc_name,
                                    'resource': resource, 'limit': val.get('limit', ''),
                                    'in_use': val.get('in_use', ''), 'reserved': val.get('reserved', ''),
                                })
                            else:
                                quota_export_rows.append({
                                    'project_name': pname, 'project_id': pid, 'service': svc_name,
                                    'resource': resource, 'limit': val, 'in_use': '', 'reserved': '',
                                })
                except Exception:
                    pass
    except Exception as e:
        print(f"    [WARN] Quota export rows collection failed: {e}")
    print(f"    [OK] Collected {len(quota_export_rows)} quota rows for export")

    # --------------------------------------------------------------
    # Masking (customer-safe view)
    # --------------------------------------------------------------
    if args.customer_view:
        for p in projects:
            p["id"] = mask_value(p.get("id"), "project")
        for s in servers:
            s["id"] = mask_value(s.get("id"), "vm")
            s["OS-EXT-SRV-ATTR:host"] = mask_value(
                s.get("OS-EXT-SRV-ATTR:host"), "host"
            )
        for v in volumes:
            v["id"] = mask_value(v.get("id"), "vol")
        for sn in snapshots:
            sn["id"] = mask_value(sn.get("id"), "snap")
        for h in hypervisors:
            h["hypervisor_hostname"] = mask_value(
                h.get("hypervisor_hostname"), "hv"
            )

    # --------------------------------------------------------------
    # Delta / trend (counts)
    # --------------------------------------------------------------
    prev = load_state()
    current = {
        "servers": len(servers),
        "volumes": len(volumes),
        "snapshots": len(snapshots),
        "errors": len(ERRORS),
        "timestamp": now_utc_str(),
    }
    delta = {
        k: current[k] - prev.get(k, 0)
        for k in ("servers", "volumes", "snapshots", "errors")
    }
    save_state(current)

    # --------------------------------------------------------------
    # DB export (if enabled)
    # --------------------------------------------------------------
    if conn and run_id is not None:
        try:
            n_domains      = upsert_domains(conn, domains, run_id=run_id)
            n_projects     = upsert_projects(conn, projects, run_id=run_id)
            
            # User Management Data (NEW)
            print("    [INFO] Writing user management data...")
            n_users        = write_users(conn, users, run_id=run_id)
            n_roles        = write_roles(conn, roles, run_id=run_id) 
            n_assignments  = write_role_assignments(conn, role_assignments, run_id=run_id)
            n_groups       = write_groups(conn, groups, run_id=run_id)
            
            # Commit after user management to prevent transaction abort from affecting subsequent operations
            conn.commit()
            
            n_flavors      = upsert_flavors(conn, flavors, run_id=run_id)
            n_images       = upsert_images(conn, images, run_id=run_id)
            n_hv           = upsert_hypervisors(conn, hypervisors, run_id=run_id)
            n_servers      = upsert_servers(conn, servers, run_id=run_id)
            n_vols         = upsert_volumes(conn, volumes, run_id=run_id)
            n_snaps        = upsert_snapshots(conn, snapshots, run_id=run_id)  # ✅ NEW
            n_nets         = upsert_networks(conn, networks, run_id=run_id)
            n_subnets      = upsert_subnets(conn, subnets, run_id=run_id)
            n_ports        = upsert_ports(conn, ports, run_id=run_id)
            n_routers      = upsert_routers(conn, routers, run_id=run_id)
            n_fips         = upsert_floating_ips(conn, floatingips, run_id=run_id)
            n_sgs          = upsert_security_groups(conn, security_groups, run_id=run_id)
            n_sg_rules     = upsert_security_group_rules(conn, security_group_rules, run_id=run_id)

            # Additional metadata (non-critical, won't fail the run)
            n_keypairs = n_sgroups = n_aggs = n_vtypes = n_quotas = 0
            try:
                n_keypairs = upsert_keypairs(conn, keypairs)
            except Exception as e:
                print(f"    [WARN] Keypairs DB write failed: {e}")
            try:
                n_sgroups = upsert_server_groups(conn, server_groups)
            except Exception as e:
                print(f"    [WARN] Server groups DB write failed: {e}")
            try:
                n_aggs = upsert_host_aggregates(conn, aggregates)
            except Exception as e:
                print(f"    [WARN] Aggregates DB write failed: {e}")
            try:
                n_vtypes = upsert_volume_types(conn, volume_types)
            except Exception as e:
                print(f"    [WARN] Volume types DB write failed: {e}")

            # Project quotas (iterate projects)
            try:
                for p in projects:
                    pid = p.get('id')
                    if not pid:
                        continue
                    try:
                        nq = nova_quotas(session, pid)
                        upsert_project_quotas(conn, pid, 'nova', nq)
                        n_quotas += len(nq)
                    except Exception:
                        pass
                    try:
                        cq = cinder_quotas(session, pid)
                        upsert_project_quotas(conn, pid, 'cinder', cq)
                        n_quotas += len(cq)
                    except Exception:
                        pass
                    try:
                        neq = neutron_quotas(session, pid)
                        upsert_project_quotas(conn, pid, 'neutron', neq)
                        n_quotas += len(neq)
                    except Exception:
                        pass
                print(f"    [OK] Stored quotas for {len(projects)} projects ({n_quotas} quota entries)")
            except Exception as e:
                print(f"    [WARN] Quotas collection failed: {e}")

            notes = (
                f"domains={n_domains}, projects={n_projects}, "
                f"users={n_users}, roles={n_roles}, role_assignments={n_assignments}, groups={n_groups}, "
                f"flavors={n_flavors}, images={n_images}, "
                f"hypervisors={n_hv}, servers={n_servers}, "
                f"volumes={n_vols}, snapshots={n_snaps}, "
                f"networks={n_nets}, subnets={n_subnets}, ports={n_ports}, "
                f"routers={n_routers}, floating_ips={n_fips}, "
                f"security_groups={n_sgs}, security_group_rules={n_sg_rules}, "
                f"keypairs={n_keypairs}, server_groups={n_sgroups}, "
                f"aggregates={n_aggs}, volume_types={n_vtypes}, quota_entries={n_quotas}"
            )
            finish_inventory_run(conn, run_id, status="success", notes=notes)
            
            # Commit the main transaction first
            conn.commit()
            
            # Clean up old records in a separate transaction
            cleanup_old_records(conn)
        except Exception as e:
            traceback.print_exc()
            try:
                conn.rollback()  # IMPORTANT: clear failed transaction
                finish_inventory_run(conn, run_id, status="failure", notes=str(e))
            except Exception:
                traceback.print_exc()

        finally:
            try:
                conn.close()
            except Exception:
                pass
            conn = None
            run_id = None

    # --------------------------------------------------------------
    # Excel + CSV Export
    # --------------------------------------------------------------
    ts = now_utc_str().replace(":", "").replace(" ", "_")
    out_xlsx = os.path.join(CFG["OUTPUT_DIR"], f"p9_rvtools_{ts}.xlsx")
    csv_dir = os.path.join(CFG["OUTPUT_DIR"], "csv")

    os.makedirs(CFG["OUTPUT_DIR"], exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # Summary sheet
    summary = [{
        "Generated_UTC": now_utc_str(),
        "Servers": current["servers"],
        "Δ_Servers": delta["servers"],
        "Volumes": current["volumes"],
        "Δ_Volumes": delta["volumes"],
        "Snapshots": current["snapshots"],
        "Δ_Snapshots": delta["snapshots"],
        "Users": len(users),
        "Roles": len(roles),
        "Role_Assignments": len(role_assignments),
        "Groups": len(groups),
        "Keypairs": len(keypairs),
        "ServerGroups": len(server_groups),
        "HostAggregates": len(aggregates),
        "VolumeTypes": len(volume_types),
        "QuotaEntries": len(quota_export_rows),
        "Errors": current["errors"],
        "Δ_Errors": delta["errors"],
        "Scope_Mode": scope_mode,
    }]
    ws_summary = wb.create_sheet("Summary")
    write_table(ws_summary, summary)

    # Detailed sheets
    sheets = [
        ("Domains", domains),
        ("Projects", projects),
        ("Users", users),                    # ✅ NEW
        ("Roles", roles),                    # ✅ NEW 
        ("Role_Assignments", role_assignments), # ✅ NEW
        ("Groups", groups),                  # ✅ NEW
        ("Servers", servers),
        ("Volumes", volumes),
        ("Snapshots", snapshots),
        ("Images", images),
        ("Hypervisors", hypervisors),
        ("Flavors", flavors),
        ("Networks", networks),
        ("Subnets", subnets),
        ("Ports", ports),
        ("Routers", routers),
        ("FloatingIPs", floatingips),
        ("SecurityGroups", security_groups),
        ("SecurityGroupRules", security_group_rules),
        ("Keypairs", keypairs),
        ("ServerGroups", server_groups),
        ("HostAggregates", aggregates),
        ("VolumeTypes", volume_types),
        ("ProjectQuotas", quota_export_rows),
    ]

    for name, data in sheets:
        if data:
            ws = wb.create_sheet(name)
            write_table(ws, data)
            export_csv(csv_dir, name, data)

    wb.save(out_xlsx)
    print(f"Saved {out_xlsx}")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        traceback.print_exc()
        raise
