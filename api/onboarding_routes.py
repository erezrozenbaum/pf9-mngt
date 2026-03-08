# ---------------------------------------------------------------------------
# Bulk Customer Onboarding API — /api/onboarding
# Version: v1.38.0
# ---------------------------------------------------------------------------
"""
Workflow:
  UPLOAD → VALIDATING → VALIDATED | INVALID
  → PENDING_APPROVAL → APPROVED | REJECTED
  → DRY_RUN_PENDING → DRY_RUN_PASSED | DRY_RUN_FAILED
  → EXECUTING → COMPLETE | PARTIALLY_FAILED

Execution is hard-locked until both:
  - approval_status == 'approved'
  - status == 'dry_run_passed'
"""

from __future__ import annotations

import io
import ipaddress
import logging
import os
import re
import secrets
import string
import threading
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

import psycopg2.extras
from fastapi import APIRouter, Body, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from auth import get_current_user, require_permission
from db_pool import get_connection

logger = logging.getLogger("onboarding")

router = APIRouter(prefix="/api/onboarding", tags=["onboarding"])

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _now() -> datetime:
    return datetime.now(timezone.utc)


def _gen_temp_password(length: int = 16) -> str:
    alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
    while True:
        pw = "".join(secrets.choice(alphabet) for _ in range(length))
        if (
            any(c.isupper() for c in pw)
            and any(c.islower() for c in pw)
            and any(c.isdigit() for c in pw)
            and any(c in "!@#$%^&*" for c in pw)
        ):
            return pw


def _fire(event: str, title: str, message: str, severity: str = "info") -> None:
    try:
        from notification_routes import _fire_notification  # type: ignore
        _fire_notification(event, title, message, severity)
    except Exception as exc:
        logger.warning("Could not fire notification %s: %s", event, exc)


def _ensure_tables() -> None:
    _MIGRATION_SQL = """
    CREATE TABLE IF NOT EXISTS onboarding_batches (
        id                  SERIAL PRIMARY KEY,
        batch_id            UUID NOT NULL DEFAULT gen_random_uuid(),
        batch_name          TEXT NOT NULL,
        uploaded_by         TEXT NOT NULL,
        status              TEXT NOT NULL DEFAULT 'validating',
        validation_errors   JSONB DEFAULT '[]'::jsonb,
        dry_run_result      JSONB,
        approval_status     TEXT DEFAULT 'not_submitted',
        approved_by         TEXT,
        approved_at         TIMESTAMPTZ,
        rejection_comment   TEXT,
        execution_result    JSONB,
        total_customers     INT DEFAULT 0,
        total_projects      INT DEFAULT 0,
        total_networks      INT DEFAULT 0,
        total_users         INT DEFAULT 0,
        created_at          TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        updated_at          TIMESTAMPTZ NOT NULL DEFAULT NOW()
    );
    CREATE UNIQUE INDEX IF NOT EXISTS uq_onboarding_batches_batch_id
        ON onboarding_batches (batch_id);
    CREATE TABLE IF NOT EXISTS onboarding_customers (
        id              SERIAL PRIMARY KEY,
        batch_id        UUID NOT NULL REFERENCES onboarding_batches(batch_id) ON DELETE CASCADE,
        domain_name     TEXT NOT NULL,
        display_name    TEXT,
        description     TEXT,
        contact_email   TEXT,
        department_tag  TEXT,
        status          TEXT NOT NULL DEFAULT 'pending',
        error_msg       TEXT,
        pcd_domain_id   TEXT,
        created_at      TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        updated_at      TIMESTAMPTZ NOT NULL DEFAULT NOW()
    );
    CREATE INDEX IF NOT EXISTS idx_onboarding_customers_batch
        ON onboarding_customers (batch_id);
    CREATE TABLE IF NOT EXISTS onboarding_projects (
        id                      SERIAL PRIMARY KEY,
        batch_id                UUID NOT NULL REFERENCES onboarding_batches(batch_id) ON DELETE CASCADE,
        customer_id             INT REFERENCES onboarding_customers(id) ON DELETE CASCADE,
        domain_name             TEXT NOT NULL,
        project_name            TEXT NOT NULL,
        subscription_id         TEXT DEFAULT '',
        description             TEXT,
        quota_vcpu              INT DEFAULT 20,
        quota_ram_mb            INT DEFAULT 51200,
        quota_instances         INT DEFAULT 10,
        quota_server_groups     INT DEFAULT 10,
        quota_networks          INT DEFAULT 10,
        quota_subnets           INT DEFAULT 20,
        quota_routers           INT DEFAULT 5,
        quota_ports             INT DEFAULT 200,
        quota_floatingips       INT DEFAULT 20,
        quota_security_groups   INT DEFAULT 10,
        quota_volumes           INT DEFAULT 20,
        quota_snapshots         INT DEFAULT 20,
        quota_disk_gb           INT DEFAULT 1000,
        status                  TEXT NOT NULL DEFAULT 'pending',
        error_msg               TEXT,
        pcd_project_id          TEXT,
        created_at              TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        updated_at              TIMESTAMPTZ NOT NULL DEFAULT NOW()
    );
    CREATE INDEX IF NOT EXISTS idx_onboarding_projects_batch
        ON onboarding_projects (batch_id);
    CREATE TABLE IF NOT EXISTS onboarding_networks (
        id                      SERIAL PRIMARY KEY,
        batch_id                UUID NOT NULL REFERENCES onboarding_batches(batch_id) ON DELETE CASCADE,
        project_id              INT REFERENCES onboarding_projects(id) ON DELETE CASCADE,
        domain_name             TEXT NOT NULL,
        project_name            TEXT NOT NULL,
        network_name            TEXT NOT NULL,
        network_kind            TEXT DEFAULT 'physical_managed',
        network_type            TEXT DEFAULT 'vlan',
        physical_network        TEXT DEFAULT 'physnet1',
        cidr                    TEXT,
        gateway                 TEXT,
        dns1                    TEXT DEFAULT '8.8.8.8',
        vlan_id                 INT,
        dhcp_enabled            BOOLEAN DEFAULT TRUE,
        is_external             BOOLEAN DEFAULT FALSE,
        shared                  BOOLEAN DEFAULT FALSE,
        allocation_pool_start   TEXT,
        allocation_pool_end     TEXT,
        status                  TEXT NOT NULL DEFAULT 'pending',
        error_msg               TEXT,
        pcd_network_id          TEXT,
        pcd_subnet_id           TEXT,
        created_at              TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        updated_at              TIMESTAMPTZ NOT NULL DEFAULT NOW()
    );
    CREATE INDEX IF NOT EXISTS idx_onboarding_networks_batch
        ON onboarding_networks (batch_id);
    CREATE TABLE IF NOT EXISTS onboarding_users (
        id                  SERIAL PRIMARY KEY,
        batch_id            UUID NOT NULL REFERENCES onboarding_batches(batch_id) ON DELETE CASCADE,
        project_id          INT REFERENCES onboarding_projects(id) ON DELETE CASCADE,
        domain_name         TEXT NOT NULL,
        project_name        TEXT NOT NULL,
        username            TEXT NOT NULL,
        email               TEXT,
        role                TEXT NOT NULL DEFAULT 'member',
        user_password       TEXT,
        send_welcome_email  BOOLEAN DEFAULT TRUE,
        status              TEXT NOT NULL DEFAULT 'pending',
        error_msg           TEXT,
        pcd_user_id         TEXT,
        temp_password       TEXT,
        created_at          TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        updated_at          TIMESTAMPTZ NOT NULL DEFAULT NOW()
    );
    CREATE INDEX IF NOT EXISTS idx_onboarding_users_batch
        ON onboarding_users (batch_id);
    """
    # Columns added after initial release — safe on existing installs via IF NOT EXISTS
    _ALTER_SQL = """
    ALTER TABLE onboarding_projects
        ADD COLUMN IF NOT EXISTS subscription_id       TEXT DEFAULT '',
        ADD COLUMN IF NOT EXISTS quota_instances       INT DEFAULT 10,
        ADD COLUMN IF NOT EXISTS quota_server_groups   INT DEFAULT 10,
        ADD COLUMN IF NOT EXISTS quota_networks        INT DEFAULT 10,
        ADD COLUMN IF NOT EXISTS quota_subnets         INT DEFAULT 20,
        ADD COLUMN IF NOT EXISTS quota_routers         INT DEFAULT 5,
        ADD COLUMN IF NOT EXISTS quota_ports           INT DEFAULT 200,
        ADD COLUMN IF NOT EXISTS quota_floatingips     INT DEFAULT 20,
        ADD COLUMN IF NOT EXISTS quota_security_groups INT DEFAULT 10,
        ADD COLUMN IF NOT EXISTS quota_volumes         INT DEFAULT 20,
        ADD COLUMN IF NOT EXISTS quota_snapshots       INT DEFAULT 20;
    ALTER TABLE onboarding_networks
        ADD COLUMN IF NOT EXISTS network_kind          TEXT DEFAULT 'physical_managed',
        ADD COLUMN IF NOT EXISTS network_type          TEXT DEFAULT 'vlan',
        ADD COLUMN IF NOT EXISTS physical_network      TEXT DEFAULT 'physnet1',
        ADD COLUMN IF NOT EXISTS is_external           BOOLEAN DEFAULT FALSE,
        ADD COLUMN IF NOT EXISTS shared                BOOLEAN DEFAULT FALSE,
        ADD COLUMN IF NOT EXISTS allocation_pool_start TEXT,
        ADD COLUMN IF NOT EXISTS allocation_pool_end   TEXT;
    ALTER TABLE onboarding_networks
        ALTER COLUMN cidr DROP NOT NULL;
    ALTER TABLE onboarding_users
        ADD COLUMN IF NOT EXISTS user_password TEXT;
    ALTER TABLE onboarding_customers
        ADD COLUMN IF NOT EXISTS department_tag TEXT;
    """
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT EXISTS (SELECT FROM information_schema.tables "
                    "WHERE table_name = 'onboarding_batches')"
                )
                if not cur.fetchone()[0]:
                    logger.info("Running onboarding DB migration …")
                    cur.execute(_MIGRATION_SQL)
                else:
                    # Tables already exist — apply column additions for existing installs
                    logger.info("Applying onboarding schema ALTER statements …")
                    cur.execute(_ALTER_SQL)
    except Exception as exc:
        logger.warning("Could not ensure onboarding tables: %s", exc)


_ensure_tables()

# ---------------------------------------------------------------------------
# Excel template generation
# ---------------------------------------------------------------------------

def _build_template_bytes() -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()

    # ── Style helpers ─────────────────────────────────────────────────────
    hdr_font     = Font(bold=True, color="FFFFFF")
    hdr_fill     = PatternFill("solid", fgColor="2F5496")
    sample_fill  = PatternFill("solid", fgColor="E8F4E8")   # light green for sample rows
    note_font    = Font(italic=True, color="595959")
    title_font   = Font(bold=True, size=14, color="1F3864")
    req_fill     = PatternFill("solid", fgColor="FFF2CC")   # amber for required columns
    req_font     = Font(bold=True, color="7F3F00")
    warn_font    = Font(bold=True, color="CC0000")

    def _set_col_width(ws, col_idx: int, width: int) -> None:
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    def add_sheet(
        name: str,
        headers: List[str],
        required_cols: List[str],
        col_notes: List[str],
        sample_rows: List[List],
    ) -> None:
        try:
            from openpyxl.comments import Comment
        except ImportError:
            Comment = None  # type: ignore

        if name == wb.active.title:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(name)

        # Header row (row 1) — data rows follow immediately at row 2
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill if h not in required_cols else PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center")
            _set_col_width(ws, col_idx, max(len(h) + 6, 22))
            # Add hint as a cell comment
            if Comment is not None and col_idx <= len(col_notes):
                reqstr = " [REQUIRED]" if h in required_cols else " [optional]"
                note = col_notes[col_idx - 1] + reqstr
                cell.comment = Comment(note, "pf9-onboarding")

        ws.freeze_panes = "A2"

        # Sample data rows start at row 2 (highlighted green — user deletes before upload)
        for row_idx, row_data in enumerate(sample_rows, 2):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = sample_fill

    # ── Instructions sheet ────────────────────────────────────────────────
    ws_info = wb.active
    ws_info.title = "README"
    ws_info.column_dimensions["A"].width = 18
    ws_info.column_dimensions["B"].width = 80

    instructions = [
        ("", ""),
        ("  BULK CUSTOMER ONBOARDING TEMPLATE", ""),
        ("", ""),
        ("How to use:", "1. Read this sheet carefully."),
        ("", "2. Fill in the 4 sheets: customers  ·  projects  ·  networks  ·  users"),
        ("", "3. Delete the green sample rows in each sheet before uploading."),
        ("", "4. Save the file and upload it via the portal."),
        ("", ""),
        ("SHEET: customers", "One row per customer (OpenStack domain). Required: domain_name."),
        ("SHEET: projects", (
            "One row per project. Required: domain_name, project_name. "
            "Use subscription_id to auto-derive the project name: "
            "{domain_name}_tenant_subid_{subscription_id}"
        )),
        ("SHEET: networks", (
            "One row per network. Required: domain_name, project_name, network_name.\n"
            "network_kind controls which other fields are used:\n"
            "  physical_managed → provider net + subnet: needs physical_network, vlan_id, cidr, gateway, dhcp/pool fields\n"
            "  physical_l2      → provider net, NO subnet: needs physical_network, vlan_id — leave cidr/gateway/dhcp/pool BLANK\n"
            "  virtual          → tenant overlay net + subnet: needs cidr, gateway, dhcp/pool — leave vlan_id/physical_network BLANK\n"
            "network_type: vlan (default) or flat.\n"
            "is_external: LEAVE AS false (default). Setting true makes the network visible across ALL tenants as an admin/external net."
        )),
        ("SHEET: users", (
            "One row per user. Required: domain_name, project_name, username, role. "
            "user_password: leave blank to auto-generate a secure password at execution time. "
            "Auto-generated passwords are stored in the DB and visible in the execution results."
        )),
        ("", ""),
        ("IMPORTANT:", "domain_name / project_name values must match exactly across sheets."),
        ("", "Valid roles: member, admin, reader"),
        ("", ""),
        ("NETWORK KIND", "FIELD REQUIREMENTS PER KIND:"),
        ("physical_managed", "physical_network ✓  vlan_id ✓  cidr ✓  gateway ✓  dhcp ✓  pool ✓   (provider net + subnet)"),
        ("physical_l2",      "physical_network ✓  vlan_id ✓  cidr ✗  gateway ✗  dhcp ✗  pool ✗   (provider net only — no subnet created)"),
        ("virtual",          "physical_network ✗  vlan_id ✗  cidr ✓  gateway ✓  dhcp ✓  pool ✓   (tenant overlay net + subnet)"),
        ("", ""),
        ("", "Valid network_type: vlan, flat  (not applicable for virtual kind)"),
        ("", "CIDR format: 10.x.x.x/24"),
    ]
    for r, (label, value) in enumerate(instructions, 1):
        c1 = ws_info.cell(row=r, column=1, value=label)
        c2 = ws_info.cell(row=r, column=2, value=value)
        if label.startswith("  BULK"):
            c1.font = title_font
        elif label in ("IMPORTANT:", "SHEET: customers", "SHEET: projects",
                       "SHEET: networks", "SHEET: users", "How to use:"):
            c1.font = Font(bold=True, color="1F3864")
        elif label == "":
            pass
        if label == "IMPORTANT:":
            c2.font = warn_font
        c2.alignment = Alignment(wrap_text=True)
        ws_info.row_dimensions[r].height = 20

    # ── customers sheet ───────────────────────────────────────────────────
    add_sheet(
        "customers",
        ["domain_name", "display_name", "description", "contact_email"],
        ["domain_name"],
        [
            "Unique OS domain key (lowercase, hyphens ok)",
            "Human-readable name shown in portal",
            "Optional description",
            "Primary contact email for this customer",
        ],
        [
            ["customer-a", "Customer A Org", "Cloud environment for Customer A", "admin@customer-a.example"],
            ["customer-b", "Customer B Org", "Cloud environment for Customer B", "admin@customer-b.example"],
        ],
    )

    # ── projects sheet ────────────────────────────────────────────────────
    # Naming convention: {domain_name}_tenant_subid_{subscription_id}
    add_sheet(
        "projects",
        [
            "domain_name", "project_name", "subscription_id", "description",
            "quota_vcpu", "quota_ram_gb", "quota_instances", "quota_server_groups",
            "quota_networks", "quota_subnets", "quota_routers", "quota_ports",
            "quota_floatingips", "quota_security_groups",
            "quota_volumes", "quota_snapshots", "quota_disk_gb",
        ],
        ["domain_name", "project_name"],
        [
            "Must match domain_name in customers sheet",
            "Project name in PCD. Tip: use naming convention below",
            "Subscription/tenant ID — used to auto-derive project name: {domain}_tenant_subid_{id}",
            "Optional description",
            "vCPU quota (default 20)",
            "RAM quota in GiB (default 50)",
            "Max VM instances (default 10)",
            "Max server groups (default 10)",
            "Neutron: max networks (default 10)",
            "Neutron: max subnets (default 20)",
            "Neutron: max routers (default 5)",
            "Neutron: max ports (default 200)",
            "Neutron: max floating IPs (default 20)",
            "Neutron: max security groups (default 10)",
            "Cinder: max volumes (default 20)",
            "Cinder: max snapshots (default 20)",
            "Cinder: total GB (default 1000)",
        ],
        [
            ["customer-a", "customer-a_tenant_subid_S001", "S001", "Production", 40, 100, 20, 10, 10, 20, 5, 200, 20, 10, 20, 20, 2000],
            ["customer-a", "customer-a_tenant_subid_S002", "S002", "Dev/Test",   20,  50, 10, 10, 10, 20, 5, 100, 10, 10, 10, 10,  500],
            ["customer-b", "customer-b_tenant_subid_T001", "T001", "Production", 40, 100, 20, 10, 10, 20, 5, 200, 20, 10, 20, 20, 2000],
        ],
    )

    # ── networks sheet ────────────────────────────────────────────────────
    # Field matrix by network_kind:
    #   physical_managed : physical_network ✓  vlan_id ✓  cidr ✓  gateway ✓  dhcp ✓  pool ✓
    #   physical_l2      : physical_network ✓  vlan_id ✓  cidr ✗  gateway ✗  dhcp ✗  pool ✗
    #   virtual          : physical_network ✗  vlan_id ✗  cidr ✓  gateway ✓  dhcp ✓  pool ✓
    add_sheet(
        "networks",
        [
            "domain_name", "project_name", "network_name",
            "network_kind", "network_type", "physical_network", "vlan_id",
            "cidr", "gateway", "dns1", "dhcp_enabled",
            "is_external", "shared",
            "allocation_pool_start", "allocation_pool_end",
        ],
        ["domain_name", "project_name", "network_name"],  # cidr required only for non-l2; validated in parser
        [
            "Must match domain_name in customers sheet",
            "Must match project_name in projects sheet",
            "Network name. Naming convention:\n"
            "  physical_managed → {domain}_tenant_extnet_vlan_{vlan}\n"
            "  physical_l2      → {domain}_tenant_L2net_vlan_{vlan}\n"
            "  virtual          → {domain}_tenant_vxlan (no VLAN in name)",
            "REQUIRED: physical_managed | physical_l2 | virtual\n"
            "  physical_managed = provider net + subnet (routable, external)\n"
            "  physical_l2      = provider net only, NO subnet (pure L2 bridging)\n"
            "  virtual          = tenant overlay net + subnet (no physical provider)",
            "vlan or flat. Default: vlan\n"
            "  N/A for virtual networks (leave blank)",
            "Provider network name in PCD (e.g. physnet1)\n"
            "  Required for: physical_managed, physical_l2\n"
            "  N/A for: virtual (leave blank)",
            "VLAN ID integer (e.g. 100)\n"
            "  Required for: physical_managed, physical_l2\n"
            "  N/A for: virtual (leave blank)",
            "Subnet CIDR, e.g. 10.10.10.0/24\n"
            "  Required for: physical_managed, virtual\n"
            "  N/A for: physical_l2 (leave blank — no subnet is created)",
            "Gateway IP. Auto-derived from CIDR if blank\n"
            "  N/A for: physical_l2 (leave blank)",
            "DNS server. Default: 8.8.8.8\n"
            "  N/A for: physical_l2 (leave blank)",
            "Enable DHCP: true or false. Default: true\n"
            "  N/A for: physical_l2 (leave blank)",
            "true = makes network external/shared across all tenants (router:external=True).\n"
            "  Default: false — customer networks should almost always be false.\n"
            "  Only set true if this is a shared provider/external network managed by the admin tenant.",
            "true = shared across all projects. Default: false",
            "DHCP pool start IP (optional — auto-derived from CIDR if blank)\n"
            "  N/A for: physical_l2 (leave blank)",
            "DHCP pool end IP (optional — auto-derived from CIDR if blank)\n"
            "  N/A for: physical_l2 (leave blank)",
        ],
        [
            # physical_managed — full provider net + subnet
            ["customer-a", "customer-a_tenant_subid_S001", "customer-a_tenant_extnet_vlan_100",
             "physical_managed", "vlan", "physnet1", 100,
             "10.100.0.0/24", "10.100.0.1", "8.8.8.8", "true", "false", "false", "10.100.0.10", "10.100.0.254"],
            # physical_l2 — no subnet; leave cidr/gateway/dhcp/pool blank
            ["customer-a", "customer-a_tenant_subid_S001", "customer-a_tenant_L2net_vlan_200",
             "physical_l2", "vlan", "physnet1", 200,
             "", "", "", "", "false", "false", "", ""],
            # virtual — tenant overlay; no vlan / physical_network
            ["customer-b", "customer-b_tenant_subid_T001", "customer-b_tenant_vxlan",
             "virtual", "", "", "",
             "192.168.100.0/24", "192.168.100.1", "8.8.8.8", "true", "false", "false",
             "192.168.100.10", "192.168.100.254"],
        ],
    )

    # ── users sheet ───────────────────────────────────────────────────────
    add_sheet(
        "users",
        ["domain_name", "project_name", "username", "email", "role", "user_password", "send_welcome_email"],
        ["domain_name", "project_name", "username", "role"],
        [
            "Must match domain_name in customers sheet",
            "Must match project_name in projects sheet",
            "Username in PCD (lowercase, no spaces)",
            "User email address",
            "member | admin | reader",
            "Leave blank to auto-generate a secure password at execution time. Auto-generated passwords are stored in execution results.",
            "true to send welcome email (default: true)",
        ],
        [
            ["customer-a", "customer-a_tenant_subid_S001", "alice", "alice@example.com", "admin", "", "true"],
            ["customer-a", "customer-a_tenant_subid_S001", "bob",   "bob@example.com",   "member", "", "true"],
            ["customer-b", "customer-b_tenant_subid_T001", "carol", "carol@example.com", "admin", "", "true"],
        ],
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Excel parsing + validation
# ---------------------------------------------------------------------------

VALID_ROLES = {"member", "admin", "reader", "_member_"}
CIDR_RE = re.compile(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}/\d{1,2}$")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _bool_val(val: Any) -> bool:
    if isinstance(val, bool):
        return val
    return _str(val).lower() not in ("false", "0", "no", "")


def _int_val(val: Any, default: int) -> int:
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def _parse_excel(raw: bytes) -> Dict[str, Any]:
    """Parse uploaded Excel bytes; return {customers, projects, networks, users, errors}."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    errors: List[Dict] = []
    customers: List[Dict] = []
    projects: List[Dict] = []
    networks: List[Dict] = []
    users: List[Dict] = []

    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(raw), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot parse Excel file: {exc}")

    required_sheets = {"customers", "projects", "networks", "users"}
    missing = required_sheets - {s.lower() for s in wb.sheetnames}
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing sheets: {', '.join(sorted(missing))}")

    def rows_as_dicts(sheet_name: str) -> List[Dict]:
        # Case-insensitive sheet lookup
        ws = next((wb[s] for s in wb.sheetnames if s.lower() == sheet_name.lower()), None)
        if ws is None:
            return []
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [_str(h).lower().replace(" ", "_") for h in rows[0]]
        result = []
        for row in rows[1:]:
            if all(v is None or _str(v) == "" for v in row):
                continue
            result.append(dict(zip(headers, row)))
        return result

    # ── customers ──
    known_domains: set = set()
    for i, row in enumerate(rows_as_dicts("customers"), 2):
        dn = _str(row.get("domain_name", ""))
        if not dn:
            errors.append({"sheet": "customers", "row": i, "field": "domain_name", "message": "Required"})
            continue
        if dn in known_domains:
            errors.append({"sheet": "customers", "row": i, "field": "domain_name", "message": f"Duplicate domain_name '{dn}'"})
            continue
        known_domains.add(dn)
        email = _str(row.get("contact_email", ""))
        if email and not EMAIL_RE.match(email):
            errors.append({"sheet": "customers", "row": i, "field": "contact_email", "message": f"Invalid email '{email}'"})
        customers.append({
            "domain_name": dn,
            "display_name": _str(row.get("display_name", "")),
            "description": _str(row.get("description", "")),
            "contact_email": email,
        })

    # ── projects ──
    known_projects: set = set()  # (domain_name, project_name)
    for i, row in enumerate(rows_as_dicts("projects"), 2):
        dn = _str(row.get("domain_name", ""))
        pn = _str(row.get("project_name", ""))
        if not dn:
            errors.append({"sheet": "projects", "row": i, "field": "domain_name", "message": "Required"})
            continue
        if not pn:
            errors.append({"sheet": "projects", "row": i, "field": "project_name", "message": "Required"})
            continue
        if dn not in known_domains:
            errors.append({"sheet": "projects", "row": i, "field": "domain_name",
                           "message": f"domain_name '{dn}' not found in customers sheet"})
        key = (dn, pn)
        if key in known_projects:
            errors.append({"sheet": "projects", "row": i, "field": "project_name",
                           "message": f"Duplicate project '{pn}' in domain '{dn}'"})
            continue
        known_projects.add(key)
        projects.append({
            "domain_name": dn,
            "project_name": pn,
            "subscription_id": _str(row.get("subscription_id", "")),
            "description": _str(row.get("description", "")),
            "quota_vcpu": _int_val(row.get("quota_vcpu"), 20),
            "quota_ram_mb": _int_val(row.get("quota_ram_gb"), 50) * 1024,
            "quota_instances": _int_val(row.get("quota_instances"), 10),
            "quota_server_groups": _int_val(row.get("quota_server_groups"), 10),
            "quota_networks": _int_val(row.get("quota_networks"), 10),
            "quota_subnets": _int_val(row.get("quota_subnets"), 20),
            "quota_routers": _int_val(row.get("quota_routers"), 5),
            "quota_ports": _int_val(row.get("quota_ports"), 200),
            "quota_floatingips": _int_val(row.get("quota_floatingips"), 20),
            "quota_security_groups": _int_val(row.get("quota_security_groups"), 10),
            "quota_volumes": _int_val(row.get("quota_volumes"), 20),
            "quota_snapshots": _int_val(row.get("quota_snapshots"), 20),
            "quota_disk_gb": _int_val(row.get("quota_disk_gb"), 1000),
        })

    # ── networks ──
    VALID_NETWORK_KINDS = {"physical_managed", "physical_l2", "virtual"}
    VALID_NETWORK_TYPES = {"vlan", "flat"}
    for i, row in enumerate(rows_as_dicts("networks"), 2):
        dn = _str(row.get("domain_name", ""))
        pn = _str(row.get("project_name", ""))
        nn = _str(row.get("network_name", ""))
        network_kind = _str(row.get("network_kind", "physical_managed")) or "physical_managed"
        cidr = _str(row.get("cidr", ""))
        if not dn:
            errors.append({"sheet": "networks", "row": i, "field": "domain_name", "message": "Required"}); continue
        if not pn:
            errors.append({"sheet": "networks", "row": i, "field": "project_name", "message": "Required"}); continue
        if not nn:
            errors.append({"sheet": "networks", "row": i, "field": "network_name", "message": "Required"}); continue
        if network_kind not in VALID_NETWORK_KINDS:
            errors.append({"sheet": "networks", "row": i, "field": "network_kind",
                           "message": f"network_kind '{network_kind}' not valid; use physical_managed, physical_l2, or virtual"}); continue
        # cidr only required for physical_managed and virtual (not physical_l2)
        if network_kind != "physical_l2":
            if not cidr:
                errors.append({"sheet": "networks", "row": i, "field": "cidr", "message": "Required"}); continue
            if not CIDR_RE.match(cidr):
                errors.append({"sheet": "networks", "row": i, "field": "cidr", "message": f"Invalid CIDR '{cidr}'"}); continue
            try:
                ipaddress.ip_network(cidr, strict=False)
            except ValueError:
                errors.append({"sheet": "networks", "row": i, "field": "cidr", "message": f"Invalid network address '{cidr}'"}); continue
        network_type = _str(row.get("network_type", "vlan")) or "vlan"
        if network_type not in VALID_NETWORK_TYPES:
            errors.append({"sheet": "networks", "row": i, "field": "network_type",
                           "message": f"network_type '{network_type}' not valid; use vlan or flat"})
        if (dn, pn) not in known_projects and dn in known_domains:
            errors.append({"sheet": "networks", "row": i, "field": "project_name",
                           "message": f"Project '{pn}' in domain '{dn}' not found in projects sheet"})
        # is_external defaults: always False — customer tenant networks are never external
        default_is_external = False
        networks.append({
            "domain_name": dn,
            "project_name": pn,
            "network_name": nn,
            "network_kind": network_kind,
            "network_type": network_type,
            "physical_network": _str(row.get("physical_network", "physnet1")) or "physnet1",
            "cidr": cidr or None,
            "gateway": _str(row.get("gateway", "")),
            "dns1": _str(row.get("dns1", "")) or "8.8.8.8",
            "vlan_id": _int_val(row.get("vlan_id"), 0) or None,
            "dhcp_enabled": _bool_val(row.get("dhcp_enabled", True)),
            "is_external": _bool_val(row.get("is_external", default_is_external)),
            "shared": _bool_val(row.get("shared", False)),
            "allocation_pool_start": _str(row.get("allocation_pool_start", "")),
            "allocation_pool_end": _str(row.get("allocation_pool_end", "")),
        })

    # ── users ──
    for i, row in enumerate(rows_as_dicts("users"), 2):
        dn = _str(row.get("domain_name", ""))
        pn = _str(row.get("project_name", ""))
        un = _str(row.get("username", ""))
        role = _str(row.get("role", "member")).lower()
        if not dn:
            errors.append({"sheet": "users", "row": i, "field": "domain_name", "message": "Required"}); continue
        if not pn:
            errors.append({"sheet": "users", "row": i, "field": "project_name", "message": "Required"}); continue
        if not un:
            errors.append({"sheet": "users", "row": i, "field": "username", "message": "Required"}); continue
        if role not in VALID_ROLES:
            errors.append({"sheet": "users", "row": i, "field": "role",
                           "message": f"Role '{role}' not valid; use member/admin/reader"})
        email = _str(row.get("email", ""))
        if email and not EMAIL_RE.match(email):
            errors.append({"sheet": "users", "row": i, "field": "email", "message": f"Invalid email '{email}'"})
        users.append({
            "domain_name": dn,
            "project_name": pn,
            "username": un,
            "email": email,
            "role": role,
            "user_password": _str(row.get("user_password", "")),  # empty = auto-generate at execute
            "send_welcome_email": _bool_val(row.get("send_welcome_email", True)),
        })

    return {
        "customers": customers,
        "projects": projects,
        "networks": networks,
        "users": users,
        "errors": errors,
    }


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def _get_batch(cur, batch_id: str) -> Dict:
    cur.execute(
        "SELECT * FROM onboarding_batches WHERE batch_id = %s",
        (batch_id,),
    )
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail=f"Batch '{batch_id}' not found")
    return dict(row)


def _batch_items(cur, batch_id: str) -> Dict:
    cur.execute("SELECT * FROM onboarding_customers WHERE batch_id = %s ORDER BY id", (batch_id,))
    customers = [dict(r) for r in cur.fetchall()]
    cur.execute("SELECT * FROM onboarding_projects  WHERE batch_id = %s ORDER BY id", (batch_id,))
    projects = [dict(r) for r in cur.fetchall()]
    cur.execute("SELECT * FROM onboarding_networks  WHERE batch_id = %s ORDER BY id", (batch_id,))
    networks = [dict(r) for r in cur.fetchall()]
    cur.execute("SELECT * FROM onboarding_users     WHERE batch_id = %s ORDER BY id", (batch_id,))
    users = [dict(r) for r in cur.fetchall()]
    return {"customers": customers, "projects": projects, "networks": networks, "users": users}


# ---------------------------------------------------------------------------
# GET /template
# ---------------------------------------------------------------------------

@router.get("/template")
def download_template(current_user: dict = Depends(get_current_user)):
    """Download the blank Excel onboarding template."""
    data = _build_template_bytes()
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="onboarding_template.xlsx"'},
    )


# ---------------------------------------------------------------------------
# POST /upload
# ---------------------------------------------------------------------------

@router.post("/upload", status_code=201)
def upload_batch(
    file: UploadFile = File(...),
    batch_name: str = "",
    current_user: dict = Depends(require_permission("onboarding", "create")),
):
    """Parse and validate uploaded Excel; create a new batch record."""
    raw = file.file.read()
    parsed = _parse_excel(raw)

    errors = parsed["errors"]
    status = "invalid" if errors else "validated"
    bn = batch_name or (file.filename or "batch").rsplit(".", 1)[0]
    actor = current_user.get("username", "unknown")

    with get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                INSERT INTO onboarding_batches
                    (batch_name, uploaded_by, status, validation_errors,
                     total_customers, total_projects, total_networks, total_users)
                VALUES (%s, %s, %s, %s::jsonb, %s, %s, %s, %s)
                RETURNING *
                """,
                (
                    bn, actor, status,
                    psycopg2.extras.Json(errors),
                    len(parsed["customers"]),
                    len(parsed["projects"]),
                    len(parsed["networks"]),
                    len(parsed["users"]),
                ),
            )
            batch = dict(cur.fetchone())
            batch_id = batch["batch_id"]

            # Insert customers
            for c in parsed["customers"]:
                cur.execute(
                    """
                    INSERT INTO onboarding_customers
                        (batch_id, domain_name, display_name, description, contact_email)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (batch_id, c["domain_name"], c["display_name"], c["description"],
                     c["contact_email"]),
                )

            # Insert projects
            for p in parsed["projects"]:
                cur.execute(
                    """
                    INSERT INTO onboarding_projects
                        (batch_id, domain_name, project_name, subscription_id, description,
                         quota_vcpu, quota_ram_mb, quota_instances, quota_server_groups,
                         quota_networks, quota_subnets, quota_routers, quota_ports,
                         quota_floatingips, quota_security_groups,
                         quota_volumes, quota_snapshots, quota_disk_gb)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (batch_id, p["domain_name"], p["project_name"], p["subscription_id"],
                     p["description"], p["quota_vcpu"], p["quota_ram_mb"],
                     p["quota_instances"], p["quota_server_groups"],
                     p["quota_networks"], p["quota_subnets"], p["quota_routers"],
                     p["quota_ports"], p["quota_floatingips"], p["quota_security_groups"],
                     p["quota_volumes"], p["quota_snapshots"], p["quota_disk_gb"]),
                )

            # Insert networks
            for n in parsed["networks"]:
                cur.execute(
                    """
                    INSERT INTO onboarding_networks
                        (batch_id, domain_name, project_name, network_name,
                         network_kind, network_type, physical_network,
                         cidr, gateway, dns1, vlan_id, dhcp_enabled,
                         is_external, shared,
                         allocation_pool_start, allocation_pool_end)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (batch_id, n["domain_name"], n["project_name"], n["network_name"],
                     n["network_kind"], n["network_type"], n["physical_network"],
                     n["cidr"], n["gateway"] or None, n["dns1"],
                     n["vlan_id"], n["dhcp_enabled"],
                     n["is_external"], n["shared"],
                     n["allocation_pool_start"] or None, n["allocation_pool_end"] or None),
                )

            # Insert users
            for u in parsed["users"]:
                cur.execute(
                    """
                    INSERT INTO onboarding_users
                        (batch_id, domain_name, project_name, username, email,
                         role, user_password, send_welcome_email)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (batch_id, u["domain_name"], u["project_name"], u["username"],
                     u["email"] or None, u["role"],
                     u["user_password"] or None, u["send_welcome_email"]),
                )

    logger.info("Onboarding batch %s created by %s — status=%s", batch_id, actor, status)
    return {
        "batch_id": str(batch_id),
        "batch_name": bn,
        "status": status,
        "total_customers": batch["total_customers"],
        "total_projects": batch["total_projects"],
        "total_networks": batch["total_networks"],
        "total_users": batch["total_users"],
        "validation_errors": errors,
    }


# ---------------------------------------------------------------------------
# GET /batches
# ---------------------------------------------------------------------------

@router.get("/batches")
def list_batches(current_user: dict = Depends(require_permission("onboarding", "read"))):
    with get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT * FROM onboarding_batches ORDER BY created_at DESC"
            )
            return [dict(r) for r in cur.fetchall()]


# ---------------------------------------------------------------------------
# GET /batches/{batch_id}
# ---------------------------------------------------------------------------

@router.get("/batches/{batch_id}")
def get_batch(batch_id: str, current_user: dict = Depends(require_permission("onboarding", "read"))):
    with get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            batch = _get_batch(cur, batch_id)
            items = _batch_items(cur, batch_id)
    return {**batch, **items}


# ---------------------------------------------------------------------------
# POST /batches/{batch_id}/dry-run
# ---------------------------------------------------------------------------

@router.post("/batches/{batch_id}/dry-run")
def run_dry_run(batch_id: str, current_user: dict = Depends(require_permission("onboarding", "create"))):
    """
    Connect to PCD and check which resources already exist.
    Sets batch status to dry_run_passed / dry_run_failed.
    Execution is hard-locked until dry_run_passed.
    """
    with get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            batch = _get_batch(cur, batch_id)
            if batch["status"] not in ("validated", "dry_run_failed", "dry_run_passed"):
                raise HTTPException(
                    status_code=400,
                    detail=f"Cannot dry-run a batch with status '{batch['status']}'. "
                           "Batch must be validated first.",
                )
            items = _batch_items(cur, batch_id)

    # Update status to dry_run_pending
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE onboarding_batches SET status = 'dry_run_pending' WHERE batch_id = %s",
                (batch_id,),
            )

    try:
        from pf9_control import get_client  # type: ignore
        client = get_client()
        client.authenticate()

        existing_domains = {d["name"]: d for d in (client.list_domains() or [])}
        existing_projects_by_domain: Dict[str, Dict[str, Any]] = {}
        existing_networks_by_project: Dict[str, List[str]] = {}

        # Prefetch projects for existing domains
        for cust in items["customers"]:
            dn = cust["domain_name"]
            if dn in existing_domains:
                dom_id = existing_domains[dn]["id"]
                projs = client.list_projects(domain_id=dom_id) or []
                existing_projects_by_domain[dn] = {p["name"]: p for p in projs}

        dry_items: List[Dict] = []
        conflicts: List[Dict] = []

        # Check customers
        for cust in items["customers"]:
            dn = cust["domain_name"]
            if dn in existing_domains:
                dry_items.append({"resource_type": "domain", "name": dn,
                                   "action": "skip", "reason": "Domain already exists"})
                conflicts.append({"resource_type": "domain", "name": dn,
                                   "reason": "Domain already exists — rename or remove from batch"})
            else:
                dry_items.append({"resource_type": "domain", "name": dn, "action": "create", "reason": ""})

        # Check projects
        for proj in items["projects"]:
            dn = proj["domain_name"]
            pn = proj["project_name"]
            dom_projs = existing_projects_by_domain.get(dn, {})
            if pn in dom_projs:
                dry_items.append({"resource_type": "project", "name": f"{dn}/{pn}",
                                   "action": "skip", "reason": "Project already exists"})
                conflicts.append({"resource_type": "project", "name": f"{dn}/{pn}",
                                   "reason": "Project already exists in PCD"})
            else:
                dry_items.append({"resource_type": "project", "name": f"{dn}/{pn}",
                                   "action": "create", "reason": ""})

        # Check networks (lightweight — just flag if project already has a network with the same name)
        try:
            all_nets = client.list_networks() or []
            net_names_all = {n.get("name", "") for n in all_nets}
        except Exception:
            net_names_all = set()

        for net in items["networks"]:
            nn = net["network_name"]
            key = f"{net['domain_name']}/{net['project_name']}/{nn}"
            if nn in net_names_all:
                dry_items.append({"resource_type": "network", "name": key,
                                   "action": "skip", "reason": "Network name already exists in PCD"})
                conflicts.append({"resource_type": "network", "name": key,
                                   "reason": "Network already exists — rename or remove from batch"})
            else:
                dry_items.append({"resource_type": "network", "name": key, "action": "create", "reason": ""})

        # Users — just log planned creation (no global user list query)
        for user in items["users"]:
            key = f"{user['domain_name']}/{user['username']}"
            dry_items.append({"resource_type": "user", "name": key, "action": "create", "reason": ""})

    except HTTPException:
        raise
    except Exception as exc:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE onboarding_batches SET status = 'dry_run_failed', "
                    "dry_run_result = %s::jsonb WHERE batch_id = %s",
                    (psycopg2.extras.Json({"error": str(exc), "items": []}), batch_id),
                )
        raise HTTPException(status_code=502, detail=f"PCD connection failed: {exc}")

    # Determine pass/fail
    has_conflicts = len(conflicts) > 0
    new_status = "dry_run_failed" if has_conflicts else "dry_run_passed"

    summary = {
        "total_would_create": sum(1 for d in dry_items if d["action"] == "create"),
        "total_would_skip": sum(1 for d in dry_items if d["action"] == "skip"),
        "conflicts": conflicts,
    }
    result = {"summary": summary, "items": dry_items}

    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE onboarding_batches SET status = %s, dry_run_result = %s::jsonb WHERE batch_id = %s",
                (new_status, psycopg2.extras.Json(result), batch_id),
            )

    logger.info("Dry-run for batch %s: status=%s conflicts=%d", batch_id, new_status, len(conflicts))
    return {"status": new_status, **result}


# ---------------------------------------------------------------------------
# POST /batches/{batch_id}/submit  (submit for approval)
# ---------------------------------------------------------------------------

@router.post("/batches/{batch_id}/submit")
def submit_for_approval(batch_id: str, current_user: dict = Depends(require_permission("onboarding", "create"))):
    with get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            batch = _get_batch(cur, batch_id)
            if batch["status"] not in ("validated", "dry_run_passed", "dry_run_failed"):
                raise HTTPException(
                    status_code=400,
                    detail=f"Cannot submit a batch with status '{batch['status']}'",
                )
            cur.execute(
                "UPDATE onboarding_batches SET approval_status = 'pending_approval' WHERE batch_id = %s",
                (batch_id,),
            )

    actor = current_user.get("username", "unknown")
    _fire(
        "onboarding_submitted",
        "Onboarding Batch Submitted",
        f"Batch '{batch['batch_name']}' submitted for approval by {actor}.",
        "info",
    )
    return {"status": "pending_approval", "batch_id": batch_id}


# ---------------------------------------------------------------------------
# POST /batches/{batch_id}/decision  (approve or reject)
# ---------------------------------------------------------------------------

class DecisionPayload(BaseModel):
    decision: str  # "approve" | "reject"
    comment: Optional[str] = None


@router.post("/batches/{batch_id}/decision")
def batch_decision(
    batch_id: str,
    payload: DecisionPayload,
    current_user: dict = Depends(require_permission("onboarding", "approve")),
):
    if payload.decision not in ("approve", "reject"):
        raise HTTPException(status_code=400, detail="decision must be 'approve' or 'reject'")

    actor = current_user.get("username", "unknown")

    with get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            batch = _get_batch(cur, batch_id)
            if batch["approval_status"] != "pending_approval":
                raise HTTPException(
                    status_code=400,
                    detail=f"Batch approval_status is '{batch['approval_status']}', not pending_approval",
                )
            new_approval = "approved" if payload.decision == "approve" else "rejected"
            cur.execute(
                """
                UPDATE onboarding_batches
                   SET approval_status = %s,
                       approved_by = %s,
                       approved_at = NOW(),
                       rejection_comment = %s
                 WHERE batch_id = %s
                """,
                (new_approval, actor, payload.comment, batch_id),
            )

    event = "onboarding_approved" if new_approval == "approved" else "onboarding_rejected"
    title = f"Onboarding Batch {'Approved' if new_approval == 'approved' else 'Rejected'}"
    msg = f"Batch '{batch['batch_name']}' was {new_approval} by {actor}."
    if payload.comment:
        msg += f" Comment: {payload.comment}"
    _fire(event, title, msg, "info" if new_approval == "approved" else "warning")

    return {"approval_status": new_approval, "batch_id": batch_id}


# ---------------------------------------------------------------------------
# POST /batches/{batch_id}/execute
# ---------------------------------------------------------------------------

def _execute_batch_sync(batch_id: str, actor: str) -> None:
    """Run in a background thread — creates all PCD resources."""
    logger.info("Starting execution of batch %s by %s", batch_id, actor)

    with get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM onboarding_batches WHERE batch_id = %s", (batch_id,))
            batch = dict(cur.fetchone())
            items = _batch_items(cur, batch_id)

    try:
        from pf9_control import get_client  # type: ignore
        client = get_client()
        client.authenticate()

        existing_domains = {d["name"]: d for d in (client.list_domains() or [])}
        # domain_id lookup cache — also pre-seeded from DB for rerun support
        domain_id_map: Dict[str, str] = {dn: d["id"] for dn, d in existing_domains.items()}
        project_id_map: Dict[tuple, str] = {}  # (domain_name, project_name) → project_id

        # Pre-seed maps from items already successfully created in a prior execution (rerun support)
        for _c in items["customers"]:
            if _c.get("status") == "created" and _c.get("pcd_domain_id"):
                domain_id_map[_c["domain_name"]] = _c["pcd_domain_id"]
        for _p in items["projects"]:
            if _p.get("status") == "created" and _p.get("pcd_project_id"):
                project_id_map[(_p["domain_name"], _p["project_name"])] = _p["pcd_project_id"]

        # Execution log list — flushed to DB after each section
        exec_log: list = []

        def _log(level: str, msg: str) -> None:
            entry = {"ts": _now().isoformat(), "level": level, "msg": msg}
            exec_log.append(entry)
            logger.info("[exec-log] %s", msg)

        def _flush_log() -> None:
            if not exec_log:
                return
            snapshot = list(exec_log)
            exec_log.clear()
            try:
                with get_connection() as _lc:
                    with _lc.cursor() as _lcur:
                        _lcur.execute(
                            "UPDATE onboarding_batches SET execution_log = execution_log || %s::jsonb WHERE batch_id = %s",
                            (psycopg2.extras.Json(snapshot), batch_id),
                        )
            except Exception as _le:
                logger.warning("Could not flush execution log: %s", _le)

        _log("info", f"Execution started by {actor}")

        # ── Create domains ──────────────────────────────────────────────────
        for cust in items["customers"]:
            dn = cust["domain_name"]
            if cust.get("status") in ("created", "skipped"):
                _log("info", f"Domain '{dn}' already handled — skipping")
                continue
            if dn in existing_domains:
                logger.warning("Domain '%s' already exists — skipping", dn)
                with get_connection() as conn2:
                    with conn2.cursor() as cur2:
                        cur2.execute(
                            "UPDATE onboarding_customers SET status = 'skipped', error_msg = %s, "
                            "pcd_domain_id = %s WHERE batch_id = %s AND domain_name = %s",
                            ("Domain already exists", existing_domains[dn]["id"], batch_id, dn),
                        )
                domain_id_map[dn] = existing_domains[dn]["id"]
                continue
            try:
                desc = cust.get("description") or ""
                result_d = client.create_domain(
                    dn,
                    description=desc,
                )
                dom_id = result_d.get("id") or result_d.get("domain", {}).get("id", "")
                domain_id_map[dn] = dom_id
                with get_connection() as conn2:
                    with conn2.cursor() as cur2:
                        cur2.execute(
                            "UPDATE onboarding_customers SET status = 'created', pcd_domain_id = %s "
                            "WHERE batch_id = %s AND domain_name = %s",
                            (dom_id, batch_id, dn),
                        )
                _log("info", f"Created domain '{dn}' (id={dom_id})")
            except Exception as exc:
                logger.error("Failed to create domain '%s': %s", dn, exc)
                _log("error", f"Failed to create domain '{dn}': {exc}")
                with get_connection() as conn2:
                    with conn2.cursor() as cur2:
                        cur2.execute(
                            "UPDATE onboarding_customers SET status = 'failed', error_msg = %s "
                            "WHERE batch_id = %s AND domain_name = %s",
                            (str(exc), batch_id, dn),
                        )
        _flush_log()

        # ── Create projects ─────────────────────────────────────────────────
        for proj in items["projects"]:
            dn = proj["domain_name"]
            pn = proj["project_name"]
            if proj.get("status") == "created":
                _log("info", f"Project '{dn}/{pn}' already created — skipping")
                continue
            dom_id = domain_id_map.get(dn)
            if not dom_id:
                with get_connection() as conn2:
                    with conn2.cursor() as cur2:
                        cur2.execute(
                            "UPDATE onboarding_projects SET status = 'failed', error_msg = %s "
                            "WHERE batch_id = %s AND domain_name = %s AND project_name = %s",
                            (f"Parent domain '{dn}' not created", batch_id, dn, pn),
                        )
                continue
            try:
                result_p = client.create_project(
                    pn,
                    domain_id=dom_id,
                    description=proj.get("description") or "",
                )
                proj_id = result_p.get("id") or result_p.get("project", {}).get("id", "")
                project_id_map[(dn, pn)] = proj_id
                # Set compute quotas
                try:
                    client.update_compute_quotas(proj_id, {
                        "cores": proj["quota_vcpu"],
                        "ram": proj["quota_ram_mb"],
                        "instances": proj["quota_instances"],
                        "server_groups": proj["quota_server_groups"],
                    })
                except Exception as qe:
                    logger.warning("Could not set compute quotas for project '%s': %s", pn, qe)
                # Set network quotas
                try:
                    client.update_network_quotas(proj_id, {
                        "network": proj["quota_networks"],
                        "subnet": proj["quota_subnets"],
                        "router": proj["quota_routers"],
                        "port": proj["quota_ports"],
                        "floatingip": proj["quota_floatingips"],
                        "security_group": proj["quota_security_groups"],
                    })
                except Exception as qe:
                    logger.warning("Could not set network quotas for project '%s': %s", pn, qe)
                # Set storage quotas
                try:
                    client.update_storage_quotas(proj_id, {
                        "gigabytes": proj["quota_disk_gb"],
                        "volumes": proj["quota_volumes"],
                        "snapshots": proj["quota_snapshots"],
                    })
                except Exception as qe:
                    logger.warning("Could not set storage quotas for project '%s': %s", pn, qe)
                with get_connection() as conn2:
                    with conn2.cursor() as cur2:
                        cur2.execute(
                            "UPDATE onboarding_projects SET status = 'created', pcd_project_id = %s "
                            "WHERE batch_id = %s AND domain_name = %s AND project_name = %s",
                            (proj_id, batch_id, dn, pn),
                        )
                logger.info("Created project '%s/%s' id=%s", dn, pn, proj_id)
                _log("info", f"Created project '{dn}/{pn}' (id={proj_id})")
            except Exception as exc:
                logger.error("Failed to create project '%s/%s': %s", dn, pn, exc)
                _log("error", f"Failed to create project '{dn}/{pn}': {exc}")
                with get_connection() as conn2:
                    with conn2.cursor() as cur2:
                        cur2.execute(
                            "UPDATE onboarding_projects SET status = 'failed', error_msg = %s "
                            "WHERE batch_id = %s AND domain_name = %s AND project_name = %s",
                            (str(exc), batch_id, dn, pn),
                        )
        _flush_log()

        # ── Create networks + subnets ────────────────────────────────────────
        for net in items["networks"]:
            dn = net["domain_name"]
            pn = net["project_name"]
            nn = net["network_name"]
            if net.get("status") == "created":
                _log("info", f"Network '{dn}/{pn}/{nn}' already created — skipping")
                continue
            proj_id = project_id_map.get((dn, pn))
            if not proj_id:
                with get_connection() as conn2:
                    with conn2.cursor() as cur2:
                        cur2.execute(
                            "UPDATE onboarding_networks SET status = 'failed', error_msg = %s "
                            "WHERE batch_id = %s AND domain_name = %s AND project_name = %s AND network_name = %s",
                            (f"Parent project '{dn}/{pn}' not created", batch_id, dn, pn, nn),
                        )
                continue
            try:
                network_kind = net.get("network_kind", "physical_managed")
                network_type = net.get("network_type", "vlan")
                physical_network = net.get("physical_network") or "physnet1"
                is_external = net.get("is_external", False)   # never external by default
                shared = net.get("shared", False)               # never shared by default

                if network_kind in ("physical_managed", "physical_l2"):
                    result_n = client.create_provider_network(
                        name=nn,
                        network_type=network_type,
                        physical_network=physical_network,
                        segmentation_id=int(net["vlan_id"]) if net.get("vlan_id") else None,
                        project_id=proj_id,
                        shared=shared,
                        external=is_external,
                    )
                else:  # virtual / tenant network
                    result_n = client.create_network(
                        name=nn,
                        project_id=proj_id,
                        shared=shared,
                        external=is_external,
                    )
                net_id = result_n.get("id") or result_n.get("network", {}).get("id", "")
                subnet_id = ""

                # physical_l2 = no subnet; physical_managed and virtual = create subnet
                if network_kind != "physical_l2" and net.get("cidr"):
                    cidr = net["cidr"]
                    net_obj = ipaddress.ip_network(cidr, strict=False)
                    gateway = net.get("gateway") or str(net_obj.network_address + 1)
                    # Use explicit pool if provided, else auto-derive
                    pool_start = net.get("allocation_pool_start") or str(net_obj.network_address + 2)
                    pool_end   = net.get("allocation_pool_end")   or str(net_obj.broadcast_address - 1)

                    subnet_kwargs: Dict = {
                        "network_id": net_id,
                        "cidr": cidr,
                        "name": f"{nn}-subnet",
                        "gateway_ip": gateway,
                        "dns_nameservers": [net.get("dns1") or "8.8.8.8"],
                        "enable_dhcp": net.get("dhcp_enabled", True),
                        "allocation_pools": [{"start": pool_start, "end": pool_end}],
                    }
                    result_sn = client.create_subnet(**subnet_kwargs)
                    subnet_id = result_sn.get("id") or result_sn.get("subnet", {}).get("id", "")

                with get_connection() as conn2:
                    with conn2.cursor() as cur2:
                        cur2.execute(
                            "UPDATE onboarding_networks SET status = 'created', "
                            "pcd_network_id = %s, pcd_subnet_id = %s "
                            "WHERE batch_id = %s AND domain_name = %s AND project_name = %s AND network_name = %s",
                            (net_id, subnet_id, batch_id, dn, pn, nn),
                        )
                logger.info("Created network '%s/%s/%s' net=%s subnet=%s", dn, pn, nn, net_id, subnet_id)
                _log("info", f"Created network '{dn}/{pn}/{nn}' net={net_id}")
            except Exception as exc:
                logger.error("Failed to create network '%s/%s/%s': %s", dn, pn, nn, exc)
                _log("error", f"Failed to create network '{dn}/{pn}/{nn}': {exc}")
                with get_connection() as conn2:
                    with conn2.cursor() as cur2:
                        cur2.execute(
                            "UPDATE onboarding_networks SET status = 'failed', error_msg = %s "
                            "WHERE batch_id = %s AND domain_name = %s AND project_name = %s AND network_name = %s",
                            (str(exc), batch_id, dn, pn, nn),
                        )
        _flush_log()

        # ── Create users ──────────────────────────────────────────────────────
        for user in items["users"]:
            dn = user["domain_name"]
            pn = user["project_name"]
            un = user["username"]
            if user.get("status") == "created":
                _log("info", f"User '{dn}/{un}' already created — skipping")
                continue
            dom_id = domain_id_map.get(dn)
            proj_id = project_id_map.get((dn, pn))
            if not dom_id or not proj_id:
                with get_connection() as conn2:
                    with conn2.cursor() as cur2:
                        cur2.execute(
                            "UPDATE onboarding_users SET status = 'failed', error_msg = %s "
                            "WHERE batch_id = %s AND domain_name = %s AND project_name = %s AND username = %s",
                            (f"Parent domain/project not created", batch_id, dn, pn, un),
                        )
                continue
            temp_pw = user.get("user_password") or _gen_temp_password()
            try:
                try:
                    result_u = client.create_user(
                        un,
                        password=temp_pw,
                        domain_id=dom_id,
                        email=user.get("email") or None,
                    )
                    user_id = result_u.get("id") or result_u.get("user", {}).get("id", "")
                    _log("info", f"Created user '{dn}/{un}' (id={user_id})")
                except Exception as create_exc:
                    # 409 Conflict → user already exists; look it up and reuse
                    status_code = getattr(getattr(create_exc, "response", None), "status_code", None)
                    if status_code == 409:
                        existing_users = client.list_users(domain_id=dom_id)
                        match_u = next((u for u in existing_users if u.get("name") == un), None)
                        if match_u:
                            user_id = match_u["id"]
                            _log("info", f"User '{dn}/{un}' already exists (id={user_id}), reusing")
                        else:
                            raise create_exc
                    else:
                        raise create_exc

                # Map role name to role ID
                role_name = user.get("role", "member")
                try:
                    role_id = client.get_role_id(role_name)
                except Exception:
                    role_id = role_name  # fallback: pass name directly

                client.assign_role_to_user_on_project(proj_id, user_id, role_id)

                with get_connection() as conn2:
                    with conn2.cursor() as cur2:
                        cur2.execute(
                            "UPDATE onboarding_users SET status = 'created', "
                            "pcd_user_id = %s, temp_password = %s "
                            "WHERE batch_id = %s AND domain_name = %s AND project_name = %s AND username = %s",
                            (user_id, temp_pw, batch_id, dn, pn, un),
                        )
                logger.info("Created/reused user '%s/%s' id=%s", dn, un, user_id)
            except Exception as exc:
                logger.error("Failed to create user '%s/%s': %s", dn, un, exc)
                _log("error", f"Failed to create user '{dn}/{un}': {exc}")
                with get_connection() as conn2:
                    with conn2.cursor() as cur2:
                        cur2.execute(
                            "UPDATE onboarding_users SET status = 'failed', error_msg = %s "
                            "WHERE batch_id = %s AND domain_name = %s AND project_name = %s AND username = %s",
                            (str(exc), batch_id, dn, pn, un),
                        )
        _flush_log()

        # ── Compute final status ─────────────────────────────────────────────
        with get_connection() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                # Count failures across all item tables
                cur.execute(
                    """
                    SELECT
                        (SELECT COUNT(*) FROM onboarding_customers WHERE batch_id=%s AND status='failed') +
                        (SELECT COUNT(*) FROM onboarding_projects  WHERE batch_id=%s AND status='failed') +
                        (SELECT COUNT(*) FROM onboarding_networks  WHERE batch_id=%s AND status='failed') +
                        (SELECT COUNT(*) FROM onboarding_users     WHERE batch_id=%s AND status='failed') AS fail_count
                    """,
                    (batch_id, batch_id, batch_id, batch_id),
                )
                fail_count = cur.fetchone()["fail_count"]

        final_status = "partially_failed" if fail_count > 0 else "complete"
        exec_result = {"failed_items": fail_count, "completed_by": actor, "completed_at": _now().isoformat()}

        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE onboarding_batches SET status = %s, execution_result = %s::jsonb WHERE batch_id = %s",
                    (final_status, psycopg2.extras.Json(exec_result), batch_id),
                )

        _log("info", f"Execution finished: {final_status} ({fail_count} failure(s))")
        _flush_log()
        event = "onboarding_completed" if final_status == "complete" else "onboarding_failed"
        severity = "info" if final_status == "complete" else "warning"
        _fire(
            event,
            f"Onboarding {'Completed' if final_status == 'complete' else 'Partially Failed'}",
            f"Batch '{batch['batch_name']}' finished with status '{final_status}'. "
            f"{fail_count} item(s) failed.",
            severity,
        )
        logger.info("Batch %s execution finished: %s (failures=%d)", batch_id, final_status, fail_count)

    except Exception as exc:
        logger.exception("Unexpected error executing batch %s: %s", batch_id, exc)
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE onboarding_batches SET status = 'partially_failed', "
                    "execution_result = %s::jsonb WHERE batch_id = %s",
                    (psycopg2.extras.Json({"error": str(exc)}), batch_id),
                )
        _fire(
            "onboarding_failed",
            "Onboarding Failed",
            f"Batch execution failed with an unexpected error: {exc}",
            "critical",
        )


@router.post("/batches/{batch_id}/execute")
def execute_batch(batch_id: str, current_user: dict = Depends(require_permission("onboarding", "execute"))):
    """
    Execute the batch. Requires:
      - approval_status == 'approved'
      - status == 'dry_run_passed'
    """
    with get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            batch = _get_batch(cur, batch_id)

    if batch["approval_status"] != "approved":
        raise HTTPException(
            status_code=400,
            detail="Batch must be approved before execution. "
                   f"Current approval_status: '{batch['approval_status']}'",
        )
    if batch["status"] != "dry_run_passed":
        raise HTTPException(
            status_code=400,
            detail="Dry-run must pass (zero conflicts) before execution is unlocked. "
                   f"Current status: '{batch['status']}'",
        )

    actor = current_user.get("username", "unknown")

    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE onboarding_batches SET status = 'executing' WHERE batch_id = %s",
                (batch_id,),
            )

    # Run async so the HTTP request returns immediately
    thread = threading.Thread(
        target=_execute_batch_sync,
        args=(batch_id, actor),
        daemon=True,
    )
    thread.start()

    logger.info("Batch %s execution started in background by %s", batch_id, actor)
    return {
        "status": "executing",
        "batch_id": batch_id,
        "message": "Execution started. Poll GET /api/onboarding/batches/{batch_id} for status.",
    }


# ---------------------------------------------------------------------------
# POST /batches/{batch_id}/rerun
# ---------------------------------------------------------------------------

@router.post("/batches/{batch_id}/rerun")
def rerun_batch(
    batch_id: str,
    current_user: dict = Depends(require_permission("onboarding", "execute")),
):
    """
    Re-execute a partially-failed batch: reset failed items to 'pending' and
    restart execution. Already-created items are skipped automatically.
    Requires approval_status=='approved'.
    """
    with get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            batch = _get_batch(cur, batch_id)

    if batch["status"] not in ("partially_failed", "complete"):
        raise HTTPException(
            status_code=400,
            detail=f"Only partially_failed or complete batches can be rerun. Current status: '{batch['status']}'",
        )
    if batch["approval_status"] != "approved":
        raise HTTPException(
            status_code=400,
            detail="Batch must be approved before rerun.",
        )

    actor = current_user.get("username", "unknown")

    with get_connection() as conn:
        with conn.cursor() as cur:
            # Reset failed items in all child tables
            for tbl in ("onboarding_customers", "onboarding_projects",
                         "onboarding_networks", "onboarding_users"):
                cur.execute(
                    f"UPDATE {tbl} SET status = 'pending', error_msg = NULL "
                    "WHERE batch_id = %s AND status = 'failed'",
                    (batch_id,),
                )
            cur.execute(
                "UPDATE onboarding_batches SET status = 'executing', "
                "rerun_count = rerun_count + 1 WHERE batch_id = %s",
                (batch_id,),
            )

    thread = threading.Thread(
        target=_execute_batch_sync,
        args=(batch_id, actor),
        daemon=True,
    )
    thread.start()

    logger.info("Batch %s rerun started by %s", batch_id, actor)
    return {
        "status": "executing",
        "batch_id": batch_id,
        "message": "Rerun started. Failed items have been reset and execution is running.",
    }


# ---------------------------------------------------------------------------
# POST /batches/{batch_id}/send-notifications  — send welcome emails to users
# ---------------------------------------------------------------------------

class SendNotificationsRequest(BaseModel):
    user_ids: List[int] = []          # empty = send to ALL created users with email
    extra_recipients: List[str] = []  # additional addresses (non-personalised summary)


def _smtp_send(msg, to_email: str, *args, **kwargs):
    """Thin wrapper — delegates to smtp_helper.send_raw."""
    from smtp_helper import send_raw
    send_raw(msg, to_email)


@router.post("/batches/{batch_id}/send-notifications")
def send_notifications(
    batch_id: str,
    body: SendNotificationsRequest = None,
    current_user: dict = Depends(require_permission("onboarding", "execute")),
):
    """Send welcome emails to selected users and/or extra recipients."""
    if body is None:
        body = SendNotificationsRequest()

    actor = current_user.get("username", "system") if current_user else "system"

    with get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            batch = _get_batch(cur, batch_id)
            if batch["status"] not in ("complete", "partially_failed"):
                raise HTTPException(status_code=400, detail="Batch must be complete or partially_failed to send notifications")

            # Build user query — filter by explicit IDs or fall back to all created
            if body.user_ids:
                id_placeholders = ",".join(["%s"] * len(body.user_ids))
                cur.execute(
                    f"""SELECT id, username, email, domain_name, project_name, role, temp_password
                        FROM onboarding_users
                        WHERE batch_id = %s AND id IN ({id_placeholders})
                          AND status = 'created' AND send_welcome_email = true""",
                    [batch_id] + body.user_ids,
                )
            else:
                cur.execute(
                    """SELECT id, username, email, domain_name, project_name, role, temp_password
                       FROM onboarding_users
                       WHERE batch_id = %s AND status = 'created'
                         AND email IS NOT NULL AND email != ''
                         AND send_welcome_email = true""",
                    (batch_id,),
                )
            users = cur.fetchall()

            # Fetch all networks for this batch (for email enrichment)
            cur.execute(
                """SELECT domain_name, project_name, network_name, network_kind,
                          network_type, vlan_id, cidr, gateway, is_external, shared, status
                   FROM onboarding_networks
                   WHERE batch_id = %s AND status = 'created'
                   ORDER BY domain_name, project_name, network_name""",
                (batch_id,),
            )
            all_networks = cur.fetchall()

            # Also fetch all domains/projects for the summary
            cur.execute(
                """SELECT domain_name, display_name, status FROM onboarding_customers
                   WHERE batch_id = %s ORDER BY domain_name""", (batch_id,)
            )
            all_domains = cur.fetchall()
            cur.execute(
                """SELECT domain_name, project_name, status FROM onboarding_projects
                   WHERE batch_id = %s AND status = 'created' ORDER BY domain_name, project_name""", (batch_id,)
            )
            all_projects = cur.fetchall()

    # Build lookup: (domain, project) -> list of networks
    from collections import defaultdict
    nets_by_proj: dict = defaultdict(list)
    for n in all_networks:
        nets_by_proj[(n["domain_name"], n["project_name"])].append(n)

    from smtp_helper import (
        SMTP_ENABLED, SMTP_HOST, SMTP_PORT, SMTP_USE_TLS,
        SMTP_USERNAME, SMTP_PASSWORD, SMTP_FROM_ADDRESS, SMTP_FROM_NAME,
    )
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    sent = 0
    skipped = 0

    # ── Personalised emails to selected users ─────────────────────────────
    for u in users:
        to_email = u.get("email") or ""
        if not to_email or not EMAIL_RE.match(to_email):
            skipped += 1
            continue

        # Build networks section for this user's project
        proj_nets = nets_by_proj.get((u["domain_name"], u["project_name"]), [])
        nets_rows = ""
        for n in proj_nets:
            kind_label = {"physical_managed": "Provider (routed)", "physical_l2": "Provider (L2 only)", "virtual": "Tenant overlay"}.get(n["network_kind"], n["network_kind"])
            vlan_str = f"VLAN {n['vlan_id']}" if n.get("vlan_id") else "—"
            cidr_str = n.get("cidr") or "—"
            nets_rows += (
                f"<tr style='border-bottom:1px solid #e2e8f0;'>"
                f"<td style='padding:6px 12px;'><strong>{n['network_name']}</strong></td>"
                f"<td style='padding:6px 12px; color:#64748b;'>{kind_label}</td>"
                f"<td style='padding:6px 12px; color:#64748b;'>{vlan_str}</td>"
                f"<td style='padding:6px 12px; color:#64748b;'>{cidr_str}</td>"
                f"</tr>"
            )
        networks_section = ""
        if nets_rows:
            networks_section = f"""
          <h3 style="color:#334155; margin-top:28px; margin-bottom:8px; font-size:14px; text-transform:uppercase; letter-spacing:.05em;">🌐 Networks in Your Project</h3>
          <table style="border-collapse:collapse; width:100%; font-size:13px;">
            <thead>
              <tr style="background:#f1f5f9; font-size:12px; color:#64748b; text-transform:uppercase;">
                <th style="padding:6px 12px; text-align:left;">Network Name</th>
                <th style="padding:6px 12px; text-align:left;">Type</th>
                <th style="padding:6px 12px; text-align:left;">VLAN</th>
                <th style="padding:6px 12px; text-align:left;">CIDR / Subnet</th>
              </tr>
            </thead>
            <tbody>{nets_rows}</tbody>
          </table>"""

        pwd_display = u['temp_password'] or '(set during onboarding — contact your administrator)'
        html = f"""<html>
        <body style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background:#f8fafc; margin:0; padding:0;">
          <div style="max-width:620px; margin:32px auto; background:#ffffff; border-radius:10px;
                      border:1px solid #e2e8f0; overflow:hidden; box-shadow:0 2px 8px rgba(0,0,0,.06);">

            <!-- Header -->
            <div style="background:linear-gradient(135deg,#1a73e8,#0d47a1); padding:28px 32px;">
              <h1 style="color:#fff; margin:0; font-size:22px; font-weight:600;">Welcome to Platform9</h1>
              <p style="color:#bfdbfe; margin:6px 0 0; font-size:14px;">Your cloud account is ready</p>
            </div>

            <!-- Body -->
            <div style="padding:28px 32px;">
              <p style="color:#1e293b; font-size:15px; margin-top:0;">
                Dear <strong>{u['username']}</strong>,<br><br>
                Your Platform9 account has been provisioned successfully. Below are your login credentials and environment details.
              </p>

              <h3 style="color:#334155; margin-bottom:8px; font-size:14px; text-transform:uppercase; letter-spacing:.05em;">🔐 Your Credentials</h3>
              <table style="border-collapse:collapse; width:100%; font-size:14px; margin-bottom:8px;">
                <tr style="border-bottom:1px solid #e2e8f0;">
                  <td style="padding:9px 12px; color:#64748b; width:180px;">Username</td>
                  <td style="padding:9px 12px;"><strong style='font-family:monospace;font-size:15px;'>{u['username']}</strong></td>
                </tr>
                <tr style="background:#fef9f0; border-bottom:1px solid #e2e8f0;">
                  <td style="padding:9px 12px; color:#64748b;">Temporary Password</td>
                  <td style="padding:9px 12px;">
                    <strong style='font-family:monospace; font-size:15px; background:#fef3c7; padding:2px 8px; border-radius:4px;'>{pwd_display}</strong>
                    <span style="color:#b45309; font-size:12px; margin-left:8px;">⚠ Change on first login</span>
                  </td>
                </tr>
                <tr style="border-bottom:1px solid #e2e8f0;">
                  <td style="padding:9px 12px; color:#64748b;">Domain / Tenant</td>
                  <td style="padding:9px 12px;"><strong>{u['domain_name']}</strong></td>
                </tr>
                <tr style="background:#f8fafc; border-bottom:1px solid #e2e8f0;">
                  <td style="padding:9px 12px; color:#64748b;">Project</td>
                  <td style="padding:9px 12px;">{u['project_name']}</td>
                </tr>
                <tr>
                  <td style="padding:9px 12px; color:#64748b;">Role</td>
                  <td style="padding:9px 12px;">{u['role']}</td>
                </tr>
              </table>

              {networks_section}

              <div style="background:#f0fdf4; border:1px solid #bbf7d0; border-radius:8px; padding:14px 16px; margin-top:24px;">
                <p style="margin:0; color:#166534; font-size:13px;">
                  💡 <strong>Login tip:</strong> On the Platform9 login page, make sure to specify your domain name
                  (<strong>{u['domain_name']}</strong>) in the domain field before entering your credentials.
                </p>
              </div>

              <p style="color:#64748b; font-size:13px; margin-top:20px;">
                If you have any questions or need assistance, contact your platform administrator.
              </p>
            </div>

            <!-- Footer -->
            <div style="background:#f1f5f9; padding:14px 32px; border-top:1px solid #e2e8f0;">
              <p style="color:#94a3b8; font-size:11px; margin:0;">
                Platform9 Management — Automated Onboarding Notification &nbsp;|&nbsp; Batch: {batch['batch_name']}
              </p>
            </div>
          </div>
        </body></html>"""

        if not SMTP_ENABLED:
            logger.info("SMTP disabled — skipping welcome email for %s", to_email)
            skipped += 1
            continue

        try:
            msg = MIMEMultipart("alternative")
            msg["From"] = f"{SMTP_FROM_NAME} <{SMTP_FROM_ADDRESS}>"
            msg["To"]   = to_email
            msg["Subject"] = "Welcome to Platform9 — Your Account is Ready"
            msg.attach(MIMEText(html, "html"))
            _smtp_send(msg, to_email, SMTP_HOST, SMTP_PORT, SMTP_USE_TLS, SMTP_USERNAME, SMTP_PASSWORD, SMTP_FROM_ADDRESS)
            logger.info("Welcome email sent to %s (%s)", to_email, u["username"])
            sent += 1
        except Exception as mail_exc:
            logger.error("Failed to send welcome email to %s: %s", to_email, mail_exc)
            skipped += 1

    # ── General summary to extra recipients ───────────────────────────────
    for extra_email in body.extra_recipients:
        if not extra_email or not EMAIL_RE.match(extra_email):
            skipped += 1
            continue

        # Build per-domain sections (domains → projects → networks + users)
        domain_sections = ""
        for dom in all_domains:
            dn = dom["domain_name"]
            dom_projects = [p for p in all_projects if p["domain_name"] == dn]
            proj_sections = ""
            for proj in dom_projects:
                pn = proj["project_name"]
                p_nets = nets_by_proj.get((dn, pn), [])
                p_users = [u for u in users if u["domain_name"] == dn and u["project_name"] == pn]

                net_rows = "".join(
                    f"<tr style='border-bottom:1px solid #e2e8f0;'>"
                    f"<td style='padding:5px 10px; font-size:13px;'>{n['network_name']}</td>"
                    f"<td style='padding:5px 10px; font-size:13px; color:#64748b;'>"
                      f"{'Provider (routed)' if n['network_kind']=='physical_managed' else 'Provider (L2)' if n['network_kind']=='physical_l2' else 'Tenant overlay'}</td>"
                    f"<td style='padding:5px 10px; font-size:13px; color:#64748b;'>{'VLAN ' + str(n['vlan_id']) if n.get('vlan_id') else '—'}</td>"
                    f"<td style='padding:5px 10px; font-size:13px; color:#64748b;'>{n.get('cidr') or '—'}</td>"
                    f"</tr>"
                    for n in p_nets
                ) if p_nets else "<tr><td colspan='4' style='padding:5px 10px; color:#94a3b8; font-size:13px;'>No networks</td></tr>"

                usr_rows = "".join(
                    f"<tr style='border-bottom:1px solid #e2e8f0;'>"
                    f"<td style='padding:5px 10px; font-size:13px;'><strong>{u['username']}</strong></td>"
                    f"<td style='padding:5px 10px; font-size:13px; color:#64748b;'>{u['email'] or '—'}</td>"
                    f"<td style='padding:5px 10px; font-size:13px; color:#64748b;'>{u['role']}</td>"
                    f"<td style='padding:5px 10px; font-size:13px; font-family:monospace; background:#fef9f0;'>{u['temp_password'] or '—'}</td>"
                    f"</tr>"
                    for u in p_users
                ) if p_users else ""

                proj_sections += f"""
                <div style="margin-left:16px; margin-top:10px; border-left:3px solid #bfdbfe; padding-left:12px;">
                  <p style="margin:0 0 6px; font-weight:600; color:#1e40af; font-size:13px;">📁 {pn}</p>
                  <p style="margin:0 0 4px; font-size:12px; color:#64748b; text-transform:uppercase; letter-spacing:.04em;">Networks</p>
                  <table style="border-collapse:collapse; width:100%; margin-bottom:8px; font-size:12px;">
                    <thead><tr style="background:#f1f5f9; color:#64748b; text-transform:uppercase; font-size:11px;">
                      <th style="padding:4px 10px; text-align:left;">Name</th>
                      <th style="padding:4px 10px; text-align:left;">Kind</th>
                      <th style="padding:4px 10px; text-align:left;">VLAN</th>
                      <th style="padding:4px 10px; text-align:left;">CIDR</th>
                    </tr></thead>
                    <tbody>{net_rows}</tbody>
                  </table>
                  {"" if not p_users else f'''
                  <p style="margin:4px 0; font-size:12px; color:#64748b; text-transform:uppercase; letter-spacing:.04em;">Users</p>
                  <table style="border-collapse:collapse; width:100%; font-size:12px;">
                    <thead><tr style="background:#f1f5f9; color:#64748b; text-transform:uppercase; font-size:11px;">
                      <th style="padding:4px 10px; text-align:left;">Username</th>
                      <th style="padding:4px 10px; text-align:left;">Email</th>
                      <th style="padding:4px 10px; text-align:left;">Role</th>
                      <th style="padding:4px 10px; text-align:left;">Temp Password</th>
                    </tr></thead>
                    <tbody>{usr_rows}</tbody>
                  </table>'''}
                </div>"""

            domain_sections += f"""
            <div style="background:#f8fafc; border:1px solid #e2e8f0; border-radius:8px; padding:14px 16px; margin-bottom:16px;">
              <p style="margin:0 0 8px; font-weight:700; color:#0f172a; font-size:15px;">🏢 {dn}
                {"<span style='color:#64748b; font-size:12px; font-weight:400;'> — " + (dom.get('display_name') or '') + "</span>" if dom.get('display_name') else ""}
              </p>
              {proj_sections}
            </div>"""

        all_users_rows = "".join(
            f"<tr style='border-bottom:1px solid #e2e8f0;'>"
            f"<td style='padding:6px 12px;'><strong>{u['username']}</strong></td>"
            f"<td style='padding:6px 12px; color:#64748b;'>{u['domain_name']}</td>"
            f"<td style='padding:6px 12px; color:#64748b;'>{u['project_name']}</td>"
            f"<td style='padding:6px 12px; color:#64748b;'>{u['role']}</td>"
            f"<td style='padding:6px 12px; font-family:monospace; font-size:13px; background:#fef9f0;'>{u['temp_password'] or '—'}</td>"
            f"</tr>"
            for u in users
        )

        summary_html = f"""<html>
        <body style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background:#f8fafc; margin:0; padding:0;">
          <div style="max-width:680px; margin:32px auto; background:#ffffff; border-radius:10px;
                      border:1px solid #e2e8f0; overflow:hidden; box-shadow:0 2px 8px rgba(0,0,0,.06);">

            <div style="background:linear-gradient(135deg,#1a73e8,#0d47a1); padding:24px 32px;">
              <h1 style="color:#fff; margin:0; font-size:20px; font-weight:600;">Platform9 Onboarding — Batch Complete</h1>
              <p style="color:#bfdbfe; margin:5px 0 0; font-size:13px;">Batch: <strong style="color:#fff;">{batch['batch_name']}</strong></p>
            </div>

            <div style="padding:24px 32px;">
              <p style="color:#1e293b; font-size:14px; margin-top:0;">
                Onboarding completed for <strong>{len(all_domains)}</strong> domain(s),
                <strong>{len(all_projects)}</strong> project(s),
                <strong>{len(all_networks)}</strong> network(s), and
                <strong>{len(users)}</strong> user(s).
              </p>

              {domain_sections}

              {"" if not users else f'''
              <h3 style="color:#334155; font-size:13px; text-transform:uppercase; letter-spacing:.05em; margin-top:24px;">All Users &amp; Credentials</h3>
              <table style="border-collapse:collapse; width:100%; font-size:13px;">
                <thead><tr style="background:#f1f5f9; color:#64748b; text-transform:uppercase; font-size:11px;">
                  <th style="padding:6px 12px; text-align:left;">Username</th>
                  <th style="padding:6px 12px; text-align:left;">Domain</th>
                  <th style="padding:6px 12px; text-align:left;">Project</th>
                  <th style="padding:6px 12px; text-align:left;">Role</th>
                  <th style="padding:6px 12px; text-align:left;">Temp Password</th>
                </tr></thead>
                <tbody>{all_users_rows}</tbody>
              </table>'''}
            </div>

            <div style="background:#f1f5f9; padding:12px 32px; border-top:1px solid #e2e8f0;">
              <p style="color:#94a3b8; font-size:11px; margin:0;">
                Platform9 Management — Onboarding Summary for {actor} &nbsp;|&nbsp; {batch['batch_name']}
              </p>
            </div>
          </div>
        </body></html>"""

        if not SMTP_ENABLED:
            skipped += 1
            continue

        try:
            msg = MIMEMultipart("alternative")
            msg["From"] = f"{SMTP_FROM_NAME} <{SMTP_FROM_ADDRESS}>"
            msg["To"]   = extra_email
            msg["Subject"] = f"Platform9 Onboarding Summary — {batch['batch_name']}"
            msg.attach(MIMEText(summary_html, "html"))
            _smtp_send(msg, extra_email, SMTP_HOST, SMTP_PORT, SMTP_USE_TLS, SMTP_USERNAME, SMTP_PASSWORD, SMTP_FROM_ADDRESS)
            logger.info("Onboarding summary sent to extra recipient %s", extra_email)
            sent += 1
        except Exception as mail_exc:
            logger.error("Failed to send summary to %s: %s", extra_email, mail_exc)
            skipped += 1

    # ── Log to activity_log ────────────────────────────────────────────────
    try:
        from migration_routes import _log_activity  # type: ignore
        _log_activity(
            actor=actor,
            action="send_welcome_notifications",
            resource_type="onboarding_notification",
            resource_id=batch_id,
            resource_name=batch.get("batch_name", batch_id),
            details={
                "sent": sent,
                "skipped": skipped,
                "user_count": len(users),
                "extra_recipients": body.extra_recipients,
                "batch_name": batch.get("batch_name"),
            },
            result="success" if sent > 0 else "skipped",
        )
    except Exception as log_exc:
        logger.warning("Could not write activity log for notification send: %s", log_exc)

    return {
        "sent": sent,
        "skipped": skipped,
        "message": f"Welcome emails: {sent} sent, {skipped} skipped.",
    }


# ---------------------------------------------------------------------------
# DELETE /batches/{batch_id}
# ---------------------------------------------------------------------------

@router.delete("/batches/{batch_id}", status_code=204)
def delete_batch(batch_id: str, current_user: dict = Depends(require_permission("onboarding", "create"))):
    with get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            batch = _get_batch(cur, batch_id)
            if batch["status"] in ("executing",):
                raise HTTPException(status_code=400, detail="Cannot delete a batch that is currently executing")
            cur.execute("DELETE FROM onboarding_batches WHERE batch_id = %s", (batch_id,))
