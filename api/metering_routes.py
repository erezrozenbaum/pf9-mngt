"""
Metering API Routes
===================
Endpoints for operational metering data: resource usage, snapshots, restores,
API usage, quotas, efficiency scores, and chargeback export.

RBAC
----
  - admin      → metering:read  (view all metering data + export)
  - superadmin → metering:read + metering:write (configure cost model, toggle)
"""

from __future__ import annotations

import csv
import io
import logging
from datetime import datetime, timezone, timedelta
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel, Field
from psycopg2.extras import RealDictCursor

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _XLSX_AVAILABLE = True
except ImportError:
    _XLSX_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    _PDF_AVAILABLE = True
except ImportError:
    _PDF_AVAILABLE = False

from auth import require_permission, get_current_user, User, get_effective_region_filter
from db_pool import get_connection
from smtp_helper import (
    SMTP_ENABLED, SMTP_HOST, SMTP_PORT, SMTP_USE_TLS,
    SMTP_USERNAME, SMTP_PASSWORD, SMTP_FROM_ADDRESS, SMTP_FROM_NAME,
)

logger = logging.getLogger("pf9.metering")

router = APIRouter(prefix="/api/metering", tags=["metering"])


# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------

class MeteringConfigResponse(BaseModel):
    enabled: bool
    collection_interval_min: int
    retention_days: int
    cost_per_vcpu_hour: float
    cost_per_gb_ram_hour: float
    cost_per_gb_storage_month: float
    cost_per_snapshot_gb_month: float
    cost_per_api_call: float
    cost_currency: str
    updated_at: Optional[str] = None


class MeteringConfigUpdate(BaseModel):
    enabled: Optional[bool] = None
    collection_interval_min: Optional[int] = Field(None, ge=5, le=1440)
    retention_days: Optional[int] = Field(None, ge=1, le=3650)
    cost_per_vcpu_hour: Optional[float] = Field(None, ge=0)
    cost_per_gb_ram_hour: Optional[float] = Field(None, ge=0)
    cost_per_gb_storage_month: Optional[float] = Field(None, ge=0)
    cost_per_snapshot_gb_month: Optional[float] = Field(None, ge=0)
    cost_per_api_call: Optional[float] = Field(None, ge=0)
    cost_currency: Optional[str] = Field(None, max_length=10)


class ResourceRecord(BaseModel):
    collected_at: str
    vm_id: str
    vm_name: Optional[str] = None
    vm_ip: Optional[str] = None
    project_name: Optional[str] = None
    domain: Optional[str] = None
    host: Optional[str] = None
    flavor: Optional[str] = None
    vcpus_allocated: Optional[int] = None
    ram_allocated_mb: Optional[int] = None
    disk_allocated_gb: Optional[int] = None
    cpu_usage_percent: Optional[float] = None
    ram_usage_mb: Optional[float] = None
    ram_usage_percent: Optional[float] = None
    disk_used_gb: Optional[float] = None
    disk_usage_percent: Optional[float] = None
    network_rx_bytes: Optional[int] = None
    network_tx_bytes: Optional[int] = None
    storage_read_bytes: Optional[int] = None
    storage_write_bytes: Optional[int] = None


class EfficiencyRecord(BaseModel):
    collected_at: str
    vm_id: str
    vm_name: Optional[str] = None
    project_name: Optional[str] = None
    domain: Optional[str] = None
    cpu_efficiency: Optional[float] = None
    ram_efficiency: Optional[float] = None
    storage_efficiency: Optional[float] = None
    overall_score: Optional[float] = None
    classification: Optional[str] = None
    recommendation: Optional[str] = None


# ---------------------------------------------------------------------------
# Config endpoints
# ---------------------------------------------------------------------------

@router.get("/config", response_model=MeteringConfigResponse)
async def get_metering_config(user: User = Depends(require_permission("metering", "read"))):
    """Return current metering configuration."""
    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM metering_config WHERE id = 1")
            row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Metering config not found")
        row["updated_at"] = row["updated_at"].isoformat() if row.get("updated_at") else None
        # Convert Decimal to float
        for k in list(row.keys()):
            if hasattr(row[k], "as_tuple"):
                row[k] = float(row[k])
        return row


@router.put("/config", response_model=MeteringConfigResponse)
async def update_metering_config(
    body: MeteringConfigUpdate,
    user: User = Depends(require_permission("metering", "write")),
):
    """Update metering configuration (superadmin only)."""
    updates = body.dict(exclude_unset=True)
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")

    set_clauses = []
    params = []
    for key, val in updates.items():
        set_clauses.append(f"{key} = %s")
        params.append(val)
    set_clauses.append("updated_at = now()")

    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                f"UPDATE metering_config SET {', '.join(set_clauses)} WHERE id = 1 RETURNING *",
                params,
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Metering config not found")
        row["updated_at"] = row["updated_at"].isoformat() if row.get("updated_at") else None
        for k in list(row.keys()):
            if hasattr(row[k], "as_tuple"):
                row[k] = float(row[k])
        return row


# ---------------------------------------------------------------------------
# Resource metering
# ---------------------------------------------------------------------------

@router.get("/resources")
async def get_resource_metering(
    project: Optional[str] = Query(None, description="Filter by project name"),
    domain: Optional[str] = Query(None, description="Filter by domain"),
    vm_id: Optional[str] = Query(None, description="Filter by VM ID"),
    hours: int = Query(24, ge=1, le=2160, description="Lookback window in hours"),
    limit: int = Query(500, ge=1, le=10000),
    region_id: Optional[str] = Query(None, description="Filter by region ID"),
    user: User = Depends(require_permission("metering", "read")),
):
    """Return resource metering records (latest per VM by default)."""
    effective_region = get_effective_region_filter(
        user["username"] if isinstance(user, dict) else user.username, region_id
    )
    with get_connection() as conn:
        where = ["collected_at > now() - interval '%s hours'"]
        params: list = [hours]
        if project:
            where.append("project_name = %s")
            params.append(project)
        if domain:
            where.append("domain = %s")
            params.append(domain)
        if vm_id:
            where.append("vm_id = %s")
            params.append(vm_id)
        if effective_region:
            where.append("region_id = %s")
            params.append(effective_region)

        # Return only the LATEST record per VM, enriched with live DB metadata
        # when the stored domain/project/vm_ip fields are stale or missing.
        sql = f"""
            SELECT DISTINCT ON (mr.vm_id) mr.*,
                COALESCE(NULLIF(mr.domain, 'Unknown'), d.name)         AS domain_enriched,
                COALESCE(NULLIF(mr.project_name, 'Unknown'), p.name)   AS project_name_enriched,
                COALESCE(
                    NULLIF(mr.vm_ip, 'Unknown'),
                    (SELECT fi->>'addr'
                     FROM jsonb_each(s.raw_json->'addresses') AS net,
                          jsonb_array_elements(net.value) AS fi
                     LIMIT 1)
                ) AS vm_ip_enriched
            FROM metering_resources mr
            LEFT JOIN servers s ON s.id = mr.vm_id
            LEFT JOIN projects p ON p.id = s.project_id
            LEFT JOIN domains d ON d.id = p.domain_id
            WHERE {' AND '.join(where)}
            ORDER BY mr.vm_id, mr.collected_at DESC
            LIMIT %s
        """
        params.append(limit)

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()
            # Apply enriched values over the originals
            for r in rows:
                enriched_domain   = r.pop("domain_enriched", None)
                enriched_project  = r.pop("project_name_enriched", None)
                enriched_ip       = r.pop("vm_ip_enriched", None)
                if enriched_domain:
                    r["domain"] = enriched_domain
                if enriched_project:
                    r["project_name"] = enriched_project
                if enriched_ip:
                    r["vm_ip"] = enriched_ip

        # Serialize
        for r in rows:
            for k in list(r.keys()):
                if hasattr(r[k], "isoformat"):
                    r[k] = r[k].isoformat()
                elif hasattr(r[k], "as_tuple"):
                    r[k] = float(r[k])

        return {"data": rows, "count": len(rows)}


# ---------------------------------------------------------------------------
# Snapshot metering
# ---------------------------------------------------------------------------

@router.get("/snapshots")
async def get_snapshot_metering(
    project: Optional[str] = Query(None),
    domain: Optional[str] = Query(None),
    hours: int = Query(24, ge=1, le=2160),
    limit: int = Query(500, ge=1, le=10000),
    region_id: Optional[str] = Query(None, description="Filter by region ID"),
    user: User = Depends(require_permission("metering", "read")),
):
    """Return snapshot metering records."""
    effective_region = get_effective_region_filter(
        user["username"] if isinstance(user, dict) else user.username, region_id
    )
    with get_connection() as conn:
        where = ["collected_at > now() - interval '%s hours'"]
        params: list = [hours]
        if project:
            where.append("project_name = %s")
            params.append(project)
        if domain:
            where.append("domain = %s")
            params.append(domain)
        if effective_region:
            where.append("region_id = %s")
            params.append(effective_region)

        sql = f"""
            SELECT * FROM metering_snapshots
            WHERE {' AND '.join(where)}
            ORDER BY collected_at DESC LIMIT %s
        """
        params.append(limit)

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()

        for r in rows:
            for k in list(r.keys()):
                if hasattr(r[k], "isoformat"):
                    r[k] = r[k].isoformat()
                elif hasattr(r[k], "as_tuple"):
                    r[k] = float(r[k])

        return {"data": rows, "count": len(rows)}


# ---------------------------------------------------------------------------
# Restore metering
# ---------------------------------------------------------------------------

@router.get("/restores")
async def get_restore_metering(
    project: Optional[str] = Query(None),
    domain: Optional[str] = Query(None),
    hours: int = Query(168, ge=1, le=2160),
    limit: int = Query(200, ge=1, le=5000),
    region_id: Optional[str] = Query(None, description="Filter by region ID"),
    user: User = Depends(require_permission("metering", "read")),
):
    """Return restore operation metering records."""
    effective_region = get_effective_region_filter(
        user["username"] if isinstance(user, dict) else user.username, region_id
    )
    with get_connection() as conn:
        where = ["collected_at > now() - interval '%s hours'"]
        params: list = [hours]
        if project:
            where.append("project_name = %s")
            params.append(project)
        if domain:
            where.append("domain = %s")
            params.append(domain)
        if effective_region:
            where.append("region_id = %s")
            params.append(effective_region)

        sql = f"""
            SELECT * FROM metering_restores
            WHERE {' AND '.join(where)}
            ORDER BY collected_at DESC LIMIT %s
        """
        params.append(limit)

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()

        for r in rows:
            for k in list(r.keys()):
                if hasattr(r[k], "isoformat"):
                    r[k] = r[k].isoformat()
                elif hasattr(r[k], "as_tuple"):
                    r[k] = float(r[k])

        return {"data": rows, "count": len(rows)}


# ---------------------------------------------------------------------------
# API usage metering
# ---------------------------------------------------------------------------

@router.get("/api-usage")
async def get_api_usage_metering(
    endpoint: Optional[str] = Query(None),
    hours: int = Query(24, ge=1, le=2160),
    limit: int = Query(500, ge=1, le=10000),
    user: User = Depends(require_permission("metering", "read")),
):
    """Return API usage metering records."""
    with get_connection() as conn:
        where = ["collected_at > now() - interval '%s hours'"]
        params: list = [hours]
        if endpoint:
            where.append("endpoint ILIKE %s")
            params.append(f"%{endpoint}%")

        sql = f"""
            SELECT * FROM metering_api_usage
            WHERE {' AND '.join(where)}
            ORDER BY collected_at DESC LIMIT %s
        """
        params.append(limit)

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()

        for r in rows:
            for k in list(r.keys()):
                if hasattr(r[k], "isoformat"):
                    r[k] = r[k].isoformat()
                elif hasattr(r[k], "as_tuple"):
                    r[k] = float(r[k])

        return {"data": rows, "count": len(rows)}


# ---------------------------------------------------------------------------
# Efficiency scores
# ---------------------------------------------------------------------------

@router.get("/efficiency")
async def get_efficiency_scores(
    project: Optional[str] = Query(None),
    domain: Optional[str] = Query(None),
    classification: Optional[str] = Query(None, description="Filter: excellent|good|fair|poor|idle"),
    hours: int = Query(24, ge=1, le=2160),
    limit: int = Query(500, ge=1, le=10000),
    region_id: Optional[str] = Query(None, description="Filter by region ID"),
    user: User = Depends(require_permission("metering", "read")),
):
    """Return VM efficiency scores (latest per VM by default)."""
    effective_region = get_effective_region_filter(
        user["username"] if isinstance(user, dict) else user.username, region_id
    )
    with get_connection() as conn:
        where = ["collected_at > now() - interval '%s hours'"]
        params: list = [hours]
        if project:
            where.append("project_name = %s")
            params.append(project)
        if domain:
            where.append("domain = %s")
            params.append(domain)
        if classification:
            where.append("classification = %s")
            params.append(classification)
        if effective_region:
            where.append("region_id = %s")
            params.append(effective_region)

        # Return only the LATEST record per VM to avoid duplication
        sql = f"""
            SELECT DISTINCT ON (vm_id) *
            FROM metering_efficiency
            WHERE {' AND '.join(where)}
            ORDER BY vm_id, collected_at DESC
            LIMIT %s
        """
        params.append(limit)

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()

        for r in rows:
            for k in list(r.keys()):
                if hasattr(r[k], "isoformat"):
                    r[k] = r[k].isoformat()
                elif hasattr(r[k], "as_tuple"):
                    r[k] = float(r[k])

        return {"data": rows, "count": len(rows)}


# ---------------------------------------------------------------------------
# Filters – dropdown values for UI
# ---------------------------------------------------------------------------

@router.get("/filters")
async def get_metering_filters(
    user: User = Depends(require_permission("metering", "read")),
):
    """Return available projects, domains, and flavors for filter dropdowns."""
    with get_connection() as conn:
        with conn.cursor() as cur:
            # Active projects & domains from metering data
            cur.execute("""
                SELECT DISTINCT project_name FROM metering_resources
                WHERE project_name IS NOT NULL AND project_name != ''
                ORDER BY project_name
            """)
            projects = [r[0] for r in cur.fetchall()]

            cur.execute("""
                SELECT DISTINCT domain FROM metering_resources
                WHERE domain IS NOT NULL AND domain != ''
                ORDER BY domain
            """)
            domains = [r[0] for r in cur.fetchall()]

            # All projects/domains from projects table (including those without metering yet)
            cur.execute("""
                SELECT DISTINCT p.name AS project, COALESCE(d.name, '') AS domain
                FROM projects p
                LEFT JOIN domains d ON d.id = p.domain_id
                ORDER BY domain, project
            """)
            all_tenants = [{"project": r[0], "domain": r[1]} for r in cur.fetchall()]

            # All flavors from flavors table
            cur.execute("SELECT name, vcpus, ram_mb, disk_gb FROM flavors ORDER BY name")
            flavors = [{"name": r[0], "vcpus": r[1], "ram_mb": r[2], "disk_gb": r[3]} for r in cur.fetchall()]

    # Merge OpenStack project/domain names so dropdowns are populated
    # even before metering_resources has any rows.
    merged_projects = sorted(set(projects) | {t["project"] for t in all_tenants if t["project"]})
    merged_domains  = sorted(set(domains)  | {t["domain"]  for t in all_tenants if t["domain"]})

    return {
        "projects": merged_projects,
        "domains": merged_domains,
        "all_tenants": all_tenants,
        "flavors": flavors,
    }


# ---------------------------------------------------------------------------
# Overview / summary
# ---------------------------------------------------------------------------

@router.get("/overview")
async def get_metering_overview(
    project: Optional[str] = Query(None),
    domain: Optional[str] = Query(None),
    region_id: Optional[str] = Query(None, description="Filter by region ID"),
    user: User = Depends(require_permission("metering", "read")),
):
    """
    High-level metering overview: totals and recent counts across all
    metering categories for the MSP dashboard.
    """
    effective_region = get_effective_region_filter(
        user["username"] if isinstance(user, dict) else user.username, region_id
    )
    with get_connection() as conn:
        pfilter = ""
        params: list = []
        if project:
            pfilter = " AND project_name = %s"
            params.append(project)
        if domain:
            pfilter = pfilter + " AND domain = %s"
            params.append(domain)
        if effective_region:
            pfilter = pfilter + " AND region_id = %s"
            params.append(effective_region)

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Total active VMs — use live servers table for accurate count.
            # Filters on project_name / domain are joined through projects/domains tables.
            live_vm_filter = "WHERE s.status NOT IN ('DELETED', 'SOFT_DELETED')"
            live_vm_params: list = []
            if project:
                live_vm_filter += " AND p.name = %s"
                live_vm_params.append(project)
            if domain:
                live_vm_filter += " AND d.name = %s"
                live_vm_params.append(domain)
            try:
                cur.execute(
                    f"""SELECT COUNT(DISTINCT s.id) AS total_vms
                          FROM servers s
                          LEFT JOIN projects p ON p.id = s.project_id
                          LEFT JOIN domains d ON d.id = p.domain_id
                          {live_vm_filter}""",
                    live_vm_params,
                )
                total_vms = cur.fetchone()["total_vms"]
            except Exception:
                # Fallback to metering records if servers table is unavailable
                cur.execute(
                    f"SELECT COUNT(DISTINCT vm_id) AS total_vms FROM metering_resources WHERE 1=1 {pfilter}",
                    params,
                )

            # Latest resource totals – use DISTINCT ON to get most recent per VM
            cur.execute(f"""
                SELECT
                    COALESCE(SUM(vcpus_allocated), 0) AS total_vcpus,
                    COALESCE(SUM(ram_allocated_mb), 0) AS total_ram_mb,
                    COALESCE(SUM(disk_allocated_gb), 0) AS total_disk_gb,
                    COALESCE(AVG(cpu_usage_percent), 0) AS avg_cpu_usage,
                    COALESCE(AVG(ram_usage_percent), 0) AS avg_ram_usage,
                    COALESCE(AVG(disk_usage_percent), 0) AS avg_disk_usage
                FROM (
                    SELECT DISTINCT ON (vm_id)
                        vcpus_allocated, ram_allocated_mb, disk_allocated_gb,
                        cpu_usage_percent, ram_usage_percent, disk_usage_percent
                    FROM metering_resources
                    WHERE collected_at > now() - interval '24 hours' {pfilter}
                    ORDER BY vm_id, collected_at DESC
                ) latest
            """, params)
            res_totals = dict(cur.fetchone())

            # Snapshot totals – use latest collection only
            cur.execute(f"""
                SELECT
                    COUNT(*) AS total_snapshots,
                    COALESCE(SUM(size_gb), 0) AS total_snapshot_gb,
                    COUNT(*) FILTER (WHERE is_compliant IS TRUE)  AS compliant_count,
                    COUNT(*) FILTER (WHERE is_compliant IS FALSE) AS non_compliant_count,
                    COUNT(*) FILTER (WHERE is_compliant IS NULL)  AS unknown_count
                FROM (
                    SELECT DISTINCT ON (snapshot_id) size_gb, is_compliant
                    FROM metering_snapshots
                    WHERE collected_at > now() - interval '24 hours' {pfilter}
                    ORDER BY snapshot_id, collected_at DESC
                ) latest
            """, params)
            snap_totals = dict(cur.fetchone())

            # Restore stats (last 7 days)
            cur.execute(f"""
                SELECT
                    COUNT(*) AS total_restores,
                    COUNT(*) FILTER (WHERE status = 'completed') AS successful,
                    COUNT(*) FILTER (WHERE status = 'failed') AS failed,
                    COALESCE(AVG(duration_seconds), 0) AS avg_duration_sec
                FROM metering_restores
                WHERE collected_at > now() - interval '7 days' {pfilter}
            """, params)
            restore_totals = dict(cur.fetchone())

            # API usage (last 24 hours)
            cur.execute("""
                SELECT
                    COALESCE(SUM(total_calls), 0) AS total_api_calls,
                    COALESCE(SUM(error_count), 0) AS total_api_errors,
                    COALESCE(AVG(avg_latency_ms), 0) AS avg_api_latency_ms
                FROM metering_api_usage
                WHERE collected_at > now() - interval '24 hours'
            """)
            api_totals = dict(cur.fetchone())

            # Efficiency summary – latest per VM
            cur.execute(f"""
                SELECT
                    COALESCE(AVG(overall_score), 0) AS avg_efficiency,
                    COUNT(*) FILTER (WHERE classification = 'excellent') AS excellent_count,
                    COUNT(*) FILTER (WHERE classification = 'good') AS good_count,
                    COUNT(*) FILTER (WHERE classification = 'fair') AS fair_count,
                    COUNT(*) FILTER (WHERE classification = 'poor') AS poor_count,
                    COUNT(*) FILTER (WHERE classification = 'idle') AS idle_count
                FROM (
                    SELECT DISTINCT ON (vm_id) overall_score, classification
                    FROM metering_efficiency
                    WHERE collected_at > now() - interval '24 hours' {pfilter}
                    ORDER BY vm_id, collected_at DESC
                ) latest
            """, params)
            eff_totals = dict(cur.fetchone())

        # Convert Decimals
        for d in [res_totals, snap_totals, restore_totals, api_totals, eff_totals]:
            for k in list(d.keys()):
                if hasattr(d[k], "as_tuple"):
                    d[k] = float(d[k])

        return {
            "total_vms_metered": total_vms,
            "resources": res_totals,
            "snapshots": snap_totals,
            "restores": restore_totals,
            "api_usage": api_totals,
            "efficiency": eff_totals,
        }


# ---------------------------------------------------------------------------
# CSV Export
# ---------------------------------------------------------------------------

def _rows_to_csv(rows: list, filename: str) -> StreamingResponse:
    """Convert list of dicts to a streaming CSV response."""
    if not rows:
        output = io.StringIO("No data\n")
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=rows[0].keys())
    writer.writeheader()
    for row in rows:
        # Stringify any non-primitive values
        clean = {}
        for k, v in row.items():
            if hasattr(v, "isoformat"):
                clean[k] = v.isoformat()
            elif hasattr(v, "as_tuple"):
                clean[k] = float(v)
            else:
                clean[k] = v
        writer.writerow(clean)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/resources")
async def export_resources(
    project: Optional[str] = Query(None),
    domain: Optional[str] = Query(None),
    hours: int = Query(24, ge=1, le=2160),
    user: User = Depends(require_permission("metering", "read")),
):
    """Export resource metering data as CSV."""
    with get_connection() as conn:
        where = ["collected_at > now() - interval '%s hours'"]
        params: list = [hours]
        if project:
            where.append("project_name = %s")
            params.append(project)
        if domain:
            where.append("domain = %s")
            params.append(domain)

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(f"SELECT * FROM metering_resources WHERE {' AND '.join(where)} ORDER BY collected_at DESC", params)
            rows = cur.fetchall()

        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
        return _rows_to_csv(rows, f"metering_resources_{ts}.csv")


@router.get("/export/snapshots")
async def export_snapshots(
    project: Optional[str] = Query(None),
    domain: Optional[str] = Query(None),
    hours: int = Query(24, ge=1, le=2160),
    user: User = Depends(require_permission("metering", "read")),
):
    """Export snapshot metering data as CSV."""
    with get_connection() as conn:
        where = ["collected_at > now() - interval '%s hours'"]
        params: list = [hours]
        if project:
            where.append("project_name = %s")
            params.append(project)
        if domain:
            where.append("domain = %s")
            params.append(domain)

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(f"SELECT * FROM metering_snapshots WHERE {' AND '.join(where)} ORDER BY collected_at DESC", params)
            rows = cur.fetchall()

        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
        return _rows_to_csv(rows, f"metering_snapshots_{ts}.csv")


@router.get("/export/restores")
async def export_restores(
    project: Optional[str] = Query(None),
    domain: Optional[str] = Query(None),
    hours: int = Query(168, ge=1, le=2160),
    user: User = Depends(require_permission("metering", "read")),
):
    """Export restore metering data as CSV."""
    with get_connection() as conn:
        where = ["collected_at > now() - interval '%s hours'"]
        params: list = [hours]
        if project:
            where.append("project_name = %s")
            params.append(project)
        if domain:
            where.append("domain = %s")
            params.append(domain)

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(f"SELECT * FROM metering_restores WHERE {' AND '.join(where)} ORDER BY collected_at DESC", params)
            rows = cur.fetchall()

        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
        return _rows_to_csv(rows, f"metering_restores_{ts}.csv")


@router.get("/export/api-usage")
async def export_api_usage(
    hours: int = Query(24, ge=1, le=2160),
    user: User = Depends(require_permission("metering", "read")),
):
    """Export API usage metering data as CSV."""
    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                "SELECT * FROM metering_api_usage WHERE collected_at > now() - interval '%s hours' ORDER BY collected_at DESC",
                [hours],
            )
            rows = cur.fetchall()

        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
        return _rows_to_csv(rows, f"metering_api_usage_{ts}.csv")


@router.get("/export/efficiency")
async def export_efficiency(
    project: Optional[str] = Query(None),
    domain: Optional[str] = Query(None),
    hours: int = Query(24, ge=1, le=2160),
    user: User = Depends(require_permission("metering", "read")),
):
    """Export efficiency scores as CSV."""
    with get_connection() as conn:
        where = ["collected_at > now() - interval '%s hours'"]
        params: list = [hours]
        if project:
            where.append("project_name = %s")
            params.append(project)
        if domain:
            where.append("domain = %s")
            params.append(domain)

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(f"SELECT * FROM metering_efficiency WHERE {' AND '.join(where)} ORDER BY collected_at DESC", params)
            rows = cur.fetchall()

        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
        return _rows_to_csv(rows, f"metering_efficiency_{ts}.csv")


@router.get("/export/chargeback")
async def export_chargeback(
    project: Optional[str] = Query(None),
    domain: Optional[str] = Query(None),
    hours: int = Query(720, ge=1, le=8760, description="Lookback hours (default 30 days)"),
    currency: Optional[str] = Query(None, description="Override currency (e.g. ILS, EUR, GBP)"),
    user: User = Depends(require_permission("metering", "read")),
):
    """
    Export chargeback report as CSV.
    Uses the unified metering_pricing table (flavor, storage_gb, snapshot_gb,
    snapshot_op, restore, volume, network, public_ip, os_license) to compute per-tenant costs.
    Counts ACTUAL volumes, networks, subnets, routers, floating IPs, and ports
    from inventory tables — not just VM-based approximations.
    Falls back to metering_config rates when no pricing entry matches.
    Currency is taken from pricing configuration (metering_config.cost_currency).
    """
    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Load global fallback config
            cur.execute("SELECT * FROM metering_config WHERE id = 1")
            cfg = cur.fetchone()

        fb_vcpu = float(cfg["cost_per_vcpu_hour"]) if cfg else 0
        fb_ram = float(cfg["cost_per_gb_ram_hour"]) if cfg else 0
        fb_storage = float(cfg["cost_per_gb_storage_month"]) if cfg else 0
        fb_snap = float(cfg["cost_per_snapshot_gb_month"]) if cfg else 0
        fb_api = float(cfg["cost_per_api_call"]) if cfg else 0

        # Currency resolution: query param → first pricing row → metering_config → USD
        config_currency = cfg["cost_currency"] if cfg else "USD"

        # Load all pricing entries into a lookup
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM metering_pricing ORDER BY id")
            pricing_rows = cur.fetchall()

        # If no override, try to pick currency from first pricing entry, else config
        if not currency:
            if pricing_rows:
                currency = pricing_rows[0].get("currency") or config_currency
            else:
                currency = config_currency

        # Build pricing lookup by category
        flavor_pricing = {}   # item_name -> {cost_per_hour, cost_per_month, disk_cost_per_gb}
        cat_pricing = {}      # category -> {cost_per_hour, cost_per_month, ...}
        for p in pricing_rows:
            entry = {
                "cost_per_hour": float(p.get("cost_per_hour") or 0),
                "cost_per_month": float(p.get("cost_per_month") or 0),
            }
            if p["category"] == "flavor":
                entry["disk_cost_per_gb"] = float(p.get("disk_cost_per_gb") or 0)
                flavor_pricing[p["item_name"]] = entry
            else:
                cat_pricing[p["category"]] = entry

        # Helper to get per-hour cost for a category
        def cat_hourly(cat: str, fallback: float = 0.0) -> float:
            e = cat_pricing.get(cat)
            if e:
                if e["cost_per_hour"] > 0:
                    return e["cost_per_hour"]
                if e["cost_per_month"] > 0:
                    return e["cost_per_month"] / 730.0
            return fallback

        # Helper to get per-month cost for a category
        def cat_monthly(cat: str, fallback: float = 0.0) -> float:
            e = cat_pricing.get(cat)
            if e:
                if e["cost_per_month"] > 0:
                    return e["cost_per_month"]
                if e["cost_per_hour"] > 0:
                    return e["cost_per_hour"] * 730.0
            return fallback

        storage_per_gb_hr = cat_hourly("storage_gb", fb_storage / 730.0 if fb_storage else 0)
        snapshot_per_gb_hr = cat_hourly("snapshot_gb", fb_snap / 730.0 if fb_snap else 0)
        snapshot_op_cost = cat_monthly("snapshot_op")       # per snapshot operation
        restore_per_op = cat_monthly("restore")             # per restore operation
        volume_per_month = cat_monthly("volume")            # per volume per month
        network_per_month = cat_monthly("network")          # per network per month
        public_ip_per_month = cat_monthly("public_ip")      # per floating IP per month

        # OS license pricing (category = 'os_license', item_name = os_distro e.g. 'windows')
        os_license_pricing = {}  # item_name (lowercase os_distro) -> {cost_per_hour, cost_per_month}
        for p in pricing_rows:
            if p["category"] == "os_license":
                os_license_pricing[p["item_name"].lower()] = {
                    "cost_per_hour": float(p.get("cost_per_hour") or 0),
                    "cost_per_month": float(p.get("cost_per_month") or 0),
                }

        # Time-based filters for metering tables
        where_r = ["collected_at > now() - interval '%s hours'"]
        params_r: list = [hours]
        if project:
            where_r.append("project_name = %s")
            params_r.append(project)
        if domain:
            where_r.append("domain = %s")
            params_r.append(domain)
        where_clause = " AND ".join(where_r)

        # Inventory filters for inventory tables (project/domain via join)
        inv_where_parts = []
        inv_params: list = []
        if project:
            inv_where_parts.append("p.name = %s")
            inv_params.append(project)
        if domain:
            inv_where_parts.append("d.name = %s")
            inv_params.append(domain)
        inv_where = (" AND " + " AND ".join(inv_where_parts)) if inv_where_parts else ""

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # ── Per-tenant resource aggregation (latest per VM) ──
            cur.execute(f"""
                WITH latest AS (
                    SELECT DISTINCT ON (vm_id) *
                    FROM metering_resources
                    WHERE {where_clause}
                    ORDER BY vm_id, collected_at DESC
                )
                SELECT
                    COALESCE(project_name, 'unknown') AS tenant,
                    COALESCE(domain, '')               AS domain,
                    COUNT(*)                           AS vm_count,
                    ROUND(AVG(vcpus_allocated)::numeric, 1)          AS avg_vcpus,
                    ROUND(AVG(ram_allocated_mb)::numeric / 1024, 2)  AS avg_ram_gb,
                    ROUND(SUM(COALESCE(disk_allocated_gb,0))::numeric, 1) AS total_disk_gb,
                    ROUND(AVG(cpu_usage_percent)::numeric, 1)        AS avg_cpu_pct,
                    ROUND(AVG(ram_usage_percent)::numeric, 1)        AS avg_ram_pct,
                    array_agg(DISTINCT flavor) FILTER (WHERE flavor IS NOT NULL) AS flavors
                FROM latest
                GROUP BY project_name, domain
                ORDER BY project_name
            """, params_r)
            resource_rows = cur.fetchall()

            # ── Snapshot storage per tenant (from metering_snapshots) ──
            cur.execute(f"""
                WITH latest_snap AS (
                    SELECT DISTINCT ON (snapshot_id) *
                    FROM metering_snapshots s
                    WHERE {where_clause}
                    ORDER BY snapshot_id, collected_at DESC
                )
                SELECT
                    COALESCE(project_name, 'unknown') AS tenant,
                    COALESCE(domain, '')               AS domain,
                    COUNT(*)                           AS snapshot_count,
                    COALESCE(SUM(size_gb), 0)          AS total_snap_gb
                FROM latest_snap
                GROUP BY project_name, domain
            """, params_r)
            snap_rows = {(r["tenant"], r["domain"]): r for r in cur.fetchall()}

            # ── Restore count per tenant (from metering_restores) ──
            cur.execute(f"""
                SELECT
                    COALESCE(project_name, 'unknown') AS tenant,
                    COALESCE(domain, '')               AS domain,
                    COUNT(*) AS restore_count
                FROM metering_restores
                WHERE {where_clause}
                GROUP BY project_name, domain
            """, params_r)
            restore_rows = {(r["tenant"], r["domain"]): r for r in cur.fetchall()}

            # ── Actual volume counts per tenant (from inventory) ──
            try:
                cur.execute(f"""
                    SELECT
                        COALESCE(p.name, 'unknown') AS tenant,
                        COALESCE(d.name, '')         AS domain,
                        COUNT(*)                     AS volume_count,
                        COALESCE(SUM(v.size_gb), 0)  AS total_volume_gb
                    FROM volumes v
                    LEFT JOIN projects p ON v.project_id = p.id
                    LEFT JOIN domains d  ON p.domain_id = d.id
                    WHERE 1=1 {inv_where}
                    GROUP BY p.name, d.name
                """, inv_params)
                vol_rows = {(r["tenant"], r["domain"]): r for r in cur.fetchall()}
            except Exception:
                conn.rollback()
                vol_rows = {}

            # ── Actual network counts per tenant (from inventory) ──
            try:
                cur.execute(f"""
                    SELECT
                        COALESCE(p.name, 'unknown') AS tenant,
                        COALESCE(d.name, '')         AS domain,
                        COUNT(DISTINCT n.id)         AS network_count,
                        COUNT(DISTINCT sub.id)       AS subnet_count,
                        COUNT(DISTINCT r.id)         AS router_count
                    FROM networks n
                    LEFT JOIN projects p  ON n.project_id = p.id
                    LEFT JOIN domains d   ON p.domain_id = d.id
                    LEFT JOIN subnets sub ON sub.network_id = n.id
                    LEFT JOIN routers r   ON r.project_id = p.id
                    WHERE 1=1 {inv_where}
                    GROUP BY p.name, d.name
                """, inv_params)
                net_rows = {(r["tenant"], r["domain"]): r for r in cur.fetchall()}
            except Exception:
                conn.rollback()
                net_rows = {}

            # ── Actual floating IP counts per tenant (from inventory) ──
            try:
                cur.execute(f"""
                    SELECT
                        COALESCE(p.name, 'unknown') AS tenant,
                        COALESCE(d.name, '')         AS domain,
                        COUNT(*)                     AS floating_ip_count
                    FROM floating_ips fi
                    LEFT JOIN projects p ON fi.project_id = p.id
                    LEFT JOIN domains d  ON p.domain_id = d.id
                    WHERE 1=1 {inv_where}
                    GROUP BY p.name, d.name
                """, inv_params)
                fip_rows = {(r["tenant"], r["domain"]): r for r in cur.fetchall()}
            except Exception:
                conn.rollback()
                fip_rows = {}

            # ── OS license counts per tenant (from servers inventory) ──
            os_license_rows: dict = {}
            if os_license_pricing:
                try:
                    cur.execute(f"""
                        SELECT
                            COALESCE(p.name, 'unknown') AS tenant,
                            COALESCE(d.name, '')         AS domain,
                            LOWER(COALESCE(s.os_distro, 'unknown')) AS os_distro,
                            COUNT(*) AS vm_count
                        FROM servers s
                        LEFT JOIN projects p ON s.project_id = p.id
                        LEFT JOIN domains d  ON p.domain_id = d.id
                        WHERE s.os_distro IS NOT NULL AND s.os_distro != ''
                        {inv_where.replace('1=1 ', '') if inv_where else ''}
                        GROUP BY p.name, d.name, LOWER(s.os_distro)
                    """, inv_params)
                    for row in cur.fetchall():
                        key = (row["tenant"], row["domain"])
                        if key not in os_license_rows:
                            os_license_rows[key] = {}
                        os_license_rows[key][row["os_distro"]] = row["vm_count"]
                except Exception:
                    conn.rollback()

        # ── Build a unified set of all tenants across all data sources ──
        all_tenants = set()
        for r in resource_rows:
            all_tenants.add((r["tenant"], r["domain"]))
        for key in list(snap_rows) + list(restore_rows) + list(vol_rows) + list(net_rows) + list(fip_rows):
            all_tenants.add(key)

        # Pre-index resource rows for lookup
        resource_by_tenant = {(r["tenant"], r["domain"]): r for r in resource_rows}

        # Derive period_months for per-month rates
        interval_hours = hours
        period_months = hours / 730.0  # 730 hours ≈ 1 month

        # Calculate chargeback
        chargeback = []
        for (tenant, dom) in sorted(all_tenants):
            r = resource_by_tenant.get((tenant, dom), {})
            vm_count = int(r.get("vm_count") or 0)
            avg_vcpus = float(r.get("avg_vcpus") or 0)
            avg_ram_gb = float(r.get("avg_ram_gb") or 0)
            total_disk_gb = float(r.get("total_disk_gb") or 0)

            # ── Compute cost: flavor pricing first, fallback to vCPU/RAM rates ──
            flavors_list = r.get("flavors") or []
            flavor_cost_hr = 0.0
            ephemeral_disk_cost = 0.0
            matched_flavors = 0
            for fl in flavors_list:
                if fl in flavor_pricing:
                    fp = flavor_pricing[fl]
                    hr = fp["cost_per_hour"] if fp["cost_per_hour"] > 0 else (fp["cost_per_month"] / 730.0 if fp["cost_per_month"] > 0 else 0)
                    flavor_cost_hr += hr
                    matched_flavors += 1
                    # Ephemeral disk cost per GB (if configured)
                    if fp.get("disk_cost_per_gb", 0) > 0:
                        # Look up disk size from the pricing entry's associated data
                        # The disk_gb is stored in flavor_pricing when synced
                        ephemeral_disk_cost += fp["disk_cost_per_gb"] * total_disk_gb / max(len(flavors_list), 1)

            if matched_flavors > 0:
                unmatched = vm_count - matched_flavors
                fallback_hr = (avg_vcpus * fb_vcpu + avg_ram_gb * fb_ram) if unmatched > 0 else 0
                compute_cost = (flavor_cost_hr + fallback_hr) * interval_hours
            else:
                compute_cost = (avg_vcpus * fb_vcpu + avg_ram_gb * fb_ram) * interval_hours

            # Add ephemeral disk cost (monthly)
            compute_cost += ephemeral_disk_cost * period_months

            # ── Storage cost (disk GB) ──
            storage_cost = total_disk_gb * storage_per_gb_hr * interval_hours

            # ── Snapshot cost (storage + per-operation) ──
            snap_info = snap_rows.get((tenant, dom), {})
            snap_count = int(snap_info.get("snapshot_count", 0))
            snap_gb = float(snap_info.get("total_snap_gb", 0))
            snap_storage_cost = snap_gb * snapshot_per_gb_hr * interval_hours
            snap_op_cost = snap_count * snapshot_op_cost * period_months
            snap_cost = snap_storage_cost + snap_op_cost

            # ── Restore cost ──
            rest_info = restore_rows.get((tenant, dom), {})
            restore_count = int(rest_info.get("restore_count", 0))
            restore_cost = restore_count * restore_per_op

            # ── Volume cost (actual count from inventory) ──
            vol_info = vol_rows.get((tenant, dom), {})
            volume_count = int(vol_info.get("volume_count", 0))
            volume_gb = float(vol_info.get("total_volume_gb", 0))
            vol_cost = volume_count * volume_per_month * period_months

            # ── Network cost (actual count from inventory) ──
            net_info = net_rows.get((tenant, dom), {})
            network_count = int(net_info.get("network_count", 0))
            subnet_count = int(net_info.get("subnet_count", 0))
            router_count = int(net_info.get("router_count", 0))
            net_cost = network_count * network_per_month * period_months

            # ── Public IP cost (actual floating IPs from inventory) ──
            fip_info = fip_rows.get((tenant, dom), {})
            floating_ip_count = int(fip_info.get("floating_ip_count", 0))
            public_ip_cost = floating_ip_count * public_ip_per_month * period_months

            # ── OS License cost (per VM with matching os_distro) ──
            os_license_cost = 0.0
            os_licensed_vms = 0
            os_distro_counts = os_license_rows.get((tenant, dom), {})
            for os_name, os_count in os_distro_counts.items():
                lp = os_license_pricing.get(os_name)
                if lp:
                    hr = lp["cost_per_hour"] if lp["cost_per_hour"] > 0 else (lp["cost_per_month"] / 730.0 if lp["cost_per_month"] > 0 else 0)
                    os_license_cost += hr * os_count * interval_hours
                    os_licensed_vms += os_count

            total_cost = (compute_cost + storage_cost + snap_cost +
                          restore_cost + vol_cost + net_cost + public_ip_cost + os_license_cost)

            chargeback.append({
                "Tenant / Project": tenant,
                "Domain": dom,
                "VM Count": vm_count,
                "Avg vCPUs": avg_vcpus,
                "Avg RAM (GB)": avg_ram_gb,
                "Total Disk (GB)": total_disk_gb,
                "Avg CPU Usage (%)": float(r.get("avg_cpu_pct") or 0),
                "Avg RAM Usage (%)": float(r.get("avg_ram_pct") or 0),
                f"Compute Cost ({currency})": round(compute_cost, 2),
                f"Storage Cost ({currency})": round(storage_cost, 2),
                "Snapshots": snap_count,
                "Snapshot GB": snap_gb,
                f"Snapshot Cost ({currency})": round(snap_cost, 2),
                "Restores": restore_count,
                f"Restore Cost ({currency})": round(restore_cost, 2),
                "Volumes": volume_count,
                "Volume GB": volume_gb,
                f"Volume Cost ({currency})": round(vol_cost, 2),
                "Networks": network_count,
                "Subnets": subnet_count,
                "Routers": router_count,
                f"Network Cost ({currency})": round(net_cost, 2),
                "Floating IPs": floating_ip_count,
                f"Public IP Cost ({currency})": round(public_ip_cost, 2),
                "OS Licensed VMs": os_licensed_vms,
                f"OS License Cost ({currency})": round(os_license_cost, 2),
                f"TOTAL Cost ({currency})": round(total_cost, 2),
                "Period (hours)": interval_hours,
                "Currency": currency,
            })

        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
        return _rows_to_csv(chargeback, f"chargeback_report_{ts}.csv")


# ---------------------------------------------------------------------------
# Chargeback Summary (JSON — for the UI Chargeback tab)
# ---------------------------------------------------------------------------

@router.get("/chargeback-summary")
async def get_chargeback_summary(
    hours: int = Query(720, ge=1, le=8760, description="Lookback hours (default 30 days)"),
    start_date: Optional[str] = Query(None, description="Custom start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="Custom end date (YYYY-MM-DD)"),
    period: Optional[str] = Query(None, description="Predefined period: 24h, 7d, 30d"),
    currency: Optional[str] = Query(None, description="Override display currency (e.g. ILS, EUR, GBP)"),
    domain: Optional[str] = Query(None),
    user: User = Depends(require_permission("metering", "read")),
):
    """
    JSON chargeback summary — total estimated cost per tenant for the UI Chargeback tab.
    Calculates costs for VMs, Storage, Snapshots, Networks, and Public IPs.
    Supports custom date ranges and predefined periods (24h, 7d, 30d).
    Returns one row per (domain, project_name) sorted by cost descending.
    """
    from datetime import datetime
    
    # Handle period and date range logic
    if period == "24h":
        hours = 24
    elif period == "7d":
        hours = 168  # 7 * 24
    elif period == "30d":
        hours = 720  # 30 * 24
    
    # Custom date range overrides hours parameter
    if start_date and end_date:
        start = datetime.fromisoformat(start_date + "T00:00:00")
        end = datetime.fromisoformat(end_date + "T23:59:59")
        hours = int((end - start).total_seconds() / 3600)
        since = start.replace(tzinfo=timezone.utc)
    else:
        since = datetime.now(timezone.utc) - timedelta(hours=hours)
    
    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM metering_config WHERE id = 1")
            cfg = cur.fetchone()
        fb_vcpu = float(cfg["cost_per_vcpu_hour"]) if cfg else 0
        fb_ram = float(cfg["cost_per_gb_ram_hour"]) if cfg else 0
        config_currency = cfg["cost_currency"] if cfg else "USD"

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM metering_pricing ORDER BY id")
            pricing_rows = cur.fetchall()

        if not currency:
            if pricing_rows:
                currency = pricing_rows[0].get("currency") or config_currency
            else:
                currency = config_currency

        # Build pricing lookups for all resource types
        flavor_pricing = {}
        storage_pricing = {"storage_per_gb_month": 0.0, "snapshot_per_gb_month": 0.0}
        network_pricing = {"network_per_month": 0.0, "public_ip_per_month": 0.0}
        
        for p in pricing_rows:
            if p["category"] == "flavor":
                flavor_pricing[p["item_name"]] = {
                    "cost_per_hour": float(p.get("cost_per_hour") or 0),
                }
            elif p["category"] == "storage_gb" and p["item_name"] in ["storage", "Storage (per GB)"]:
                storage_pricing["storage_per_gb_month"] = float(p.get("cost_per_month") or 0)
            elif p["category"] == "snapshot_gb" and p["item_name"] in ["snapshot", "Snapshot Storage (per GB)"]:
                storage_pricing["snapshot_per_gb_month"] = float(p.get("cost_per_month") or 0)
            elif p["category"] == "network" and p["item_name"] in ["network", "Network"]:
                network_pricing["network_per_month"] = float(p.get("cost_per_month") or 0)
            elif p["category"] == "public_ip" and p["item_name"] in ["public_ip", "Public IP"]:
                network_pricing["public_ip_per_month"] = float(p.get("cost_per_month") or 0)

        where_parts = ["collected_at > %s"]
        params: list = [since]
        if domain:
            where_parts.append("domain = %s")
            params.append(domain)

        # Get VMs with their latest metrics
        vm_sql = f"""
            SELECT DISTINCT ON (vm_id)
                vm_id, vm_name, project_name, domain,
                vcpus_allocated AS vcpus, ram_allocated_mb AS ram_mb, 
                flavor AS flavor_name, disk_allocated_gb, collected_at
            FROM metering_resources
            WHERE {' AND '.join(where_parts)}
            ORDER BY vm_id, collected_at DESC
        """
        
        # Get storage data (volumes and snapshots)
        storage_sql = f"""
            SELECT project_name, domain,
                   SUM(COALESCE(size_gb, 0)) as total_storage_gb
            FROM metering_snapshots ms
            WHERE {' AND '.join(where_parts)}
            GROUP BY project_name, domain
        """
        
        # Get network data (approximate - we'll use VM count as proxy for now)
        # In a real implementation, you'd query network tables
        
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(vm_sql, params)
            vms = cur.fetchall()
            
            cur.execute(storage_sql, params)
            storage_data = {(row["domain"], row["project_name"]): row for row in cur.fetchall()}

    tenant_totals: dict = {}
    vm_details: list = []
    period_months = hours / 730.0  # Convert hours to months for monthly pricing
    
    for vm in vms:
        key = (vm.get("domain") or "unknown", vm.get("project_name") or "unknown")
        
        # Calculate VM compute costs
        fp = flavor_pricing.get(vm.get("flavor_name") or "")
        if fp and fp["cost_per_hour"] > 0:
            compute_cost_per_hour = fp["cost_per_hour"]
            cost_breakdown = {"flavor": compute_cost_per_hour, "vcpu": 0.0, "ram": 0.0}
        else:
            vcpu_cost = (vm.get("vcpus") or 0) * fb_vcpu
            ram_cost = ((vm.get("ram_mb") or 0) / 1024) * fb_ram
            compute_cost_per_hour = vcpu_cost + ram_cost
            cost_breakdown = {"flavor": 0.0, "vcpu": vcpu_cost, "ram": ram_cost}

        # Calculate VM storage costs (allocated disk)
        vm_disk_gb = vm.get("disk_allocated_gb") or 0
        vm_storage_cost = vm_disk_gb * storage_pricing["storage_per_gb_month"] * period_months
        
        total_vm_cost = (compute_cost_per_hour * hours) + vm_storage_cost
        cost_breakdown["storage"] = vm_storage_cost
        
        # Add VM details for per-VM breakdown
        vm_details.append({
            "vm_id": vm.get("vm_id"),
            "vm_name": vm.get("vm_name"),
            "domain": vm.get("domain") or "unknown",
            "project_name": vm.get("project_name") or "unknown",
            "flavor_name": vm.get("flavor_name"),
            "vcpus": vm.get("vcpus") or 0,
            "ram_gb": round((vm.get("ram_mb") or 0) / 1024, 2),
            "disk_gb": vm_disk_gb,
            "compute_cost_per_hour": round(compute_cost_per_hour, 4),
            "storage_cost": round(vm_storage_cost, 2),
            "total_cost": round(total_vm_cost, 2),
            "cost_breakdown": cost_breakdown,
        })

        entry = tenant_totals.setdefault(key, {
            "domain": key[0],
            "project_name": key[1],
            "vm_count": 0,
            "total_vcpus": 0,
            "total_ram_gb": 0.0,
            "total_disk_gb": 0.0,
            "compute_cost": 0.0,
            "storage_cost": 0.0,
            "snapshot_cost": 0.0,
            "network_cost": 0.0,
            "estimated_cost": 0.0,
        })
        entry["vm_count"] += 1
        entry["total_vcpus"] += vm.get("vcpus") or 0
        entry["total_ram_gb"] += round((vm.get("ram_mb") or 0) / 1024, 2)
        entry["total_disk_gb"] += vm_disk_gb
        entry["compute_cost"] += compute_cost_per_hour * hours
        entry["storage_cost"] += vm_storage_cost
        entry["estimated_cost"] += total_vm_cost

    # Add snapshot and network costs to each tenant
    for key, entry in tenant_totals.items():
        # Snapshot costs from storage data
        storage_info = storage_data.get(key, {})
        snapshot_gb = storage_info.get("total_storage_gb") or 0
        snapshot_cost = snapshot_gb * storage_pricing["snapshot_per_gb_month"] * period_months
        entry["snapshot_cost"] = round(snapshot_cost, 2)
        
        # Network costs (simplified: 1 network per project + potential public IPs)
        # In production, you'd query actual network/IP usage
        network_cost = network_pricing["network_per_month"] * period_months
        public_ip_cost = network_pricing["public_ip_per_month"] * period_months * entry["vm_count"]  # Assume 1 IP per VM
        entry["network_cost"] = round(network_cost + public_ip_cost, 2)
        
        # Update total cost
        entry["estimated_cost"] += snapshot_cost + network_cost + public_ip_cost
    
    rows = sorted(tenant_totals.values(), key=lambda x: x["estimated_cost"], reverse=True)
    for r in rows:
        for cost_field in ["compute_cost", "storage_cost", "snapshot_cost", "network_cost", "estimated_cost"]:
            r[cost_field] = round(r[cost_field], 2)

    return {
        "currency": currency,
        "period_hours": hours,
        "period_description": period or f"Last {hours} hours",
        "start_date": start_date,
        "end_date": end_date,
        "tenants": rows,
        "vm_details": vm_details,
        "total_cost": round(sum(r["estimated_cost"] for r in rows), 2),
        "cost_breakdown": {
            "compute": round(sum(r["compute_cost"] for r in rows), 2),
            "storage": round(sum(r["storage_cost"] for r in rows), 2),
            "snapshots": round(sum(r["snapshot_cost"] for r in rows), 2),
            "network": round(sum(r["network_cost"] for r in rows), 2),
        },
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


# ---------------------------------------------------------------------------
# Per-VM Chargeback Details (JSON — detailed breakdown per VM)
# ---------------------------------------------------------------------------

@router.get("/chargeback-details")
async def get_chargeback_details(
    hours: int = Query(720, ge=1, le=8760, description="Lookback hours (default 30 days)"),
    currency: Optional[str] = Query(None, description="Override display currency"),
    domain: Optional[str] = Query(None, description="Filter by domain"),
    project: Optional[str] = Query(None, description="Filter by project"),
    user: User = Depends(require_permission("metering", "read")),
):
    """
    Detailed per-VM chargeback breakdown with cost attribution by usage type
    (compute, storage, snapshots, etc.) and currency selection.
    """
    with get_connection() as conn:
        # Get config and pricing
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM metering_config WHERE id = 1")
            cfg = cur.fetchone()
        fb_vcpu = float(cfg["cost_per_vcpu_hour"]) if cfg else 0
        fb_ram = float(cfg["cost_per_gb_ram_hour"]) if cfg else 0
        config_currency = cfg["cost_currency"] if cfg else "USD"

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM metering_pricing ORDER BY id")
            pricing_rows = cur.fetchall()

        # Determine currency
        if not currency:
            currency = pricing_rows[0].get("currency") if pricing_rows else config_currency

        # Build pricing lookup
        flavor_pricing = {}
        storage_pricing = {"storage_per_gb_month": 0.0, "snapshot_per_gb_month": 0.0}
        for p in pricing_rows:
            if p["category"] == "flavor":
                flavor_pricing[p["item_name"]] = {
                    "cost_per_hour": float(p.get("cost_per_hour") or 0),
                }
            elif p["category"] == "storage" and p["item_name"] == "storage":
                storage_pricing["storage_per_gb_month"] = float(p.get("cost_per_month") or 0)
            elif p["category"] == "storage" and p["item_name"] == "snapshot":
                storage_pricing["snapshot_per_gb_month"] = float(p.get("cost_per_month") or 0)

        # Get VM data
        since = datetime.now(timezone.utc) - timedelta(hours=hours)
        where_parts = ["mr.collected_at > %s"]
        params: list = [since]
        if domain:
            where_parts.append("mr.domain = %s")
            params.append(domain)
        if project:
            where_parts.append("mr.project_name = %s")
            params.append(project)

        sql = f"""
            SELECT DISTINCT ON (mr.vm_id)
                mr.vm_id, mr.vm_name, mr.project_name, mr.domain,
                mr.vcpus_allocated AS vcpus, mr.ram_allocated_mb AS ram_mb, 
                mr.flavor AS flavor_name, mr.storage_allocated_gb,
                mr.collected_at, s.status AS vm_status
            FROM metering_resources mr
            LEFT JOIN servers s ON s.id = mr.vm_id
            WHERE {' AND '.join(where_parts)}
            ORDER BY mr.vm_id, mr.collected_at DESC
        """
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql, params)
            vms = cur.fetchall()

        # Get storage and snapshot data per VM
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT 
                    v.server_id AS vm_id,
                    SUM(v.size_gb) AS total_volume_gb,
                    COUNT(v.id) AS volume_count
                FROM volumes v 
                WHERE v.server_id IS NOT NULL
                GROUP BY v.server_id
            """)
            storage_by_vm = {r["vm_id"]: r for r in cur.fetchall()}

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT 
                    v.server_id AS vm_id,
                    COUNT(s.id) AS snapshot_count,
                    SUM(s.size_gb) AS total_snapshot_gb
                FROM volumes v
                LEFT JOIN snapshots s ON s.volume_id = v.id
                WHERE v.server_id IS NOT NULL AND s.id IS NOT NULL
                GROUP BY v.server_id
            """)
            snapshots_by_vm = {r["vm_id"]: r for r in cur.fetchall()}

        # Calculate detailed breakdown per VM
        vm_details = []
        period_months = hours / 730.0  # 730 hours ≈ 1 month

        for vm in vms:
            vm_id = vm.get("vm_id")
            storage_info = storage_by_vm.get(vm_id, {})
            snapshot_info = snapshots_by_vm.get(vm_id, {})
            
            # Compute costs
            fp = flavor_pricing.get(vm.get("flavor_name") or "")
            if fp and fp["cost_per_hour"] > 0:
                compute_cost_per_hour = fp["cost_per_hour"]
                cost_attribution = {"compute": "flavor"}
            else:
                vcpu_cost = (vm.get("vcpus") or 0) * fb_vcpu
                ram_cost = ((vm.get("ram_mb") or 0) / 1024) * fb_ram
                compute_cost_per_hour = vcpu_cost + ram_cost
                cost_attribution = {"compute": "vcpu_ram"}

            # Storage costs
            volume_gb = storage_info.get("total_volume_gb") or vm.get("storage_allocated_gb") or 0
            storage_cost = volume_gb * storage_pricing["storage_per_gb_month"] * period_months
            
            # Snapshot costs
            snapshot_gb = snapshot_info.get("total_snapshot_gb") or 0
            snapshot_cost = snapshot_gb * storage_pricing["snapshot_per_gb_month"] * period_months
            
            # Total costs
            total_compute_cost = compute_cost_per_hour * hours
            total_cost = total_compute_cost + storage_cost + snapshot_cost

            vm_details.append({
                "vm_id": vm_id,
                "vm_name": vm.get("vm_name"),
                "vm_status": vm.get("vm_status"),
                "domain": vm.get("domain") or "unknown",
                "project_name": vm.get("project_name") or "unknown",
                "flavor_name": vm.get("flavor_name"),
                "vcpus": vm.get("vcpus") or 0,
                "ram_gb": round((vm.get("ram_mb") or 0) / 1024, 2),
                "storage_gb": round(volume_gb, 2),
                "volume_count": storage_info.get("volume_count") or 0,
                "snapshot_gb": round(snapshot_gb, 2),
                "snapshot_count": snapshot_info.get("snapshot_count") or 0,
                "costs": {
                    "compute_per_hour": round(compute_cost_per_hour, 4),
                    "compute_total": round(total_compute_cost, 2),
                    "storage_total": round(storage_cost, 2),
                    "snapshot_total": round(snapshot_cost, 2),
                    "total": round(total_cost, 2),
                },
                "cost_attribution": cost_attribution,
            })

        # Sort by total cost descending
        vm_details.sort(key=lambda x: x["costs"]["total"], reverse=True)

        return {
            "currency": currency,
            "period_hours": hours,
            "period_months": round(period_months, 2),
            "vm_details": vm_details,
            "total_cost": round(sum(vm["costs"]["total"] for vm in vm_details), 2),
            "breakdown_totals": {
                "compute": round(sum(vm["costs"]["compute_total"] for vm in vm_details), 2),
                "storage": round(sum(vm["costs"]["storage_total"] for vm in vm_details), 2),
                "snapshots": round(sum(vm["costs"]["snapshot_total"] for vm in vm_details), 2),
            },
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }


# ---------------------------------------------------------------------------
# Tenant Growth (JSON — monthly VM counts per tenant)
# ---------------------------------------------------------------------------

@router.get("/tenant-growth")
async def get_tenant_growth(
    months: int = Query(6, ge=1, le=24, description="Number of months to look back"),
    domain: Optional[str] = Query(None),
    user: User = Depends(require_permission("metering", "read")),
):
    """
    Returns monthly VM counts per tenant for the last N months.
    Used by the Tenant Growth tab in the Metering section.
    """
    with get_connection() as conn:
        where_parts = [
            "collected_at > now() - (%s * interval '1 month')"
        ]
        params: list = [months]
        if domain:
            where_parts.append("domain = %s")
            params.append(domain)

        sql = f"""
            SELECT
                TO_CHAR(DATE_TRUNC('month', collected_at), 'YYYY-MM') AS month,
                domain,
                project_name,
                COUNT(DISTINCT vm_id) AS vm_count
            FROM metering_resources
            WHERE {' AND '.join(where_parts)}
            GROUP BY 1, 2, 3
            ORDER BY 1, 2, 3
        """
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()

    # Build sorted month list
    months_seen: list = sorted({r["month"] for r in rows})
    tenants_seen: list = sorted({
        f"{r['domain'] or 'unknown'}/{r['project_name'] or 'unknown'}"
        for r in rows
    })

    # Pivot: tenant label → {month: vm_count}
    pivot: dict = {t: {} for t in tenants_seen}
    for r in rows:
        label = f"{r['domain'] or 'unknown'}/{r['project_name'] or 'unknown'}"
        pivot[label][r["month"]] = r["vm_count"]

    series = [
        {
            "tenant": t,
            "data": [pivot[t].get(m, 0) for m in months_seen],
        }
        for t in tenants_seen
    ]

    return {
        "months": months_seen,
        "series": series,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


# ---------------------------------------------------------------------------
# Excel export — row per VM with summary row (chargeback)
# ---------------------------------------------------------------------------

def _build_chargeback_xlsx(rows: list, currency: str, summary: dict) -> bytes:
    """Build an Excel workbook: one row per VM, a summary sheet, styled."""
    if not _XLSX_AVAILABLE:
        raise RuntimeError("openpyxl not installed")

    wb = openpyxl.Workbook()

    # ── Sheet 1: VM Details ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "VM Details"

    hdr_fill = PatternFill("solid", fgColor="1D4ED8")
    hdr_font = Font(bold=True, color="FFFFFF")
    sum_fill = PatternFill("solid", fgColor="DBEAFE")
    sum_font = Font(bold=True, color="1E3A8A")

    vm_headers = [
        "Domain", "Project / Tenant", "VM Name", "VM ID",
        "vCPUs", "RAM (GB)", "Disk (GB)", "CPU Usage %", "RAM Usage %",
        "Flavor", f"Est. Compute Cost ({currency})",
    ]
    ws.append(vm_headers)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([
            r.get("domain", ""),
            r.get("project_name", ""),
            r.get("vm_name", ""),
            r.get("vm_id", ""),
            r.get("vcpus"),
            round((r.get("ram_mb") or 0) / 1024, 2),
            r.get("disk_gb"),
            r.get("cpu_usage_percent"),
            r.get("ram_usage_percent"),
            r.get("flavor_name", ""),
            r.get("estimated_cost", 0),
        ])

    # Summary row
    total_cost = sum(r.get("estimated_cost", 0) for r in rows)
    summary_row = [
        "TOTAL", "", f"{len(rows)} VMs", "",
        sum(r.get("vcpus") or 0 for r in rows), "", "", "", "", "",
        round(total_cost, 2),
    ]
    ws.append(summary_row)
    for cell in ws[ws.max_row]:
        cell.fill = sum_fill
        cell.font = sum_font

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # ── Sheet 2: Summary by Tenant ───────────────────────────────────────
    ws2 = wb.create_sheet("Tenant Summary")
    sum_headers = [
        "Domain", "Project / Tenant", "VM Count",
        "Total vCPUs", "Total RAM (GB)", f"Est. Cost ({currency})",
    ]
    ws2.append(sum_headers)
    for cell in ws2[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    from collections import defaultdict
    by_tenant: dict = defaultdict(lambda: {"vm_count": 0, "total_vcpus": 0, "total_ram_gb": 0.0, "total_cost": 0.0})
    for r in rows:
        key = (r.get("domain", ""), r.get("project_name", ""))
        by_tenant[key]["vm_count"] += 1
        by_tenant[key]["total_vcpus"] += r.get("vcpus") or 0
        by_tenant[key]["total_ram_gb"] += (r.get("ram_mb") or 0) / 1024
        by_tenant[key]["total_cost"] += r.get("estimated_cost", 0)

    for (domain, project), v in sorted(by_tenant.items()):
        ws2.append([domain, project, v["vm_count"], v["total_vcpus"],
                    round(v["total_ram_gb"], 2), round(v["total_cost"], 2)])

    ws2.append(["GRAND TOTAL", "", sum(v["vm_count"] for v in by_tenant.values()), "", "", round(total_cost, 2)])
    for cell in ws2[ws2.max_row]:
        cell.fill = sum_fill
        cell.font = sum_font

    for col in ws2.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/export/chargeback-excel")
async def export_chargeback_excel(
    project: Optional[str] = Query(None),
    domain: Optional[str] = Query(None),
    hours: int = Query(720, ge=1, le=8760),
    currency: Optional[str] = Query(None),
    user: User = Depends(require_permission("metering", "read")),
):
    """Export chargeback as an Excel file with one row per VM + a tenant summary sheet."""
    if not _XLSX_AVAILABLE:
        raise HTTPException(status_code=501, detail="openpyxl not installed")

    rows, resolved_currency = await _collect_vm_rows_for_export(project, domain, hours, currency)
    xlsx_bytes = _build_chargeback_xlsx(rows, resolved_currency, {})
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    filename = f"chargeback_vm_{ts}.xlsx"
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def _collect_vm_rows_for_export(
    project: Optional[str],
    domain: Optional[str],
    hours: int,
    currency: Optional[str],
) -> tuple:
    """Shared query: return (rows, currency) where rows has per-VM data + estimated_cost."""
    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM metering_config WHERE id = 1")
            cfg = cur.fetchone()
        fb_vcpu = float(cfg["cost_per_vcpu_hour"]) if cfg else 0
        fb_ram = float(cfg["cost_per_gb_ram_hour"]) if cfg else 0
        config_currency = (cfg["cost_currency"] if cfg else "USD") or "USD"

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM metering_pricing ORDER BY id")
            pricing_rows = cur.fetchall()

        if not currency:
            if pricing_rows:
                currency = pricing_rows[0].get("currency") or config_currency
            else:
                currency = config_currency

        flavor_pricing = {p["item_name"]: float(p.get("cost_per_hour") or 0)
                          for p in pricing_rows if p["category"] == "flavor"}

        where_parts = ["collected_at > now() - (%s * interval '1 hour')"]
        params: list = [hours]
        if project:
            where_parts.append("project_name = %s")
            params.append(project)
        if domain:
            where_parts.append("domain = %s")
            params.append(domain)

        sql = f"""
            SELECT DISTINCT ON (vm_id)
                vm_id, vm_name, project_name, domain,
                vcpus_allocated AS vcpus, ram_allocated_mb AS ram_mb,
                disk_allocated_gb AS disk_gb,
                cpu_usage_percent, ram_usage_percent, flavor AS flavor_name,
                collected_at
            FROM metering_resources
            WHERE {' AND '.join(where_parts)}
            ORDER BY vm_id, collected_at DESC
        """
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql, params)
            vms = cur.fetchall()

    result = []
    for vm in vms:
        fp_cost = flavor_pricing.get(vm.get("flavor_name") or "")
        if fp_cost and fp_cost > 0:
            cost = fp_cost * hours
        else:
            cost = ((vm.get("vcpus") or 0) * fb_vcpu + (vm.get("ram_mb") or 0) / 1024 * fb_ram) * hours
        row = dict(vm)
        row["estimated_cost"] = round(cost, 2)
        result.append(row)

    return result, currency


# ---------------------------------------------------------------------------
# PDF export — chargeback summary
# ---------------------------------------------------------------------------

@router.get("/export/chargeback-pdf")
async def export_chargeback_pdf(
    project: Optional[str] = Query(None),
    domain: Optional[str] = Query(None),
    hours: int = Query(720, ge=1, le=8760),
    currency: Optional[str] = Query(None),
    user: User = Depends(require_permission("metering", "read")),
):
    """Export chargeback per-tenant summary as PDF."""
    if not _PDF_AVAILABLE:
        raise HTTPException(status_code=501, detail="reportlab not installed")

    rows, resolved_currency = await _collect_vm_rows_for_export(project, domain, hours, currency)

    # Aggregate by tenant
    from collections import defaultdict
    by_tenant: dict = defaultdict(lambda: {"vm_count": 0, "total_vcpus": 0, "total_ram_gb": 0.0, "total_cost": 0.0})
    for r in rows:
        key = (r.get("domain", ""), r.get("project_name", ""))
        by_tenant[key]["vm_count"] += 1
        by_tenant[key]["total_vcpus"] += r.get("vcpus") or 0
        by_tenant[key]["total_ram_gb"] += (r.get("ram_mb") or 0) / 1024
        by_tenant[key]["total_cost"] += r.get("estimated_cost", 0)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []

    ts_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    elements.append(Paragraph(f"Chargeback Report — {ts_str}", styles["Title"]))
    elements.append(Paragraph(f"Period: {hours // 24} days | Currency: {resolved_currency}", styles["Normal"]))
    elements.append(Spacer(1, 6*mm))

    # Table header
    data = [["Domain", "Project / Tenant", "VMs", "vCPUs", "RAM (GB)", f"Est. Cost ({resolved_currency})"]]
    total_cost = 0.0
    for (domain_v, project_v), v in sorted(by_tenant.items()):
        cost_val = round(v["total_cost"], 2)
        total_cost += cost_val
        data.append([
            domain_v, project_v,
            str(v["vm_count"]), str(v["total_vcpus"]),
            f"{v['total_ram_gb']:.1f}", f"{cost_val:,.2f}",
        ])
    data.append(["TOTAL", "", str(sum(v["vm_count"] for v in by_tenant.values())),
                 "", "", f"{total_cost:,.2f}"])

    col_widths = [55*mm, 70*mm, 20*mm, 20*mm, 25*mm, 40*mm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1D4ED8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("BACKGROUND", (0, -1), (-1, -1), rl_colors.HexColor("#DBEAFE")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [rl_colors.white, rl_colors.HexColor("#F1F5F9")]),
        ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#CBD5E1")),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    doc.build(elements)

    ts_file = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="chargeback_{ts_file}.pdf"'},
    )


# ---------------------------------------------------------------------------
# Email export — send chargeback as attachment
# ---------------------------------------------------------------------------

class ExportEmailRequest(BaseModel):
    to_email: str = Field(..., description="Recipient email address")
    format: str = Field("excel", description="excel | pdf | csv")
    project: Optional[str] = None
    domain: Optional[str] = None
    hours: int = Field(720, ge=1, le=8760)
    currency: Optional[str] = None


@router.post("/export/send-email")
async def send_chargeback_email(
    req: ExportEmailRequest,
    user: User = Depends(require_permission("metering", "read")),
):
    """Send a chargeback report as an email attachment."""
    import smtplib
    from email.mime.multipart import MIMEMultipart as _MIMEMulti
    from email.mime.text import MIMEText as _MIMEText
    from email.mime.base import MIMEBase as _MIMEBase
    from email import encoders as _encoders

    if not SMTP_ENABLED or not SMTP_HOST:
        raise HTTPException(status_code=503, detail="SMTP is not configured on this server")

    rows, resolved_currency = await _collect_vm_rows_for_export(
        req.project, req.domain, req.hours, req.currency
    )
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")

    if req.format == "excel":
        if not _XLSX_AVAILABLE:
            raise HTTPException(status_code=501, detail="openpyxl not installed")
        content = _build_chargeback_xlsx(rows, resolved_currency, {})
        filename = f"chargeback_{ts}.xlsx"
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif req.format == "pdf":
        if not _PDF_AVAILABLE:
            raise HTTPException(status_code=501, detail="reportlab not installed")
        # Reuse PDF endpoint logic inline
        from collections import defaultdict
        by_tenant: dict = defaultdict(lambda: {"vm_count": 0, "total_vcpus": 0, "total_ram_gb": 0.0, "total_cost": 0.0})
        for r in rows:
            key = (r.get("domain", ""), r.get("project_name", ""))
            by_tenant[key]["vm_count"] += 1
            by_tenant[key]["total_vcpus"] += r.get("vcpus") or 0
            by_tenant[key]["total_ram_gb"] += (r.get("ram_mb") or 0) / 1024
            by_tenant[key]["total_cost"] += r.get("estimated_cost", 0)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=15*mm, rightMargin=15*mm,
                                topMargin=15*mm, bottomMargin=15*mm)
        styles = getSampleStyleSheet()
        elements = []
        ts_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        elements.append(Paragraph(f"Chargeback Report — {ts_str}", styles["Title"]))
        elements.append(Paragraph(f"Period: {req.hours // 24} days | Currency: {resolved_currency}", styles["Normal"]))
        elements.append(Spacer(1, 6*mm))
        data = [["Domain", "Project / Tenant", "VMs", "vCPUs", "RAM (GB)", f"Est. Cost ({resolved_currency})"]]
        total_cost = 0.0
        for (d_, p_), v in sorted(by_tenant.items()):
            c_ = round(v["total_cost"], 2)
            total_cost += c_
            data.append([d_, p_, str(v["vm_count"]), str(v["total_vcpus"]), f"{v['total_ram_gb']:.1f}", f"{c_:,.2f}"])
        data.append(["TOTAL", "", str(sum(v["vm_count"] for v in by_tenant.values())), "", "", f"{total_cost:,.2f}"])
        col_widths = [55*mm, 70*mm, 20*mm, 20*mm, 25*mm, 40*mm]
        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1D4ED8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [rl_colors.white, rl_colors.HexColor("#F1F5F9")]),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), rl_colors.HexColor("#DBEAFE")),
            ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#CBD5E1")),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ]))
        elements.append(tbl)
        doc.build(elements)
        content = buf.getvalue()
        filename = f"chargeback_{ts}.pdf"
        mime_type = "application/pdf"
    else:
        # CSV
        out = io.StringIO()
        writer = csv.DictWriter(out, fieldnames=["domain", "project_name", "vm_name", "vm_id", "vcpus", "ram_mb", "flavor_name", f"estimated_cost_{resolved_currency}"])
        writer.writeheader()
        for r in rows:
            row_dict = {k: r.get(k) for k in ["domain", "project_name", "vm_name", "vm_id", "vcpus", "ram_mb", "flavor_name"]}
            row_dict[f"estimated_cost_{resolved_currency}"] = r.get("estimated_cost", 0)
            writer.writerow(row_dict)
        content = out.getvalue().encode("utf-8")
        filename = f"chargeback_{ts}.csv"
        mime_type = "text/csv"

    # Build email with attachment
    msg = _MIMEMulti()
    msg["From"] = f"{SMTP_FROM_NAME} <{SMTP_FROM_ADDRESS}>"
    msg["To"] = req.to_email
    msg["Subject"] = f"Chargeback Report — {datetime.now(timezone.utc).strftime('%Y-%m-%d')} ({resolved_currency})"

    body = _MIMEText(
        f"<p>Please find the chargeback report attached (<strong>{filename}</strong>).</p>"
        f"<p>Period: {req.hours // 24} days | Currency: {resolved_currency}</p>",
        "html",
    )
    msg.attach(body)

    part = _MIMEBase("application", "octet-stream")
    part.set_payload(content if isinstance(content, bytes) else content)
    _encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    part.add_header("Content-Type", mime_type)
    msg.attach(part)

    try:
        import ssl as _ssl
        if SMTP_USE_TLS:
            ctx = _ssl.create_default_context()
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=ctx, timeout=30) as server:
                if SMTP_USERNAME:
                    server.login(SMTP_USERNAME, SMTP_PASSWORD or "")
                server.sendmail(SMTP_FROM_ADDRESS, [req.to_email], msg.as_string())
        else:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as server:
                if SMTP_USERNAME:
                    server.login(SMTP_USERNAME, SMTP_PASSWORD or "")
                server.sendmail(SMTP_FROM_ADDRESS, [req.to_email], msg.as_string())
    except Exception as exc:
        logger.error("Failed to send metering report email: %s", exc)
        raise HTTPException(status_code=500, detail="Failed to send email")

    return {"ok": True, "message": f"Report sent to {req.to_email}", "filename": filename}


# ---------------------------------------------------------------------------
# Pricing CRUD (metering_pricing – unified pricing table)
# ---------------------------------------------------------------------------
# Categories: flavor, storage_gb, snapshot_gb, snapshot_op, restore, volume, network, public_ip, os_license, custom
# Each row: category, item_name, unit, cost_per_hour, cost_per_month, currency

class PricingItemCreate(BaseModel):
    category: str = Field(..., description="flavor|storage_gb|snapshot_gb|snapshot_op|restore|volume|network|public_ip|custom")
    item_name: str = Field(..., min_length=1, max_length=200)
    unit: str = Field("per hour", max_length=50, description="e.g. per hour, per GB/month, per operation")
    cost_per_hour: float = Field(0, ge=0)
    cost_per_month: float = Field(0, ge=0)
    currency: str = Field("USD", max_length=10)
    notes: Optional[str] = None
    # Flavor-specific fields (populated from flavors table)
    vcpus: Optional[int] = None
    ram_gb: Optional[float] = None
    disk_gb: Optional[float] = None
    disk_cost_per_gb: Optional[float] = Field(None, ge=0, description="Cost per GB for ephemeral disk in flavors")


class PricingItemUpdate(BaseModel):
    category: Optional[str] = None
    item_name: Optional[str] = Field(None, min_length=1, max_length=200)
    unit: Optional[str] = Field(None, max_length=50)
    cost_per_hour: Optional[float] = Field(None, ge=0)
    cost_per_month: Optional[float] = Field(None, ge=0)
    currency: Optional[str] = Field(None, max_length=10)
    notes: Optional[str] = None
    vcpus: Optional[int] = None
    ram_gb: Optional[float] = None
    disk_gb: Optional[float] = None
    disk_cost_per_gb: Optional[float] = Field(None, ge=0)


@router.get("/pricing")
async def list_pricing(
    user: User = Depends(require_permission("metering", "read")),
):
    """List all pricing entries grouped by category."""
    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM metering_pricing ORDER BY category, item_name")
            rows = cur.fetchall()
    result = []
    for r in rows:
        d = dict(r)
        for k in list(d.keys()):
            if hasattr(d[k], "isoformat"):
                d[k] = d[k].isoformat()
            elif hasattr(d[k], "as_tuple"):
                d[k] = float(d[k])
        result.append(d)
    return result


@router.post("/pricing", status_code=201)
async def create_pricing(
    body: PricingItemCreate,
    user: User = Depends(require_permission("metering", "write")),
):
    """Create a new pricing entry."""
    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Check for duplicate item_name in other categories (cross-category overlap)
            if body.category == "custom":
                cur.execute(
                    "SELECT category FROM metering_pricing WHERE item_name = %s AND category != 'custom' LIMIT 1",
                    (body.item_name,),
                )
                existing = cur.fetchone()
                if existing:
                    raise HTTPException(
                        status_code=409,
                        detail=f"An item named '{body.item_name}' already exists in category '{existing['category']}'. Use a different name for custom items.",
                    )
            try:
                cur.execute("""
                    INSERT INTO metering_pricing
                        (category, item_name, unit, cost_per_hour, cost_per_month, currency, notes, vcpus, ram_gb, disk_gb, disk_cost_per_gb)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING *
                """, (body.category, body.item_name, body.unit, body.cost_per_hour,
                      body.cost_per_month, body.currency, body.notes,
                      body.vcpus, body.ram_gb, body.disk_gb, body.disk_cost_per_gb or 0))
                row = cur.fetchone()
            except Exception as e:
                conn.rollback()
                if "uq_pricing_category_name" in str(e):
                    raise HTTPException(status_code=409, detail=f"A pricing entry for '{body.item_name}' in category '{body.category}' already exists.")
                raise
    d = dict(row)
    for k in list(d.keys()):
        if hasattr(d[k], "isoformat"):
            d[k] = d[k].isoformat()
        elif hasattr(d[k], "as_tuple"):
            d[k] = float(d[k])
    return d


@router.put("/pricing/{pricing_id}")
async def update_pricing(
    pricing_id: int,
    body: PricingItemUpdate,
    user: User = Depends(require_permission("metering", "write")),
):
    """Update a pricing entry."""
    fields = []
    values = []
    for field_name, value in body.dict(exclude_unset=True).items():
        fields.append(f"{field_name} = %s")
        values.append(value)
    if not fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    fields.append("updated_at = now()")
    values.append(pricing_id)

    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                f"UPDATE metering_pricing SET {', '.join(fields)} WHERE id = %s RETURNING *",
                values,
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Pricing entry not found")
    d = dict(row)
    for k in list(d.keys()):
        if hasattr(d[k], "isoformat"):
            d[k] = d[k].isoformat()
        elif hasattr(d[k], "as_tuple"):
            d[k] = float(d[k])
    return d


@router.delete("/pricing/{pricing_id}")
async def delete_pricing(
    pricing_id: int,
    user: User = Depends(require_permission("metering", "write")),
):
    """Delete a pricing entry."""
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM metering_pricing WHERE id = %s", (pricing_id,))
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="Pricing entry not found")
    return {"detail": "Deleted"}


@router.post("/pricing/sync-flavors")
async def sync_flavors_to_pricing(
    user: User = Depends(require_permission("metering", "write")),
):
    """
    Sync flavors from live OpenStack into the flavors table AND metering_pricing.
    1. Fetches live flavors from Nova via pf9_control.
    2. Updates the flavors DB table: inserts new, removes deleted.
    3. Adds new flavor pricing entries (cost defaults to 0).
    4. Removes stale pricing entries for flavors that no longer exist.
    """
    from cluster_registry import get_registry

    try:
        client = get_registry().get_default_region()
        live_flavors = client.list_flavors()
    except Exception as e:
        logger.error("Failed to fetch live flavors from OpenStack: %s", e)
        raise HTTPException(status_code=502, detail=f"Cannot reach OpenStack Nova to list flavors: {e}")

    live_names = {f["name"] for f in live_flavors}

    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # ---- 1. Sync the flavors table ----
            # Upsert live flavors
            for f in live_flavors:
                vcpus = f.get("vcpus", 0) or 0
                ram_mb = f.get("ram", 0) or 0
                disk_gb = f.get("disk", 0) or 0
                ephemeral_gb = f.get("OS-FLV-EXT-DATA:ephemeral", 0) or 0
                swap_mb = f.get("swap", 0) or 0
                is_public = f.get("os-flavor-access:is_public", True)
                cur.execute("""
                    INSERT INTO flavors (id, name, vcpus, ram_mb, disk_gb, ephemeral_gb, swap_mb, is_public)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (id) DO UPDATE SET
                        name = EXCLUDED.name,
                        vcpus = EXCLUDED.vcpus,
                        ram_mb = EXCLUDED.ram_mb,
                        disk_gb = EXCLUDED.disk_gb,
                        ephemeral_gb = EXCLUDED.ephemeral_gb,
                        swap_mb = EXCLUDED.swap_mb,
                        is_public = EXCLUDED.is_public
                """, (f["id"], f["name"], vcpus, ram_mb, disk_gb, ephemeral_gb, swap_mb, is_public))

            # Remove flavors that no longer exist in OpenStack
            if live_names:
                cur.execute("SELECT name FROM flavors")
                db_flavor_names = {r["name"] for r in cur.fetchall()}
                stale_flavor_names = db_flavor_names - live_names
                if stale_flavor_names:
                    cur.execute("DELETE FROM flavors WHERE name = ANY(%s)", (list(stale_flavor_names),))
            else:
                stale_flavor_names = set()

            # ---- 2. Sync metering_pricing ----
            cur.execute("SELECT item_name FROM metering_pricing WHERE category = 'flavor'")
            existing_pricing = {r["item_name"] for r in cur.fetchall()}

            # Add new
            inserted = 0
            for f in live_flavors:
                if f["name"] not in existing_pricing:
                    vcpus = f.get("vcpus", 0) or 0
                    ram_mb = f.get("ram", 0) or 0
                    disk_gb = f.get("disk", 0) or 0
                    cur.execute("""
                        INSERT INTO metering_pricing
                            (category, item_name, unit, cost_per_hour, cost_per_month, currency, vcpus, ram_gb, disk_gb)
                        VALUES ('flavor', %s, 'per hour', 0, 0, 'USD', %s, %s, %s)
                    """, (f["name"], vcpus, round(ram_mb / 1024, 2), disk_gb))
                    inserted += 1

            # Remove stale pricing
            stale_pricing = existing_pricing - live_names
            removed = 0
            if stale_pricing:
                cur.execute(
                    "DELETE FROM metering_pricing WHERE category = 'flavor' AND item_name = ANY(%s)",
                    (list(stale_pricing),),
                )
                removed = cur.rowcount


    return {
        "detail": f"Synced from OpenStack: {inserted} added, {removed} removed, {len(existing_pricing) - removed} unchanged",
        "inserted": inserted,
        "removed": removed,
        "live_count": len(live_names),
    }


# Legacy compatibility endpoint (redirects to new pricing)
@router.get("/flavor-pricing")
async def list_flavor_pricing_legacy(
    user: User = Depends(require_permission("metering", "read")),
):
    """Legacy: list flavor pricing entries. Use /pricing instead."""
    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM metering_pricing WHERE category = 'flavor' ORDER BY item_name")
            rows = cur.fetchall()
    result = []
    for r in rows:
        d = dict(r)
        for k in list(d.keys()):
            if hasattr(d[k], "isoformat"):
                d[k] = d[k].isoformat()
            elif hasattr(d[k], "as_tuple"):
                d[k] = float(d[k])
        result.append(d)
    return result
