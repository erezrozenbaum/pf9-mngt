"""
Migration Planner API  —  Phase 1 + Phase 2
=============================================
CRUD for migration projects, RVTools XLSX upload/parse, tenant detection,
VM listing, risk scoring, project lifecycle management.

Phase 1 Routes:
  POST   /api/migration/projects              Create project
  GET    /api/migration/projects              List projects
  GET    /api/migration/projects/{id}         Get project detail
  PATCH  /api/migration/projects/{id}         Update project settings
  DELETE /api/migration/projects/{id}         Delete project (cascade)
  POST   /api/migration/projects/{id}/archive Archive + purge

  POST   /api/migration/projects/{id}/upload  Upload RVTools XLSX
  POST   /api/migration/projects/{id}/reparse Re-parse from stored data

  GET    /api/migration/projects/{id}/vms     List VMs (filters, sort, pagination)
  PATCH  /api/migration/projects/{id}/vms/{vm_id}  Update VM overrides

  GET    /api/migration/projects/{id}/tenants       Detected tenants
  PATCH  /api/migration/projects/{id}/tenants/{tid} Confirm/edit/scope tenant

  GET    /api/migration/projects/{id}/hosts         Source hosts
  GET    /api/migration/projects/{id}/clusters       Source clusters
  GET    /api/migration/projects/{id}/stats          Summary stats

  POST   /api/migration/projects/{id}/assess        Run risk + mode scoring
  GET    /api/migration/projects/{id}/risk-config    Get risk rules
  PUT    /api/migration/projects/{id}/risk-config    Update risk rules

  POST   /api/migration/projects/{id}/reset-assessment  Clear computed results
  POST   /api/migration/projects/{id}/reset-plan        Clear wave/prep data

  POST   /api/migration/projects/{id}/approve       Approve project (gate)

  GET    /api/migration/projects/{id}/bandwidth      Bandwidth model
  GET    /api/migration/projects/{id}/agent-recommendation  Agent sizing

  GET    /api/migration/projects/{id}/export-plan   JSON migration plan
  GET    /api/migration/projects/{id}/export-report.xlsx  Excel report
  GET    /api/migration/projects/{id}/export-report.pdf   PDF report

Phase 2A — Tenant Scoping:
  PATCH  /api/migration/projects/{id}/tenants/bulk-scope   Bulk include/exclude
  GET    /api/migration/projects/{id}/tenant-filters       List auto-exclude patterns
  POST   /api/migration/projects/{id}/tenant-filters       Add auto-exclude pattern
  DELETE /api/migration/projects/{id}/tenant-filters/{fid} Remove pattern

Phase 2C — Quota & Overcommit:
  GET    /api/migration/overcommit-profiles                   List profiles
  PATCH  /api/migration/projects/{id}/overcommit-profile      Set active profile
  GET    /api/migration/projects/{id}/quota-requirements      Compute per-tenant quota

Phase 2D — PCD Node Sizing:
  GET    /api/migration/projects/{id}/node-profiles           List node profiles
  POST   /api/migration/projects/{id}/node-profiles           Create/update profile
  DELETE /api/migration/projects/{id}/node-profiles/{pid}     Delete profile
  GET    /api/migration/projects/{id}/node-inventory          Get current inventory
  PUT    /api/migration/projects/{id}/node-inventory          Upsert inventory
  GET    /api/migration/projects/{id}/node-sizing             HA-aware sizing result

Phase 2E — PCD Readiness:
  PATCH  /api/migration/projects/{id}/pcd-settings            Store PCD connection settings
  POST   /api/migration/projects/{id}/pcd-gap-analysis        Run gap analysis
  GET    /api/migration/projects/{id}/pcd-gaps                List gap results
  PATCH  /api/migration/projects/{id}/pcd-gaps/{gid}/resolve  Mark gap resolved

Phase 2.8 — Pre-Phase 3 Polish:
  GET    /api/migration/projects/{id}/pcd-auto-detect-profile Detect dominant PCD node type from hypervisors
  GET    /api/migration/projects/{id}/export-gaps-report.xlsx PCD Readiness action report (Excel)
  GET    /api/migration/projects/{id}/export-gaps-report.pdf  PCD Readiness action report (PDF)
"""

import os
import io
import json
import logging
import re
import traceback
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional

import psycopg2
from psycopg2.extras import RealDictCursor, Json
from fastapi import APIRouter, HTTPException, Depends, Query, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, validator

from auth import require_permission, get_current_user

from migration_engine import (
    build_column_map,
    extract_row,
    classify_os_family,
    extract_os_version,
    extract_vlan_id,
    classify_network_type,
    assign_tenant,
    compute_risk,
    classify_migration_mode,
    compute_bandwidth_model,
    recommend_agent_sizing,
    estimate_vm_time,
    generate_migration_plan,
    summarize_rvtools_stats,
    COLUMN_ALIASES,
    # Phase 2
    compute_quota_requirements,
    compute_node_sizing,
    analyze_pcd_gaps,
    OVERCOMMIT_PRESETS,
)
from export_reports import (
    generate_excel_report, generate_pdf_report,
    generate_gaps_excel_report, generate_gaps_pdf_report,
)

logger = logging.getLogger("migration_planner")

router = APIRouter(prefix="/api/migration", tags=["migration"])

# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------
def _get_conn():
    from db_pool import get_connection
    return get_connection()


def _ensure_tables():
    """Run migration if tables don't exist yet."""
    from db_pool import get_connection
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables
                    WHERE table_name = 'migration_projects'
                )
            """)
            if not cur.fetchone()[0]:
                migration = os.path.join(
                    os.path.dirname(__file__), "..", "db", "migrate_migration_planner.sql"
                )
                if os.path.exists(migration):
                    with open(migration) as f:
                        cur.execute(f.read())
                    conn.commit()
                    logger.info("Migration planner tables created")


def _ensure_activity_log_table():
    from db_pool import get_connection
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables
                    WHERE table_name = 'activity_log'
                )
            """)
            if not cur.fetchone()[0]:
                migration = os.path.join(
                    os.path.dirname(__file__), "..", "db", "migrate_activity_log.sql"
                )
                if os.path.exists(migration):
                    with open(migration) as f:
                        cur.execute(f.read())
                    conn.commit()


# Run on import
try:
    _ensure_tables()
except Exception as e:
    logger.warning(f"Could not ensure migration tables on startup: {e}")

try:
    _ensure_activity_log_table()
except Exception as e:
    logger.warning(f"Could not ensure activity_log table: {e}")


# ---------------------------------------------------------------------------
# Activity log helper
# ---------------------------------------------------------------------------
def _log_activity(
    actor: str, action: str, resource_type: str,
    resource_id: str = None, resource_name: str = None,
    details: dict = None, result: str = "success", error_message: str = None,
):
    try:
        from db_pool import get_connection
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO activity_log
                        (actor, action, resource_type, resource_id, resource_name,
                         details, result, error_message)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (actor, action, resource_type, resource_id, resource_name,
                      Json(details or {}), result, error_message))
    except Exception as e:
        logger.error(f"Failed to write activity log: {e}")


# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------
class CreateProjectRequest(BaseModel):
    name: str
    description: Optional[str] = ""
    topology_type: Optional[str] = "local"
    # Source side
    source_nic_speed_gbps: Optional[float] = 10.0
    source_usable_pct: Optional[float] = 40.0
    # Transport link
    link_speed_gbps: Optional[float] = None
    link_usable_pct: Optional[float] = 60.0
    source_upload_mbps: Optional[float] = None
    dest_download_mbps: Optional[float] = None
    estimated_rtt_ms: Optional[float] = None
    rtt_category: Optional[str] = None
    # PCD / Agent side
    target_ingress_speed_gbps: Optional[float] = 10.0
    target_usable_pct: Optional[float] = 40.0
    pcd_storage_write_mbps: Optional[float] = 500.0
    # Agent profile
    agent_count: Optional[int] = 2
    agent_concurrent_vms: Optional[int] = 5
    agent_vcpu_per_slot: Optional[int] = 2
    agent_ram_base_gb: Optional[float] = 2.0
    agent_ram_per_slot_gb: Optional[float] = 1.0
    agent_nic_speed_gbps: Optional[float] = 10.0
    agent_nic_usable_pct: Optional[float] = 70.0
    agent_disk_buffer_factor: Optional[float] = 1.2
    # Warm migration
    daily_change_rate_pct: Optional[float] = 5.0

    @validator("name")
    def name_not_empty(cls, v):
        if not v or not v.strip():
            raise ValueError("Project name cannot be empty")
        return v.strip()

    @validator("topology_type")
    def valid_topology(cls, v):
        allowed = ("local", "cross_site_dedicated", "cross_site_internet")
        if v and v not in allowed:
            raise ValueError(f"topology must be one of {allowed}")
        return v


class UpdateProjectRequest(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    topology_type: Optional[str] = None
    source_nic_speed_gbps: Optional[float] = None
    source_usable_pct: Optional[float] = None
    link_speed_gbps: Optional[float] = None
    link_usable_pct: Optional[float] = None
    source_upload_mbps: Optional[float] = None
    dest_download_mbps: Optional[float] = None
    estimated_rtt_ms: Optional[float] = None
    rtt_category: Optional[str] = None
    target_ingress_speed_gbps: Optional[float] = None
    target_usable_pct: Optional[float] = None
    pcd_storage_write_mbps: Optional[float] = None
    agent_count: Optional[int] = None
    agent_concurrent_vms: Optional[int] = None
    agent_vcpu_per_slot: Optional[int] = None
    agent_ram_base_gb: Optional[float] = None
    agent_ram_per_slot_gb: Optional[float] = None
    agent_nic_speed_gbps: Optional[float] = None
    agent_nic_usable_pct: Optional[float] = None
    agent_disk_buffer_factor: Optional[float] = None
    daily_change_rate_pct: Optional[float] = None
    migration_duration_days: Optional[int] = None
    working_hours_per_day: Optional[float] = None
    working_days_per_week: Optional[int] = None
    target_vms_per_day: Optional[int] = None


class UpdateVMRequest(BaseModel):
    exclude_from_migration: Optional[bool] = None
    exclude_reason: Optional[str] = None
    manual_mode_override: Optional[str] = None
    priority: Optional[int] = None
    tenant_name: Optional[str] = None
    app_group: Optional[str] = None


class UpdateTenantRequest(BaseModel):
    tenant_name: Optional[str] = None
    org_vdc: Optional[str] = None
    confirmed: Optional[bool] = None
    # Phase 2A — scoping
    include_in_plan: Optional[bool] = None
    exclude_reason: Optional[str] = None
    # Phase 2B — target mapping
    target_domain_name: Optional[str] = None
    target_project_name: Optional[str] = None
    target_display_name: Optional[str] = None         # project description / friendly name
    target_domain_description: Optional[str] = None  # domain description
    target_confirmed: Optional[bool] = None  # true = user reviewed/confirmed the target names
    # Legacy aliases (kept for backwards compat)
    target_domain: Optional[str] = None
    target_project: Optional[str] = None
    notes: Optional[str] = None
    # Phase 2.10D — priority
    migration_priority: Optional[int] = None
    # Phase 2.10G — cohort assignment
    cohort_id: Optional[int] = None


# ---------------------------------------------------------------------------
# Phase 2.10 request models
# ---------------------------------------------------------------------------

class VMStatusUpdateRequest(BaseModel):
    status: str  # not_started|assigned|in_progress|migrated|failed|skipped
    status_note: Optional[str] = None


class VMBulkStatusRequest(BaseModel):
    vm_ids: List[int]
    status: str
    status_note: Optional[str] = None


class VMModeOverrideRequest(BaseModel):
    override: Optional[str] = None  # 'warm' | 'cold' | None (= clear override)


class VMDependencyRequest(BaseModel):
    depends_on_vm_id: int
    dependency_type: Optional[str] = "must_complete_before"
    notes: Optional[str] = None


class NetworkMappingCreateRequest(BaseModel):
    source_network_name: str
    target_network_name: Optional[str] = None
    target_network_id: Optional[str] = None
    vlan_id: Optional[int] = None
    notes: Optional[str] = None


class NetworkMappingUpdateRequest(BaseModel):
    target_network_name: Optional[str] = None
    target_network_id: Optional[str] = None
    vlan_id: Optional[int] = None
    notes: Optional[str] = None
    confirmed: Optional[bool] = None


class CreateCohortRequest(BaseModel):
    name: str
    description: Optional[str] = None
    cohort_order: Optional[int] = 999
    scheduled_start: Optional[str] = None  # ISO date string
    scheduled_end: Optional[str] = None
    owner_name: Optional[str] = None
    depends_on_cohort_id: Optional[int] = None
    overcommit_profile_override: Optional[str] = None
    agent_slots_override: Optional[int] = None
    schedule_duration_days: Optional[int] = None   # planned working days for this cohort
    target_vms_per_day: Optional[int] = None        # per-cohort VMs/day override
    notes: Optional[str] = None


class UpdateCohortRequest(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    cohort_order: Optional[int] = None
    status: Optional[str] = None
    scheduled_start: Optional[str] = None
    scheduled_end: Optional[str] = None
    owner_name: Optional[str] = None
    depends_on_cohort_id: Optional[int] = None
    overcommit_profile_override: Optional[str] = None
    agent_slots_override: Optional[int] = None
    schedule_duration_days: Optional[int] = None
    target_vms_per_day: Optional[int] = None
    notes: Optional[str] = None


class AssignTenantsToCohortRequest(BaseModel):
    tenant_ids: List[int]


# ---------------------------------------------------------------------------
# Helper: get project or 404
# ---------------------------------------------------------------------------
def _get_project(project_id: str, conn) -> Dict[str, Any]:
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("SELECT * FROM migration_projects WHERE project_id = %s", (project_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Migration project not found")
        return dict(row)


# =====================================================================
# PROJECT CRUD
# =====================================================================

@router.post("/projects", dependencies=[Depends(require_permission("migration", "write"))])
async def create_project(req: CreateProjectRequest, user = Depends(get_current_user)):
    """Create a new migration project."""
    actor = user.username if user else "system"

    with _get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                INSERT INTO migration_projects (
                    name, description, topology_type,
                    source_nic_speed_gbps, source_usable_pct,
                    link_speed_gbps, link_usable_pct,
                    source_upload_mbps, dest_download_mbps,
                    estimated_rtt_ms, rtt_category,
                    target_ingress_speed_gbps, target_usable_pct,
                    pcd_storage_write_mbps,
                    agent_count, agent_concurrent_vms,
                    agent_vcpu_per_slot, agent_ram_base_gb, agent_ram_per_slot_gb,
                    agent_nic_speed_gbps, agent_nic_usable_pct, agent_disk_buffer_factor,
                    daily_change_rate_pct,
                    created_by
                ) VALUES (
                    %s, %s, %s,
                    %s, %s,
                    %s, %s,
                    %s, %s,
                    %s, %s,
                    %s, %s,
                    %s,
                    %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s,
                    %s
                )
                RETURNING *
            """, (
                req.name, req.description, req.topology_type,
                req.source_nic_speed_gbps, req.source_usable_pct,
                req.link_speed_gbps, req.link_usable_pct,
                req.source_upload_mbps, req.dest_download_mbps,
                req.estimated_rtt_ms, req.rtt_category,
                req.target_ingress_speed_gbps, req.target_usable_pct,
                req.pcd_storage_write_mbps,
                req.agent_count, req.agent_concurrent_vms,
                req.agent_vcpu_per_slot, req.agent_ram_base_gb, req.agent_ram_per_slot_gb,
                req.agent_nic_speed_gbps, req.agent_nic_usable_pct, req.agent_disk_buffer_factor,
                req.daily_change_rate_pct,
                actor,
            ))
            project = dict(cur.fetchone())

            # Seed default risk config
            cur.execute("""
                INSERT INTO migration_risk_config (project_id) VALUES (%s)
            """, (project["project_id"],))

            # Seed default tenant detection rules
            cur.execute("""
                INSERT INTO migration_tenant_rules (project_id) VALUES (%s)
            """, (project["project_id"],))

            conn.commit()

    _log_activity(actor=actor, action="create", resource_type="migration_project",
                  resource_id=project["project_id"], resource_name=req.name)

    return {"status": "ok", "project": _serialize_project(project)}


@router.get("/projects", dependencies=[Depends(require_permission("migration", "read"))])
async def list_projects(
    status: Optional[str] = Query(None),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
):
    """List migration projects with optional status filter."""
    with _get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            where = "WHERE 1=1"
            params: list = []
            if status:
                where += " AND status = %s"
                params.append(status)

            cur.execute(f"""
                SELECT mp.*,
                       (SELECT count(*) FROM migration_vms mv WHERE mv.project_id = mp.project_id) as vm_count,
                       (SELECT count(*) FROM migration_tenants mt WHERE mt.project_id = mp.project_id) as tenant_count
                FROM migration_projects mp
                {where}
                ORDER BY mp.created_at DESC
                LIMIT %s OFFSET %s
            """, params + [limit, offset])
            projects = [dict(r) for r in cur.fetchall()]

            cur.execute(f"SELECT count(*) as cnt FROM migration_projects {where}", params)
            total = cur.fetchone()["cnt"]

    return {"status": "ok", "total": total, "projects": [_serialize_project(p) for p in projects]}


@router.get("/projects/{project_id}", dependencies=[Depends(require_permission("migration", "read"))])
async def get_project(project_id: str):
    """Get detailed project info."""
    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # VM counts by risk
            cur.execute("""
                SELECT risk_category, count(*) as cnt
                FROM migration_vms WHERE project_id = %s AND NOT exclude_from_migration
                GROUP BY risk_category
            """, (project_id,))
            risk_summary = {r["risk_category"]: r["cnt"] for r in cur.fetchall() if r["risk_category"]}

            # VM counts by mode
            cur.execute("""
                SELECT migration_mode, count(*) as cnt
                FROM migration_vms WHERE project_id = %s AND NOT exclude_from_migration
                GROUP BY migration_mode
            """, (project_id,))
            mode_summary = {r["migration_mode"]: r["cnt"] for r in cur.fetchall() if r["migration_mode"]}

            # Totals
            cur.execute("""
                SELECT count(*) as total_vms,
                       coalesce(sum(total_disk_gb), 0) as total_disk_gb,
                       coalesce(sum(ram_mb), 0) as total_ram_mb,
                       coalesce(sum(cpu_count), 0) as total_vcpu,
                       count(*) FILTER (WHERE exclude_from_migration) as excluded_vms
                FROM migration_vms WHERE project_id = %s
            """, (project_id,))
            totals = dict(cur.fetchone())

            # Tenant count
            cur.execute("SELECT count(*) as cnt FROM migration_tenants WHERE project_id = %s", (project_id,))
            totals["tenant_count"] = cur.fetchone()["cnt"]

    result = _serialize_project(project)
    result["risk_summary"] = risk_summary
    result["mode_summary"] = mode_summary
    result["totals"] = _serialize_row(totals)
    return {"status": "ok", "project": result}


@router.patch("/projects/{project_id}", dependencies=[Depends(require_permission("migration", "write"))])
async def update_project(project_id: str, req: UpdateProjectRequest, user = Depends(get_current_user)):
    """Update project settings (topology, bandwidth, agents)."""
    actor = user.username if user else "system"

    updates = {k: v for k, v in req.dict().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")

    set_clauses = []
    params = []
    for key, val in updates.items():
        set_clauses.append(f"{key} = %s")
        params.append(val)
    set_clauses.append("updated_at = now()")

    with _get_conn() as conn:
        _get_project(project_id, conn)  # 404 check
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(f"""
                UPDATE migration_projects
                SET {', '.join(set_clauses)}
                WHERE project_id = %s
                RETURNING *
            """, params + [project_id])
            project = dict(cur.fetchone())
            conn.commit()

    _log_activity(actor=actor, action="update", resource_type="migration_project",
                  resource_id=project_id, details=updates)

    return {"status": "ok", "project": _serialize_project(project)}


@router.delete("/projects/{project_id}", dependencies=[Depends(require_permission("migration", "admin"))])
async def delete_project(project_id: str, user = Depends(get_current_user)):
    """Delete a migration project and all associated data (CASCADE)."""
    actor = user.username if user else "system"

    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        with conn.cursor() as cur:
            cur.execute("DELETE FROM migration_projects WHERE project_id = %s", (project_id,))
            conn.commit()

    _log_activity(actor=actor, action="delete", resource_type="migration_project",
                  resource_id=project_id, resource_name=project["name"])

    return {"status": "ok", "message": f"Project '{project['name']}' and all data deleted"}


@router.post("/projects/{project_id}/archive", dependencies=[Depends(require_permission("migration", "admin"))])
async def archive_project(project_id: str, user = Depends(get_current_user)):
    """Archive project summary then purge full data."""
    actor = user.username if user else "system"

    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Build summary
            cur.execute("""
                SELECT count(*) as total_vms,
                       coalesce(sum(total_disk_gb), 0) as total_disk_gb,
                       count(*) FILTER (WHERE risk_category = 'GREEN') as green,
                       count(*) FILTER (WHERE risk_category = 'YELLOW') as yellow,
                       count(*) FILTER (WHERE risk_category = 'RED') as red
                FROM migration_vms WHERE project_id = %s
            """, (project_id,))
            vm_stats = dict(cur.fetchone())

            cur.execute("SELECT count(*) as cnt FROM migration_tenants WHERE project_id = %s", (project_id,))
            tenant_count = cur.fetchone()["cnt"]

            summary = {
                "name": project["name"],
                "status_at_archive": project["status"],
                "topology_type": project["topology_type"],
                "vm_count": vm_stats["total_vms"],
                "total_disk_tb": round(float(vm_stats["total_disk_gb"]) / 1024, 2),
                "tenant_count": tenant_count,
                "risk_green": vm_stats["green"],
                "risk_yellow": vm_stats["yellow"],
                "risk_red": vm_stats["red"],
                "created_at": project["created_at"].isoformat() if project["created_at"] else None,
                "archived_at": datetime.now(timezone.utc).isoformat(),
            }

            cur.execute("""
                INSERT INTO migration_project_archives (original_project_id, name, summary, archived_by)
                VALUES (%s, %s, %s, %s)
            """, (project_id, project["name"], Json(summary), actor))

            cur.execute("DELETE FROM migration_projects WHERE project_id = %s", (project_id,))
            conn.commit()

    _log_activity(actor=actor, action="archive", resource_type="migration_project",
                  resource_id=project_id, resource_name=project["name"], details=summary)

    return {"status": "ok", "message": f"Project archived and data purged", "summary": summary}


# =====================================================================
# RVTOOLS UPLOAD & PARSE
# =====================================================================

@router.post("/projects/{project_id}/upload", dependencies=[Depends(require_permission("migration", "write"))])
async def upload_rvtools(project_id: str, file: UploadFile = File(...), user = Depends(get_current_user)):
    """Upload and parse an RVTools XLSX file."""
    actor = user.username if user else "system"

    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xls files accepted")

    # Read file into memory
    contents = await file.read()
    if len(contents) > 100 * 1024 * 1024:  # 100 MB limit
        raise HTTPException(status_code=400, detail="File too large (max 100 MB)")

    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        if project["status"] not in ("draft", "assessment"):
            raise HTTPException(
                status_code=400,
                detail=f"Cannot upload RVTools in status '{project['status']}'. Reset to draft first."
            )

        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Cannot read XLSX: {exc}")

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Clear existing data for re-import
            for table in (
                "migration_wave_vms", "migration_waves", "migration_prep_tasks",
                "migration_target_gaps", "migration_vm_snapshots", "migration_vm_nics",
                "migration_vm_disks", "migration_vms", "migration_hosts",
                "migration_clusters", "migration_tenants",
            ):
                cur.execute(f"DELETE FROM {table} WHERE project_id = %s", (project_id,))

            # ----- Parse vHost -----
            vhost_count = 0
            if "vHost" in wb.sheetnames:
                vhost_count = _parse_vhost_sheet(wb["vHost"], project_id, cur)

            # ----- Parse vCluster -----
            vcluster_count = 0
            if "vCluster" in wb.sheetnames:
                vcluster_count = _parse_vcluster_sheet(wb["vCluster"], project_id, cur)

            # ----- Parse vInfo (main) -----
            vinfo_count = 0
            if "vInfo" in wb.sheetnames:
                vinfo_count = _parse_vinfo_sheet(wb["vInfo"], project_id, cur)
            else:
                raise HTTPException(status_code=400, detail="vInfo sheet not found in XLSX")

            # ----- Parse vDisk -----
            vdisk_count = 0
            if "vDisk" in wb.sheetnames:
                vdisk_count = _parse_vdisk_sheet(wb["vDisk"], project_id, cur)

            # ----- Parse vNIC -----
            vnic_count = 0
            vnetwork_infrastructure_count = 0
            if "vNetwork" in wb.sheetnames:
                # Parse NICs/adapters from vNetwork sheet
                vnic_count = _parse_vnic_sheet(wb["vNetwork"], project_id, cur)
                # Parse network infrastructure from vNetwork sheet
                vnetwork_infrastructure_count = _parse_vnetwork_sheet(wb["vNetwork"], project_id, cur)
            elif "vNIC" in wb.sheetnames:
                vnic_count = _parse_vnic_sheet(wb["vNIC"], project_id, cur)

            # ----- Parse vSnapshot -----
            vsnapshot_count = 0
            if "vSnapshot" in wb.sheetnames:
                vsnapshot_count = _parse_vsnapshot_sheet(wb["vSnapshot"], project_id, cur)

            # ----- Parse vPartition (used disk space) -----
            vpartition_count = 0
            if "vPartition" in wb.sheetnames:
                vpartition_count = _parse_vpartition_sheet(wb["vPartition"], project_id, cur)

            # ----- Parse vCPU (CPU usage metrics) -----
            vcpu_count = 0
            if "vCPU" in wb.sheetnames:
                vcpu_count = _parse_vcpu_sheet(wb["vCPU"], project_id, cur)

            # ----- Parse vMemory (Memory usage metrics) -----
            vmemory_count = 0
            if "vMemory" in wb.sheetnames:
                vmemory_count = _parse_vmemory_sheet(wb["vMemory"], project_id, cur)

            # Update disk/nic/snapshot summaries on VMs
            _update_vm_summaries(project_id, cur)

            # Build network infrastructure summary
            _build_network_summary(project_id, cur)

            # Run tenant detection
            _run_tenant_detection(project_id, cur)

            # Clean up network mappings for source networks no longer in VM data
            cur.execute("""
                DELETE FROM migration_network_mappings
                WHERE project_id = %s
                  AND source_network_name NOT IN (
                      SELECT DISTINCT network_name
                      FROM migration_vms
                      WHERE project_id = %s
                        AND network_name IS NOT NULL
                        AND network_name != ''
                  )
            """, (project_id, project_id))

            # ---- Collect enriched stats from DB ----
            # Power state breakdown
            cur.execute("""
                SELECT coalesce(lower(power_state), 'unknown') as state, count(*) as cnt
                FROM migration_vms WHERE project_id = %s AND (template IS NOT TRUE)
                GROUP BY coalesce(lower(power_state), 'unknown')
            """, (project_id,))
            power_state_counts = {r["state"]: r["cnt"] for r in cur.fetchall()}

            # OS family breakdown
            cur.execute("""
                SELECT coalesce(os_family, 'unknown') as family, count(*) as cnt
                FROM migration_vms WHERE project_id = %s AND (template IS NOT TRUE)
                GROUP BY coalesce(os_family, 'unknown')
            """, (project_id,))
            os_family_counts = {r["family"]: r["cnt"] for r in cur.fetchall()}

            # Tenant count
            cur.execute("""
                SELECT count(*) as cnt FROM migration_tenants WHERE project_id = %s
            """, (project_id,))
            tenant_count = cur.fetchone()["cnt"]

            # Template count
            cur.execute("""
                SELECT count(*) as cnt FROM migration_vms
                WHERE project_id = %s AND template IS TRUE
            """, (project_id,))
            template_count = cur.fetchone()["cnt"]

            # Detect if vCD environment (any tenant detected via vcd_folder or vapp_name with org_vdc)
            cur.execute("""
                SELECT count(*) as cnt FROM migration_tenants
                WHERE project_id = %s AND detection_method IN ('vcd_folder', 'vapp_name')
                  AND org_vdc IS NOT NULL
            """, (project_id,))
            vcd_detected = cur.fetchone()["cnt"] > 0

            # Update project metadata
            stats = summarize_rvtools_stats(
                vinfo_count, vdisk_count, vnic_count, vhost_count, vcluster_count, vsnapshot_count,
                vpartition_count=vpartition_count,
                vcpu_count=vcpu_count,
                vmemory_count=vmemory_count,
                power_state_counts=power_state_counts,
                os_family_counts=os_family_counts,
                tenant_count=tenant_count,
                vcd_detected=vcd_detected,
                template_count=template_count,
            )
            cur.execute("""
                UPDATE migration_projects
                SET rvtools_filename = %s,
                    rvtools_uploaded_at = now(),
                    rvtools_sheet_stats = %s,
                    status = CASE WHEN status = 'draft' THEN 'assessment' ELSE status END,
                    updated_at = now()
                WHERE project_id = %s
            """, (file.filename, Json(stats), project_id))

            conn.commit()

        wb.close()

    _log_activity(actor=actor, action="upload_rvtools", resource_type="migration_project",
                  resource_id=project_id, details={"filename": file.filename, "stats": stats})

    return {
        "status": "ok",
        "message": f"Parsed {file.filename}",
        "stats": stats,
    }


# ---------------------------------------------------------------------------
# Re-parse ONLY the vMemory sheet (to fix active vs consumed without full re-upload)
# ---------------------------------------------------------------------------

@router.post("/projects/{project_id}/reparse-memory",
             dependencies=[Depends(require_permission("migration", "write"))])
async def reparse_memory(project_id: str, file: UploadFile = File(...), user = Depends(get_current_user)):
    """
    Re-parse ONLY the vMemory sheet from a fresh RVTools XLSX to update
    memory_usage_percent and memory_usage_mb on existing VMs — without
    wiping any other data (VMs, tenants, disks, etc.).

    Use this when the initial upload captured 'Consumed' instead of 'Active',
    giving all VMs ~100% memory utilisation.
    """
    actor = user.username if user else "system"

    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xls files accepted")

    contents = await file.read()
    if len(contents) > 100 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 100 MB)")

    with _get_conn() as conn:
        _get_project(project_id, conn)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Cannot read XLSX: {exc}")

        if "vMemory" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="vMemory sheet not found in the uploaded file")

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Reset existing memory columns so COALESCE picks up fresh values
            cur.execute("""
                UPDATE migration_vms
                SET memory_usage_percent = NULL, memory_usage_mb = NULL
                WHERE project_id = %s
            """, (project_id,))

            updated = _parse_vmemory_sheet(wb["vMemory"], project_id, cur)

        conn.commit()
        wb.close()

    _log_activity(actor=actor, action="reparse_memory", resource_type="migration_project",
                  resource_id=project_id, details={"filename": file.filename, "updated": updated})

    return {
        "status": "ok",
        "message": f"Memory metrics updated for {updated} VMs from {file.filename}",
        "updated_vms": updated,
    }


# ---------------------------------------------------------------------------
# Clear uploaded RVTools data (keep project settings)
# ---------------------------------------------------------------------------

@router.delete("/projects/{project_id}/rvtools",
               dependencies=[Depends(require_permission("migration", "write"))])
async def clear_rvtools_data(project_id: str, user = Depends(get_current_user)):
    """Delete all uploaded RVTools data for a project (VMs, disks, NICs, etc.)
    while preserving the project itself and its settings."""
    actor = user.username if user else "system"

    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        with conn.cursor() as cur:
            for table in (
                "migration_wave_vms", "migration_waves", "migration_prep_tasks",
                "migration_target_gaps", "migration_vm_snapshots", "migration_vm_nics",
                "migration_vm_disks", "migration_vms", "migration_hosts",
                "migration_clusters", "migration_tenants", "migration_networks",
                "migration_network_mappings",
            ):
                cur.execute(f"DELETE FROM {table} WHERE project_id = %s", (project_id,))

            cur.execute("""
                UPDATE migration_projects SET
                    rvtools_filename = NULL,
                    rvtools_uploaded_at = NULL,
                    rvtools_sheet_stats = '{}',
                    status = 'draft',
                    updated_at = now()
                WHERE project_id = %s
            """, (project_id,))
            conn.commit()

    _log_activity(actor=actor, action="clear_rvtools", resource_type="migration_project",
                  resource_id=project_id, resource_name=project["name"])

    return {"status": "ok", "message": "All RVTools data cleared. Project reset to draft."}


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------

def _sheet_headers(sheet) -> List[str]:
    """Extract header row from an openpyxl sheet."""
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        return [str(c) if c else "" for c in row]
    return []


def _sheet_rows(sheet) -> List[list]:
    """All data rows (skip header)."""
    rows = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            continue
        rows.append(list(row))
    return rows


def _safe_int(val, default=0) -> int:
    if val is None or val == "":
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def _safe_float(val, default=0.0) -> float:
    if val is None or val == "":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _safe_bool(val) -> Optional[bool]:
    if val is None or val == "":
        return None
    if isinstance(val, bool):
        return val
    return str(val).lower() in ("true", "yes", "1", "enabled")


def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _parse_vinfo_sheet(sheet, project_id: str, cur) -> int:
    headers = _sheet_headers(sheet)
    col_map = build_column_map(headers)
    rows = _sheet_rows(sheet)
    count = 0

    for row_vals in rows:
        d = extract_row(row_vals, col_map)
        vm_name = _safe_str(d.get("vm_name"))
        if not vm_name:
            continue

        guest_os = _safe_str(d.get("guest_os"))
        guest_os_tools = _safe_str(d.get("guest_os_tools"))
        os_family = classify_os_family(guest_os, guest_os_tools)
        os_version = extract_os_version(guest_os, guest_os_tools)

        # Disk sizes from vInfo (rough — refined after vDisk parse)
        prov_mb = _safe_float(d.get("provisioned_mb"))
        in_use_raw = _safe_float(d.get("in_use_mb"))
        total_disk_gb = round(prov_mb / 1024, 2) if prov_mb else 0
        in_use_gb = round(in_use_raw / 1024, 2) if in_use_raw else 0

        cur.execute("""
            INSERT INTO migration_vms (
                project_id, vm_name, power_state, template,
                guest_os, guest_os_tools, os_family, os_version,
                folder_path, resource_pool, vapp_name, annotation,
                cpu_count, ram_mb, total_disk_gb,
                provisioned_mb, in_use_mb, in_use_gb,
                host_name, cluster, datacenter,
                vm_uuid, firmware, change_tracking, connection_state,
                dns_name, primary_ip,
                raw_data
            ) VALUES (
                %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s,
                %s
            )
            ON CONFLICT (project_id, vm_name) DO UPDATE SET
                power_state = EXCLUDED.power_state,
                template = EXCLUDED.template,
                guest_os = EXCLUDED.guest_os,
                guest_os_tools = EXCLUDED.guest_os_tools,
                os_family = EXCLUDED.os_family,
                os_version = EXCLUDED.os_version,
                folder_path = EXCLUDED.folder_path,
                resource_pool = EXCLUDED.resource_pool,
                vapp_name = EXCLUDED.vapp_name,
                annotation = EXCLUDED.annotation,
                cpu_count = EXCLUDED.cpu_count,
                ram_mb = EXCLUDED.ram_mb,
                total_disk_gb = EXCLUDED.total_disk_gb,
                provisioned_mb = EXCLUDED.provisioned_mb,
                in_use_mb = EXCLUDED.in_use_mb,
                in_use_gb = EXCLUDED.in_use_gb,
                host_name = EXCLUDED.host_name,
                cluster = EXCLUDED.cluster,
                datacenter = EXCLUDED.datacenter,
                vm_uuid = EXCLUDED.vm_uuid,
                firmware = EXCLUDED.firmware,
                change_tracking = EXCLUDED.change_tracking,
                connection_state = EXCLUDED.connection_state,
                dns_name = EXCLUDED.dns_name,
                primary_ip = EXCLUDED.primary_ip,
                raw_data = EXCLUDED.raw_data,
                updated_at = now()
        """, (
            project_id, vm_name,
            _safe_str(d.get("power_state")),
            _safe_bool(d.get("template")),
            guest_os, guest_os_tools, os_family, os_version,
            _safe_str(d.get("folder_path")),
            _safe_str(d.get("resource_pool")),
            _safe_str(d.get("vapp_name")),
            _safe_str(d.get("annotation")),
            _safe_int(d.get("cpu_count")),
            _safe_int(d.get("ram_mb")),
            total_disk_gb,
            _safe_int(prov_mb),
            _safe_int(in_use_raw),
            in_use_gb,
            _safe_str(d.get("host_name")),
            _safe_str(d.get("cluster")),
            _safe_str(d.get("datacenter")),
            _safe_str(d.get("vm_uuid")),
            _safe_str(d.get("firmware")),
            _safe_bool(d.get("change_tracking")),
            _safe_str(d.get("connection_state")),
            _safe_str(d.get("dns_name")),
            _safe_str(d.get("primary_ip")),
            Json({k: str(v) if v is not None else "" for k, v in d.items()}),
        ))
        count += 1

    return count


def _parse_vdisk_sheet(sheet, project_id: str, cur) -> int:
    headers = _sheet_headers(sheet)
    col_map = build_column_map(headers, prefix="disk_")
    rows = _sheet_rows(sheet)
    count = 0

    for row_vals in rows:
        d = extract_row(row_vals, col_map)
        vm_name = _safe_str(d.get("disk_vm_name") or d.get("vm_name"))
        if not vm_name:
            continue

        cap_mb = _safe_float(d.get("capacity_mb"))
        cap_gb = round(cap_mb / 1024, 2) if cap_mb else 0

        cur.execute("""
            INSERT INTO migration_vm_disks (
                project_id, vm_name, disk_label, disk_path,
                capacity_gb, thin_provisioned, eagerly_scrub, datastore,
                raw_data
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            project_id, vm_name,
            _safe_str(d.get("disk_label")),
            _safe_str(d.get("disk_path")),
            cap_gb,
            _safe_bool(d.get("thin")),
            _safe_bool(d.get("eagerly_scrub")),
            _safe_str(d.get("datastore")),
            Json({k: str(v) if v is not None else "" for k, v in d.items()}),
        ))
        count += 1

    return count


def _parse_vnic_sheet(sheet, project_id: str, cur) -> int:
    headers = _sheet_headers(sheet)
    col_map = build_column_map(headers, prefix="nic_")
    rows = _sheet_rows(sheet)
    count = 0

    for row_vals in rows:
        d = extract_row(row_vals, col_map)
        vm_name = _safe_str(d.get("nic_vm_name") or d.get("vm_name"))
        if not vm_name:
            continue

        cur.execute("""
            INSERT INTO migration_vm_nics (
                project_id, vm_name, nic_label, adapter_type,
                network_name, connected, mac_address, ip_address,
                raw_data
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            project_id, vm_name,
            _safe_str(d.get("nic_label")),
            _safe_str(d.get("adapter_type")),
            _safe_str(d.get("network_name")),
            _safe_bool(d.get("nic_connected")),
            _safe_str(d.get("mac_address")),
            _safe_str(d.get("ip_address")),
            Json({k: str(v) if v is not None else "" for k, v in d.items()}),
        ))
        count += 1

    return count


def _parse_vhost_sheet(sheet, project_id: str, cur) -> int:
    headers = _sheet_headers(sheet)
    col_map = build_column_map(headers, prefix="host_")
    rows = _sheet_rows(sheet)
    count = 0

    for row_vals in rows:
        d = extract_row(row_vals, col_map)
        host_name = _safe_str(d.get("host_host_name") or d.get("vm_name"))
        if not host_name:
            continue

        # Try to parse NIC speed (could be "10000" Mbps or "10 Gbit")
        nic_speed_raw = _safe_str(d.get("host_nic_speed"))
        nic_speed_mbps = _safe_int(nic_speed_raw)
        if "gbit" in nic_speed_raw.lower() or "gbps" in nic_speed_raw.lower():
            nic_speed_mbps = _safe_int(nic_speed_raw.split()[0]) * 1000

        cur.execute("""
            INSERT INTO migration_hosts (
                project_id, host_name, cluster, datacenter,
                cpu_model, cpu_count, cpu_cores, cpu_threads,
                ram_mb, nic_count, nic_speed_mbps,
                esx_version, raw_data
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (project_id, host_name) DO UPDATE SET
                cluster = EXCLUDED.cluster,
                datacenter = EXCLUDED.datacenter,
                cpu_model = EXCLUDED.cpu_model,
                cpu_count = EXCLUDED.cpu_count,
                cpu_cores = EXCLUDED.cpu_cores,
                cpu_threads = EXCLUDED.cpu_threads,
                ram_mb = EXCLUDED.ram_mb,
                nic_count = EXCLUDED.nic_count,
                nic_speed_mbps = EXCLUDED.nic_speed_mbps,
                esx_version = EXCLUDED.esx_version,
                raw_data = EXCLUDED.raw_data
        """, (
            project_id, host_name,
            _safe_str(d.get("host_cluster")),
            _safe_str(d.get("host_datacenter")),
            _safe_str(d.get("host_cpu_model")),
            _safe_int(d.get("host_cpu_count")),
            _safe_int(d.get("host_cpu_cores")),
            _safe_int(d.get("host_cpu_threads")),
            _safe_int(d.get("host_ram_mb")),
            _safe_int(d.get("host_nic_count")),
            nic_speed_mbps,
            _safe_str(d.get("host_esx_version")),
            Json({k: str(v) if v is not None else "" for k, v in d.items()}),
        ))
        count += 1

    return count


def _parse_vcluster_sheet(sheet, project_id: str, cur) -> int:
    headers = _sheet_headers(sheet)
    col_map = build_column_map(headers, prefix="cluster_")
    rows = _sheet_rows(sheet)
    count = 0

    for row_vals in rows:
        d = extract_row(row_vals, col_map)
        name = _safe_str(d.get("cluster_name_col") or d.get("vm_name"))
        if not name:
            continue

        cur.execute("""
            INSERT INTO migration_clusters (
                project_id, cluster_name, datacenter,
                host_count, total_cpu_mhz, total_ram_mb,
                ha_enabled, drs_enabled, raw_data
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (project_id, cluster_name) DO UPDATE SET
                datacenter = EXCLUDED.datacenter,
                host_count = EXCLUDED.host_count,
                total_cpu_mhz = EXCLUDED.total_cpu_mhz,
                total_ram_mb = EXCLUDED.total_ram_mb,
                ha_enabled = EXCLUDED.ha_enabled,
                drs_enabled = EXCLUDED.drs_enabled,
                raw_data = EXCLUDED.raw_data
        """, (
            project_id, name,
            _safe_str(d.get("cluster_datacenter")),
            _safe_int(d.get("cluster_host_count")),
            _safe_int(d.get("cluster_total_cpu")),
            _safe_int(d.get("cluster_total_ram")),
            _safe_bool(d.get("cluster_ha")),
            _safe_bool(d.get("cluster_drs")),
            Json({k: str(v) if v is not None else "" for k, v in d.items()}),
        ))
        count += 1

    return count


def _parse_vsnapshot_sheet(sheet, project_id: str, cur) -> int:
    headers = _sheet_headers(sheet)
    col_map = build_column_map(headers, prefix="snap_")
    rows = _sheet_rows(sheet)
    count = 0

    for row_vals in rows:
        d = extract_row(row_vals, col_map)
        vm_name = _safe_str(d.get("snap_vm_name") or d.get("vm_name"))
        if not vm_name:
            continue

        # Parse snapshot date
        snap_date = d.get("snap_created")
        if snap_date and not isinstance(snap_date, datetime):
            try:
                snap_date = datetime.fromisoformat(str(snap_date))
            except (ValueError, TypeError):
                snap_date = None

        size_raw = _safe_float(d.get("snap_size_mb"))
        size_gb = round(size_raw / 1024, 2) if size_raw else 0

        cur.execute("""
            INSERT INTO migration_vm_snapshots (
                project_id, vm_name, snapshot_name, description,
                created_date, size_gb, is_current, raw_data
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            project_id, vm_name,
            _safe_str(d.get("snap_name")),
            _safe_str(d.get("snap_description")),
            snap_date,
            size_gb,
            _safe_bool(d.get("snap_is_current")),
            Json({k: str(v) if v is not None else "" for k, v in d.items()}),
        ))
        count += 1

    return count


# ---------------------------------------------------------------------------
# vPartition parser — actual used disk space per VM partition
# ---------------------------------------------------------------------------

def _parse_vpartition_sheet(sheet, project_id: str, cur) -> int:
    """
    Parse the vPartition sheet from RVTools.
    Aggregates consumed space per VM and updates migration_vms.partition_used_gb.
    Also updates in_use_gb on migration_vms with partition data (more accurate than vInfo in_use_mb).
    """
    headers = _sheet_headers(sheet)
    col_map = build_column_map(headers, prefix="part_")
    rows = _sheet_rows(sheet)
    count = 0

    # Aggregate consumed MB per VM
    vm_consumed: Dict[str, float] = {}

    for row_vals in rows:
        d = extract_row(row_vals, col_map)
        vm_name = _safe_str(d.get("part_vm_name") or d.get("vm_name"))
        if not vm_name:
            continue

        consumed_mb = _safe_float(d.get("part_consumed_mb"))
        capacity_mb = _safe_float(d.get("part_capacity_mb"))
        free_mb = _safe_float(d.get("part_free_space_mb"))

        # Some RVTools exports only have Free Space — derive consumed
        if consumed_mb <= 0 and capacity_mb > 0 and free_mb >= 0:
            consumed_mb = capacity_mb - free_mb

        if consumed_mb > 0:
            vm_consumed[vm_name] = vm_consumed.get(vm_name, 0) + consumed_mb
        count += 1

    # Bulk-update partition_used_gb and in_use_gb on VMs
    for vm_name, total_consumed_mb in vm_consumed.items():
        consumed_gb = round(total_consumed_mb / 1024, 2)
        cur.execute("""
            UPDATE migration_vms
            SET partition_used_gb = %s,
                in_use_gb = %s
            WHERE project_id = %s AND vm_name = %s
        """, (consumed_gb, consumed_gb, project_id, vm_name))

    return count


def _find_col(headers_lower: List[str], candidates: List[str]) -> int:
    """Find column index by matching normalized header against candidate names. Returns -1 if not found."""
    for candidate in candidates:
        c = candidate.lower().strip()
        for i, h in enumerate(headers_lower):
            if h == c or h.replace(" ", "") == c.replace(" ", ""):
                return i
    return -1


def _parse_vcpu_sheet(sheet, project_id: str, cur) -> int:
    """
    Parse the vCPU sheet from RVTools to extract CPU usage metrics.
    Updates migration_vms with cpu_usage_percent and cpu_demand_mhz.
    Uses direct column matching against known RVTools vCPU sheet header names.
    """
    headers = _sheet_headers(sheet)
    headers_lower = [h.lower().strip() for h in headers]

    # Find column indices by direct header matching - RVTools vCPU sheet uses these exact names
    vm_col = _find_col(headers_lower, ["vm", "name", "vm name"])
    # 'overall' = current CPU usage in MHz; 'cpus' = vCPU count for computing %
    usage_col = _find_col(headers_lower, ["% usage", "usage %", "cpu usage %", "cpu usage", "average % usage", "avg % usage", "average usage %"])
    demand_col = _find_col(headers_lower, ["demand mhz", "demand (mhz)", "cpu demand mhz", "demand", "cpu demand", "overall"])
    cpus_col = _find_col(headers_lower, ["cpus", "cpu count", "vcpu count", "num cpus", "# cpus", "num vcpus"])


    rows = _sheet_rows(sheet)
    count = 0

    for row_vals in rows:
        if vm_col < 0 or vm_col >= len(row_vals):
            continue
        vm_name = _safe_str(row_vals[vm_col])
        if not vm_name:
            continue

        cpu_usage_pct = None
        cpu_demand_mhz = None

        if usage_col >= 0 and usage_col < len(row_vals):
            v = row_vals[usage_col]
            if v is not None and str(v).strip() not in ("", "None"):
                try:
                    cpu_usage_pct = float(v)
                except (ValueError, TypeError):
                    pass

        if demand_col >= 0 and demand_col < len(row_vals):
            v = row_vals[demand_col]
            if v is not None and str(v).strip() not in ("", "None"):
                try:
                    cpu_demand_mhz = int(float(v))
                except (ValueError, TypeError):
                    pass

        # If no direct % column, compute from demand_mhz / (cpus * ~2400 MHz) * 100
        if cpu_usage_pct is None and cpu_demand_mhz is not None:
            vcpus = None
            if cpus_col >= 0 and cpus_col < len(row_vals):
                cpuv = row_vals[cpus_col]
                if cpuv is not None and str(cpuv).strip() not in ("", "None"):
                    try:
                        vcpus = int(float(cpuv))
                    except (ValueError, TypeError):
                        pass
            if vcpus and vcpus > 0:
                # 2400 MHz per vCPU is a reasonable ESXi average (modern hardware 2.0-3.5GHz)
                cpu_usage_pct = round(min(cpu_demand_mhz / (vcpus * 2400.0) * 100, 100.0), 1)

        if cpu_usage_pct is None and cpu_demand_mhz is None:
            continue

        cur.execute("""
            UPDATE migration_vms
            SET cpu_usage_percent = COALESCE(%s, cpu_usage_percent),
                cpu_demand_mhz = COALESCE(%s, cpu_demand_mhz)
            WHERE project_id = %s AND vm_name = %s
        """, (cpu_usage_pct, cpu_demand_mhz, project_id, vm_name))

        if cur.rowcount > 0:
            count += 1

    return count


def _parse_vmemory_sheet(sheet, project_id: str, cur) -> int:
    """
    Parse the vMemory sheet from RVTools to extract memory usage metrics.
    Updates migration_vms with memory_usage_percent and memory_usage_mb.
    Uses direct column matching against known RVTools vMemory sheet header names.
    """
    headers = _sheet_headers(sheet)
    headers_lower = [h.lower().strip() for h in headers]

    # Find column indices by direct header matching - RVTools vMemory sheet uses these exact names
    vm_col = _find_col(headers_lower, ["vm", "name", "vm name"])

    # Explicit % column from vSphere real-time stats (may be absent in static RVTools exports)
    usage_col = _find_col(headers_lower, ["% usage", "usage %", "memory usage %", "mem usage %",
                                          "memory usage", "average % usage", "avg % usage",
                                          "average usage %", "% mem", "mem %"])

    # Prefer "Active (MiB)" — the guest working set actually being read/written.
    # "Consumed (MiB)" is configured RAM + VMkernel overhead and is always ~100% for
    # powered-on VMs regardless of real guest utilisation — do NOT use it for the % indicator.
    usage_mb_col = _find_col(headers_lower, [
        "active (mib)", "active mib", "active mb", "active",   # real guest working set ← preferred
        "usage mb", "usage (mb)", "memory usage mb", "mem usage mb", "used mb",
        "consumed (mib)", "consumed mib", "consumed",           # fallback — always ≈ configured RAM
    ])

    # Separate lookup for consumed so we can still log/store it for reference (optional)
    consumed_col = _find_col(headers_lower, ["consumed (mib)", "consumed mib", "consumed"])

    # Configured memory size — needed to compute % when no explicit % column exists
    # RVTools vMemory uses "Memory" (MB) or "Size MiB"
    size_mib_col = _find_col(headers_lower, ["size mib", "size (mib)", "size mb",
                                              "configured size", "memory size", "memory"])


    rows = _sheet_rows(sheet)
    count = 0

    for row_vals in rows:
        if vm_col < 0 or vm_col >= len(row_vals):
            continue
        vm_name = _safe_str(row_vals[vm_col])
        if not vm_name:
            continue

        memory_usage_pct = None
        memory_usage_mb = None

        if usage_col >= 0 and usage_col < len(row_vals):
            v = row_vals[usage_col]
            if v is not None and str(v).strip() not in ("", "None"):
                try:
                    memory_usage_pct = float(v)
                except (ValueError, TypeError):
                    pass

        if usage_mb_col >= 0 and usage_mb_col < len(row_vals):
            v = row_vals[usage_mb_col]
            if v is not None and str(v).strip() not in ("", "None"):
                try:
                    memory_usage_mb = int(float(v))
                except (ValueError, TypeError):
                    pass

        # Compute % from active / configured when no direct % column exists.
        # If usage_mb came from "consumed" this will still be ~100% (expected); if from
        # "active" it will correctly reflect the guest's real working-set utilisation.
        if memory_usage_pct is None and memory_usage_mb is not None and size_mib_col >= 0 and size_mib_col < len(row_vals):
            sz = row_vals[size_mib_col]
            if sz is not None and str(sz).strip() not in ("", "None"):
                try:
                    size_mib = float(sz)
                    if size_mib > 0:
                        memory_usage_pct = round(memory_usage_mb / size_mib * 100, 1)
                except (ValueError, TypeError):
                    pass

        if memory_usage_pct is None and memory_usage_mb is None:
            continue

        cur.execute("""
            UPDATE migration_vms
            SET memory_usage_percent = COALESCE(%s, memory_usage_percent),
                memory_usage_mb = COALESCE(%s, memory_usage_mb)
            WHERE project_id = %s AND vm_name = %s
        """, (memory_usage_pct, memory_usage_mb, project_id, vm_name))

        if cur.rowcount > 0:
            count += 1

    return count


def _parse_vnetwork_sheet(sheet, project_id: str, cur) -> int:
    """
    Parse the vNetwork sheet from RVTools to extract network infrastructure data.
    Updates migration_networks with subnet, gateway, DNS, and IP range information.
    """
    headers = _sheet_headers(sheet)
    col_map = build_column_map(headers, prefix="net_")
    rows = _sheet_rows(sheet)
    count = 0

    for row_vals in rows:
        d = extract_row(row_vals, col_map)
        network_name = _safe_str(d.get("net_network_name") or d.get("network_name"))
        if not network_name:
            continue

        # Extract network infrastructure data
        vlan_id = _safe_str(d.get("net_vlan_id"))
        subnet = _safe_str(d.get("net_subnet"))
        gateway = _safe_str(d.get("net_gateway"))
        dns_servers = _safe_str(d.get("net_dns_servers"))
        ip_range = _safe_str(d.get("net_ip_range"))

        # Update existing network record or skip if not found
        # (Networks should already exist from NIC parsing)
        cur.execute("""
            UPDATE migration_networks
            SET subnet = %s,
                gateway = %s,
                dns_servers = %s,
                ip_range = %s
            WHERE project_id = %s AND network_name = %s
        """, (subnet, gateway, dns_servers, ip_range, project_id, network_name))
        
        if cur.rowcount > 0:
            count += 1

    return count


# ---------------------------------------------------------------------------
# Post-parse: update VM summaries (disk/nic/snapshot counts from detail tables)
# ---------------------------------------------------------------------------

def _update_vm_summaries(project_id: str, cur):
    """Update disk_count, nic_count, snapshot_count etc from detail tables."""
    # Disk totals
    cur.execute("""
        UPDATE migration_vms mv SET
            total_disk_gb = sub.sum_gb,
            disk_count = sub.cnt
        FROM (
            SELECT vm_name, coalesce(sum(capacity_gb), 0) as sum_gb, count(*) as cnt
            FROM migration_vm_disks WHERE project_id = %s
            GROUP BY vm_name
        ) sub
        WHERE mv.project_id = %s AND mv.vm_name = sub.vm_name
    """, (project_id, project_id))

    # NIC count
    cur.execute("""
        UPDATE migration_vms mv SET
            nic_count = sub.cnt
        FROM (
            SELECT vm_name, count(*) as cnt
            FROM migration_vm_nics WHERE project_id = %s
            GROUP BY vm_name
        ) sub
        WHERE mv.project_id = %s AND mv.vm_name = sub.vm_name
    """, (project_id, project_id))

    # Snapshot count + oldest
    cur.execute("""
        UPDATE migration_vms mv SET
            snapshot_count = sub.cnt,
            snapshot_oldest_days = sub.oldest_days
        FROM (
            SELECT vm_name,
                   count(*) as cnt,
                   EXTRACT(DAY FROM now() - min(created_date))::int as oldest_days
            FROM migration_vm_snapshots WHERE project_id = %s
            GROUP BY vm_name
        ) sub
        WHERE mv.project_id = %s AND mv.vm_name = sub.vm_name
    """, (project_id, project_id))

    # Host VM counts
    cur.execute("""
        UPDATE migration_hosts mh SET
            vm_count = sub.cnt
        FROM (
            SELECT host_name, count(*) as cnt
            FROM migration_vms WHERE project_id = %s
            GROUP BY host_name
        ) sub
        WHERE mh.project_id = %s AND mh.host_name = sub.host_name
    """, (project_id, project_id))

    # Network name from first NIC (primary network for each VM)
    cur.execute("""
        UPDATE migration_vms mv SET
            network_name = sub.first_network
        FROM (
            SELECT DISTINCT ON (vm_name) vm_name, network_name as first_network
            FROM migration_vm_nics
            WHERE project_id = %s AND network_name IS NOT NULL AND network_name != ''
            ORDER BY vm_name, id ASC
        ) sub
        WHERE mv.project_id = %s AND mv.vm_name = sub.vm_name
    """, (project_id, project_id))

    # If vPartition was NOT parsed, derive in_use_gb from vInfo in_use_mb
    # (only for VMs where partition_used_gb is still 0)
    cur.execute("""
        UPDATE migration_vms
        SET in_use_gb = CASE
                WHEN in_use_mb > 0 THEN round(in_use_mb / 1024.0, 2)
                ELSE total_disk_gb
            END
        WHERE project_id = %s
          AND (partition_used_gb IS NULL OR partition_used_gb = 0)
          AND (in_use_gb IS NULL OR in_use_gb = 0)
    """, (project_id,))


# ---------------------------------------------------------------------------
# Network summary builder
# ---------------------------------------------------------------------------

def _build_network_summary(project_id: str, cur):
    """Aggregate unique networks from NIC data, classify type, extract VLAN IDs."""
    cur.execute("""
        SELECT network_name, count(DISTINCT vm_name) as vm_count
        FROM migration_vm_nics
        WHERE project_id = %s AND network_name IS NOT NULL AND network_name != ''
        GROUP BY network_name ORDER BY vm_count DESC
    """, (project_id,))
    networks = cur.fetchall()

    for net in networks:
        name = net["network_name"]
        vlan_id = extract_vlan_id(name)
        net_type = classify_network_type(name)
        cur.execute("""
            INSERT INTO migration_networks (project_id, network_name, vlan_id, network_type, vm_count)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (project_id, network_name) DO UPDATE SET
                vlan_id = EXCLUDED.vlan_id,
                network_type = EXCLUDED.network_type,
                vm_count = EXCLUDED.vm_count,
                updated_at = now()
        """, (project_id, name, vlan_id, net_type, net["vm_count"]))


# ---------------------------------------------------------------------------
# Tenant detection
# ---------------------------------------------------------------------------

def _run_tenant_detection(project_id: str, cur):
    """Run tenant detection for all VMs in a project."""
    # Get detection config
    cur.execute("""
        SELECT detection_config FROM migration_tenant_rules
        WHERE project_id = %s ORDER BY id DESC LIMIT 1
    """, (project_id,))
    row = cur.fetchone()
    if row:
        raw = row["detection_config"]
        detection_config = raw if isinstance(raw, dict) else json.loads(raw)
    else:
        detection_config = {}

    # Default detection: try vCD folder first, then vApp, folder path, resource pool
    default_methods = [
        {"method": "vcd_folder", "enabled": True},
        {"method": "vapp_name", "enabled": True},
        {"method": "folder_path", "enabled": True, "depth": 2},
        {"method": "resource_pool", "enabled": True},
        {"method": "cluster", "enabled": True},
    ]

    # Use stored methods, but ensure vcd_folder is present (may be missing in old configs)
    methods = detection_config.get("methods", default_methods)
    has_vcd_folder = any(m.get("method") == "vcd_folder" for m in methods)
    if not has_vcd_folder:
        methods.insert(0, {"method": "vcd_folder", "enabled": True})
    detection_config["methods"] = methods

    if "fallback_tenant" not in detection_config:
        detection_config["fallback_tenant"] = "Unassigned"
    if "orgvdc_detection" not in detection_config:
        detection_config["orgvdc_detection"] = {"use_resource_pool": True, "use_folder_depth3": True}

    # Get all VMs
    cur.execute("""
        SELECT vm_name, folder_path, resource_pool, vapp_name, annotation, cluster
        FROM migration_vms WHERE project_id = %s
    """, (project_id,))
    vms = [dict(row) for row in cur.fetchall()]

    # Detect tenants
    tenant_map: Dict[str, Dict] = {}
    for vm in vms:
        assignment = assign_tenant(vm, detection_config)

        # Update VM
        cur.execute("""
            UPDATE migration_vms SET
                tenant_name = %s,
                org_vdc = %s,
                app_group = %s,
                updated_at = now()
            WHERE project_id = %s AND vm_name = %s
        """, (assignment.tenant_name, assignment.org_vdc, assignment.app_group,
              project_id, vm["vm_name"]))

        # Aggregate tenant info
        key = (assignment.tenant_name, assignment.org_vdc or "")
        if key not in tenant_map:
            tenant_map[key] = {
                "tenant_name": assignment.tenant_name,
                "org_vdc": assignment.org_vdc,
                "detection_method": assignment.detection_method,
                "vm_count": 0,
            }
        tenant_map[key]["vm_count"] += 1

    # Upsert tenants — pre-seed target names from source name (confirmed=false = needs review)
    for (tname, ovdc), info in tenant_map.items():
        target_project = info["org_vdc"] or info["tenant_name"]
        cur.execute("""
            INSERT INTO migration_tenants
                (project_id, tenant_name, org_vdc, detection_method, vm_count,
                 target_domain_name, target_domain_description,
                 target_project_name, target_display_name, target_confirmed)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, false)
            ON CONFLICT (project_id, tenant_name, org_vdc) DO UPDATE SET
                detection_method = EXCLUDED.detection_method,
                vm_count = EXCLUDED.vm_count,
                updated_at = now()
                -- Do NOT overwrite target names / target_confirmed if already set by user
        """, (project_id, info["tenant_name"], info["org_vdc"] or None,
              info["detection_method"], info["vm_count"],
              info["tenant_name"],
              # target_domain_description: seeded as a description hint from domain name
              info["tenant_name"],
              # target_project_name: use OrgVDC if available (maps to PCD Project),
              # else fall back to tenant_name (non-vCloud or unknown VDC)
              target_project,
              # target_display_name: seeded as a description hint from target_project_name
              target_project))

    # Update tenant disk/ram/vcpu/in_use totals
    cur.execute("""
        UPDATE migration_tenants mt SET
            total_disk_gb = sub.disk_gb,
            total_in_use_gb = sub.in_use_gb,
            total_ram_mb = sub.ram_mb,
            total_vcpu = sub.vcpu
        FROM (
            SELECT tenant_name, coalesce(org_vdc, '') as ovdc,
                   coalesce(sum(total_disk_gb), 0) as disk_gb,
                   coalesce(sum(in_use_gb), 0) as in_use_gb,
                   coalesce(sum(ram_mb), 0) as ram_mb,
                   coalesce(sum(cpu_count), 0) as vcpu
            FROM migration_vms WHERE project_id = %s
            GROUP BY tenant_name, coalesce(org_vdc, '')
        ) sub
        WHERE mt.project_id = %s
          AND mt.tenant_name = sub.tenant_name
          AND coalesce(mt.org_vdc, '') = sub.ovdc
    """, (project_id, project_id))


# =====================================================================
# VM ENDPOINTS
# =====================================================================

@router.get("/projects/{project_id}/vms", dependencies=[Depends(require_permission("migration", "read"))])
async def list_vms(
    project_id: str,
    tenant: Optional[str] = Query(None),
    risk: Optional[str] = Query(None),
    mode: Optional[str] = Query(None),
    host: Optional[str] = Query(None),
    cluster: Optional[str] = Query(None),
    os_family: Optional[str] = Query(None),
    os_version: Optional[str] = Query(None),
    power_state: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    exclude_templates: bool = Query(True),
    sort_by: str = Query("vm_name"),
    sort_dir: str = Query("asc"),
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
):
    """List VMs in a migration project with filters."""
    allowed_sorts = {
        "vm_name", "risk_score", "total_disk_gb", "ram_mb", "cpu_count",
        "tenant_name", "host_name", "cluster", "migration_mode", "risk_category",
        "power_state", "priority", "nic_count", "disk_count", "in_use_mb",
        "in_use_gb", "partition_used_gb", "network_name", "os_version",
        "primary_ip", "dns_name", "os_family",
    }
    if sort_by not in allowed_sorts:
        sort_by = "vm_name"
    if sort_dir.lower() not in ("asc", "desc"):
        sort_dir = "asc"

    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            where = "WHERE project_id = %s"
            params: list = [project_id]

            if exclude_templates:
                where += " AND (template IS NULL OR template = false)"
            if tenant:
                where += " AND tenant_name = %s"
                params.append(tenant)
            if risk:
                where += " AND risk_category = %s"
                params.append(risk.upper())
            if mode:
                where += " AND migration_mode = %s"
                params.append(mode)
            if host:
                where += " AND host_name = %s"
                params.append(host)
            if cluster:
                where += " AND cluster = %s"
                params.append(cluster)
            if os_family:
                where += " AND os_family = %s"
                params.append(os_family)
            if os_version:
                where += " AND os_version ILIKE %s"
                params.append(f"%{os_version}%")
            if power_state:
                where += " AND power_state = %s"
                params.append(power_state)
            if search:
                where += " AND (vm_name ILIKE %s OR dns_name ILIKE %s OR primary_ip ILIKE %s)"
                pattern = f"%{search}%"
                params.extend([pattern, pattern, pattern])

            cur.execute(f"""
                SELECT * FROM migration_vms
                {where}
                ORDER BY {sort_by} {sort_dir}
                LIMIT %s OFFSET %s
            """, params + [limit, offset])
            vms = [_serialize_row(dict(r)) for r in cur.fetchall()]

            cur.execute(f"SELECT count(*) as cnt FROM migration_vms {where}", params)
            total = cur.fetchone()["cnt"]

    return {"status": "ok", "total": total, "vms": vms}


@router.patch("/projects/{project_id}/vms/{vm_id}",
              dependencies=[Depends(require_permission("migration", "write"))])
async def update_vm(project_id: str, vm_id: str, req: UpdateVMRequest, user = Depends(get_current_user)):
    """Update VM overrides (exclude, mode, priority, tenant)."""
    actor = user.username if user else "system"

    updates = {k: v for k, v in req.dict().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")

    set_clauses = [f"{k} = %s" for k in updates]
    set_clauses.append("updated_at = now()")
    params = list(updates.values())

    with _get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(f"""
                UPDATE migration_vms
                SET {', '.join(set_clauses)}
                WHERE project_id = %s AND vm_id = %s
                RETURNING *
            """, params + [project_id, vm_id])
            vm = cur.fetchone()
            if not vm:
                raise HTTPException(status_code=404, detail="VM not found")
            conn.commit()

    _log_activity(actor=actor, action="update_vm", resource_type="migration_vm",
                  resource_id=vm_id, details=updates)

    return {"status": "ok", "vm": _serialize_row(dict(vm))}


# =====================================================================
# TENANT ENDPOINTS
# =====================================================================

@router.get("/projects/{project_id}/tenants", dependencies=[Depends(require_permission("migration", "read"))])
async def list_tenants(project_id: str):
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT * FROM migration_tenants
                WHERE project_id = %s ORDER BY vm_count DESC
            """, (project_id,))
            tenants = [_serialize_row(dict(r)) for r in cur.fetchall()]
    return {"status": "ok", "tenants": tenants}


@router.patch("/projects/{project_id}/tenants/{tenant_id:int}",
              dependencies=[Depends(require_permission("migration", "write"))])
async def update_tenant(project_id: str, tenant_id: int, req: UpdateTenantRequest, user = Depends(get_current_user)):
    actor = user.username if user else "system"

    updates = {k: v for k, v in req.dict().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")

    set_clauses = [f"{k} = %s" for k in updates]
    set_clauses.append("updated_at = now()")
    params = list(updates.values())

    with _get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Get old tenant name first (for VM cascade)
            cur.execute("SELECT tenant_name FROM migration_tenants WHERE id = %s AND project_id = %s",
                        (tenant_id, project_id))
            old_row = cur.fetchone()
            if not old_row:
                raise HTTPException(status_code=404, detail="Tenant not found")
            old_tenant_name = old_row["tenant_name"]

            cur.execute(f"""
                UPDATE migration_tenants
                SET {', '.join(set_clauses)}
                WHERE id = %s AND project_id = %s
                RETURNING *
            """, params + [tenant_id, project_id])
            tenant = cur.fetchone()

            # If tenant_name changed, update VMs too (use OLD name for the WHERE)
            if "tenant_name" in updates and updates["tenant_name"] != old_tenant_name:
                cur.execute("""
                    UPDATE migration_vms SET tenant_name = %s, updated_at = now()
                    WHERE project_id = %s AND tenant_name = %s
                """, (updates["tenant_name"], project_id, old_tenant_name))

            conn.commit()

    _log_activity(actor=actor, action="update_tenant", resource_type="migration_tenant",
                  resource_id=str(tenant_id), details=updates)

    return {"status": "ok", "tenant": _serialize_row(dict(tenant))}


# =====================================================================
# NETWORK ENDPOINTS
# =====================================================================

@router.get("/projects/{project_id}/networks", dependencies=[Depends(require_permission("migration", "read"))])
async def list_networks(project_id: str):
    """List network infrastructure summary for a migration project."""
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT mn.*,
                       string_agg(DISTINCT vm.tenant_name, ', ' ORDER BY vm.tenant_name) as tenant_names,
                       count(DISTINCT vm.tenant_name) as tenant_count
                FROM migration_networks mn
                LEFT JOIN migration_vm_nics nic ON mn.project_id = nic.project_id 
                    AND mn.network_name = nic.network_name
                LEFT JOIN migration_vms vm ON nic.project_id = vm.project_id 
                    AND nic.vm_name = vm.vm_name
                WHERE mn.project_id = %s
                GROUP BY mn.id, mn.project_id, mn.network_name, mn.vlan_id, mn.network_type, 
                         mn.vm_count, mn.subnet, mn.gateway, mn.dns_servers, mn.ip_range, 
                         mn.pcd_target, mn.notes, mn.created_at, mn.updated_at
                ORDER BY mn.vm_count DESC
            """, (project_id,))
            networks = [_serialize_row(dict(r)) for r in cur.fetchall()]
    return {"status": "ok", "networks": networks}


@router.patch("/projects/{project_id}/networks/{network_id}",
              dependencies=[Depends(require_permission("migration", "write"))])
async def update_network(project_id: str, network_id: int, req: Request, user=Depends(get_current_user)):
    """Update editable network fields (subnet, gateway, dns_servers, ip_range, pcd_target, notes, network_type, vlan_id)."""
    actor = user.username if user else "system"
    body = await req.json()
    allowed = {"subnet", "gateway", "dns_servers", "ip_range", "pcd_target", "notes", "network_type", "vlan_id"}
    updates = {k: v for k, v in body.items() if k in allowed}
    if not updates:
        raise HTTPException(status_code=400, detail="No valid fields to update")

    set_clauses = [f"{k} = %s" for k in updates]
    set_clauses.append("updated_at = now()")
    params = list(updates.values())

    with _get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(f"""
                UPDATE migration_networks
                SET {', '.join(set_clauses)}
                WHERE id = %s AND project_id = %s
                RETURNING *
            """, params + [network_id, project_id])
            net = cur.fetchone()
            if not net:
                raise HTTPException(status_code=404, detail="Network not found")
            conn.commit()

    _log_activity(actor=actor, action="update_network", resource_type="migration_network",
                  resource_id=str(network_id), details=updates)

    return {"status": "ok", "network": _serialize_row(dict(net))}


# =====================================================================
# HOST / CLUSTER / STATS
# =====================================================================

@router.get("/projects/{project_id}/hosts", dependencies=[Depends(require_permission("migration", "read"))])
async def list_hosts(project_id: str):
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT * FROM migration_hosts
                WHERE project_id = %s ORDER BY host_name
            """, (project_id,))
            hosts = [_serialize_row(dict(r)) for r in cur.fetchall()]
    return {"status": "ok", "hosts": hosts}


@router.get("/projects/{project_id}/clusters", dependencies=[Depends(require_permission("migration", "read"))])
async def list_clusters(project_id: str):
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT * FROM migration_clusters
                WHERE project_id = %s ORDER BY cluster_name
            """, (project_id,))
            clusters = [_serialize_row(dict(r)) for r in cur.fetchall()]
    return {"status": "ok", "clusters": clusters}


@router.get("/projects/{project_id}/stats", dependencies=[Depends(require_permission("migration", "read"))])
async def get_stats(project_id: str):
    """Summary statistics for a migration project."""
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT
                    count(*) as total_vms,
                    count(*) FILTER (WHERE template = true) as templates,
                    count(*) FILTER (WHERE NOT coalesce(exclude_from_migration, false) AND NOT coalesce(template, false)) as migratable_vms,
                    count(*) FILTER (WHERE exclude_from_migration) as excluded_vms,
                    coalesce(sum(total_disk_gb), 0) as total_disk_gb,
                    coalesce(sum(ram_mb), 0) as total_ram_mb,
                    coalesce(sum(cpu_count), 0) as total_vcpus,
                    round(coalesce(sum(in_use_mb), 0) / 1024.0, 2) as total_in_use_gb,
                    count(DISTINCT host_name) as host_count,
                    count(DISTINCT cluster) as cluster_count,
                    count(DISTINCT tenant_name) as tenant_count,
                    count(*) FILTER (WHERE risk_category = 'GREEN') as risk_green,
                    count(*) FILTER (WHERE risk_category = 'YELLOW') as risk_yellow,
                    count(*) FILTER (WHERE risk_category = 'RED') as risk_red,
                    count(*) FILTER (WHERE migration_mode = 'warm_eligible') as mode_warm,
                    count(*) FILTER (WHERE migration_mode = 'warm_risky') as mode_warm_risky,
                    count(*) FILTER (WHERE migration_mode = 'cold_required') as mode_cold,
                    count(*) FILTER (WHERE power_state ILIKE '%%poweredOn%%') as powered_on,
                    count(*) FILTER (WHERE power_state ILIKE '%%poweredOff%%') as powered_off
                FROM migration_vms WHERE project_id = %s
            """, (project_id,))
            stats = _serialize_row(dict(cur.fetchone()))

            # ---- Build UI-friendly aliases ----
            stats["total_vcpu"] = stats["total_vcpus"]  # backward compat
            stats["total_provisioned_gb"] = stats["total_disk_gb"]

            # Risk & migration-mode distributions as dicts
            stats["risk_distribution"] = {
                "GREEN": stats.get("risk_green", 0),
                "YELLOW": stats.get("risk_yellow", 0),
                "RED": stats.get("risk_red", 0),
            }
            stats["mode_distribution"] = {
                "warm_eligible": stats.get("mode_warm", 0),
                "warm_risky": stats.get("mode_warm_risky", 0),
                "cold_required": stats.get("mode_cold", 0),
            }

            # OS family distribution
            cur.execute("""
                SELECT os_family, count(*) as cnt
                FROM migration_vms WHERE project_id = %s
                GROUP BY os_family ORDER BY cnt DESC
            """, (project_id,))
            os_fam = {r["os_family"]: r["cnt"] for r in cur.fetchall()}
            stats["os_families"] = os_fam
            stats["os_distribution"] = os_fam  # alias for UI

            # Top 10 largest VMs
            cur.execute("""
                SELECT vm_name, total_disk_gb, ram_mb, cpu_count, risk_category
                FROM migration_vms WHERE project_id = %s
                ORDER BY total_disk_gb DESC NULLS LAST LIMIT 10
            """, (project_id,))
            stats["largest_vms"] = [_serialize_row(dict(r)) for r in cur.fetchall()]

    return {"status": "ok", "stats": stats}


# =====================================================================
# RISK SCORING & ASSESSMENT
# =====================================================================

@router.get("/projects/{project_id}/risk-config",
            dependencies=[Depends(require_permission("migration", "read"))])
async def get_risk_config(project_id: str):
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT * FROM migration_risk_config
                WHERE project_id = %s ORDER BY id DESC LIMIT 1
            """, (project_id,))
            config = cur.fetchone()
            if not config:
                raise HTTPException(status_code=404, detail="Risk config not found")
    return {"status": "ok", "config": _serialize_row(dict(config))}


@router.put("/projects/{project_id}/risk-config",
            dependencies=[Depends(require_permission("migration", "write"))])
async def update_risk_config(project_id: str, request: Request, user = Depends(get_current_user)):
    actor = user.username if user else "system"
    body = await request.json()
    rules = body.get("rules")
    if not rules:
        raise HTTPException(status_code=400, detail="rules field required")

    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                UPDATE migration_risk_config SET rules = %s
                WHERE project_id = %s
                RETURNING *
            """, (Json(rules), project_id))
            config = cur.fetchone()
            if not config:
                cur.execute("""
                    INSERT INTO migration_risk_config (project_id, rules)
                    VALUES (%s, %s) RETURNING *
                """, (project_id, Json(rules)))
                config = cur.fetchone()
            conn.commit()

    _log_activity(actor=actor, action="update_risk_config", resource_type="migration_project",
                  resource_id=project_id)

    return {"status": "ok", "config": _serialize_row(dict(config))}


@router.post("/projects/{project_id}/assess",
             dependencies=[Depends(require_permission("migration", "write"))])
async def run_assessment(project_id: str, user = Depends(get_current_user)):
    """Run risk scoring and warm/cold classification for all VMs."""
    actor = user.username if user else "system"

    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Get risk rules
            cur.execute("""
                SELECT rules FROM migration_risk_config
                WHERE project_id = %s ORDER BY id DESC LIMIT 1
            """, (project_id,))
            config_row = cur.fetchone()
            rules = config_row["rules"] if config_row else {}
            if isinstance(rules, str):
                rules = json.loads(rules)

            # Get all VMs
            cur.execute("""
                SELECT * FROM migration_vms WHERE project_id = %s
            """, (project_id,))
            vms = [dict(r) for r in cur.fetchall()]

            scored = 0
            for vm in vms:
                # Risk scoring
                risk = compute_risk(vm, rules)
                # Migration mode
                mode = classify_migration_mode(vm, rules)

                cur.execute("""
                    UPDATE migration_vms SET
                        risk_score = %s,
                        risk_category = %s,
                        risk_reasons = %s,
                        migration_mode = %s,
                        mode_reasons = %s,
                        updated_at = now()
                    WHERE project_id = %s AND vm_id = %s
                """, (
                    risk.score, risk.category, Json(risk.reasons),
                    mode.mode, Json(mode.reasons),
                    project_id, vm["vm_id"],
                ))
                scored += 1

            # Update project status
            cur.execute("""
                UPDATE migration_projects SET
                    status = CASE WHEN status IN ('draft', 'assessment') THEN 'assessment' ELSE status END,
                    updated_at = now()
                WHERE project_id = %s
            """, (project_id,))

            conn.commit()

    _log_activity(actor=actor, action="run_assessment", resource_type="migration_project",
                  resource_id=project_id, details={"vms_scored": scored})

    return {"status": "ok", "message": f"Assessment complete: {scored} VMs scored"}


# =====================================================================
# RESET ENDPOINTS
# =====================================================================

@router.post("/projects/{project_id}/reset-assessment",
             dependencies=[Depends(require_permission("migration", "write"))])
async def reset_assessment(project_id: str, user = Depends(get_current_user)):
    """Clear all computed risk/mode scores. Keep source data."""
    actor = user.username if user else "system"

    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE migration_vms SET
                    risk_score = NULL, risk_category = NULL, risk_reasons = '[]',
                    migration_mode = NULL, mode_reasons = '[]',
                    phase1_duration_hours = NULL, cutover_downtime_hours = NULL,
                    total_migration_hours = NULL, production_impact = NULL,
                    target_flavor = NULL, target_flavor_id = NULL,
                    target_network = NULL, target_project = NULL,
                    updated_at = now()
                WHERE project_id = %s
            """, (project_id,))

            # Also clear waves, gaps, prep tasks
            for table in ("migration_wave_vms", "migration_waves",
                          "migration_prep_tasks", "migration_target_gaps"):
                cur.execute(f"DELETE FROM {table} WHERE project_id = %s", (project_id,))

            cur.execute("""
                UPDATE migration_projects SET status = 'assessment', updated_at = now()
                WHERE project_id = %s
            """, (project_id,))
            conn.commit()

    _log_activity(actor=actor, action="reset_assessment", resource_type="migration_project",
                  resource_id=project_id)

    return {"status": "ok", "message": "Assessment data cleared"}


@router.post("/projects/{project_id}/reset-plan",
             dependencies=[Depends(require_permission("migration", "write"))])
async def reset_plan(project_id: str, user = Depends(get_current_user)):
    """Clear wave/prep data only. Keep assessment results."""
    actor = user.username if user else "system"

    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor() as cur:
            for table in ("migration_wave_vms", "migration_waves",
                          "migration_prep_tasks", "migration_target_gaps"):
                cur.execute(f"DELETE FROM {table} WHERE project_id = %s", (project_id,))

            cur.execute("""
                UPDATE migration_projects
                SET status = CASE WHEN status IN ('planned', 'approved', 'preparing', 'ready')
                                  THEN 'assessment' ELSE status END,
                    updated_at = now()
                WHERE project_id = %s
            """, (project_id,))
            conn.commit()

    _log_activity(actor=actor, action="reset_plan", resource_type="migration_project",
                  resource_id=project_id)

    return {"status": "ok", "message": "Plan data cleared"}


# =====================================================================
# APPROVAL GATE
# =====================================================================

@router.post("/projects/{project_id}/approve",
             dependencies=[Depends(require_permission("migration", "admin"))])
async def approve_project(project_id: str, user = Depends(get_current_user)):
    """Approve a project — gate before any PCD write operations."""
    actor = user.username if user else "system"

    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        if project["status"] not in ("planned", "assessment"):
            raise HTTPException(
                status_code=400,
                detail=f"Cannot approve project in status '{project['status']}'. Must be 'planned' or 'assessment'."
            )

        with conn.cursor() as cur:
            cur.execute("""
                UPDATE migration_projects SET
                    status = 'approved',
                    approved_by = %s,
                    approved_at = now(),
                    updated_at = now()
                WHERE project_id = %s
            """, (actor, project_id))
            conn.commit()

    _log_activity(actor=actor, action="approve", resource_type="migration_project",
                  resource_id=project_id, resource_name=project["name"])

    return {"status": "ok", "message": f"Project approved by {actor}"}


# =====================================================================
# BANDWIDTH & AGENT RECOMMENDATION
# =====================================================================

@router.get("/projects/{project_id}/bandwidth",
            dependencies=[Depends(require_permission("migration", "read"))])
async def get_bandwidth_model(project_id: str):
    """Compute the bandwidth constraint model for the project."""
    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        model = compute_bandwidth_model(project)

    return {
        "status": "ok",
        "bandwidth": {
            "source_effective_mbps": round(model.source_effective_mbps, 1),
            "link_effective_mbps": round(model.link_effective_mbps, 1) if model.link_effective_mbps else None,
            "agent_effective_mbps": round(model.agent_effective_mbps, 1),
            "storage_effective_mbps": round(model.storage_effective_mbps, 1),
            "bottleneck": model.bottleneck,
            "bottleneck_mbps": round(model.bottleneck_mbps, 1),
            "topology_type": project["topology_type"],
        },
    }


@router.get("/projects/{project_id}/agent-recommendation",
            dependencies=[Depends(require_permission("migration", "read"))])
async def get_agent_recommendation(project_id: str):
    """Compute recommended vJailbreak agent sizing."""
    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT count(*) as vm_count,
                       coalesce(max(total_disk_gb), 0) as max_disk_gb,
                       coalesce(sum(total_disk_gb), 0) as total_disk_gb
                FROM migration_vms
                WHERE project_id = %s AND NOT coalesce(exclude_from_migration, false)
                  AND NOT coalesce(template, false)
            """, (project_id,))
            summary = dict(cur.fetchone())

            cur.execute("""
                SELECT total_disk_gb FROM migration_vms
                WHERE project_id = %s AND NOT coalesce(exclude_from_migration, false)
                  AND NOT coalesce(template, false)
                ORDER BY total_disk_gb DESC NULLS LAST LIMIT 10
            """, (project_id,))
            top_disks = [float(r["total_disk_gb"] or 0) for r in cur.fetchall()]

    rec = recommend_agent_sizing(
        vm_count=summary["vm_count"],
        largest_disk_gb=float(summary["max_disk_gb"]),
        top5_disk_sizes_gb=top_disks[:5],
        project_settings=project,
        total_disk_gb=float(summary["total_disk_gb"]),
    )

    return {
        "status": "ok",
        "recommendation": {
            "recommended_agent_count": rec.recommended_count,
            "vcpu_per_agent": rec.vcpu_per_agent,
            "ram_gb_per_agent": rec.ram_gb_per_agent,
            "disk_gb_per_agent": round(rec.disk_gb_per_agent, 0),
            "max_concurrent_vms": rec.max_concurrent_vms,
            "reasoning": rec.reasoning,
        },
    }


# =====================================================================
# VM DETAIL (Disks + NICs per VM)
# =====================================================================

@router.get("/projects/{project_id}/vms/{vm_name}/details",
            dependencies=[Depends(require_permission("migration", "read"))])
async def get_vm_details(project_id: str, vm_name: str):
    """Get detailed disk and NIC info for a specific VM."""
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT disk_label, disk_path, capacity_gb, thin_provisioned,
                       eagerly_scrub, datastore
                FROM migration_vm_disks
                WHERE project_id = %s AND vm_name = %s
                ORDER BY disk_label
            """, (project_id, vm_name))
            disks = [_serialize_row(dict(r)) for r in cur.fetchall()]

            cur.execute("""
                SELECT nic_label, adapter_type, network_name, connected,
                       mac_address, ip_address
                FROM migration_vm_nics
                WHERE project_id = %s AND vm_name = %s
                ORDER BY nic_label
            """, (project_id, vm_name))
            nics = [_serialize_row(dict(r)) for r in cur.fetchall()]

    return {"status": "ok", "disks": disks, "nics": nics}


# =====================================================================
# MIGRATION PLAN EXPORT
# =====================================================================

@router.get("/projects/{project_id}/export-plan",
            dependencies=[Depends(require_permission("migration", "read"))])
async def export_migration_plan(project_id: str):
    """
    Generate a comprehensive migration plan with per-tenant assessment,
    daily VM schedule, and per-VM time estimates (warm/cold).
    Only includes tenants and VMs where include_in_plan = true.
    """
    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Included tenants only
            cur.execute("""
                SELECT * FROM migration_tenants
                WHERE project_id = %s AND include_in_plan = true
                ORDER BY vm_count DESC
            """, (project_id,))
            tenants = [dict(r) for r in cur.fetchall()]
            included_names = {t["tenant_name"] for t in tenants}

            # Excluded tenants count for project summary
            cur.execute("""
                SELECT count(*) FROM migration_tenants
                WHERE project_id = %s AND include_in_plan = false
            """, (project_id,))
            excluded_count = cur.fetchone()["count"]

            # Only VMs belonging to included tenants
            cur.execute("""
                SELECT v.* FROM migration_vms v
                JOIN migration_tenants t ON t.project_id = v.project_id
                    AND t.tenant_name = v.tenant_name
                WHERE v.project_id = %s
                  AND NOT coalesce(v.template, false)
                  AND t.include_in_plan = true
                ORDER BY v.tenant_name, v.priority, v.total_disk_gb DESC NULLS LAST
            """, (project_id,))
            vms = [dict(r) for r in cur.fetchall()]

    bw_model = compute_bandwidth_model(project)
    plan = generate_migration_plan(
        vms=vms,
        tenants=tenants,
        project_settings=project,
        bottleneck_mbps=bw_model.bottleneck_mbps,
    )
    plan["project_summary"]["excluded_tenants"] = int(excluded_count)

    return {
        "status": "ok",
        "project_name": project["name"],
        "topology_type": project["topology_type"],
        **plan,
    }


@router.get("/projects/{project_id}/export-report.xlsx",
            dependencies=[Depends(require_permission("migration", "read"))])
async def export_report_excel(project_id: str):
    """Download a full migration plan report as an Excel workbook (4 sheets)."""
    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT * FROM migration_tenants
                WHERE project_id = %s AND include_in_plan = true
                ORDER BY vm_count DESC
            """, (project_id,))
            tenants = [dict(r) for r in cur.fetchall()]
            cur.execute("""
                SELECT v.* FROM migration_vms v
                JOIN migration_tenants t ON t.project_id = v.project_id
                    AND t.tenant_name = v.tenant_name
                WHERE v.project_id = %s AND NOT coalesce(v.template, false)
                  AND t.include_in_plan = true
                ORDER BY v.tenant_name, v.priority, v.total_disk_gb DESC NULLS LAST
            """, (project_id,))
            vms = [dict(r) for r in cur.fetchall()]

    bw_model = compute_bandwidth_model(project)
    plan = generate_migration_plan(
        vms=vms, tenants=tenants, project_settings=project,
        bottleneck_mbps=bw_model.bottleneck_mbps,
    )
    plan["bandwidth_model"] = {
        "bottleneck": bw_model.bottleneck,
        "source_effective_mbps": round(bw_model.source_effective_mbps, 1),
        "link_effective_mbps": round(bw_model.link_effective_mbps, 1),
        "agent_effective_mbps": round(bw_model.agent_effective_mbps, 1),
        "storage_effective_mbps": round(bw_model.storage_effective_mbps, 1),
    }

    xlsx_bytes = generate_excel_report(plan, project["name"])
    safe_name = project["name"].replace(" ", "_")
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="migration-plan-{safe_name}.xlsx"'},
    )


@router.get("/projects/{project_id}/export-report.pdf",
            dependencies=[Depends(require_permission("migration", "read"))])
async def export_report_pdf(project_id: str):
    """Download a full migration plan report as a PDF (landscape A4)."""
    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT * FROM migration_tenants
                WHERE project_id = %s AND include_in_plan = true
                ORDER BY vm_count DESC
            """, (project_id,))
            tenants = [dict(r) for r in cur.fetchall()]
            cur.execute("""
                SELECT v.* FROM migration_vms v
                JOIN migration_tenants t ON t.project_id = v.project_id
                    AND t.tenant_name = v.tenant_name
                WHERE v.project_id = %s AND NOT coalesce(v.template, false)
                  AND t.include_in_plan = true
                ORDER BY v.tenant_name, v.priority, v.total_disk_gb DESC NULLS LAST
            """, (project_id,))
            vms = [dict(r) for r in cur.fetchall()]

    bw_model = compute_bandwidth_model(project)
    plan = generate_migration_plan(
        vms=vms, tenants=tenants, project_settings=project,
        bottleneck_mbps=bw_model.bottleneck_mbps,
    )
    plan["bandwidth_model"] = {
        "bottleneck": bw_model.bottleneck,
        "source_effective_mbps": round(bw_model.source_effective_mbps, 1),
        "link_effective_mbps": round(bw_model.link_effective_mbps, 1),
        "agent_effective_mbps": round(bw_model.agent_effective_mbps, 1),
        "storage_effective_mbps": round(bw_model.storage_effective_mbps, 1),
    }

    pdf_bytes = generate_pdf_report(plan, project["name"])
    safe_name = project["name"].replace(" ", "_")
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="migration-plan-{safe_name}.pdf"'},
    )


# =====================================================================
# Phase 2A — Tenant Exclusion / Include-in-plan
# =====================================================================

class BulkScopeRequest(BaseModel):
    tenant_ids: List[int]
    include_in_plan: bool
    exclude_reason: Optional[str] = None


class BulkReplaceTargetRequest(BaseModel):
    field: str                       # "target_domain_name" | "target_project_name"
    find: str                        # literal substring to find
    replace: str                     # replacement string (empty string = strip)
    case_sensitive: bool = False     # default: case-insensitive
    unconfirmed_only: bool = False   # if true, only affect rows not yet confirmed
    preview_only: bool = False       # if true, return preview without writing


class BulkReplaceNetworkRequest(BaseModel):
    find: str                        # literal substring to find
    replace: str                     # replacement string (empty string = strip)
    case_sensitive: bool = False     # default: case-insensitive
    unconfirmed_only: bool = False   # if true, only affect rows where confirmed=false
    preview_only: bool = False       # if true, return preview without writing


@router.patch("/projects/{project_id}/tenants/bulk-scope",
              dependencies=[Depends(require_permission("migration", "write"))])
async def bulk_scope_tenants(project_id: str, req: BulkScopeRequest, user=Depends(get_current_user)):
    """Include or exclude multiple tenants from the migration plan at once."""
    actor = user.username if user else "system"
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor() as cur:
            for tid in req.tenant_ids:
                cur.execute("""
                    UPDATE migration_tenants
                    SET include_in_plan = %s,
                        exclude_reason  = %s,
                        updated_at      = now()
                    WHERE id = %s AND project_id = %s
                """, (req.include_in_plan, req.exclude_reason if not req.include_in_plan else None,
                      tid, project_id))
        conn.commit()
    _log_activity(actor=actor, action="bulk_scope_tenants", resource_type="migration_project",
                  resource_id=project_id, details={"tenant_ids": req.tenant_ids, "include": req.include_in_plan})
    return {"status": "ok", "updated": len(req.tenant_ids)}


@router.post("/projects/{project_id}/tenants/bulk-replace-target",
             dependencies=[Depends(require_permission("migration", "write"))])
async def bulk_replace_target(project_id: str, req: BulkReplaceTargetRequest,
                               user=Depends(get_current_user)):
    """
    Find-and-replace in target_domain_name or target_project_name across all tenants.
    Supports case-insensitive matching (default) and preview mode (dry run).
    Matching is literal substring — not regex — to keep it safe and predictable.
    After apply, affected rows are marked target_confirmed=false so operator reviews the result.
    """
    ALLOWED_FIELDS = {"target_domain_name", "target_domain_description",
                       "target_project_name", "target_display_name"}
    if req.field not in ALLOWED_FIELDS:
        raise HTTPException(status_code=400, detail=f"field must be one of {ALLOWED_FIELDS}")
    if not req.find:
        raise HTTPException(status_code=400, detail="find string must not be empty")

    actor = user.username if user else "system"
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Fetch candidates — rows where the target field contains the find substring
            if req.unconfirmed_only:
                filter_sql = "AND target_confirmed = false"
            else:
                filter_sql = ""

            cur.execute(f"""
                SELECT id, tenant_name, org_vdc,
                       {req.field} AS current_value
                FROM migration_tenants
                WHERE project_id = %s
                  AND {req.field} IS NOT NULL
                  AND {req.field} != ''
                  {filter_sql}
                ORDER BY tenant_name
            """, (project_id,))
            rows = cur.fetchall()

            # Compute replacements in Python (safe from SQL injection, predictable)
            preview = []
            for r in rows:
                old = r["current_value"] or ""
                if req.case_sensitive:
                    new = old.replace(req.find, req.replace)
                else:
                    new = re.sub(re.escape(req.find), req.replace,
                                 old, flags=re.IGNORECASE)
                if old != new:
                    preview.append({
                        "id": r["id"],
                        "tenant_name": r["tenant_name"],
                        "org_vdc": r["org_vdc"],
                        "old_value": old,
                        "new_value": new,
                    })

            if req.preview_only:
                return {"status": "ok", "preview": preview, "affected_count": len(preview)}

            # Apply updates
            for item in preview:
                cur.execute(f"""
                    UPDATE migration_tenants
                    SET {req.field} = %s,
                        target_confirmed = false,
                        updated_at = now()
                    WHERE id = %s AND project_id = %s
                """, (item["new_value"], item["id"], project_id))
            conn.commit()

    _log_activity(actor=actor, action="bulk_replace_target", resource_type="migration_project",
                  resource_id=project_id,
                  details={"field": req.field, "find": req.find, "replace": req.replace,
                           "affected": len(preview)})
    return {"status": "ok", "affected_count": len(preview), "preview": preview}


@router.post("/projects/{project_id}/network-mappings/bulk-replace",
             dependencies=[Depends(require_permission("migration", "write"))])
async def bulk_replace_network(project_id: str, req: BulkReplaceNetworkRequest,
                               user=Depends(get_current_user)):
    """
    Find-and-replace in target_network_name across all network mappings.
    Supports case-insensitive matching (default) and preview mode (dry run).
    Matching is literal substring — not regex — to keep it safe and predictable.
    After apply, affected rows are marked confirmed=false so operator reviews the result.
    """
    if not req.find:
        raise HTTPException(status_code=400, detail="find string must not be empty")

    actor = user.username if user else "system"
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            filter_sql = "AND confirmed = false" if req.unconfirmed_only else ""
            cur.execute(f"""
                SELECT id, source_network_name, target_network_name
                FROM migration_network_mappings
                WHERE project_id = %s
                  AND target_network_name IS NOT NULL
                  AND target_network_name != ''
                  {filter_sql}
                ORDER BY source_network_name
            """, (project_id,))
            rows = cur.fetchall()

            preview = []
            for r in rows:
                old = r["target_network_name"] or ""
                if req.case_sensitive:
                    new = old.replace(req.find, req.replace)
                else:
                    new = re.sub(re.escape(req.find), req.replace,
                                 old, flags=re.IGNORECASE)
                if old != new:
                    preview.append({
                        "id": r["id"],
                        "source_network_name": r["source_network_name"],
                        "old_value": old,
                        "new_value": new,
                    })

            if req.preview_only:
                return {"status": "ok", "preview": preview, "affected_count": len(preview)}

            for item in preview:
                cur.execute("""
                    UPDATE migration_network_mappings
                    SET target_network_name = %s,
                        confirmed = false,
                        updated_at = now()
                    WHERE id = %s AND project_id = %s
                """, (item["new_value"], item["id"], project_id))
            conn.commit()

    _log_activity(actor=actor, action="bulk_replace_network", resource_type="migration_project",
                  resource_id=project_id,
                  details={"find": req.find, "replace": req.replace, "affected": len(preview)})
    return {"status": "ok", "affected_count": len(preview), "preview": preview}


@router.post("/projects/{project_id}/tenants/confirm-all",
             dependencies=[Depends(require_permission("migration", "write"))])
async def confirm_all_tenants(project_id: str, user=Depends(get_current_user)):
    """Mark target_confirmed=true on every tenant in the project that is currently unconfirmed."""
    actor = user.username if user else "system"
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE migration_tenants
                SET target_confirmed = true, updated_at = now()
                WHERE project_id = %s AND target_confirmed = false
            """, (project_id,))
            affected = cur.rowcount
        conn.commit()
    _log_activity(actor=actor, action="confirm_all_tenants", resource_type="migration_project",
                  resource_id=project_id, details={"affected": affected})
    return {"status": "ok", "affected_count": affected}


@router.post("/projects/{project_id}/network-mappings/confirm-all",
             dependencies=[Depends(require_permission("migration", "write"))])
async def confirm_all_networks(project_id: str, user=Depends(get_current_user)):
    """Mark confirmed=true on every network mapping in the project that is currently unconfirmed."""
    actor = user.username if user else "system"
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE migration_network_mappings
                SET confirmed = true, updated_at = now()
                WHERE project_id = %s AND confirmed = false
            """, (project_id,))
            affected = cur.rowcount
        conn.commit()
    _log_activity(actor=actor, action="confirm_all_networks", resource_type="migration_project",
                  resource_id=project_id, details={"affected": affected})
    return {"status": "ok", "affected_count": affected}


# ── Auto-exclude filter patterns ─────────────────────────────────────────────

class TenantFilterRequest(BaseModel):
    pattern: str
    reason: Optional[str] = None
    auto_exclude: Optional[bool] = True


@router.get("/projects/{project_id}/tenant-filters",
            dependencies=[Depends(require_permission("migration", "read"))])
async def list_tenant_filters(project_id: str):
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM migration_tenant_filters WHERE project_id = %s ORDER BY id", (project_id,))
            filters = [_serialize_row(dict(r)) for r in cur.fetchall()]
    return {"status": "ok", "filters": filters}


@router.post("/projects/{project_id}/tenant-filters",
             dependencies=[Depends(require_permission("migration", "write"))])
async def add_tenant_filter(project_id: str, req: TenantFilterRequest, user=Depends(get_current_user)):
    """Add an auto-exclude pattern (e.g. 'LAB-%') and optionally apply it now."""
    actor = user.username if user else "system"
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                INSERT INTO migration_tenant_filters (project_id, pattern, reason, auto_exclude)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (project_id, pattern) DO UPDATE
                  SET reason = EXCLUDED.reason, auto_exclude = EXCLUDED.auto_exclude
                RETURNING *
            """, (project_id, req.pattern.strip(), req.reason, req.auto_exclude))
            row = _serialize_row(dict(cur.fetchone()))

            if req.auto_exclude:
                # Apply the pattern immediately: exclude matching tenants
                # Simple glob: only supports % wildcard → convert to SQL LIKE
                sql_pattern = req.pattern.strip().replace("*", "%")
                cur.execute("""
                    UPDATE migration_tenants
                    SET include_in_plan = false, exclude_reason = %s, updated_at = now()
                    WHERE project_id = %s AND tenant_name ILIKE %s AND include_in_plan = true
                """, (req.reason or f"Matched pattern: {req.pattern}", project_id, sql_pattern))
                affected = cur.rowcount
            else:
                affected = 0

        conn.commit()
    _log_activity(actor=actor, action="add_tenant_filter", resource_type="migration_project",
                  resource_id=project_id, details={"pattern": req.pattern, "affected": affected})
    return {"status": "ok", "filter": row, "tenants_excluded": affected}


@router.delete("/projects/{project_id}/tenant-filters/{filter_id}",
               dependencies=[Depends(require_permission("migration", "write"))])
async def delete_tenant_filter(project_id: str, filter_id: int, user=Depends(get_current_user)):
    actor = user.username if user else "system"
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor() as cur:
            cur.execute("DELETE FROM migration_tenant_filters WHERE id = %s AND project_id = %s",
                        (filter_id, project_id))
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="Filter not found")
        conn.commit()
    _log_activity(actor=actor, action="delete_tenant_filter", resource_type="migration_project",
                  resource_id=project_id, details={"filter_id": filter_id})
    return {"status": "ok"}


# =====================================================================
# Phase 2C — Quota & Overcommit Modeling
# =====================================================================

@router.get("/overcommit-profiles",
            dependencies=[Depends(require_permission("migration", "read"))])
async def list_overcommit_profiles():
    """List all overcommit profiles (seeded presets + any custom ones)."""
    with _get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM migration_overcommit_profiles ORDER BY id")
            profiles = [_serialize_row(dict(r)) for r in cur.fetchall()]
    return {"status": "ok", "profiles": profiles}


class UpdateOvercommitRequest(BaseModel):
    overcommit_profile_name: str


@router.patch("/projects/{project_id}/overcommit-profile",
              dependencies=[Depends(require_permission("migration", "write"))])
async def set_overcommit_profile(project_id: str, req: UpdateOvercommitRequest, user=Depends(get_current_user)):
    """Set the active overcommit profile for a project."""
    actor = user.username if user else "system"
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM migration_overcommit_profiles WHERE profile_name = %s",
                        (req.overcommit_profile_name,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail=f"Profile '{req.overcommit_profile_name}' not found")
            cur.execute("""
                UPDATE migration_projects SET overcommit_profile_name = %s, updated_at = now()
                WHERE project_id = %s
            """, (req.overcommit_profile_name, project_id))
        conn.commit()
    _log_activity(actor=actor, action="set_overcommit_profile", resource_type="migration_project",
                  resource_id=project_id, details={"profile": req.overcommit_profile_name})
    return {"status": "ok"}


@router.get("/projects/{project_id}/quota-requirements",
            dependencies=[Depends(require_permission("migration", "read"))])
async def get_quota_requirements(project_id: str):
    """
    Compute per-tenant and aggregate quota requirements.
    Uses the project's selected overcommit profile.
    Only includes tenants where include_in_plan = true.
    """
    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT * FROM migration_tenants
                WHERE project_id = %s AND include_in_plan = true
                ORDER BY vm_count DESC
            """, (project_id,))
            tenants = [dict(r) for r in cur.fetchall()]

            # Actual RAM usage per tenant+org_vdc from vMemory telemetry
            # Key by (tenant_name, org_vdc) so that tenants split across multiple
            # org_vdcs (e.g. Autosoft2) don't each get the full aggregate total.
            cur.execute("""
                SELECT tenant_name,
                       COALESCE(org_vdc, '') AS org_vdc,
                       ROUND(SUM(COALESCE(memory_usage_mb, 0)) / 1024.0, 1) AS ram_used_gb
                FROM migration_vms
                WHERE project_id = %s
                  AND memory_usage_mb IS NOT NULL AND memory_usage_mb > 0
                GROUP BY tenant_name, COALESCE(org_vdc, '')
            """, (project_id,))
            ram_used_map = {
                (r["tenant_name"], r["org_vdc"]): float(r["ram_used_gb"])
                for r in cur.fetchall()
            }
            for t in tenants:
                key = (t.get("tenant_name", ""), t.get("org_vdc") or "")
                t["ram_used_gb"] = ram_used_map.get(key, 0.0)

            profile_name = project.get("overcommit_profile_name") or "balanced"
            cur.execute("SELECT * FROM migration_overcommit_profiles WHERE profile_name = %s", (profile_name,))
            profile_row = cur.fetchone()
            if profile_row:
                profile = dict(profile_row)
            else:
                profile = {"profile_name": "balanced", "cpu_ratio": 4.0, "ram_ratio": 1.5, "disk_snapshot_factor": 1.5}

    result = compute_quota_requirements(tenants, profile)
    return {"status": "ok", "quota": result}


# =====================================================================
# Phase 2D — PCD Hardware Node Sizing
# =====================================================================

class NodeProfileRequest(BaseModel):
    profile_name: str
    cpu_cores: Optional[int] = 48
    cpu_threads: Optional[int] = 96
    ram_gb: Optional[float] = 384.0
    storage_tb: Optional[float] = 20.0
    max_cpu_util_pct: Optional[float] = 70.0
    max_ram_util_pct: Optional[float] = 75.0
    max_disk_util_pct: Optional[float] = 70.0
    is_default: Optional[bool] = False


@router.get("/projects/{project_id}/node-profiles",
            dependencies=[Depends(require_permission("migration", "read"))])
async def list_node_profiles(project_id: str):
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM migration_pcd_node_profiles WHERE project_id = %s ORDER BY id",
                        (project_id,))
            profiles = [_serialize_row(dict(r)) for r in cur.fetchall()]
    return {"status": "ok", "profiles": profiles}


@router.post("/projects/{project_id}/node-profiles",
             dependencies=[Depends(require_permission("migration", "write"))])
async def create_node_profile(project_id: str, req: NodeProfileRequest, user=Depends(get_current_user)):
    actor = user.username if user else "system"
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            if req.is_default:
                cur.execute("UPDATE migration_pcd_node_profiles SET is_default = false WHERE project_id = %s",
                            (project_id,))
            cur.execute("""
                INSERT INTO migration_pcd_node_profiles
                  (project_id, profile_name, cpu_cores, cpu_threads, ram_gb, storage_tb,
                   max_cpu_util_pct, max_ram_util_pct, max_disk_util_pct, is_default)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (project_id, profile_name) DO UPDATE
                  SET cpu_cores=EXCLUDED.cpu_cores, cpu_threads=EXCLUDED.cpu_threads,
                      ram_gb=EXCLUDED.ram_gb, storage_tb=EXCLUDED.storage_tb,
                      max_cpu_util_pct=EXCLUDED.max_cpu_util_pct,
                      max_ram_util_pct=EXCLUDED.max_ram_util_pct,
                      max_disk_util_pct=EXCLUDED.max_disk_util_pct,
                      is_default=EXCLUDED.is_default, updated_at=now()
                RETURNING *
            """, (project_id, req.profile_name, req.cpu_cores, req.cpu_threads, req.ram_gb,
                  req.storage_tb, req.max_cpu_util_pct, req.max_ram_util_pct, req.max_disk_util_pct,
                  req.is_default))
            profile = _serialize_row(dict(cur.fetchone()))
        conn.commit()
    _log_activity(actor=actor, action="create_node_profile", resource_type="migration_project",
                  resource_id=project_id, details={"profile_name": req.profile_name})
    return {"status": "ok", "profile": profile}


@router.delete("/projects/{project_id}/node-profiles/{profile_id}",
               dependencies=[Depends(require_permission("migration", "write"))])
async def delete_node_profile(project_id: str, profile_id: int, user=Depends(get_current_user)):
    actor = user.username if user else "system"
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor() as cur:
            cur.execute("DELETE FROM migration_pcd_node_profiles WHERE id = %s AND project_id = %s",
                        (profile_id, project_id))
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="Node profile not found")
        conn.commit()
    _log_activity(actor=actor, action="delete_node_profile", resource_type="migration_project",
                  resource_id=project_id, details={"profile_id": profile_id})
    return {"status": "ok"}


class NodeInventoryRequest(BaseModel):
    profile_id: Optional[int] = None
    current_nodes: int = 0
    current_vcpu_used: Optional[int] = 0
    current_ram_gb_used: Optional[float] = 0.0
    current_disk_tb_used: Optional[float] = 0.0
    notes: Optional[str] = None


@router.get("/projects/{project_id}/node-inventory",
            dependencies=[Depends(require_permission("migration", "read"))])
async def get_node_inventory(project_id: str):
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM migration_pcd_node_inventory WHERE project_id = %s", (project_id,))
            row = cur.fetchone()
    return {"status": "ok", "inventory": _serialize_row(dict(row)) if row else None}


@router.get("/projects/{project_id}/pcd-live-inventory",
            dependencies=[Depends(require_permission("migration", "read"))])
async def get_pcd_live_inventory(project_id: str):
    """
    Auto-discover current PCD cluster capacity from the hypervisors inventory table
    (populated by pf9_rvtools.py).  Also pulls committed vCPU/RAM from servers+flavors.
    Returns a 'live' snapshot the UI can offer to sync into the manual inventory form.
    """
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # ── Hypervisor totals (PCD compute nodes) ───────────────────────
            cur.execute("""
                SELECT
                  COUNT(*)::int                                              AS node_count,
                  COALESCE(SUM(vcpus),       0)::int                        AS total_vcpus,
                  ROUND(COALESCE(SUM(memory_mb), 0) / 1024.0, 1)           AS total_ram_gb,
                  ROUND(COALESCE(SUM(local_gb),  0) / 1024.0, 3)           AS total_local_disk_tb,
                  COALESCE(SUM(running_vms),  0)::int                       AS running_vms
                FROM hypervisors
                WHERE state = 'up' AND status = 'enabled'
            """)
            cluster = dict(cur.fetchone() or {})

            # ── vCPU / RAM already committed (active servers × flavor) ──────
            cur.execute("""
                SELECT
                  COUNT(s.id)::int                                           AS vm_count,
                  COALESCE(SUM(f.vcpus), 0)::int                            AS vcpus_used,
                  ROUND(COALESCE(SUM(f.ram_mb), 0) / 1024.0, 1)            AS ram_gb_used
                FROM servers s
                LEFT JOIN flavors f ON f.id = s.flavor_id
                WHERE s.status NOT IN ('DELETED', 'ERROR')
            """)
            usage = dict(cur.fetchone() or {})

            # ── Cinder block-storage already allocated ───────────────────────
            cur.execute("""
                SELECT ROUND(COALESCE(SUM(size_gb), 0) / 1024.0, 3) AS disk_tb_used
                FROM volumes
                WHERE status NOT IN ('deleting', 'error')
            """)
            disk_row = dict(cur.fetchone() or {})

    return {
        "status": "ok",
        "live": {
            "node_count":         int(cluster.get("node_count",         0)),
            "total_vcpus":        int(cluster.get("total_vcpus",        0)),
            "total_ram_gb":       float(cluster.get("total_ram_gb",     0) or 0),
            "total_local_disk_tb": float(cluster.get("total_local_disk_tb", 0) or 0),
            "running_vms":        int(cluster.get("running_vms",        0)),
            "vcpus_used":         int(usage.get("vcpus_used",           0)),
            "ram_gb_used":        float(usage.get("ram_gb_used",        0) or 0),
            "vm_count":           int(usage.get("vm_count",             0)),
            "disk_tb_used":       float(disk_row.get("disk_tb_used",    0) or 0),
        },
    }


@router.put("/projects/{project_id}/node-inventory",
            dependencies=[Depends(require_permission("migration", "write"))])
async def upsert_node_inventory(project_id: str, req: NodeInventoryRequest, user=Depends(get_current_user)):
    actor = user.username if user else "system"
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                INSERT INTO migration_pcd_node_inventory
                  (project_id, profile_id, current_nodes, current_vcpu_used, current_ram_gb_used, current_disk_tb_used, notes)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (project_id) DO UPDATE
                  SET profile_id=EXCLUDED.profile_id, current_nodes=EXCLUDED.current_nodes,
                      current_vcpu_used=EXCLUDED.current_vcpu_used,
                      current_ram_gb_used=EXCLUDED.current_ram_gb_used,
                      current_disk_tb_used=EXCLUDED.current_disk_tb_used,
                      notes=EXCLUDED.notes, updated_at=now()
                RETURNING *
            """, (project_id, req.profile_id, req.current_nodes, req.current_vcpu_used,
                  req.current_ram_gb_used, req.current_disk_tb_used, req.notes))
            inv = _serialize_row(dict(cur.fetchone()))
        conn.commit()
    _log_activity(actor=actor, action="upsert_node_inventory", resource_type="migration_project",
                  resource_id=project_id, details={"current_nodes": req.current_nodes})
    return {"status": "ok", "inventory": inv}


@router.get("/projects/{project_id}/node-sizing",
            dependencies=[Depends(require_permission("migration", "read"))])
async def get_node_sizing(
    project_id: str,
    max_util_pct: float = Query(70.0, ge=10.0, le=100.0,
        description="Maximum cluster utilisation target (%). HA headroom lives inside this cap."),
    peak_buffer_pct: float = Query(15.0, ge=0.0, le=100.0,
        description="Extra buffer added to migration workload for traffic spikes (%)."),
):
    """
    Compute PCD node sizing based on quota requirements and the current live cluster.

    Uses a 70%-utilisation HA model: the utilisation cap IS the HA strategy —
    no separate spare nodes are added on top.
    """
    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT * FROM migration_tenants
                WHERE project_id = %s AND include_in_plan = true
            """, (project_id,))
            tenants = [dict(r) for r in cur.fetchall()]

            profile_name = project.get("overcommit_profile_name") or "balanced"
            cur.execute("SELECT * FROM migration_overcommit_profiles WHERE profile_name = %s", (profile_name,))
            profile_row = cur.fetchone()
            profile = dict(profile_row) if profile_row else {"cpu_ratio": 4.0, "ram_ratio": 1.5, "disk_snapshot_factor": 1.5}

            cur.execute("""
                SELECT * FROM migration_pcd_node_profiles
                WHERE project_id = %s
                ORDER BY is_default DESC, id ASC LIMIT 1
            """, (project_id,))
            node_profile_row = cur.fetchone()

            cur.execute("SELECT * FROM migration_pcd_node_inventory WHERE project_id = %s", (project_id,))
            inv_row = cur.fetchone()

            # ── Actual VM footprint from RVtools (source of truth for HW sizing) ──
            # Prefer actual performance data (cpu_usage_percent/memory_usage_percent)
            # when available — these reflect real running consumption, not allocation.
            # SUM(cpu_count * cpu_usage_percent/100) = actual physical vCPU demand:
            # no overcommit division needed because utilisation already represents
            # real scheduler pressure on the host (not theoretical peak).
            cur.execute("""
                SELECT
                    COUNT(*) FILTER (WHERE power_state = 'poweredOn')                            AS powered_on_count,
                    -- Performance-based actual demand
                    COUNT(*) FILTER (WHERE power_state = 'poweredOn'
                                     AND cpu_usage_percent IS NOT NULL)                          AS vms_with_cpu_perf,
                    COUNT(*) FILTER (WHERE power_state = 'poweredOn'
                                     AND memory_usage_percent IS NOT NULL)                       AS vms_with_ram_perf,
                    COALESCE(SUM(cpu_count * cpu_usage_percent / 100.0)
                             FILTER (WHERE power_state = 'poweredOn'
                                      AND cpu_usage_percent IS NOT NULL), 0)                    AS actual_vcpu_used,
                    COALESCE(SUM(ram_mb * memory_usage_percent / 100.0)
                             FILTER (WHERE power_state = 'poweredOn'
                                      AND memory_usage_percent IS NOT NULL), 0) / 1024.0        AS actual_ram_gb,
                    -- Allocation-based fallback
                    COALESCE(SUM(cpu_count)    FILTER (WHERE power_state = 'poweredOn'), 0)     AS vm_vcpu_alloc,
                    COALESCE(SUM(ram_mb)       FILTER (WHERE power_state = 'poweredOn'), 0)
                        / 1024.0                                                                AS vm_ram_gb_alloc,
                    COALESCE(SUM(total_disk_gb)FILTER (WHERE power_state = 'poweredOn'), 0)
                        / 1024.0                                                                AS vm_disk_tb,
                    COUNT(DISTINCT host_name)                                                    AS source_node_count
                FROM migration_vms
                WHERE project_id = %s
            """, (project_id,))
            vm_row = dict(cur.fetchone() or {})

    if not node_profile_row:
        raise HTTPException(status_code=400,
                            detail="No node profile configured. Add a PCD node profile first.")

    # Fetch live cluster state (preferred over manual inventory)
    live_cluster = None
    try:
        live_resp = await get_pcd_live_inventory(project_id)
        live = live_resp.get("live", {})
        if live.get("node_count", 0) > 0:
            live_cluster = live
    except Exception:
        pass  # Fall back to manual inventory

    quota = compute_quota_requirements(tenants, profile)

    # ── Physical demand for HW sizing ──────────────────────────────────────────────
    # Priority 1 — Actual performance data (cpu_usage_percent / memory_usage_percent)
    #   SUM(cpu_count × cpu_usage_percent/100) = physical vCPU actually consumed.
    #   No overcommit division needed: utilisation is already physical scheduler load.
    #
    # Priority 2 — Allocation ÷ overcommit ratio (fallback when <50% VMs have perf data)
    #   SUM(cpu_count) is vCPU configured, not running. Divide by PCD overcommit ratio
    #   to convert to physical core demand.
    #
    # Priority 3 — Tenant quota (if no VM data at all, e.g. RVtools not yet imported)
    #   Quota already has overcommit baked in; using fallback avoids a zero-node result.
    # ─────────────────────────────────────────────────────────────────────────────────
    cpu_ratio          = float(profile.get("cpu_ratio", 4.0))
    powered_on         = int(vm_row.get("powered_on_count") or 0)
    vms_with_cpu_perf  = int(vm_row.get("vms_with_cpu_perf") or 0)
    vms_with_ram_perf  = int(vm_row.get("vms_with_ram_perf") or 0)
    actual_vcpu_used   = float(vm_row.get("actual_vcpu_used") or 0)
    actual_ram_gb      = float(vm_row.get("actual_ram_gb") or 0)
    vm_vcpu_alloc      = float(vm_row.get("vm_vcpu_alloc") or 0)
    vm_ram_gb_alloc    = float(vm_row.get("vm_ram_gb_alloc") or 0)
    vm_disk_tb         = float(vm_row.get("vm_disk_tb") or 0)
    source_node_count  = int(vm_row.get("source_node_count") or 0)

    perf_threshold = 0.5  # require ≥50% of powered-on VMs to have perf data
    cpu_perf_coverage = vms_with_cpu_perf / max(powered_on, 1)
    ram_perf_coverage = vms_with_ram_perf / max(powered_on, 1)

    if powered_on > 0 and cpu_perf_coverage >= perf_threshold:
        # ✅ Use actual performance data — most accurate physical demand
        phys_vcpu  = round(actual_vcpu_used)
        phys_ram   = round(actual_ram_gb if ram_perf_coverage >= perf_threshold else vm_ram_gb_alloc, 1)
        sizing_basis = "actual_performance"
    elif vm_vcpu_alloc > 0:
        # ⚠️ Fallback: allocation ÷ overcommit
        phys_vcpu  = round(vm_vcpu_alloc / cpu_ratio)
        phys_ram   = round(vm_ram_gb_alloc, 1)
        sizing_basis = "allocation"
    else:
        # 🔴 Last resort: tenant quota totals
        phys_vcpu  = quota["totals_recommended"]["vcpu"]
        phys_ram   = quota["totals_recommended"]["ram_gb"]
        sizing_basis = "quota"

    physical_totals = {
        "vcpu":    phys_vcpu,
        "ram_gb":  phys_ram,
        "disk_tb": quota["totals_recommended"]["disk_tb"],  # quota disk includes snapshot factor
    }

    sizing = compute_node_sizing(
        totals=physical_totals,
        node_profile=dict(node_profile_row),
        existing_inventory=dict(inv_row) if inv_row else None,
        live_cluster=live_cluster,
        max_util_pct=max_util_pct,
        peak_buffer_pct=peak_buffer_pct,
    )

    # Attach metadata so the UI can explain what basis was used
    sizing["sizing_basis"]         = sizing_basis
    sizing["vm_powered_on_count"]   = powered_on
    sizing["vm_vcpu_actual"]        = round(actual_vcpu_used) if sizing_basis == "actual_performance" else round(vm_vcpu_alloc / max(cpu_ratio, 1))
    sizing["vm_ram_gb_actual"]      = round(actual_ram_gb)    if sizing_basis == "actual_performance" else round(vm_ram_gb_alloc)
    sizing["vm_vcpu_alloc"]         = round(vm_vcpu_alloc)
    sizing["vm_ram_gb_alloc"]       = round(vm_ram_gb_alloc)
    sizing["source_node_count"]     = source_node_count
    sizing["perf_coverage_pct"]     = round(cpu_perf_coverage * 100)

    return {
        "status": "ok",
        "quota_summary": quota["totals_recommended"],
        "sizing": sizing,
        "live_cluster": live_cluster,
    }


# =====================================================================
# Phase 2E — PCD Readiness & Gap Analysis
# =====================================================================

class PcdSettingsRequest(BaseModel):
    pcd_auth_url: Optional[str] = None
    pcd_username: Optional[str] = None
    pcd_password_hint: Optional[str] = None
    pcd_region: Optional[str] = "region-one"


@router.patch("/projects/{project_id}/pcd-settings",
              dependencies=[Depends(require_permission("migration", "write"))])
async def update_pcd_settings(project_id: str, req: PcdSettingsRequest, user=Depends(get_current_user)):
    """Store PCD connection info on the migration project. Password is stored as hint only."""
    actor = user.username if user else "system"
    updates = {k: v for k, v in req.dict().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="Nothing to update")

    set_parts = [f"{k} = %s" for k in updates]
    set_parts.append("updated_at = now()")
    params = list(updates.values())

    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor() as cur:
            cur.execute(f"UPDATE migration_projects SET {', '.join(set_parts)} WHERE project_id = %s",
                        params + [project_id])
        conn.commit()
    _log_activity(actor=actor, action="update_pcd_settings", resource_type="migration_project",
                  resource_id=project_id, details=list(updates.keys()))
    return {"status": "ok"}


@router.post("/projects/{project_id}/pcd-gap-analysis",
             dependencies=[Depends(require_permission("migration", "write"))])
async def run_pcd_gap_analysis(project_id: str, user=Depends(get_current_user)):
    """
    Connect to PCD (using credentials from .env or project settings),
    fetch flavors/networks/images, compare against VMware workload, and store gaps.
    Falls back to offline analysis (only mapping-gap checks) if PCD is unreachable.
    """
    actor = user.username if user else "system"

    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT * FROM migration_tenants
                WHERE project_id = %s AND include_in_plan = true
            """, (project_id,))
            tenants = [dict(r) for r in cur.fetchall()]

            cur.execute("""
                SELECT vm_name, cpu_count, ram_mb, os_family, network_name, tenant_name
                FROM migration_vms WHERE project_id = %s
            """, (project_id,))
            vms = [dict(r) for r in cur.fetchall()]

    # ── Try to connect to PCD ─────────────────────────────────────────────────
    pcd_flavors: List[Dict[str, Any]] = []
    pcd_networks: List[Dict[str, Any]] = []
    pcd_images: List[Dict[str, Any]] = []
    pcd_connected = False
    pcd_error = None

    try:
        import sys as _sys
        _sys.path.insert(0, os.path.dirname(__file__))
        import p9_common  # type: ignore

        # Override config with project-level PCD settings if provided
        pcd_auth_url = project.get("pcd_auth_url") or p9_common.CFG.get("KEYSTONE_URL", "")
        pcd_username  = project.get("pcd_username") or p9_common.CFG.get("USERNAME", "")
        pcd_password  = os.getenv("PF9_PASSWORD", p9_common.CFG.get("PASSWORD", ""))

        if pcd_auth_url and pcd_username and pcd_password:
            # Temporarily patch global CFG if project has custom PCD settings
            _orig_cfg = {}
            if project.get("pcd_auth_url"):
                _orig_cfg["KEYSTONE_URL"] = p9_common.CFG["KEYSTONE_URL"]
                p9_common.CFG["KEYSTONE_URL"] = pcd_auth_url
            if project.get("pcd_username"):
                _orig_cfg["USERNAME"] = p9_common.CFG["USERNAME"]
                p9_common.CFG["USERNAME"] = pcd_username
            if project.get("pcd_region"):
                _orig_cfg["REGION_NAME"] = p9_common.CFG.get("REGION_NAME", "region-one")
                p9_common.CFG["REGION_NAME"] = project["pcd_region"]

            try:
                session = p9_common.get_session_best_scope()
                pcd_flavors  = p9_common.nova_flavors(session)
                pcd_networks = p9_common.neutron_list(session, "networks")
                pcd_images   = p9_common.glance_images(session)
                pcd_connected = True
            finally:
                # Restore original config
                for k, v in _orig_cfg.items():
                    p9_common.CFG[k] = v
        else:
            pcd_error = "PCD credentials not configured. Set PF9_AUTH_URL, PF9_USERNAME, PF9_PASSWORD in .env or project PCD settings."
    except Exception as e:
        pcd_error = str(e)
        logger.warning(f"PCD gap analysis: could not connect to PCD: {e}")

    # ── Run gap analysis ──────────────────────────────────────────────────────
    gaps = analyze_pcd_gaps(tenants, pcd_flavors, pcd_networks, pcd_images, vms)

    # ── Store gaps in DB ──────────────────────────────────────────────────────
    with _get_conn() as conn:
        with conn.cursor() as cur:
            # Clear old gaps for this project
            cur.execute("DELETE FROM migration_pcd_gaps WHERE project_id = %s", (project_id,))
            for g in gaps:
                cur.execute("""
                    INSERT INTO migration_pcd_gaps
                      (project_id, gap_type, resource_name, tenant_name, details, severity, resolution)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (project_id, gap_type, resource_name, tenant_name)
                    DO UPDATE SET details=EXCLUDED.details, severity=EXCLUDED.severity,
                                  resolution=EXCLUDED.resolution, checked_at=now(), resolved=false
                """, (project_id, g["gap_type"], g["resource_name"], g.get("tenant_name"),
                      Json(g.get("details", {})), g.get("severity", "warning"), g.get("resolution")))

            # Update readiness score on project (100 − penalties)
            critical_count = sum(1 for g in gaps if g.get("severity") == "critical")
            warning_count  = sum(1 for g in gaps if g.get("severity") == "warning")
            score = max(0, 100 - critical_count * 15 - warning_count * 5)
            cur.execute("""
                UPDATE migration_projects
                SET pcd_readiness_score = %s, pcd_last_checked_at = now(), updated_at = now()
                WHERE project_id = %s
            """, (score, project_id))

        conn.commit()

    _log_activity(actor=actor, action="pcd_gap_analysis", resource_type="migration_project",
                  resource_id=project_id,
                  details={"gaps": len(gaps), "pcd_connected": pcd_connected, "score": score})

    return {
        "status": "ok",
        "pcd_connected": pcd_connected,
        "pcd_error": pcd_error,
        "gaps": gaps,
        "gap_counts": {
            "critical": critical_count,
            "warning": warning_count,
            "total": len(gaps),
        },
        "readiness_score": score,
    }


@router.get("/projects/{project_id}/pcd-gaps",
            dependencies=[Depends(require_permission("migration", "read"))])
async def get_pcd_gaps(
    project_id: str,
    resolved: Optional[bool] = Query(None),
    gap_type: Optional[str] = Query(None),
):
    """List stored PCD gap analysis results for a project."""
    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            conditions = ["project_id = %s"]
            params: List[Any] = [project_id]
            if resolved is not None:
                conditions.append("resolved = %s"); params.append(resolved)
            if gap_type:
                conditions.append("gap_type = %s"); params.append(gap_type)
            cur.execute(f"""
                SELECT * FROM migration_pcd_gaps
                WHERE {' AND '.join(conditions)}
                ORDER BY severity DESC, gap_type, resource_name
            """, params)
            gaps = [_serialize_row(dict(r)) for r in cur.fetchall()]

    return {
        "status": "ok",
        "gaps": gaps,
        "readiness_score": project.get("pcd_readiness_score"),
        "last_checked_at": project.get("pcd_last_checked_at"),
    }


@router.patch("/projects/{project_id}/pcd-gaps/{gap_id}/resolve",
              dependencies=[Depends(require_permission("migration", "write"))])
async def resolve_pcd_gap(project_id: str, gap_id: int, user=Depends(get_current_user)):
    """Mark a specific PCD gap as resolved."""
    actor = user.username if user else "system"
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE migration_pcd_gaps SET resolved = true
                WHERE id = %s AND project_id = %s
            """, (gap_id, project_id))
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="Gap not found")
        conn.commit()
    _log_activity(actor=actor, action="resolve_pcd_gap", resource_type="migration_project",
                  resource_id=project_id, details={"gap_id": gap_id})
    return {"status": "ok"}


# =====================================================================
# Phase 2.8 — Auto-detect PCD node profile from hypervisors inventory
# =====================================================================

@router.get("/projects/{project_id}/pcd-auto-detect-profile",
            dependencies=[Depends(require_permission("migration", "read"))])
async def auto_detect_pcd_node_profile(project_id: str):
    """
    Auto-detect the dominant PCD compute node type from the hypervisors
    inventory table (populated by pf9_rvtools.py nightly sync).
    Groups active hypervisors by (vcpus, memory_mb) and returns the most
    common configuration as a ready-to-use node-profile suggestion.
    """
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT
                  vcpus,
                  memory_mb,
                  COUNT(*)::int                         AS node_count,
                  ROUND(AVG(local_gb), 0)::int          AS avg_local_disk_gb
                FROM hypervisors
                WHERE state = 'up' AND status = 'enabled'
                GROUP BY vcpus, memory_mb
                ORDER BY node_count DESC
                LIMIT 5
            """)
            groups = [dict(r) for r in cur.fetchall()]

            cur.execute("""
                SELECT COUNT(*)::int AS total_nodes
                FROM hypervisors
                WHERE state = 'up' AND status = 'enabled'
            """)
            total_row = cur.fetchone()

    if not groups:
        return {
            "status": "no_data",
            "message": (
                "No active hypervisors found in inventory. "
                "Ensure the PCD inventory sync has run (pf9_rvtools.py) and "
                "hypervisors are in state=up / status=enabled."
            ),
            "profiles": [],
        }

    total_nodes = int((total_row or {}).get("total_nodes") or 0)
    suggestions = []
    for g in groups:
        vcpus     = int(g["vcpus"]   or 0)
        ram_gb    = round(int(g["memory_mb"] or 0) / 1024)
        disk_gb   = int(float(g["avg_local_disk_gb"] or 0))
        storage_tb = round(disk_gb / 1024, 2)
        # Assume HT: cpu_threads = vcpus, cpu_cores = vcpus / 2
        cpu_cores   = max(1, vcpus // 2)
        cpu_threads = vcpus
        suggestions.append({
            "suggested_name":  f"{vcpus}vCPU-{ram_gb}GB",
            "cpu_cores":       cpu_cores,
            "cpu_threads":     cpu_threads,
            "ram_gb":          ram_gb,
            "storage_tb":      storage_tb,
            "node_count":      int(g["node_count"]),
        })

    dominant = suggestions[0]
    return {
        "status":           "ok",
        "total_nodes":      total_nodes,
        "dominant_profile": dominant,
        "all_configurations": suggestions,
        "message": (
            f"Detected {len(groups)} node configuration(s) across "
            f"{total_nodes} active hypervisor(s). "
            f"Dominant: {dominant['cpu_threads']} vCPU · {dominant['ram_gb']} GB RAM · "
            f"{total_nodes} node(s)."
        ),
    }


# =====================================================================
# Phase 2.8 — Gap Analysis Action Report export (Excel + PDF)
# =====================================================================

@router.get("/projects/{project_id}/export-gaps-report.xlsx",
            dependencies=[Depends(require_permission("migration", "read"))])
async def export_gaps_report_excel(project_id: str):
    """Download PCD readiness gap analysis as a styled Excel workbook."""
    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT * FROM migration_pcd_gaps
                WHERE project_id = %s
                ORDER BY severity DESC, gap_type, resource_name
            """, (project_id,))
            gaps = [_serialize_row(dict(r)) for r in cur.fetchall()]

    readiness_score = project.get("pcd_readiness_score")
    excel_bytes = generate_gaps_excel_report(
        gaps=gaps,
        project_name=project["name"],
        readiness_score=float(readiness_score) if readiness_score is not None else None,
    )
    safe_name = project["name"].replace(" ", "_")
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="pcd-readiness-{safe_name}.xlsx"'},
    )


@router.get("/projects/{project_id}/export-gaps-report.pdf",
            dependencies=[Depends(require_permission("migration", "read"))])
async def export_gaps_report_pdf(project_id: str):
    """Download PCD readiness gap analysis as a landscape A4 PDF."""
    with _get_conn() as conn:
        project = _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT * FROM migration_pcd_gaps
                WHERE project_id = %s
                ORDER BY severity DESC, gap_type, resource_name
            """, (project_id,))
            gaps = [_serialize_row(dict(r)) for r in cur.fetchall()]

    readiness_score = project.get("pcd_readiness_score")
    pdf_bytes = generate_gaps_pdf_report(
        gaps=gaps,
        project_name=project["name"],
        readiness_score=float(readiness_score) if readiness_score is not None else None,
    )
    safe_name = project["name"].replace(" ", "_")
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="pcd-readiness-{safe_name}.pdf"'},
    )


# =====================================================================
# Serialization helpers
# =====================================================================

# =====================================================================
# PHASE 2.10 — VM STATUS & MODE OVERRIDE
# =====================================================================

VALID_VM_STATUSES = {"not_started", "assigned", "in_progress", "migrated", "failed", "skipped"}
VALID_MODE_OVERRIDES = {"warm", "cold", None}


@router.patch("/projects/{project_id}/vms/{vm_id}/status",
              dependencies=[Depends(require_permission("migration", "write"))])
async def update_vm_status(project_id: str, vm_id: int, req: VMStatusUpdateRequest,
                           user=Depends(get_current_user)):
    """Update the migration status of a single VM."""
    actor = user.username if user else "system"
    if req.status not in VALID_VM_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status. Valid: {sorted(VALID_VM_STATUSES)}")

    with _get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                UPDATE migration_vms
                SET migration_status = %s,
                    migration_status_note = %s,
                    migration_status_updated_at = now(),
                    migration_status_updated_by = %s,
                    updated_at = now()
                WHERE project_id = %s AND id = %s
                RETURNING id, vm_name, migration_status, migration_status_note,
                          migration_status_updated_at, migration_status_updated_by
            """, (req.status, req.status_note, actor, project_id, vm_id))
            vm = cur.fetchone()
            if not vm:
                raise HTTPException(status_code=404, detail="VM not found")
            conn.commit()

    _log_activity(actor=actor, action="update_vm_status", resource_type="migration_vm",
                  resource_id=str(vm_id), details={"status": req.status, "note": req.status_note})
    return {"status": "ok", "vm": _serialize_row(dict(vm))}


@router.patch("/projects/{project_id}/vms/bulk-status",
              dependencies=[Depends(require_permission("migration", "write"))])
async def bulk_update_vm_status(project_id: str, req: VMBulkStatusRequest,
                                user=Depends(get_current_user)):
    """Bulk-update migration status for multiple VMs."""
    actor = user.username if user else "system"
    if req.status not in VALID_VM_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status. Valid: {sorted(VALID_VM_STATUSES)}")
    if not req.vm_ids:
        raise HTTPException(status_code=400, detail="vm_ids must not be empty")

    with _get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                UPDATE migration_vms
                SET migration_status = %s,
                    migration_status_note = %s,
                    migration_status_updated_at = now(),
                    migration_status_updated_by = %s,
                    updated_at = now()
                WHERE project_id = %s AND id = ANY(%s)
            """, (req.status, req.status_note, actor, project_id, req.vm_ids))
            updated = cur.rowcount
            conn.commit()

    _log_activity(actor=actor, action="bulk_update_vm_status", resource_type="migration_vm",
                  resource_id=project_id, details={"status": req.status, "count": updated, "vm_ids": req.vm_ids})
    return {"status": "ok", "updated_count": updated}


@router.patch("/projects/{project_id}/vms/{vm_id}/mode-override",
              dependencies=[Depends(require_permission("migration", "write"))])
async def update_vm_mode_override(project_id: str, vm_id: int, req: VMModeOverrideRequest,
                                  user=Depends(get_current_user)):
    """Set or clear a per-VM migration mode override (warm/cold/null=auto)."""
    actor = user.username if user else "system"
    if req.override not in VALID_MODE_OVERRIDES:
        raise HTTPException(status_code=400, detail="override must be 'warm', 'cold', or null")

    with _get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                UPDATE migration_vms
                SET migration_mode_override = %s, updated_at = now()
                WHERE project_id = %s AND id = %s
                RETURNING id, vm_name, migration_mode, migration_mode_override
            """, (req.override, project_id, vm_id))
            vm = cur.fetchone()
            if not vm:
                raise HTTPException(status_code=404, detail="VM not found")
            conn.commit()

    _log_activity(actor=actor, action="update_vm_mode_override", resource_type="migration_vm",
                  resource_id=str(vm_id), details={"override": req.override})
    return {"status": "ok", "vm": _serialize_row(dict(vm))}


# =====================================================================
# PHASE 2.10 — VM DEPENDENCIES
# =====================================================================

@router.get("/projects/{project_id}/vm-dependencies",
            dependencies=[Depends(require_permission("migration", "read"))])
async def list_vm_dependencies(project_id: str, vm_id: Optional[int] = Query(None)):
    """List VM dependencies for a project, optionally filtered by vm_id."""
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            if vm_id is not None:
                cur.execute("""
                    SELECT d.*,
                           v1.vm_name AS vm_name,
                           v2.vm_name AS depends_on_vm_name
                    FROM migration_vm_dependencies d
                    JOIN migration_vms v1 ON d.vm_id = v1.id
                    JOIN migration_vms v2 ON d.depends_on_vm_id = v2.id
                    WHERE d.project_id = %s AND d.vm_id = %s
                    ORDER BY d.created_at
                """, (project_id, vm_id))
            else:
                cur.execute("""
                    SELECT d.*,
                           v1.vm_name AS vm_name,
                           v2.vm_name AS depends_on_vm_name
                    FROM migration_vm_dependencies d
                    JOIN migration_vms v1 ON d.vm_id = v1.id
                    JOIN migration_vms v2 ON d.depends_on_vm_id = v2.id
                    WHERE d.project_id = %s
                    ORDER BY v1.vm_name, d.created_at
                """, (project_id,))
            deps = [_serialize_row(dict(r)) for r in cur.fetchall()]
    return {"status": "ok", "dependencies": deps}


@router.post("/projects/{project_id}/vms/{vm_id}/dependencies",
             dependencies=[Depends(require_permission("migration", "write"))])
async def add_vm_dependency(project_id: str, vm_id: int, req: VMDependencyRequest,
                            user=Depends(get_current_user)):
    """Add a dependency: this VM must wait for depends_on_vm_id to complete first."""
    actor = user.username if user else "system"
    if vm_id == req.depends_on_vm_id:
        raise HTTPException(status_code=400, detail="A VM cannot depend on itself")

    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Circular dependency check (A→B and B→A)
            cur.execute("""
                SELECT 1 FROM migration_vm_dependencies
                WHERE project_id = %s AND vm_id = %s AND depends_on_vm_id = %s
            """, (project_id, req.depends_on_vm_id, vm_id))
            if cur.fetchone():
                raise HTTPException(status_code=409,
                    detail="Circular dependency detected: the target VM already depends on this VM")

            cur.execute("""
                INSERT INTO migration_vm_dependencies
                    (project_id, vm_id, depends_on_vm_id, dependency_type, notes)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT (vm_id, depends_on_vm_id) DO UPDATE
                    SET dependency_type = EXCLUDED.dependency_type,
                        notes = EXCLUDED.notes
                RETURNING *
            """, (project_id, vm_id, req.depends_on_vm_id, req.dependency_type, req.notes))
            dep = cur.fetchone()
            conn.commit()

    _log_activity(actor=actor, action="add_vm_dependency", resource_type="migration_vm_dependency",
                  resource_id=str(dep["id"]), details={"vm_id": vm_id, "depends_on": req.depends_on_vm_id})
    return {"status": "ok", "dependency": _serialize_row(dict(dep))}


@router.delete("/projects/{project_id}/vms/{vm_id}/dependencies/{dep_id}",
               dependencies=[Depends(require_permission("migration", "write"))])
async def delete_vm_dependency(project_id: str, vm_id: int, dep_id: int,
                               user=Depends(get_current_user)):
    """Remove a VM dependency."""
    actor = user.username if user else "system"
    with _get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                DELETE FROM migration_vm_dependencies
                WHERE id = %s AND project_id = %s AND vm_id = %s
            """, (dep_id, project_id, vm_id))
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="Dependency not found")
            conn.commit()
    _log_activity(actor=actor, action="delete_vm_dependency", resource_type="migration_vm_dependency",
                  resource_id=str(dep_id), details={"vm_id": vm_id})
    return {"status": "ok"}


# =====================================================================
# PHASE 2.10 — NETWORK MAPPINGS (Source → PCD)
# =====================================================================

@router.get("/projects/{project_id}/network-mappings",
            dependencies=[Depends(require_permission("migration", "read"))])
async def list_network_mappings(project_id: str):
    """
    List all source network → PCD network mappings for this project.
    Auto-seeds unmapped entries from distinct VM network_name values.
    """
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Auto-seed: insert any source networks from VMs that don't have a mapping yet.
            # Default target = source name ("best guess, confirm or override"), confirmed=false.
            cur.execute("""
                INSERT INTO migration_network_mappings
                    (project_id, source_network_name, target_network_name, confirmed)
                SELECT DISTINCT %s, vm.network_name, vm.network_name, false
                FROM migration_vms vm
                WHERE vm.project_id = %s
                  AND vm.network_name IS NOT NULL
                  AND vm.network_name != ''
                  AND vm.power_state = 'poweredOn'
                  AND NOT EXISTS (
                      SELECT 1 FROM migration_network_mappings m
                      WHERE m.project_id = %s AND m.source_network_name = vm.network_name
                  )
                ON CONFLICT DO NOTHING
            """, (project_id, project_id, project_id))
            # Backfill VLAN ID from network name pattern (e.g. "mynet_vlan_3399" → 3399)
            cur.execute("""
                UPDATE migration_network_mappings
                SET vlan_id = CAST(
                    NULLIF(substring(source_network_name from '[Vv][Ll][Aa][Nn][_-]?([0-9]+)'), '')
                    AS INTEGER
                )
                WHERE project_id = %s
                  AND vlan_id IS NULL
                  AND source_network_name ~ '[Vv][Ll][Aa][Nn][_-]?[0-9]+'
            """, (project_id,))
            conn.commit()

            # Fetch all mappings with VM count per source network
            cur.execute("""
                SELECT m.*,
                       COUNT(DISTINCT v.id) AS vm_count
                FROM migration_network_mappings m
                LEFT JOIN migration_vms v
                    ON v.project_id = m.project_id
                    AND v.network_name = m.source_network_name
                    AND v.power_state = 'poweredOn'
                WHERE m.project_id = %s
                GROUP BY m.id, m.project_id, m.source_network_name, m.target_network_name,
                         m.target_network_id, m.vlan_id, m.notes, m.created_at, m.updated_at
                ORDER BY vm_count DESC, m.source_network_name
            """, (project_id,))
            mappings = [_serialize_row(dict(r)) for r in cur.fetchall()]

    # unconfirmed = auto-seeded rows (target = source name but not yet reviewed by user)
    unconfirmed_count = sum(1 for m in mappings if not m.get("confirmed"))
    return {"status": "ok", "mappings": mappings, "unconfirmed_count": unconfirmed_count,
            "total": len(mappings)}


@router.post("/projects/{project_id}/network-mappings",
             dependencies=[Depends(require_permission("migration", "write"))])
async def create_network_mapping(project_id: str, req: NetworkMappingCreateRequest,
                                 user=Depends(get_current_user)):
    """Create a source→PCD network mapping entry."""
    actor = user.username if user else "system"
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                INSERT INTO migration_network_mappings
                    (project_id, source_network_name, target_network_name,
                     target_network_id, vlan_id, notes)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (project_id, source_network_name) DO UPDATE
                    SET target_network_name = EXCLUDED.target_network_name,
                        target_network_id = EXCLUDED.target_network_id,
                        vlan_id = EXCLUDED.vlan_id,
                        notes = EXCLUDED.notes,
                        updated_at = now()
                RETURNING *
            """, (project_id, req.source_network_name, req.target_network_name,
                  req.target_network_id, req.vlan_id, req.notes))
            mapping = cur.fetchone()
            conn.commit()
    _log_activity(actor=actor, action="create_network_mapping", resource_type="migration_network_mapping",
                  resource_id=str(mapping["id"]), details={"source": req.source_network_name})
    return {"status": "ok", "mapping": _serialize_row(dict(mapping))}


@router.patch("/projects/{project_id}/network-mappings/{mapping_id}",
              dependencies=[Depends(require_permission("migration", "write"))])
async def update_network_mapping(project_id: str, mapping_id: int,
                                 req: NetworkMappingUpdateRequest, user=Depends(get_current_user)):
    """Update target network name/ID for a mapping entry."""
    actor = user.username if user else "system"
    updates = {k: v for k, v in req.dict().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    set_clauses = [f"{k} = %s" for k in updates] + ["updated_at = now()"]
    params = list(updates.values())

    with _get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(f"""
                UPDATE migration_network_mappings
                SET {', '.join(set_clauses)}
                WHERE id = %s AND project_id = %s
                RETURNING *
            """, params + [mapping_id, project_id])
            mapping = cur.fetchone()
            if not mapping:
                raise HTTPException(status_code=404, detail="Network mapping not found")
            conn.commit()
    _log_activity(actor=actor, action="update_network_mapping", resource_type="migration_network_mapping",
                  resource_id=str(mapping_id), details=updates)
    return {"status": "ok", "mapping": _serialize_row(dict(mapping))}


@router.delete("/projects/{project_id}/network-mappings/{mapping_id}",
               dependencies=[Depends(require_permission("migration", "admin"))])
async def delete_network_mapping(project_id: str, mapping_id: int,
                                 user=Depends(get_current_user)):
    """Delete a network mapping entry."""
    actor = user.username if user else "system"
    with _get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM migration_network_mappings WHERE id = %s AND project_id = %s",
                        (mapping_id, project_id))
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="Network mapping not found")
            conn.commit()
    _log_activity(actor=actor, action="delete_network_mapping", resource_type="migration_network_mapping",
                  resource_id=str(mapping_id), details={})
    return {"status": "ok"}


# =====================================================================
# PHASE 2.10 — MIGRATION COHORTS
# =====================================================================

@router.get("/projects/{project_id}/cohorts",
            dependencies=[Depends(require_permission("migration", "read"))])
async def list_cohorts(project_id: str):
    """List all cohorts for a project with per-cohort VM and tenant counts."""
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT c.*,
                       COUNT(DISTINCT t.id) AS tenant_count,
                       COALESCE(SUM(t.vm_count), 0) AS vm_count,
                       COALESCE(SUM(t.total_vcpu), 0) AS total_vcpu,
                       COALESCE(ROUND(SUM(t.total_ram_mb / 1024.0)::numeric, 1), 0) AS total_ram_gb,
                       COALESCE(ROUND(SUM(t.total_disk_gb)::numeric, 1), 0) AS total_disk_gb
                FROM migration_cohorts c
                LEFT JOIN migration_tenants t ON t.cohort_id = c.id
                WHERE c.project_id = %s
                GROUP BY c.id
                ORDER BY c.cohort_order, c.created_at
            """, (project_id,))
            cohorts = [_serialize_row(dict(r)) for r in cur.fetchall()]

            # Count unassigned tenants (in this project with no cohort)
            cur.execute("""
                SELECT COUNT(*) AS cnt FROM migration_tenants
                WHERE project_id = %s AND cohort_id IS NULL AND include_in_plan = true
            """, (project_id,))
            unassigned_count = cur.fetchone()["cnt"]

    return {"status": "ok", "cohorts": cohorts, "unassigned_tenant_count": unassigned_count}


@router.post("/projects/{project_id}/cohorts",
             dependencies=[Depends(require_permission("migration", "write"))])
async def create_cohort(project_id: str, req: CreateCohortRequest,
                        user=Depends(get_current_user)):
    """Create a new migration cohort."""
    actor = user.username if user else "system"
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                INSERT INTO migration_cohorts
                    (project_id, name, description, cohort_order,
                     scheduled_start, scheduled_end, owner_name,
                     depends_on_cohort_id, overcommit_profile_override,
                     agent_slots_override, schedule_duration_days, target_vms_per_day, notes)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING *
            """, (project_id, req.name, req.description, req.cohort_order,
                  req.scheduled_start, req.scheduled_end, req.owner_name,
                  req.depends_on_cohort_id, req.overcommit_profile_override,
                  req.agent_slots_override, req.schedule_duration_days,
                  req.target_vms_per_day, req.notes))
            cohort = cur.fetchone()
            conn.commit()
    _log_activity(actor=actor, action="create_cohort", resource_type="migration_cohort",
                  resource_id=str(cohort["id"]), details={"name": req.name})
    return {"status": "ok", "cohort": _serialize_row(dict(cohort))}


@router.patch("/projects/{project_id}/cohorts/{cohort_id}",
              dependencies=[Depends(require_permission("migration", "write"))])
async def update_cohort(project_id: str, cohort_id: int, req: UpdateCohortRequest,
                        user=Depends(get_current_user)):
    """Update cohort metadata."""
    actor = user.username if user else "system"
    updates = {k: v for k, v in req.dict().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    set_clauses = [f"{k} = %s" for k in updates] + ["updated_at = now()"]
    params = list(updates.values())

    with _get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(f"""
                UPDATE migration_cohorts
                SET {', '.join(set_clauses)}
                WHERE id = %s AND project_id = %s
                RETURNING *
            """, params + [cohort_id, project_id])
            cohort = cur.fetchone()
            if not cohort:
                raise HTTPException(status_code=404, detail="Cohort not found")
            conn.commit()
    _log_activity(actor=actor, action="update_cohort", resource_type="migration_cohort",
                  resource_id=str(cohort_id), details=updates)
    return {"status": "ok", "cohort": _serialize_row(dict(cohort))}


@router.delete("/projects/{project_id}/cohorts/{cohort_id}",
               dependencies=[Depends(require_permission("migration", "admin"))])
async def delete_cohort(project_id: str, cohort_id: int, user=Depends(get_current_user)):
    """Delete a cohort. Tenants are unassigned (cohort_id = NULL), not deleted."""
    actor = user.username if user else "system"
    with _get_conn() as conn:
        with conn.cursor() as cur:
            # Unassign tenants first
            cur.execute("UPDATE migration_tenants SET cohort_id = NULL WHERE cohort_id = %s",
                        (cohort_id,))
            cur.execute("DELETE FROM migration_cohorts WHERE id = %s AND project_id = %s",
                        (cohort_id, project_id))
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="Cohort not found")
            conn.commit()
    _log_activity(actor=actor, action="delete_cohort", resource_type="migration_cohort",
                  resource_id=str(cohort_id), details={})
    return {"status": "ok"}


@router.post("/projects/{project_id}/cohorts/{cohort_id}/assign-tenants",
             dependencies=[Depends(require_permission("migration", "write"))])
async def assign_tenants_to_cohort(project_id: str, cohort_id: int,
                                   req: AssignTenantsToCohortRequest,
                                   user=Depends(get_current_user)):
    """Assign a list of tenants to a cohort (or unassign with cohort_id=0)."""
    actor = user.username if user else "system"
    if not req.tenant_ids:
        raise HTTPException(status_code=400, detail="tenant_ids must not be empty")

    # cohort_id=0 means unassign
    target_cohort_id = None if cohort_id == 0 else cohort_id

    with _get_conn() as conn:
        _get_project(project_id, conn)
        if target_cohort_id is not None:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT id FROM migration_cohorts WHERE id = %s AND project_id = %s",
                            (target_cohort_id, project_id))
                if not cur.fetchone():
                    raise HTTPException(status_code=404, detail="Cohort not found")

        with conn.cursor() as cur:
            cur.execute("""
                UPDATE migration_tenants
                SET cohort_id = %s, updated_at = now()
                WHERE project_id = %s AND id = ANY(%s)
            """, (target_cohort_id, project_id, req.tenant_ids))
            updated = cur.rowcount
            conn.commit()

    _log_activity(actor=actor, action="assign_tenants_to_cohort", resource_type="migration_cohort",
                  resource_id=str(cohort_id),
                  details={"tenant_ids": req.tenant_ids, "updated": updated})
    return {"status": "ok", "updated_count": updated}


@router.get("/projects/{project_id}/cohorts/{cohort_id}/summary",
            dependencies=[Depends(require_permission("migration", "read"))])
async def get_cohort_summary(project_id: str, cohort_id: int):
    """Detailed summary for one cohort: tenant list, VM counts, resource totals, status breakdown."""
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM migration_cohorts WHERE id = %s AND project_id = %s",
                        (cohort_id, project_id))
            cohort = cur.fetchone()
            if not cohort:
                raise HTTPException(status_code=404, detail="Cohort not found")

            # Tenants in this cohort
            cur.execute("""
                SELECT id, tenant_name, include_in_plan, migration_priority,
                       vm_count, total_vcpu, total_ram_mb, total_disk_gb,
                       target_domain_name, target_project_name
                FROM migration_tenants
                WHERE cohort_id = %s AND project_id = %s
                ORDER BY migration_priority, tenant_name
            """, (cohort_id, project_id))
            tenants = [_serialize_row(dict(r)) for r in cur.fetchall()]
            tenant_ids = [t["id"] for t in tenants]

            # VM status breakdown for this cohort
            status_breakdown = {}
            if tenant_ids:
                placeholders = ', '.join(['%s'] * len(tenant_ids))
                cur.execute(f"""
                    SELECT migration_status, COUNT(*) as cnt,
                           COALESCE(SUM(cpu_count), 0) as vcpu,
                           COALESCE(ROUND(SUM(ram_mb / 1024.0)::numeric, 1), 0) as ram_gb
                    FROM migration_vms
                    WHERE project_id = %s
                      AND tenant_name IN (
                          SELECT tenant_name FROM migration_tenants
                          WHERE id IN ({placeholders})
                      )
                      AND power_state = 'poweredOn'
                    GROUP BY migration_status
                """, [project_id] + tenant_ids)
                for row in cur.fetchall():
                    status_breakdown[row["migration_status"] or "not_started"] = {
                        "count": row["cnt"], "vcpu": row["vcpu"], "ram_gb": float(row["ram_gb"])
                    }

    return {
        "status": "ok",
        "cohort": _serialize_row(dict(cohort)),
        "tenants": tenants,
        "status_breakdown": status_breakdown,
        "tenant_count": len(tenants),
        "vm_count": sum(t.get("vm_count") or 0 for t in tenants),
        "total_vcpu": sum(t.get("total_vcpu") or 0 for t in tenants),
        "total_ram_gb": round(sum(float(t.get("total_ram_mb") or 0) / 1024.0 for t in tenants), 1),
    }


# =====================================================================
# PHASE 2.10 — AUTO-ASSIGN TENANTS TO COHORTS
# =====================================================================

@router.post("/projects/{project_id}/cohorts/auto-assign",
             dependencies=[Depends(require_permission("migration", "write"))])
async def auto_assign_tenants_to_cohorts(project_id: str,
                                         strategy: str = Query("priority",
                                             description="priority | risk | equal_split"),
                                         user=Depends(get_current_user)):
    """
    Auto-assign all in-scope unassigned tenants across existing cohorts.
    strategy=priority:    assign by migration_priority order to cohorts in order
    strategy=risk:        high-risk tenants → last cohort, low-risk → first
    strategy=equal_split: distribute evenly across cohorts by VM count
    """
    actor = user.username if user else "system"
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Get ordered cohorts
            cur.execute("""
                SELECT id, cohort_order FROM migration_cohorts
                WHERE project_id = %s ORDER BY cohort_order, id
            """, (project_id,))
            cohorts = cur.fetchall()
            if not cohorts:
                raise HTTPException(status_code=400,
                    detail="No cohorts exist. Create cohorts first.")

            cohort_ids = [c["id"] for c in cohorts]

            # Get unassigned in-scope tenants
            if strategy == "risk":
                cur.execute("""
                    SELECT t.id, t.tenant_name, t.vm_count,
                           AVG(v.risk_score) AS avg_risk
                    FROM migration_tenants t
                    LEFT JOIN migration_vms v ON v.project_id = t.project_id
                        AND v.tenant_name = t.tenant_name
                    WHERE t.project_id = %s AND t.cohort_id IS NULL
                      AND t.include_in_plan = true
                    GROUP BY t.id, t.tenant_name, t.vm_count
                    ORDER BY avg_risk ASC NULLS LAST
                """, (project_id,))
            else:
                cur.execute("""
                    SELECT id, tenant_name, vm_count
                    FROM migration_tenants
                    WHERE project_id = %s AND cohort_id IS NULL AND include_in_plan = true
                    ORDER BY migration_priority ASC, tenant_name
                """, (project_id,))
            tenants = cur.fetchall()

            if not tenants:
                return {"status": "ok", "message": "All in-scope tenants already assigned",
                        "updated_count": 0}

            # Assign tenants round-robin (equal_split) or sequentially (priority/risk)
            assignments = []
            if strategy == "equal_split":
                # Round-robin by VM count to balance cohorts
                cohort_vm_counts = {cid: 0 for cid in cohort_ids}
                for t in tenants:
                    lightest = min(cohort_vm_counts, key=cohort_vm_counts.get)
                    assignments.append((lightest, t["id"]))
                    cohort_vm_counts[lightest] += (t["vm_count"] or 0)
            else:
                # Divide tenants into len(cohorts) equal chunks
                chunk = max(1, -(-len(tenants) // len(cohort_ids)))  # ceiling division
                for i, t in enumerate(tenants):
                    cid = cohort_ids[min(i // chunk, len(cohort_ids) - 1)]
                    assignments.append((cid, t["id"]))

            # Apply assignments
            for cid, tid in assignments:
                cur.execute("""
                    UPDATE migration_tenants SET cohort_id = %s, updated_at = now()
                    WHERE id = %s
                """, (cid, tid))
            conn.commit()

    _log_activity(actor=actor, action="auto_assign_cohorts", resource_type="migration_cohort",
                  resource_id=project_id, details={"strategy": strategy, "assigned": len(assignments)})
    return {"status": "ok", "updated_count": len(assignments), "strategy": strategy}


# =====================================================================
# PHASE 2.10 — TENANT READINESS CHECKS
# =====================================================================

def _compute_tenant_readiness(project_id: str, tenant_id: int, conn) -> List[Dict]:
    """Compute readiness check results for one tenant. Returns list of check dicts."""
    results = []
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("SELECT * FROM migration_tenants WHERE id = %s AND project_id = %s",
                    (tenant_id, project_id))
        tenant = cur.fetchone()
        if not tenant:
            return []

        # Check 1: target_mapped — requires confirmed=true (user reviewed, not just auto-seeded)
        results.append({
            "check_name": "target_mapped",
            "check_status": "pass" if (tenant.get("target_domain_name") and tenant.get("target_confirmed"))
                            else "pending" if tenant.get("target_domain_name")
                            else "fail",
            "notes": "Target PCD domain confirmed" if (tenant.get("target_domain_name") and tenant.get("target_confirmed"))
                     else "Target pre-filled from source name — review and confirm in Tenants tab" if tenant.get("target_domain_name")
                     else "No target domain assigned — set in Tenants tab"
        })

        # Check 2: network_mapped — requires confirmed=true (user reviewed, not just auto-seeded)
        cur.execute("""
            SELECT COUNT(DISTINCT v.network_name) AS total,
                   COUNT(DISTINCT CASE WHEN m.confirmed = true
                         THEN v.network_name END) AS confirmed
            FROM migration_vms v
            LEFT JOIN migration_network_mappings m
                ON m.project_id = v.project_id AND m.source_network_name = v.network_name
            WHERE v.project_id = %s AND v.tenant_name = %s
              AND v.power_state = 'poweredOn' AND v.network_name IS NOT NULL
        """, (project_id, tenant["tenant_name"]))
        nm = cur.fetchone()
        total_nets = nm["total"] or 0
        confirmed_nets = nm["confirmed"] or 0
        results.append({
            "check_name": "network_mapped",
            "check_status": "pass" if (total_nets == 0 or total_nets == confirmed_nets) else "pending",
            "notes": f"{confirmed_nets}/{total_nets} networks confirmed " +
                     ("" if total_nets == confirmed_nets else "— review Network Map tab and confirm each mapping")
        })

        # Check 3: quota_sufficient — target_project_name or quota model exists
        results.append({
            "check_name": "quota_sufficient",
            "check_status": "pass" if tenant.get("target_project_name") else "pending",
            "notes": "Target PCD project set" if tenant.get("target_project_name")
                     else "Set target PCD project to enable quota validation"
        })

        # Check 4: no_critical_gaps — check migration_pcd_gaps for this project
        cur.execute("""
            SELECT COUNT(*) AS cnt FROM migration_pcd_gaps
            WHERE project_id = %s AND severity = 'critical' AND status != 'resolved'
        """, (project_id,))
        critical_gaps = cur.fetchone()["cnt"]
        results.append({
            "check_name": "no_critical_gaps",
            "check_status": "pass" if critical_gaps == 0 else "fail",
            "notes": f"{critical_gaps} unresolved critical PCD gap(s)" if critical_gaps > 0
                     else "No critical gaps"
        })

        # Check 5: vms_classified — all powered-on VMs have a migration mode
        cur.execute("""
            SELECT COUNT(*) AS total,
                   COUNT(CASE WHEN migration_mode IS NOT NULL THEN 1 END) AS classified
            FROM migration_vms
            WHERE project_id = %s AND tenant_name = %s AND power_state = 'poweredOn'
        """, (project_id, tenant["tenant_name"]))
        vc = cur.fetchone()
        results.append({
            "check_name": "vms_classified",
            "check_status": "pass" if (vc["total"] == 0 or vc["classified"] == vc["total"]) else "fail",
            "notes": f"{vc['classified']}/{vc['total']} VMs have migration mode assigned"
        })

    return results


@router.get("/projects/{project_id}/tenants/{tenant_id}/readiness",
            dependencies=[Depends(require_permission("migration", "read"))])
async def get_tenant_readiness(project_id: str, tenant_id: int):
    """Compute and return readiness check results for a single tenant."""
    with _get_conn() as conn:
        _get_project(project_id, conn)
        checks = _compute_tenant_readiness(project_id, tenant_id, conn)
        if not checks:
            raise HTTPException(status_code=404, detail="Tenant not found")
        # Persist latest results
        with conn.cursor() as cur:
            for c in checks:
                cur.execute("""
                    INSERT INTO migration_tenant_readiness
                        (tenant_id, check_name, check_status, checked_at, notes)
                    VALUES (%s, %s, %s, now(), %s)
                    ON CONFLICT (tenant_id, check_name) DO UPDATE
                        SET check_status = EXCLUDED.check_status,
                            checked_at = EXCLUDED.checked_at,
                            notes = EXCLUDED.notes
                """, (tenant_id, c["check_name"], c["check_status"], c["notes"]))
            conn.commit()

    passed = sum(1 for c in checks if c["check_status"] == "pass")
    overall = "pass" if passed == len(checks) else ("fail" if any(c["check_status"] == "fail" for c in checks) else "pending")
    return {"status": "ok", "tenant_id": tenant_id, "checks": checks,
            "overall": overall, "score": f"{passed}/{len(checks)}"}


@router.get("/projects/{project_id}/cohorts/{cohort_id}/readiness-summary",
            dependencies=[Depends(require_permission("migration", "read"))])
async def get_cohort_readiness_summary(project_id: str, cohort_id: int):
    """Readiness summary for all tenants in a cohort."""
    with _get_conn() as conn:
        _get_project(project_id, conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT id, tenant_name FROM migration_tenants WHERE cohort_id = %s AND project_id = %s",
                        (cohort_id, project_id))
            tenants = cur.fetchall()

        tenant_results = []
        for t in tenants:
            checks = _compute_tenant_readiness(project_id, t["id"], conn)
            passed = sum(1 for c in checks if c["check_status"] == "pass")
            overall = "pass" if passed == len(checks) else (
                "fail" if any(c["check_status"] == "fail" for c in checks) else "pending")
            tenant_results.append({
                "tenant_id": t["id"],
                "tenant_name": t["tenant_name"],
                "overall": overall,
                "score": f"{passed}/{len(checks)}",
                "checks": checks
            })

    ready = sum(1 for t in tenant_results if t["overall"] == "pass")
    return {
        "status": "ok",
        "cohort_id": cohort_id,
        "tenants": tenant_results,
        "ready_count": ready,
        "total_count": len(tenant_results),
        "all_ready": (ready == len(tenant_results) and len(tenant_results) > 0)
    }




def _serialize_project(row: Dict[str, Any]) -> Dict[str, Any]:
    """Convert DB row to JSON-safe dict."""
    return _serialize_row(row)


def _serialize_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """Convert a single DB row dict to JSON-safe format."""
    result = {}
    for k, v in row.items():
        if isinstance(v, datetime):
            result[k] = v.isoformat()
        elif isinstance(v, (int, float, str, bool, list, dict)) or v is None:
            result[k] = v
        else:
            result[k] = str(v)
    return result

