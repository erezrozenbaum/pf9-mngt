"""
Export helpers: generate Excel (.xlsx) and PDF reports from a migration plan dict.
Called by migration_routes.py export endpoints.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List

# ── Excel ────────────────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

# ── PDF ──────────────────────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

# ══════════════════════════════════════════════════════════════════════════════
# Colour palette
# ══════════════════════════════════════════════════════════════════════════════
_BLUE    = "1D4ED8"  # header bg
_BLUE_LT = "DBEAFE"  # section header
_GREY    = "F3F4F6"  # alternating row
_GREEN   = "DCFCE7"
_RED     = "FEE2E2"
_YELLOW  = "FEF9C3"
_WHITE   = "FFFFFF"

def _thin_border():
    s = Side(border_style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


# ══════════════════════════════════════════════════════════════════════════════
# Excel report
# ══════════════════════════════════════════════════════════════════════════════

def generate_excel_report(plan: Dict[str, Any], project_name: str) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    _sheet_summary(wb, plan, project_name)
    _sheet_per_tenant(wb, plan)
    _sheet_daily_schedule(wb, plan)
    _sheet_all_vms(wb, plan)
    _sheet_methodology(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── helpers ──────────────────────────────────────────────────────────────────

def _hdr_font(wb):
    return Font(bold=True, color=_WHITE, name="Calibri", size=11)

def _hdr_fill():
    return PatternFill("solid", fgColor=_BLUE)

def _hdr_fill_lt():
    return PatternFill("solid", fgColor=_BLUE_LT)

def _grey_fill():
    return PatternFill("solid", fgColor=_GREY)

def _set_col_widths(ws, widths: List[float]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _write_table_header(ws, row: int, cols: List[str]):
    for ci, label in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=label)
        c.font = Font(bold=True, color=_WHITE, name="Calibri", size=9)
        c.fill = _hdr_fill()
        c.border = _thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _write_row(ws, row: int, values: List[Any], alt: bool = False, fill_hex: str | None = None):
    fill = PatternFill("solid", fgColor=fill_hex) if fill_hex else (_grey_fill() if alt else PatternFill())
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.fill = fill
        c.border = _thin_border()
        c.font = Font(name="Calibri", size=9)
        c.alignment = Alignment(vertical="top", wrap_text=True)

# ── Sheet 1: Summary ──────────────────────────────────────────────────────────

def _sheet_summary(wb, plan: Dict, project_name: str):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = f"Migration Plan Report — {project_name}"
    t.font = Font(bold=True, name="Calibri", size=16, color=_BLUE)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    sub = ws["A2"]
    sub.value = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    sub.font = Font(name="Calibri", size=10, color="6B7280")

    ws.append([])  # blank row 3

    ps = plan.get("project_summary", {})
    bm = plan.get("bandwidth_model", {})

    kv_pairs = [
        ("Total VMs",            ps.get("total_vms", "")),
        ("Total Tenants",        ps.get("total_tenants", "")),
        ("Total Disk (TB)",      ps.get("total_disk_tb", "")),
        ("Warm Eligible",        ps.get("warm_eligible", "")),
        ("Cold Required",        ps.get("cold_required", "")),
        ("Warm Risky",           ps.get("warm_risky", "")),
        ("vJailbreak Agents",    ps.get("agent_count", "")),
        ("Concurrent Slots",     ps.get("total_concurrent_slots", "")),
        ("Schedule Days",        ps.get("estimated_schedule_days", "")),
        ("Project Duration",     f"{ps.get('project_duration_days', '')} days"),
        ("Working Hours/Day",    ps.get("working_hours_per_day", "")),
        ("Total Downtime (h)",   ps.get("total_downtime_hours", "")),
        ("Bottleneck",           bm.get("bottleneck", "")),
        ("Bottleneck (Mbps)",    ps.get("bottleneck_mbps", "")),
        ("Source NIC (Mbps)",    bm.get("source_effective_mbps", "")),
        ("Link (Mbps)",          bm.get("link_effective_mbps", "")),
        ("Agent Ingest (Mbps)",  bm.get("agent_effective_mbps", "")),
        ("Storage Write (Mbps)", bm.get("storage_effective_mbps", "")),
    ]

    ws.cell(row=4, column=1, value="Metric").font = Font(bold=True, name="Calibri", size=10)
    ws.cell(row=4, column=2, value="Value").font = Font(bold=True, name="Calibri", size=10)
    ws.cell(row=4, column=1).fill = _hdr_fill_lt()
    ws.cell(row=4, column=2).fill = _hdr_fill_lt()

    for i, (k, v) in enumerate(kv_pairs, 5):
        alt = (i % 2 == 0)
        ws.cell(row=i, column=1, value=k).fill = _grey_fill() if alt else PatternFill()
        ws.cell(row=i, column=2, value=v).fill = _grey_fill() if alt else PatternFill()
        for ci in (1, 2):
            ws.cell(row=i, column=ci).font = Font(name="Calibri", size=10)
            ws.cell(row=i, column=ci).border = _thin_border()

    _set_col_widths(ws, [28, 22, 10, 10, 10])


# ── Sheet 2: Per-Tenant Assessment ────────────────────────────────────────────

def _sheet_per_tenant(wb, plan: Dict):
    ws = wb.create_sheet("Per-Tenant Assessment")
    ws.sheet_view.showGridLines = False

    cols = [
        "Tenant", "OrgVDC", "VMs", "vCPU", "RAM (GB)",
        "Disk (GB)", "In Use (GB)", "Warm", "Warm Risky", "Cold",
        "Phase1 (h)", "Cutover (h)", "Downtime (h)", "Avg Risk Score",
    ]
    _write_table_header(ws, 1, cols)

    for i, tp in enumerate(plan.get("tenant_plans", []), 2):
        mode_fill = None
        if tp.get("cold_count", 0) > 0:
            mode_fill = _RED
        elif tp.get("warm_risky_count", 0) > 0:
            mode_fill = _YELLOW
        _write_row(ws, i, [
            tp.get("tenant_name", ""),
            tp.get("org_vdc", ""),
            tp.get("vm_count", 0),
            tp.get("total_vcpu", 0),
            round(tp.get("total_ram_mb", 0) / 1024, 1) if tp.get("total_ram_mb") else 0,
            round(float(tp.get("total_disk_gb", 0)), 1),
            round(float(tp.get("total_in_use_gb", 0)), 1),
            tp.get("warm_count", 0),
            tp.get("warm_risky_count", 0),
            tp.get("cold_count", 0),
            round(float(tp.get("total_warm_phase1_hours", 0)), 2),
            round(float(tp.get("total_warm_cutover_hours", 0)), 2),
            round(float(tp.get("total_downtime_hours", 0)), 2),
            round(float(tp.get("avg_risk_score", 0)), 1),
        ], alt=(i % 2 == 0), fill_hex=mode_fill)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws.freeze_panes = "A2"
    _set_col_widths(ws, [22, 16, 6, 6, 8, 9, 10, 7, 9, 6, 9, 9, 10, 11])


# ── Sheet 3: Daily Schedule ───────────────────────────────────────────────────

def _sheet_daily_schedule(wb, plan: Dict):
    ws = wb.create_sheet("Daily Schedule")
    ws.sheet_view.showGridLines = False

    cols = ["Day", "Tenant", "VM Name", "Mode", "OS Family", "Disk (GB)", "In Use (GB)", "VM Est. Hours"]
    _write_table_header(ws, 1, cols)

    row = 2
    for day in plan.get("daily_schedule", []):
        day_num = day.get("day", "")
        wall_h  = day.get("wall_clock_hours", "")
        vm_cnt  = day.get("vm_count", len(day.get("vms", [])))
        # Day summary row (light blue)
        _write_row(ws, row, [
            f"Day {day_num}",
            f"{vm_cnt} VM{'s' if vm_cnt != 1 else ''}",
            f"Wall-clock: {wall_h}h",
            "", "", "", "", "",
        ], alt=False, fill_hex="DBEAFE")
        row += 1
        for vm in day.get("vms", []):
            mode = vm.get("mode", "")
            fill = _RED if mode == "cold_required" else (_YELLOW if mode == "warm_risky" else None)
            _write_row(ws, row, [
                "",
                vm.get("tenant_name", ""),
                vm.get("vm_name", ""),
                mode.replace("_", " "),
                vm.get("os_family", ""),
                round(float(vm.get("disk_gb", 0)), 1),
                round(float(vm.get("in_use_gb", 0)), 1),
                round(float(vm.get("estimated_hours", 0)), 2),
            ], alt=(row % 2 == 0), fill_hex=fill)
            row += 1
        # blank row between days
        row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws.freeze_panes = "A2"
    _set_col_widths(ws, [6, 22, 30, 14, 12, 9, 10, 9])


# ── Sheet 4: All VMs ─────────────────────────────────────────────────────────

def _sheet_all_vms(wb, plan: Dict):
    ws = wb.create_sheet("All VMs")
    ws.sheet_view.showGridLines = False

    cols = [
        "VM Name", "Tenant", "OrgVDC", "vCPU", "CPU %", "RAM (MB)", "Mem %",
        "Disk (GB)", "In Use (GB)", "OS Family", "OS Version", "Power",
        "NICs", "Network", "IP", "Mode", "Risk", "Risk Score",
        "Phase1 (h)", "Cutover (h)", "Downtime (h)",
    ]
    _write_table_header(ws, 1, cols)

    row = 2
    for tp in plan.get("tenant_plans", []):
        for vm in tp.get("vms", []):
            mode = vm.get("migration_mode", "")
            risk = vm.get("risk_category", "")
            fill = _RED if mode == "cold_required" else (
                _YELLOW if mode == "warm_risky" else (
                    _RED if risk == "RED" else (_YELLOW if risk == "YELLOW" else None)
                )
            )
            _write_row(ws, row, [
                vm.get("vm_name", ""),
                tp.get("tenant_name", ""),
                tp.get("org_vdc", ""),
                vm.get("cpu_count", ""),
                vm.get("cpu_usage_percent", ""),
                vm.get("ram_mb", ""),
                vm.get("memory_usage_percent", ""),
                round(float(vm.get("total_disk_gb", 0) or 0), 1),
                round(float(vm.get("in_use_gb", 0) or 0), 1),
                vm.get("os_family", ""),
                vm.get("os_version", ""),
                vm.get("power_state", ""),
                vm.get("nic_count", ""),
                vm.get("network_name", ""),
                vm.get("primary_ip", ""),
                mode.replace("_", " "),
                risk,
                round(float(vm.get("risk_score", 0) or 0), 1),
                round(float(vm.get("warm_phase1_hours", 0) or 0), 2),
                round(float(vm.get("warm_cutover_hours", 0) or 0), 2),
                round(float(vm.get("warm_downtime_hours", 0) or 0), 2),
            ], alt=(row % 2 == 0), fill_hex=fill)
            row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws.freeze_panes = "A2"
    _set_col_widths(ws, [28, 20, 16, 5, 6, 8, 6, 8, 9, 10, 22, 8, 4, 20, 14, 13, 6, 8, 9, 9, 10])


# ── Sheet 5: Methodology ─────────────────────────────────────────────────────

def _sheet_methodology(wb):
    """Add a Methodology reference sheet explaining all calculations."""
    ws = wb.create_sheet("Methodology")
    ws.sheet_view.showGridLines = False

    def _title_row(ws, row: int, text: str):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, name="Calibri", size=12, color=_WHITE)
        c.fill = PatternFill("solid", fgColor=_BLUE)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 20

    def _sub_row(ws, row: int, label: str, detail: str, alt: bool = False):
        fill = PatternFill("solid", fgColor="F3F4F6") if alt else PatternFill()
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(bold=True, name="Calibri", size=9)
        c1.fill = fill
        c1.border = _thin_border()
        c1.alignment = Alignment(vertical="top", wrap_text=True)

        ws.merge_cells(f"B{row}:D{row}")
        c2 = ws.cell(row=row, column=2, value=detail)
        c2.font = Font(name="Calibri", size=9)
        c2.fill = fill
        c2.border = _thin_border()
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 36

    # Title
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "Platform9 Migration Planner — Calculation Methodology"
    t.font = Font(bold=True, name="Calibri", size=14, color=_BLUE)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.append([])  # blank row 2

    rows = [
        # (section_title, [(label, detail), ...])
        ("1. Warm vs Cold Classification", [
            ("Warm Eligible",
             "OS does NOT appear in the OS Cold Required list AND VM is powered ON. "
             "Live data-copy runs while the VM is active; only the cutover window causes downtime."),
            ("Cold Required",
             "OS matches an entry in the Risk Config → Cold Migration Rules list, OR the VM is "
             "poweredOff / suspended. The VM is shut down before the copy begins, so the entire "
             "copy duration adds to downtime."),
            ("Warm Risky",
             "Warm-eligible but the VM's risk score meets or exceeds the Yellow/Red threshold. "
             "Treated identically to Warm for timing but flagged for extra review."),
        ]),
        ("2. Bandwidth Model & Bottleneck", [
            ("Formula",
             "effective_mbps = min(source_nic_mbps, link_mbps, agent_mbps, storage_mbps)\n"
             "Each layer has an efficiency factor applied (e.g. source NIC × 0.70 for TCP overhead)."),
            ("Example",
             "Source NIC = 4,800 Mbps · Link = 4,000 Mbps · Agent = 6,000 Mbps · Storage = 5,500 Mbps"
             " → Bottleneck = link at 4,000 Mbps"),
        ]),
        ("3. Data-Copy Time", [
            ("Formula",
             "data_copy_hours = total_in_use_gb ÷ (bottleneck_mbps ÷ 8 × 3600 ÷ 1024)\n"
             "Where ÷8 converts Mb→MB and ×3600÷1024 converts MB/s → GB/h."),
            ("Example",
             "172,617 GB at 4,000 Mbps:\n"
             " 4,000 ÷ 8 × 3,600 ÷ 1,024 = 1,757.8 GB/h\n"
             " 172,617 ÷ 1,757.8 = 98.2 hours total data-copy time"),
        ]),
        ("4. Tech Fix Time Scoring", [
            ("Factor Weights (defaults)",
             "Windows OS: +20 min | Extra volume (each): +15 min | Extra NIC (each): +10 min\n"
             "Cold mode: +15 min | Risk YELLOW: +15 min | Risk RED: +25 min\n"
             "Has snapshots: +10 min | Cross-tenant dep: +15 min | Unknown OS: +5 min"),
            ("OS Fix Rates (defaults)",
             "Windows 50% · Linux 20% · Other 40%\n"
             "Global override: when set, replaces all per-OS rates."),
            ("Formula",
             "raw_score = Σ (factor_present × factor_weight)\n"
             "fix_minutes = raw_score × fix_rate"),
            ("Example",
             "Windows VM, cold mode, 3 volumes (2 extra), 2 NICs (1 extra), risk YELLOW:\n"
             " Raw = 20 + 15×2 + 10×1 + 15 + 15 = 95\n"
             " Fix = 95 × 0.50 = 47.5 min → Total downtime = 47.5 + 30 (cutover) = 77.5 min"),
        ]),
        ("5. Downtime Estimate", [
            ("Warm VMs",
             "downtime = cutover_minutes + fix_minutes\n"
             "Data copy is live (no user-visible downtime during Phase 1)."),
            ("Cold VMs",
             "downtime = data_copy_hours × 60 + cutover_minutes + fix_minutes\n"
             "VM is shut down for the entire copy phase."),
            ("Total Project Downtime",
             "Sum of individual VM downtime hours across all VMs.\n"
             "These are concurrent across tenants; wall-clock calendar time is much shorter."),
        ]),
        ("6. Daily Schedule Packing", [
            ("Algorithm",
             "1. Sort VMs: Cohort order → priority within cohort → disk size (largest first)\n"
             "2. Assign VMs to current day until wall-clock hours exceed working_hours_per_day\n"
             "3. Wall-clock per day = max(estimated_hours) among VMs assigned that day\n"
             "4. concurrent_slots = total agent slots across all vJailbreak agents"),
            ("Example",
             "12 concurrent slots, 8-hour working day, 30 VMs of varying size:\n"
             " Day 1: 12 VMs → largest takes 6.2h → wall-clock = 6.2h\n"
             " Day 2: next 12 VMs → largest = 4.8h\n"
             " Day 3: remaining 6 VMs"),
        ]),
        ("7. Risk Score & Categories", [
            ("Thresholds (default)",
             "GREEN: 0 – 29 → standard migration\n"
             "YELLOW: 30 – 59 → pre-migration review required\n"
             "RED: 60+ → architecture review + rollback plan required"),
            ("Configuration",
             "All thresholds and factor weights are configurable in the Risk Config tab."),
        ]),
    ]

    r = 3
    for section_title, entries in rows:
        _title_row(ws, r, section_title)
        r += 1
        for idx, (label, detail) in enumerate(entries):
            _sub_row(ws, r, label, detail, alt=(idx % 2 == 0))
            r += 1
        ws.append([])
        r += 1

    _set_col_widths(ws, [28, 30, 30, 20])


# ══════════════════════════════════════════════════════════════════════════════
# PDF report
# ══════════════════════════════════════════════════════════════════════════════

def generate_pdf_report(plan: Dict[str, Any], project_name: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2.0 * cm,
        bottomMargin=1.5 * cm,
        title=f"Migration Plan — {project_name}",
    )

    styles = getSampleStyleSheet()
    s_title   = ParagraphStyle("Title2",   parent=styles["Title"],   fontSize=18, spaceAfter=4, textColor=colors.HexColor("#1D4ED8"))
    s_h2      = ParagraphStyle("H2",       parent=styles["Heading2"],fontSize=12, spaceAfter=4, textColor=colors.HexColor("#1E40AF"), spaceBefore=14)
    s_body    = ParagraphStyle("Body",     parent=styles["Normal"],  fontSize=9,  spaceAfter=4)
    s_caption = ParagraphStyle("Caption",  parent=styles["Normal"],  fontSize=8,  textColor=colors.grey, spaceAfter=8)
    s_footer  = ParagraphStyle("Footer",   parent=styles["Normal"],  fontSize=8,  textColor=colors.grey, alignment=TA_CENTER)
    # Small wrap styles for table cells with potentially long content
    s_cell7   = ParagraphStyle("Cell7",    parent=styles["Normal"],  fontSize=7,  leading=8)
    s_cell8   = ParagraphStyle("Cell8",    parent=styles["Normal"],  fontSize=8,  leading=9)
    s_legend  = ParagraphStyle("Legend",   parent=styles["Normal"],  fontSize=7.5, textColor=colors.HexColor("#374151"), spaceAfter=6, leading=11)

    _HDR  = colors.HexColor("#1D4ED8")
    _HDR_LT = colors.HexColor("#DBEAFE")
    _GREY_P = colors.HexColor("#F3F4F6")
    _RED_P  = colors.HexColor("#FEE2E2")
    _YEL_P  = colors.HexColor("#FEF9C3")
    _GRN_P  = colors.HexColor("#DCFCE7")
    _BLK    = colors.black

    story: list = []

    # ── Title ──────────────────────────────────────────────────────────────
    story.append(Paragraph(f"Migration Plan Report", s_title))
    story.append(Paragraph(f"<b>Project:</b> {project_name}", s_body))
    story.append(Paragraph(
        f"Generated: {datetime.utcnow().strftime('%B %d, %Y  %H:%M UTC')}",
        s_caption,
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=_HDR, spaceAfter=10))

    # ── Executive Summary (if present) ────────────────────────────────────
    exec_summary = plan.get("executive_summary", "") or ""
    if exec_summary.strip():
        story.append(Paragraph("Executive Summary", s_h2))
        story.append(Paragraph(exec_summary, s_body))
        story.append(Spacer(1, 0.3 * cm))

    # ── Technical Notes (if present) ──────────────────────────────────────
    tech_notes = plan.get("technical_notes", "") or ""
    if tech_notes.strip():
        story.append(Paragraph("Technical Notes", s_h2))
        story.append(Paragraph(tech_notes, s_body))
        story.append(Spacer(1, 0.3 * cm))

    # ── Project Summary ────────────────────────────────────────────────────
    story.append(Paragraph("Project Summary", s_h2))
    ps = plan.get("project_summary", {})
    bm = plan.get("bandwidth_model", {})

    summary_data = [
        ["Metric", "Value", "Metric", "Value"],
        ["Total VMs",            str(ps.get("total_vms", "—")),
         "vJailbreak Agents",    str(ps.get("agent_count", "—"))],
        ["Total Tenants",        str(ps.get("total_tenants", "—")),
         "Concurrent Slots",     str(ps.get("total_concurrent_slots", "—"))],
        ["Total Disk (TB)",      str(ps.get("total_disk_tb", "—")),
         "Schedule Days",        str(ps.get("estimated_schedule_days", "—"))],
        ["Warm Eligible",        str(ps.get("warm_eligible", "—")),
         "Project Duration",     f"{ps.get('project_duration_days', '—')} days"],
        ["Cold Required",        str(ps.get("cold_required", "—")),
         "Total Downtime (h)",   str(ps.get("total_downtime_hours", "—"))],
        ["Bottleneck",           str(bm.get("bottleneck", "—")).replace("_", " "),
         "Bottleneck (Mbps)",    str(ps.get("bottleneck_mbps", "—"))],
    ]
    t_sum = Table(summary_data, colWidths=[5.5*cm, 3.5*cm, 5.5*cm, 3.5*cm])
    t_sum.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _HDR),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 1), (-1, -1), _GREY_P),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _GREY_P]),
        ("FONTNAME",   (0, 1), (0, -1), "Helvetica-Bold"),
        ("FONTNAME",   (2, 1), (2, -1), "Helvetica-Bold"),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
    ]))
    story.append(t_sum)
    story.append(Paragraph(
        "<b>What the numbers mean:</b>  "
        "<b>Warm Eligible</b> — VM is copied live while running; downtime = cutover window + tech fix time only.  "
        "<b>Cold Required</b> — VM is powered off before copy; downtime = full copy + cutover + tech fix time.  "
        "<b>Warm Risky</b> — live copy is possible but carries elevated risk flags requiring extra review; same downtime model as Warm.  "
        "<b>Concurrent Slots</b> — how many VMs can migrate simultaneously, one slot per vJailbreak agent channel.  "
        "<b>Total Downtime (h)</b> — sum of per-VM downtime hours (cutover + fix for warm; copy + cutover + fix for cold); "
        "not a wall-clock duration since VMs migrate in parallel.",

        s_legend,
    ))
    story.append(Spacer(1, 0.3 * cm))

    # ── Per-Tenant Assessment ──────────────────────────────────────────────
    story.append(Paragraph("Per-Tenant Assessment", s_h2))
    story.append(Paragraph(
        "One row per source tenant/OrgVDC.  "
        "<b>Phase1 (h)</b> — initial data-copy duration (VM stays running for warm; offline for cold).  "
        "<b>Cutover (h)</b> — final switchover window. For warm VMs: cutover + tech fix time = total downtime. "
        "For cold VMs: copy + cutover + tech fix time = total downtime.  "
        "<b>Avg Risk</b> — mean risk score across all VMs in this tenant (see Risk Config tab for thresholds).  "
        "Rows shaded red = tenant has cold-required VMs; yellow = has warm-risky VMs.",
        s_legend,
    ))

    tenant_header = [
        "Tenant", "OrgVDC", "VMs", "vCPU", "RAM (GB)",
        "Disk (GB)", "In Use\n(GB)", "Warm", "Cold", "Phase1 (h)", "Cutover (h)", "Avg Risk",
    ]
    tenant_rows = [tenant_header]
    for tp in plan.get("tenant_plans", []):
        has_cold = tp.get("cold_count", 0) > 0
        has_risky = tp.get("warm_risky_count", 0) > 0
        tenant_rows.append([
            Paragraph(tp.get("tenant_name", ""), s_cell8),
            Paragraph(tp.get("org_vdc", "") or "—", s_cell8),
            tp.get("vm_count", 0),
            tp.get("total_vcpu", 0),
            round(tp.get("total_ram_mb", 0) / 1024, 1) if tp.get("total_ram_mb") else 0,
            round(float(tp.get("total_disk_gb", 0)), 1),
            round(float(tp.get("total_in_use_gb", 0)), 1),
            tp.get("warm_count", 0),
            tp.get("cold_count", 0),
            round(float(tp.get("total_warm_phase1_hours", 0)), 1),
            round(float(tp.get("total_warm_cutover_hours", 0)), 1),
            round(float(tp.get("avg_risk_score", 0)), 1),
        ])

    # Total = 26.3 cm (landscape A4 usable = 26.7 cm)
    cw_tenant = [4.5*cm, 3.5*cm, 1.3*cm, 1.5*cm, 2*cm, 2*cm, 2*cm, 1.3*cm, 1.3*cm, 2.2*cm, 2.2*cm, 2*cm]
    t_tenant = Table(tenant_rows, colWidths=cw_tenant, repeatRows=1)
    style_tenant = [
        ("BACKGROUND", (0, 0), (-1, 0), _HDR),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _GREY_P]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",      (2, 0), (-1, -1), "CENTER"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
    ]
    # Highlight rows with cold VMs
    for ri, tp in enumerate(plan.get("tenant_plans", []), 1):
        if tp.get("cold_count", 0) > 0:
            style_tenant.append(("BACKGROUND", (0, ri), (-1, ri), _RED_P))
        elif tp.get("warm_risky_count", 0) > 0:
            style_tenant.append(("BACKGROUND", (0, ri), (-1, ri), _YEL_P))
    t_tenant.setStyle(TableStyle(style_tenant))
    story.append(t_tenant)
    story.append(Spacer(1, 0.4 * cm))

    # ── Daily Schedule ─────────────────────────────────────────────────────
    story.append(Paragraph("Daily Migration Schedule (Estimated)", s_h2))
    story.append(Paragraph(
        f"VMs are ordered by cohort → priority → disk size and packed into working days. "
        f"Each day runs up to {ps.get('total_concurrent_slots', '?')} VMs concurrently (one slot per vJailbreak agent channel). "
        f"Wall-clock time = time to push all that day\u2019s data through the shared network pipe.",
        s_caption,
    ))
    story.append(Paragraph(
        "<b>Column guide:</b>  "
        "<b>Mode</b> — warm eligible (VM stays online; only the cutover window = downtime) / "
        "warm risky (same, flagged for review) / cold required (VM shut down before copy; full copy is downtime).  "
        "<b>Power</b> — current power state of the VM in the source environment (On / Off / Susp).  "
        "<b>OS</b> — operating system family and version (e.g. Windows Server 2022, Ubuntu 22.04).  "
        "<b>In Use (GB)</b> — actual data bytes transferred over the network for this VM.  "
        "<b>Copy (h)</b> — time to transfer data: Phase\u00a01 live copy for warm VMs; full offline copy for cold VMs.  "
        "<b>Cutover (h)</b> — final switchover window (VM briefly offline). This is the <i>only</i> migration downtime for warm VMs.  "
        "<b>Fix (h)</b> — estimated post-migration tech fix time (NIC rename, disk UUID, app validation \u2014 weighted by OS fix rate).  "
        "<b>Downtime (h)</b> — total business-impact period: cutover + fix for warm; copy + cutover + fix for cold.",
        s_legend,
    ))

    # Build VM timing lookup from tenant_plans for enriching daily schedule rows
    _vm_timing: Dict[tuple, Dict] = {}
    for _tp in plan.get("tenant_plans", []):
        for _vm in _tp.get("vms", []):
            _vm_timing[(_vm.get("vm_name"), _tp.get("tenant_name"))] = _vm

    def _pwr_label(raw: str) -> str:
        r = (raw or "").lower()
        if "on" in r:   return "On"
        if "off" in r:  return "Off"
        if "sus" in r:  return "Susp"
        return raw or "—"

    # Cols: Day | Tenant | VM Name | Mode | Power | OS | In Use(GB) | Copy(h) | Cutover(h) | Fix(h) | Downtime(h)
    sched_header = ["Day", "Tenant", "VM Name", "Mode", "Power", "OS",
                    "In Use\n(GB)", "Copy\n(h)", "Cutover\n(h)", "Fix\n(h)", "Downtime\n(h)"]
    sched_rows = [sched_header]
    day_header_row_indices: list = []
    for day in plan.get("daily_schedule", []):
        day_num = day.get("day", "")
        wall_h  = day.get("wall_clock_hours", "?")
        vm_cnt  = day.get("vm_count", len(day.get("vms", [])))
        day_header_row_indices.append(len(sched_rows))
        sched_rows.append([
            f"Day {day_num}",
            f"{vm_cnt} VM{'s' if vm_cnt != 1 else ''}  \u2014  Wall-clock: {wall_h}h",
            "", "", "", "", "", "", "", "", "",
        ])
        for vm in day.get("vms", []):
            v_data    = _vm_timing.get((vm.get("vm_name"), vm.get("tenant_name")), {})
            is_cold   = "cold" in (vm.get("mode") or "")
            copy_h    = round(float(v_data.get("cold_total_hours", 0) or 0), 2) if is_cold else round(float(v_data.get("warm_phase1_hours", 0) or 0), 2)
            cutover_h = round(float(v_data.get("warm_cutover_hours", 0) or 0), 2)
            mig_dn_h  = round(float(v_data.get("cold_downtime_hours", 0) or 0), 2) if is_cold else round(float(v_data.get("warm_downtime_hours", 0) or 0), 2)
            fix_h     = round(float(v_data.get("fix_hours", 0) or 0), 2)
            os_str    = f"{v_data.get('os_family', '') or ''} {v_data.get('os_version', '') or ''}".strip() or (vm.get("os_family") or "\u2014")
            pwr_str   = _pwr_label(v_data.get("power_state") or vm.get("power_state"))
            sched_rows.append([
                "",
                Paragraph(vm.get("tenant_name", ""), s_cell8),
                Paragraph(vm.get("vm_name", ""), s_cell8),
                (vm.get("mode", "") or "").replace("_", " "),
                pwr_str,
                Paragraph(os_str, s_cell8),
                round(float(vm.get("in_use_gb", 0)), 1),
                copy_h,
                cutover_h,
                fix_h,
                round(mig_dn_h + fix_h, 2),
            ])

    # Landscape A4 usable ~26.5cm; 11 cols: 1.2+3.0+5.0+2.2+1.5+4.0+1.5+1.5+1.5+1.4+1.7 = 24.5cm
    cw_sched = [1.2*cm, 3.0*cm, 5.0*cm, 2.2*cm, 1.5*cm, 4.0*cm, 1.5*cm, 1.5*cm, 1.5*cm, 1.4*cm, 1.7*cm]
    t_sched = Table(sched_rows, colWidths=cw_sched, repeatRows=1)
    style_sched = [
        ("BACKGROUND", (0, 0), (-1, 0), _HDR),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 7.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _GREY_P]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
        ("ALIGN",      (6, 0), (-1, -1), "RIGHT"),   # numeric cols shifted by 1 (Power added)
        ("LEFTPADDING",  (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING",   (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 2),
    ]
    # Day header rows: light blue background, bold, span cols 1 onward
    for dhi in day_header_row_indices:
        style_sched.append(("BACKGROUND", (0, dhi), (-1, dhi), _HDR_LT))
        style_sched.append(("FONTNAME",   (0, dhi), (-1, dhi), "Helvetica-Bold"))
        style_sched.append(("TEXTCOLOR",  (0, dhi), (-1, dhi), colors.HexColor("#1e40af")))
        style_sched.append(("SPAN",       (1, dhi), (-1, dhi)))
    for ri, row in enumerate(sched_rows[1:], 1):
        if ri in day_header_row_indices:
            continue
        mode_str = row[3] if len(row) > 3 else ""
        if "cold" in mode_str:
            style_sched.append(("BACKGROUND", (0, ri), (-1, ri), _RED_P))
        elif "risky" in mode_str:
            style_sched.append(("BACKGROUND", (0, ri), (-1, ri), _YEL_P))
    t_sched.setStyle(TableStyle(style_sched))
    story.append(t_sched)
    story.append(Spacer(1, 0.4 * cm))

    # ── All VMs ────────────────────────────────────────────────────────────
    from reportlab.platypus import PageBreak
    story.append(PageBreak())
    story.append(Paragraph("All VMs — Full Inventory with Timing", s_h2))
    story.append(Paragraph(
        "Complete VM-by-VM assessment. Full hardware detail (vCPU, RAM, NICs, IP) is available in the Excel export.",
        s_caption,
    ))
    story.append(Paragraph(
        "<b>Column guide:</b>  "
        "<b>OS</b> — operating system family and version (e.g. Windows Server 2022, Ubuntu 22.04). Full hardware details available in the Excel export.  "
        "<b>Mode</b> — warm eligible / warm risky / cold required (row shading matches: red = cold, yellow = warm risky).  "
        "<b>Risk</b> — GREEN / YELLOW / RED complexity score (cell shaded accordingly).  "
        "<b>Disk (GB)</b> — total provisioned storage.  "
        "<b>In Use (GB)</b> — actual data transferred over the wire.  "
        "<b>Copy (h)</b> — Phase 1 live copy for warm VMs; full offline copy for cold VMs.  "
        "<b>Cutover (h)</b> — final switchover window (only migration downtime for warm VMs).  "
        "<b>Fix (h)</b> — estimated post-migration tech fix time (NIC, disk, app validation) weighted by OS fix rate.  "
        "<b>Downtime (h)</b> — total business-impact: cutover + fix for warm; copy + cutover + fix for cold.",
        s_legend,
    ))

    vm_all_header = [
        "VM Name", "Tenant", "OS", "Mode", "Risk",
        "Disk\n(GB)", "In Use\n(GB)", "Copy\n(h)", "Cutover\n(h)", "Fix\n(h)", "Downtime\n(h)",
    ]
    vm_all_rows = [vm_all_header]
    for tp in plan.get("tenant_plans", []):
        for vm in tp.get("vms", []):
            mode    = vm.get("migration_mode", "")
            risk    = vm.get("risk_category", "")
            is_cold = "cold" in mode
            os_str  = f"{vm.get('os_family','') or ''} {vm.get('os_version','') or ''}".strip() or "—"
            copy_h    = round(float(vm.get("cold_total_hours",    0) or 0), 2) if is_cold else round(float(vm.get("warm_phase1_hours",  0) or 0), 2)
            cutover_h = round(float(vm.get("warm_cutover_hours",  0) or 0), 2)
            mig_dn_h  = round(float(vm.get("cold_downtime_hours", 0) or 0), 2) if is_cold else round(float(vm.get("warm_downtime_hours", 0) or 0), 2)
            fix_h     = round(float(vm.get("fix_hours",           0) or 0), 2)
            vm_all_rows.append([
                Paragraph(vm.get("vm_name", ""), s_cell7),
                Paragraph(tp.get("tenant_name", ""), s_cell7),
                Paragraph(os_str, s_cell7),
                (mode or "").replace("_", " "),
                risk or "—",
                round(float(vm.get("total_disk_gb", 0) or 0), 1),
                round(float(vm.get("in_use_gb",    0) or 0), 1),
                copy_h,
                cutover_h,
                fix_h,
                round(mig_dn_h + fix_h, 2),
            ])

    # 11 cols — total ~22.1 cm, fits landscape A4 usable width of 26.7 cm
    cw_all = [4.0*cm, 2.5*cm, 2.5*cm, 2.2*cm, 1.3*cm, 1.5*cm, 1.5*cm, 1.8*cm, 1.8*cm, 1.5*cm, 1.7*cm]
    t_all = Table(vm_all_rows, colWidths=cw_all, repeatRows=1)
    style_all = [
        ("BACKGROUND", (0, 0), (-1, 0), _HDR),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _GREY_P]),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",      (5, 0), (10, -1), "RIGHT"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING",   (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 2),
    ]
    # Colour rows by mode (col 3) / risk cell (col 4)
    for ri, vm_row in enumerate(vm_all_rows[1:], 1):
        mode_str = vm_row[3] if len(vm_row) > 3 else ""
        risk_str = vm_row[4] if len(vm_row) > 4 else ""
        if "cold" in mode_str:
            style_all.append(("BACKGROUND", (0, ri), (-1, ri), _RED_P))
        elif "risky" in mode_str:
            style_all.append(("BACKGROUND", (0, ri), (-1, ri), _YEL_P))
        elif risk_str == "RED":
            style_all.append(("BACKGROUND", (4, ri), (4, ri), _RED_P))
        elif risk_str == "YELLOW":
            style_all.append(("BACKGROUND", (4, ri), (4, ri), _YEL_P))
        elif risk_str == "GREEN":
            style_all.append(("BACKGROUND", (4, ri), (4, ri), _GRN_P))
    t_all.setStyle(TableStyle(style_all))
    story.append(t_all)

    # ── Methodology ────────────────────────────────────────────────────────
    from reportlab.platypus import PageBreak
    story.append(PageBreak())
    story += _pdf_methodology_section(s_h2, s_body, s_caption, _HDR, _GREY_P, _HDR_LT, colors)

    # ── Footer page numbers via onLaterPages ───────────────────────────────
    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.grey)
        txt = f"Platform9 Migration Planner  ·  {project_name}  ·  Page {doc.page}"
        canvas.drawCentredString(landscape(A4)[0] / 2, 0.8 * cm, txt)
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# Migration Summary Report  (Excel + PDF)
# ══════════════════════════════════════════════════════════════════════════════

def generate_summary_excel_report(summary: Dict[str, Any], project_name: str) -> bytes:
    """
    Generate a 4-sheet Excel workbook from a migration-summary API response dict.

    Sheets:
      1. Summary KPIs      — key metrics
      2. Daily Schedule    — per_day[] with over-capacity highlighting
      3. OS Breakdown      — per_os_breakdown[]
      4. Cohort Breakdown  — per_cohort[]
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Summary KPIs ─────────────────────────────────────────────
    ws1 = wb.create_sheet("Summary KPIs")
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:C1")
    t = ws1["A1"]
    t.value = f"Migration Summary — {project_name}"
    t.font = Font(bold=True, name="Calibri", size=16, color=_BLUE)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:C2")
    sub = ws1["A2"]
    sub.value = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    sub.font = Font(name="Calibri", size=10, color="6B7280")

    ws1.append([])  # blank row 3

    migration_days = len(summary.get("per_day") or [])
    kv_pairs = [
        ("Total VMs",              summary.get("total_vms", "")),
        ("In-Use Data (GB)",       summary.get("total_data_gb", "")),
        ("Total Provisioned (GB)", summary.get("total_provisioned_gb", "")),
        ("Migration Days",         migration_days),
        ("Est. Data-Copy (h)",     round(float(summary.get("data_copy_hours") or 0), 2)),
        ("Est. Tech Fix (h)",      round(float(summary.get("total_fix_hours") or 0), 2)),
        ("Est. Copy Hours (h)",    round(float(summary.get("total_copy_hours") or 0), 2)),
        ("Est. Total Downtime (h)", round(float(summary.get("total_downtime_hours") or 0), 2)),
        ("Bandwidth (Mbps)",       summary.get("bandwidth_mbps", "")),
    ]

    ws1.cell(row=4, column=1, value="Metric").font = Font(bold=True, name="Calibri", size=10)
    ws1.cell(row=4, column=2, value="Value").font  = Font(bold=True, name="Calibri", size=10)
    ws1.cell(row=4, column=1).fill = _hdr_fill_lt()
    ws1.cell(row=4, column=2).fill = _hdr_fill_lt()
    for i, (k, v) in enumerate(kv_pairs, 5):
        alt = (i % 2 == 0)
        ws1.cell(row=i, column=1, value=k).fill  = _grey_fill() if alt else PatternFill()
        ws1.cell(row=i, column=2, value=v).fill  = _grey_fill() if alt else PatternFill()
        for ci in (1, 2):
            ws1.cell(row=i, column=ci).font   = Font(name="Calibri", size=10)
            ws1.cell(row=i, column=ci).border = _thin_border()
    _set_col_widths(ws1, [30, 22, 10])

    # ── Sheet 2: Daily Schedule ───────────────────────────────────────────
    ws2 = wb.create_sheet("Daily Schedule")
    ws2.sheet_view.showGridLines = False

    day_cols = [
        "Day", "Cohort", "Tenants", "VMs",
        "Total GB", "Wall-Clock (h)", "Total Agent (h)",
        "Cold", "Warm",
        "Risk Green", "Risk Yellow", "Risk Red", "Over Capacity",
    ]
    _write_table_header(ws2, 1, day_cols)

    _RED_FILL  = PatternFill("solid", fgColor="FEE2E2")
    _YELL_FILL = PatternFill("solid", fgColor="FEF9C3")

    for i, day in enumerate(summary.get("per_day") or [], 2):
        over = bool(day.get("over_capacity"))
        fill_hex = "FEE2E2" if over else None
        _write_row(ws2, i, [
            day.get("day", i - 1),
            day.get("cohort_name", ""),
            day.get("tenant_count", 0),
            day.get("vm_count", 0),
            day.get("total_gb", 0),
            round(float(day.get("wall_clock_hours") or 0), 2),
            round(float(day.get("total_agent_hours") or 0), 2),
            day.get("cold_count", 0),
            day.get("warm_count", 0),
            day.get("risk_green", 0),
            day.get("risk_yellow", 0),
            day.get("risk_red", 0),
            "YES" if over else "no",
        ], alt=(i % 2 == 0), fill_hex=fill_hex)

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(day_cols))}1"
    ws2.freeze_panes = "A2"
    _set_col_widths(ws2, [5, 20, 8, 6, 10, 14, 14, 6, 6, 11, 12, 10, 13])

    # ── Sheet 3: OS Breakdown ─────────────────────────────────────────────
    ws3 = wb.create_sheet("OS Breakdown")
    ws3.sheet_view.showGridLines = False

    os_cols = ["OS Family", "VM Count", "Fix Rate", "Fix Hours (Total)"]
    _write_table_header(ws3, 1, os_cols)
    os_breakdown = summary.get("per_os_breakdown") or {}
    os_items = os_breakdown.items() if isinstance(os_breakdown, dict) else [(e.get("os_family",""), e) for e in os_breakdown]
    for i, (fam, entry) in enumerate(os_items, 2):
        _write_row(ws3, i, [
            fam,
            entry.get("vm_count", 0),
            f"{float(entry.get('fix_rate', 0)):.0%}",
            round(float(entry.get("fix_hours") or 0), 2),
        ], alt=(i % 2 == 0))
    ws3.freeze_panes = "A2"
    _set_col_widths(ws3, [18, 12, 12, 18])

    # ── Sheet 4: Cohort Breakdown ─────────────────────────────────────────
    ws4 = wb.create_sheet("Cohort Breakdown")
    ws4.sheet_view.showGridLines = False

    coh_cols = ["Cohort", "VMs", "Data (GB)", "Data-Copy (h)", "Fix Hours", "Downtime (h)"]
    _write_table_header(ws4, 1, coh_cols)
    for i, coh in enumerate(summary.get("per_cohort") or [], 2):
        _write_row(ws4, i, [
            coh.get("cohort_name", ""),
            coh.get("vm_count", 0),
            round(float(coh.get("data_gb") or 0), 1),
            round(float(coh.get("data_copy_hours") or 0), 2),
            round(float(coh.get("fix_hours") or 0), 2),
            round(float(coh.get("downtime_hours") or 0), 2),
        ], alt=(i % 2 == 0))
    ws4.freeze_panes = "A2"
    _set_col_widths(ws4, [22, 8, 12, 14, 12, 14])

    _sheet_methodology(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_summary_pdf_report(summary: Dict[str, Any], project_name: str) -> bytes:
    """
    Management-level migration summary PDF (landscape A4).
    Sections:
      1. Project Scope — VM mode distribution + risk distribution + infrastructure stats
      2. Key Migration Estimates — KPI table with description column
      3. Per-Tenant Migration Plan — tenant-by-tenant breakdown
      4. Daily Migration Schedule
      5. Cohort Breakdown
      6. OS Family Breakdown & Fix Rates
      7. Calculation Methodology
      Appendix: Detailed Methodology Reference
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2.0 * cm,
        bottomMargin=1.5 * cm,
        title=f"Migration Summary — {project_name}",
    )

    styles = getSampleStyleSheet()
    s_title  = ParagraphStyle("ST",  parent=styles["Title"],   fontSize=18, spaceAfter=4,   textColor=colors.HexColor("#1D4ED8"))
    s_h2     = ParagraphStyle("SH2", parent=styles["Heading2"],fontSize=12, spaceAfter=4,   textColor=colors.HexColor("#1E40AF"), spaceBefore=14)
    s_caption= ParagraphStyle("SC",  parent=styles["Normal"],  fontSize=8,  textColor=colors.grey, spaceAfter=5)
    s_legend = ParagraphStyle("SL",  parent=styles["Normal"],  fontSize=8,  textColor=colors.HexColor("#374151"), spaceAfter=6, leading=11)
    s_cell   = ParagraphStyle("SK",  parent=styles["Normal"],  fontSize=8,  leading=10)

    _HDR    = colors.HexColor("#1D4ED8")
    _HDR_LT = colors.HexColor("#DBEAFE")
    _GREY_P = colors.HexColor("#F3F4F6")
    _RED_P  = colors.HexColor("#FEE2E2")
    _YEL_P  = colors.HexColor("#FEF9C3")
    _GRN_P  = colors.HexColor("#DCFCE7")

    story: list = []

    def _tbl_base(extra=None):
        base = [
            ("BACKGROUND",    (0, 0), (-1, 0), _HDR),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 8.5),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, _GREY_P]),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING",   (0, 0), (-1, -1), 4),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        return TableStyle(base + (extra or []))

    # ── Pull summary data ─────────────────────────────────────────────────────
    total_vms        = int(summary.get("total_vms", 0) or 0)
    total_prov_gb    = round(float(summary.get("total_provisioned_gb") or 0), 1)
    total_data_gb    = round(float(summary.get("total_data_gb") or 0), 1)
    data_copy_h      = round(float(summary.get("data_copy_hours") or 0), 2)
    fix_h_total      = round(float(summary.get("total_fix_hours") or 0), 2)
    downtime_h_total = round(float(summary.get("total_downtime_hours") or 0), 2)
    bandwidth        = summary.get("bandwidth_mbps", "—")
    warm_eligible    = int(summary.get("warm_eligible", 0) or 0)
    warm_risky       = int(summary.get("warm_risky", 0) or 0)
    cold_required    = int(summary.get("cold_required", 0) or 0)
    total_tenants    = int(summary.get("total_tenants", 0) or 0)
    agent_count      = int(summary.get("agent_count", 0) or 0)
    total_slots      = int(summary.get("total_concurrent_slots", 0) or 0)
    migration_days   = len(summary.get("per_day") or [])
    vm_risk_dist     = summary.get("vm_risk_dist", {})
    risk_green       = int(vm_risk_dist.get("GREEN", 0) or 0)
    risk_yellow      = int(vm_risk_dist.get("YELLOW", 0) or 0)
    risk_red         = int(vm_risk_dist.get("RED", 0) or 0)

    def _pct(n):
        return f"{n / max(total_vms, 1) * 100:.0f}%"

    # ── Title ─────────────────────────────────────────────────────────────────
    story.append(Paragraph("Migration Summary Report", s_title))
    story.append(Paragraph(
        f"<b>Project:</b> {project_name}     "
        f"<b>Generated:</b> {datetime.utcnow().strftime('%B %d, %Y  %H:%M UTC')}",
        s_caption,
    ))
    story.append(HRFlowable(width="100%", thickness=1.5, color=_HDR, spaceAfter=10))

    # ── Section 1: Project Scope ──────────────────────────────────────────────
    story.append(Paragraph("1  ·  Project Scope", s_h2))
    story.append(Paragraph(
        "Infrastructure overview, VM migration mode distribution, and risk classification at a glance. "
        "<b>Warm Eligible</b> — live copy while VM stays online; only the brief cutover window is downtime. "
        "<b>Warm Risky</b> — same approach, but flagged for extra review due to risk factors. "
        "<b>Cold Required</b> — VM shut down before copy; the full copy duration is downtime.",
        s_legend,
    ))

    # Infrastructure stats (left panel)
    infra_data = [
        ["Infrastructure", ""],
        ["Total VMs in Scope",    str(total_vms)],
        ["Total Tenants",         str(total_tenants)],
        ["Migration Cohorts",     str(len(summary.get("per_cohort") or []))],
        ["Migration Days (est.)", str(migration_days)],
        ["Migration Agents",      str(agent_count)],
        ["Concurrent Slots",      str(total_slots)],
        ["Bandwidth (Mbps)",      str(bandwidth)],
    ]
    t_infra = Table(infra_data, colWidths=[4.5*cm, 3.0*cm])
    infra_extra = [
        ("SPAN",       (0, 0), (-1, 0)),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME",   (0, 1), (0, -1), "Helvetica-Bold"),
        ("ALIGN",      (1, 1), (1, -1), "RIGHT"),
    ]
    t_infra.setStyle(_tbl_base(infra_extra))

    # VM Mode distribution (centre panel)
    mode_data = [
        ["VM Migration Mode", "Count", "%"],
        ["Warm Eligible",     str(warm_eligible), _pct(warm_eligible)],
        ["Warm Risky",        str(warm_risky),    _pct(warm_risky)],
        ["Cold Required",     str(cold_required), _pct(cold_required)],
        ["TOTAL",             str(total_vms),     "100%"],
    ]
    t_mode = Table(mode_data, colWidths=[5.0*cm, 2.5*cm, 2.0*cm])
    mode_style = _tbl_base([
        ("ALIGN",     (1, 0), (-1, -1), "CENTER"),
        ("FONTNAME",  (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND",(0, -1), (-1, -1), _HDR_LT),
    ])
    if warm_risky > 0:
        mode_style.add("BACKGROUND", (0, 2), (-1, 2), _YEL_P)
    if cold_required > 0:
        mode_style.add("BACKGROUND", (0, 3), (-1, 3), _RED_P)
    t_mode.setStyle(mode_style)

    # Risk distribution (right panel)
    risk_data = [
        ["Risk Category",     "Count", "%"],
        ["GREEN — low risk",  str(risk_green),  _pct(risk_green)],
        ["YELLOW — review",   str(risk_yellow), _pct(risk_yellow)],
        ["RED — high risk",   str(risk_red),    _pct(risk_red)],
        ["TOTAL",             str(total_vms),   "100%"],
    ]
    t_risk = Table(risk_data, colWidths=[5.5*cm, 2.5*cm, 2.0*cm])
    risk_style = _tbl_base([
        ("ALIGN",     (1, 0), (-1, -1), "CENTER"),
        ("FONTNAME",  (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND",(0, -1), (-1, -1), _HDR_LT),
    ])
    if risk_green > 0:
        risk_style.add("BACKGROUND", (0, 1), (-1, 1), _GRN_P)
    if risk_yellow > 0:
        risk_style.add("BACKGROUND", (0, 2), (-1, 2), _YEL_P)
    if risk_red > 0:
        risk_style.add("BACKGROUND", (0, 3), (-1, 3), _RED_P)
    t_risk.setStyle(risk_style)

    # Side-by-side outer table (infra: 7.5cm | gap 0.5 | mode: 9.5cm | gap 0.5 | risk: 10.0cm) = 28cm…
    # Keep to 26.0cm: infra(7.5) + gap(0.5) + mode(9.5) + gap(0.5) + risk(10.0) = 28 — too wide
    # Adjust: infra(7.5) + gap(0.3) + mode(8.5) + gap(0.3) + risk(10.0) = 26.6cm ✓
    outer_scope = Table([[t_infra, "", t_mode, "", t_risk]],
                        colWidths=[7.5*cm, 0.3*cm, 9.5*cm, 0.3*cm, 10.0*cm])
    outer_scope.setStyle(TableStyle([
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(outer_scope)
    story.append(Spacer(1, 0.3 * cm))

    # ── Section 2: Key Migration Estimates ────────────────────────────────────
    story.append(Paragraph("2  ·  Key Migration Estimates", s_h2))
    story.append(Paragraph(
        "All time figures are cumulative totals across all VMs. "
        "Because VMs migrate concurrently, elapsed calendar time is significantly less — see the Daily Schedule for wall-clock hours.",
        s_legend,
    ))
    kpi_data = [
        ["Metric", "Value", "What this means"],
        ["Total In-Use Data (GB)",      str(total_data_gb),
         "Actual bytes to transfer over the wire. Excludes unallocated disk space. "
         f"Storage utilisation: {round(total_data_gb/max(total_prov_gb,1)*100,0):.0f}% of {total_prov_gb} GB provisioned."],
        ["Est. Data-Copy Time (h)",     str(data_copy_h),
         f"Total in-use GB \u00f7 {bandwidth} Mbps bandwidth (converted to GB/h). "
         "Cumulative pipeline time — parallel copies overlap in calendar time."],
        ["Est. Tech Fix Time (h)",      str(fix_h_total),
         "Hands-on remediation after migration: Windows NIC/driver rename, disk-UUID updates, "
         "app validation. Weighted by OS-family fix rate. Configurable in Risk Config."],
        ["Est. Total Downtime (h)",     str(downtime_h_total),
         "Sum of per-VM business-impact periods. "
         "Warm VM downtime = cutover + fix only. Cold VM downtime = full copy + cutover + fix. "
         "Parallel execution reduces real calendar exposure."],
        ["Effective Bandwidth (Mbps)",  str(bandwidth),
         "Bottleneck throughput across source NIC, WAN/LAN link, migration agent, and storage I/O. "
         "Adjust in Project Settings to model real-world conditions."],
    ]
    t_kpi = Table(kpi_data, colWidths=[6.5*cm, 2.5*cm, 17.5*cm])
    kpi_style = _tbl_base([
        ("FONTNAME",   (0, 1), (0, -1), "Helvetica-Bold"),
        ("ALIGN",      (1, 0), (1, -1), "RIGHT"),
        ("FONTSIZE",   (2, 1), (2, -1), 8),
        ("TEXTCOLOR",  (2, 1), (2, -1), colors.HexColor("#374151")),
    ])
    kpi_style.add("BACKGROUND", (0, 4), (-1, 4), _RED_P)   # downtime row highlighted
    t_kpi.setStyle(kpi_style)
    story.append(t_kpi)
    story.append(Spacer(1, 0.2 * cm))

    # ── Section 3: Per-Tenant Migration Plan ──────────────────────────────────
    per_tenant = summary.get("per_tenant", [])
    if per_tenant:
        story.append(PageBreak())
        story.append(Paragraph("3  ·  Per-Tenant Migration Plan", s_h2))
        story.append(Paragraph(
            "Each row is one tenant (business unit or vCD organization). "
            "<b>Cold</b> — VMs migrated offline (full shutdown + copy). "
            "<b>Warm</b> — VMs migrated live with cutover-only migration downtime. "
            "<b>Fix (h)</b> — estimated post-migration remediation time for this tenant. "
            "<b>Downtime (h)</b> — total business-impact hours (cutover + fix for warm; copy + cutover + fix for cold). "
            "Rows shaded red have \u226550% cold VMs; yellow have any cold VMs.",
            s_legend,
        ))
        ten_hdr = [["Tenant", "Cohort", "VMs", "Cold", "Warm", "Data (GB)", "Fix (h)", "Downtime (h)"]]
        ten_rows = [[
            Paragraph(t.get("tenant_name", ""), s_cell),
            Paragraph(t.get("cohort_name", "") or "Unassigned", s_cell),
            t.get("vm_count", 0),
            t.get("cold_count", 0),
            t.get("warm_count", 0),
            str(round(float(t.get("data_gb") or 0), 1)),
            str(round(float(t.get("fix_hours") or 0), 2)),
            str(round(float(t.get("downtime_hours") or 0), 2)),
        ] for t in per_tenant]
        t_ten = Table(ten_hdr + ten_rows,
                      colWidths=[5.5*cm, 4.5*cm, 1.5*cm, 1.5*cm, 1.5*cm, 2.5*cm, 2.5*cm, 3.0*cm],
                      repeatRows=1)
        ten_style = _tbl_base([
            ("ALIGN",   (2, 0), (-1, -1), "RIGHT"),
            ("FONTNAME",(0, 1), (0, -1), "Helvetica-Bold"),
        ])
        for ri, r in enumerate(ten_rows, 1):
            cold_c = int(r[3] or 0)
            vm_c   = int(r[2] or 0)
            if cold_c > 0 and vm_c > 0:
                ten_style.add("BACKGROUND", (0, ri), (-1, ri),
                              _RED_P if cold_c / vm_c >= 0.5 else _YEL_P)
        t_ten.setStyle(ten_style)
        story.append(t_ten)
        story.append(Spacer(1, 0.3 * cm))

    # ── Section 4: Daily Migration Schedule ──────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph("4  ·  Daily Migration Schedule", s_h2))
    story.append(Paragraph(
        "One row per migration day. "
        "<b>Wall-h</b> — elapsed wall-clock hours from first VM start to last VM finish on that day. "
        "<b>Agent-h</b> — cumulative agent hours consumed (VMs may run in parallel so this can exceed Wall-h). "
        "<b>Cold / Warm</b> — VM counts by migration mode on that day. "
        "<b>Fix (h)</b> — total estimated post-migration tech fix hours for all VMs on that day. "
        "<b>Downtime (h)</b> — total business-impact hours on that day: cutover + fix for warm VMs; copy + cutover + fix for cold VMs. "
        "<b>Risk Red / Yel / Grn</b> — VM counts by risk category. "
        "<b>Cap!</b> — YES if this day\u2019s VM count exceeds the configured concurrent-slot capacity.",
        s_legend,
    ))
    day_hdr = [["Day", "Cohort", "Tenants", "VMs", "Data\n(GB)", "Wall-h", "Agent-h",
                "Cold", "Warm", "Fix\n(h)", "Downtime\n(h)", "Risk\nRed", "Risk\nYel", "Risk\nGrn", "Cap!"]]
    day_rows: list = []
    over_rows: list = []
    for day in summary.get("per_day") or []:
        over = bool(day.get("over_capacity"))
        ri = len(day_rows) + 1
        if over:
            over_rows.append(ri)
        day_rows.append([
            str(day.get("day", "")),
            Paragraph((day.get("cohort_name") or "")[:24], s_cell),
            day.get("tenant_count", 0),
            day.get("vm_count", 0),
            str(round(float(day.get("total_gb") or 0), 1)),
            str(round(float(day.get("wall_clock_hours") or 0), 2)),
            str(round(float(day.get("total_agent_hours") or 0), 2)),
            day.get("cold_count", 0),
            day.get("warm_count", 0),
            str(round(float(day.get("fix_hours") or 0), 2)),
            str(round(float(day.get("downtime_hours") or 0), 2)),
            day.get("risk_red", 0),
            day.get("risk_yellow", 0),
            day.get("risk_green", 0),
            "YES" if over else "",
        ])
    # Landscape A4 usable ~26.5cm; 15 cols total
    cw_day = [1.0*cm, 5.0*cm, 1.3*cm, 1.2*cm, 1.8*cm, 1.5*cm, 1.8*cm,
              1.2*cm, 1.2*cm, 1.8*cm, 2.2*cm, 1.2*cm, 1.2*cm, 1.2*cm, 1.0*cm]
    t_day = Table(day_hdr + day_rows, colWidths=cw_day, repeatRows=1)
    day_style = _tbl_base([("ALIGN", (2, 0), (-1, -1), "CENTER")])
    for ri in over_rows:
        day_style.add("BACKGROUND", (0, ri), (-1, ri), _RED_P)
        day_style.add("TEXTCOLOR",  (12, ri), (12, ri), colors.HexColor("#DC2626"))
    t_day.setStyle(day_style)
    story.append(t_day)
    story.append(Spacer(1, 0.3 * cm))

    # ── Section 5: Cohort Breakdown ───────────────────────────────────────────
    cohorts = summary.get("per_cohort") or []
    if cohorts:
        story.append(Paragraph("5  ·  Cohort Breakdown", s_h2))
        story.append(Paragraph(
            "A cohort is a logical grouping of tenants migrated in the same wave. "
            "<b>Data-Copy (h)</b> — pipeline time to copy all cohort VMs. "
            "<b>Fix (h)</b> — post-migration remediation hours. "
            "<b>Downtime (h)</b> — total per-VM business-impact hours across this cohort.",
            s_legend,
        ))
        coh_hdr = [["Cohort", "VMs", "Data (GB)", "Data-Copy (h)", "Fix (h)", "Downtime (h)"]]
        coh_rows = [[
            Paragraph(c.get("cohort_name", ""), s_cell),
            c.get("vm_count", 0),
            str(round(float(c.get("data_gb") or 0), 1)),
            str(round(float(c.get("data_copy_hours") or 0), 2)),
            str(round(float(c.get("fix_hours") or 0), 2)),
            str(round(float(c.get("downtime_hours") or 0), 2)),
        ] for c in cohorts]
        t_coh = Table(coh_hdr + coh_rows, colWidths=[8.0*cm, 2.5*cm, 3.5*cm, 4.0*cm, 3.5*cm, 4.0*cm])
        t_coh.setStyle(_tbl_base([
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("ALIGN",    (1, 0), (-1, -1), "RIGHT"),
        ]))
        story.append(t_coh)
        story.append(Spacer(1, 0.3 * cm))

    # ── Section 6: OS Family Breakdown ───────────────────────────────────────
    os_raw = summary.get("per_os_breakdown") or {}
    os_items_pdf = list(os_raw.items() if isinstance(os_raw, dict) else
                        [(e.get("os_family", ""), e) for e in os_raw])
    if os_items_pdf:
        story.append(Paragraph("6  ·  OS Family Breakdown & Fix Rates", s_h2))
        story.append(Paragraph(
            "<b>Fix Rate</b> — percentage of VMs in this OS family expected to require post-migration remediation. "
            "Configurable under Risk Config. "
            "<b>Fix (h)</b> — total estimated remediation hours for all affected VMs in this OS family.",
            s_legend,
        ))
        os_notes = {
            "windows": "NIC adapter rename (vmxnet3\u2192virtio), disk drive-letter reassignment, possible service re-activation",
            "linux":   "Minimal changes: usually only fstab UUID updates and network interface name rebinding",
        }
        os_hdr = [["OS Family", "VM Count", "Fix Rate", "Fix (h)", "Typical remediation steps"]]
        os_rows = [[
            Paragraph(fam.title(), s_cell),
            entry.get("vm_count", 0),
            f"{float(entry.get('fix_rate', 0)):.0%}",
            str(round(float(entry.get("fix_hours") or 0), 2)),
            Paragraph(os_notes.get(fam.lower(), "OS-specific remediation may apply"), s_cell),
        ] for fam, entry in os_items_pdf]
        t_os = Table(os_hdr + os_rows, colWidths=[3.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 15.5*cm])
        t_os.setStyle(_tbl_base([
            ("ALIGN",    (1, 0), (3, -1), "CENTER"),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ]))
        story.append(t_os)
        story.append(Spacer(1, 0.3 * cm))

    # ── Section 7: Methodology (summary) ─────────────────────────────────────
    methodology = summary.get("methodology") or {}
    if methodology:
        story.append(PageBreak())
        story.append(Paragraph("7  ·  Calculation Methodology", s_h2))
        s_method = ParagraphStyle("SM", parent=styles["Normal"], fontSize=9, spaceAfter=6, leading=13)
        labels = {
            "data_copy": "Data-Copy Time",
            "fix_time":  "Tech Fix Time",
            "downtime":  "Downtime Estimate",
            "fix_rates": "OS Fix Rates",
        }
        for key, label in labels.items():
            if key in methodology:
                story.append(Paragraph(f"<b>{label}:</b> {methodology[key]}", s_method))
        story.append(Spacer(1, 0.3 * cm))

    # ── Appendix: Detailed Methodology Reference ──────────────────────────────
    story.append(PageBreak())
    s_h2_m  = ParagraphStyle("SH2m", parent=styles["Heading2"], fontSize=11, spaceAfter=4,
                              textColor=colors.HexColor("#1E40AF"), spaceBefore=12)
    s_cap_m = ParagraphStyle("SCm",  parent=styles["Normal"],  fontSize=8,
                              textColor=colors.grey, spaceAfter=6)
    story.append(Paragraph("Appendix: Detailed Methodology Reference", s_h2_m))
    story += _pdf_methodology_section(s_h2_m, styles["Normal"], s_cap_m, _HDR, _GREY_P, _HDR_LT, colors)

    # ── Footer ────────────────────────────────────────────────────────────────
    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.grey)
        txt = f"Platform9 Migration Summary  \u00b7  {project_name}  \u00b7  Page {doc.page}"
        canvas.drawCentredString(landscape(A4)[0] / 2, 0.7 * cm, txt)
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# Shared PDF helper — Methodology section
# ══════════════════════════════════════════════════════════════════════════════

def _pdf_methodology_section(s_h2, s_body, s_caption, _HDR, _GREY_P, _HDR_LT, colors_mod) -> list:
    """Return a list of ReportLab flowables explaining all calculation methodologies."""
    story = []
    story.append(Paragraph("Calculation Methodology", s_h2))
    story.append(Paragraph(
        "This section documents every formula used to produce the estimates in this report.",
        s_caption,
    ))

    _W1, _W2 = 5.5 * cm, 13.5 * cm  # label / detail column widths

    sections = [
        ("1. Warm vs Cold Classification", [
            ("Warm Eligible",
             "OS does NOT appear in the OS Cold Required list AND VM is powered ON. "
             "Live data-copy runs while the VM is active. "
             "Downtime = cutover window + tech fix time (no downtime during Phase 1 data copy)."),
            ("Cold Required",
             "OS matches an entry in the Risk Config → Cold Migration Rules list, OR the VM is "
             "poweredOff / suspended. VM is shut down before copy. "
             "Downtime = full copy duration + cutover window + tech fix time."),
            ("Warm Risky",
             "Warm-eligible but risk score meets/exceeds Yellow or Red threshold. "
             "Same downtime model as Warm Eligible: cutover + fix time."),
        ]),
        ("2. Bandwidth Model", [
            ("Formula",
             "effective_mbps = min(source_nic_mbps, link_mbps, agent_mbps, storage_mbps)"),
            ("Example",
             "Source=4,800 · Link=4,000 · Agent=6,000 · Storage=5,500 Mbps → bottleneck = link at 4,000 Mbps"),
        ]),
        ("3. Data-Copy Time", [
            ("Formula",
             "data_copy_hours = total_in_use_gb ÷ (bottleneck_mbps ÷ 8 × 3600 ÷ 1024)"),
            ("Example",
             "172,617 GB at 4,000 Mbps → 1,757.8 GB/h → 98.2 h total data-copy time"),
        ]),
        ("4. Tech Fix Scoring", [
            ("Factor weights",
             "Windows: +20 min | Extra volume ea: +15 | Extra NIC ea: +10 | Cold mode: +15 | "
             "Risk YELLOW: +15 | Risk RED: +25 | Snapshots: +10 | Cross-tenant dep: +15 | Unknown OS: +5"),
            ("OS Fix Rates",
             "Windows 50% · Linux 20% · Other 40% (global override replaces all when set)"),
            ("Formula",
             "raw = Σ factor_weights; fix_minutes = raw × fix_rate"),
            ("Example",
             "Windows, cold, 2 extra vols, 1 extra NIC, YELLOW: raw=20+30+10+15+15=90; fix=90×0.50=45 min"),
        ]),
        ("5. Downtime Estimate", [
            ("Warm VMs",  "downtime = cutover_minutes + fix_minutes  (copy is live, no user impact)"),
            ("Cold VMs",  "downtime = data_copy_hours×60 + cutover_minutes + fix_minutes  (VM offline during copy)"),
            ("Total",     "Sum of all VM downtime hours. Concurrent across tenants; wall-clock is much shorter."),
        ]),
        ("6. Daily Schedule", [
            ("Algorithm",
             "VMs ordered Cohort→Priority→disk size (largest first). Assigned to current day until "
             "wall-clock time (max VM hours) exceeds working_hours_per_day; then new day starts."),
            ("Concurrency",
             "concurrent_slots = total agent slots across all vJailbreak agents. Each VM uses 1 slot."),
        ]),
        ("7. Risk Categories", [
            ("GREEN",  "Risk score 0–29 → standard migration"),
            ("YELLOW", "Risk score 30–59 → pre-migration review required"),
            ("RED",    "Risk score 60+ → architecture review + rollback plan required"),
        ]),
    ]

    for sec_title, entries in sections:
        story.append(Paragraph(f"<b>{sec_title}</b>", s_body))
        tbl_data = [["Topic", "Detail"]]
        for label, detail in entries:
            tbl_data.append([label, detail])
        t = Table(tbl_data, colWidths=[_W1, _W2])
        style = [
            ("BACKGROUND",    (0, 0), (-1, 0), _HDR),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors_mod.white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("FONTNAME",      (0, 1), (0, -1), "Helvetica-Bold"),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors_mod.white, _GREY_P]),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors_mod.HexColor("#D1D5DB")),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("WORDWRAP",      (0, 0), (-1, -1), 1),
        ]
        t.setStyle(TableStyle(style))
        story.append(t)
        story.append(Spacer(1, 0.25 * cm))

    story.append(Paragraph(
        "All default values are configurable in Risk Config and Migration Summary → Settings.",
        s_caption,
    ))
    return story


# ══════════════════════════════════════════════════════════════════════════════
# Gap Analysis Action Report  (Excel + PDF)
# ══════════════════════════════════════════════════════════════════════════════

_EFFORT_MAP = {
    "missing_flavor":  "Medium — create matching PCD flavor",
    "missing_network": "High   — provision VLAN / network segment in PCD",
    "missing_image":   "Medium — upload/register OS image in Glance",
    "unmapped_tenant": "Low    — set Target Domain/Project in Tenant Scoping",
}

def _gap_effort(gap_type: str) -> str:
    for key, val in _EFFORT_MAP.items():
        if key in (gap_type or ""):
            return val
    return "Low — review and mark resolved"


def _gap_action_steps(gap: Dict[str, Any]) -> str:
    """Return a short set of recommended steps for a gap."""
    gt = (gap.get("gap_type") or "").lower()
    res = gap.get("resolution", "")
    details = gap.get("details") or {}
    if "flavor" in gt:
        vcpu = details.get("required_vcpu") or details.get("required vcpu", "?")
        ram  = details.get("required_ram") or details.get("ram", "?")
        return f"1. Open PCD → Admin → Flavors\n2. Create flavor: vCPU={vcpu}, RAM={ram} GB\n3. Mark gap Resolved"
    if "network" in gt:
        name = details.get("network_name") or gap.get("resource_name", "?")
        return f"1. Verify VLAN/segment '{name}' exists in your SDN\n2. Create Neutron network in PCD with matching name/VLAN\n3. Assign to relevant project(s)\n4. Mark gap Resolved"
    if "image" in gt:
        name = gap.get("resource_name", "?")
        return f"1. Download/prepare OS image for '{name}'\n2. Upload to Glance: openstack image create ...\n3. Set correct properties (disk_format, container_format)\n4. Mark gap Resolved"
    if "tenant" in gt or "unmapped" in gt:
        tenant = gap.get("tenant_name") or gap.get("resource_name", "?")
        return f"1. Open Migration Planner → Tenants → Target Mapping\n2. Set Target Domain + Project for '{tenant}'\n3. Mark gap Resolved"
    return res or "Review gap details and take corrective action, then mark Resolved."


# ── Excel ────────────────────────────────────────────────────────────────────

def generate_gaps_excel_report(
    gaps: List[Dict[str, Any]],
    project_name: str,
    readiness_score: float | None = None,
    quota_summary: Dict | None = None,
    sizing: Dict | None = None,
) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _gaps_sheet_summary(wb, gaps, project_name, readiness_score, quota_summary, sizing)
    _gaps_sheet_action_items(wb, gaps)
    _gaps_sheet_all_gaps(wb, gaps)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _gaps_sheet_summary(wb, gaps, project_name, score, quota, sizing):
    ws = wb.create_sheet("Executive Summary")
    _set_col_widths(ws, [28, 22, 22, 22])

    # Title
    ws.merge_cells("A1:D1")
    t = ws["A1"]; t.value = f"PCD Readiness Report — {project_name}"
    t.font = Font(bold=True, color=_WHITE, size=14, name="Calibri")
    t.fill = _hdr_fill(); t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ts = ws["A2"]; ts.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ts.font = Font(size=9, color="6B7280", name="Calibri")
    ts.alignment = Alignment(horizontal="center")

    row = 4
    # Score
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = "Readiness Score"
    ws[f"A{row}"].font = Font(bold=True, size=11, name="Calibri")
    ws[f"C{row}"] = f"{score:.1f} / 100" if score is not None else "Not analysed"
    score_val = float(score or 0)
    score_fill = "DCFCE7" if score_val >= 80 else ("FFFBEB" if score_val >= 50 else "FEE2E2")
    ws[f"C{row}"].fill = PatternFill("solid", fgColor=score_fill)
    ws[f"C{row}"].font = Font(bold=True, size=14, name="Calibri")
    row += 2

    # Gap counts by severity
    criticals  = [g for g in gaps if g.get("severity") == "critical"  and not g.get("resolved")]
    warnings   = [g for g in gaps if g.get("severity") == "warning"   and not g.get("resolved")]
    infos      = [g for g in gaps if g.get("severity") == "info"      and not g.get("resolved")]
    resolved_n = sum(1 for g in gaps if g.get("resolved"))

    for label, val, fill in [
        ("Open Critical Gaps", len(criticals), _RED),
        ("Open Warning Gaps",  len(warnings),  _YELLOW[1:] if len(_YELLOW) == 6 else "FEF9C3"),
        ("Open Info Gaps",     len(infos),     _GREY),
        ("Resolved Gaps",      resolved_n,     _GREEN),
    ]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri", size=10)
        ws[f"B{row}"] = val
        ws[f"B{row}"].fill = PatternFill("solid", fgColor=fill)
        ws[f"B{row}"].font = Font(bold=True, size=12, name="Calibri")
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        row += 1

    row += 1
    # Gap type breakdown
    ws[f"A{row}"] = "Gap Type Breakdown"
    ws[f"A{row}"].font = Font(bold=True, size=11, name="Calibri", color=_WHITE)
    ws[f"A{row}"].fill = _hdr_fill()
    ws[f"B{row}"] = "Open"
    ws[f"B{row}"].font = Font(bold=True, size=11, name="Calibri", color=_WHITE)
    ws[f"B{row}"].fill = _hdr_fill()
    ws[f"B{row}"].alignment = Alignment(horizontal="center")
    row += 1
    from collections import Counter
    type_counts = Counter(g.get("gap_type", "unknown") for g in gaps if not g.get("resolved"))
    for gtype, cnt in sorted(type_counts.items(), key=lambda x: -x[1]):
        ws[f"A{row}"] = gtype.replace("_", " ").title()
        ws[f"B{row}"] = cnt
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        row += 1

    if quota and (quota.get("vcpu") or quota.get("ram_gb")):
        row += 1
        ws[f"A{row}"] = "Migration Quota Requirements"
        ws[f"A{row}"].font = Font(bold=True, size=11, name="Calibri", color=_WHITE)
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"].fill = _hdr_fill()
        row += 1
        for label, val in [
            ("vCPU Required", quota.get("vcpu", "—")),
            ("RAM Required (GB)", quota.get("ram_gb", "—")),
            ("Disk Required (TB)", quota.get("disk_tb", "—")),
        ]:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = val
            row += 1


def _gaps_sheet_action_items(wb, gaps):
    ws = wb.create_sheet("Action Items (Unresolved)")
    _set_col_widths(ws, [22, 20, 16, 12, 12, 35, 28])
    cols = ["Gap Type", "Resource", "Tenant", "Severity", "Effort", "Action Steps", "Resolution"]
    _write_table_header(ws, 1, cols)
    ws.freeze_panes = "A2"

    unresolved = [g for g in gaps if not g.get("resolved")]
    unresolved.sort(key=lambda g: (0 if g.get("severity") == "critical" else 1
                                    if g.get("severity") == "warning" else 2))

    sev_fills = {"critical": _RED, "warning": "FEF9C3", "info": _GREY}
    for i, g in enumerate(unresolved, 1):
        sev  = g.get("severity", "info")
        fill = sev_fills.get(sev, _GREY)
        _write_row(ws, i + 1, [
            (g.get("gap_type") or "").replace("_", " ").title(),
            g.get("resource_name", ""),
            g.get("tenant_name") or "—",
            sev,
            _gap_effort(g.get("gap_type", "")),
            _gap_action_steps(g),
            g.get("resolution", ""),
        ], fill_hex=fill if sev == "critical" else None, alt=(i % 2 == 0))
        ws.row_dimensions[i + 1].height = 48
    ws.auto_filter.ref = f"A1:G{len(unresolved) + 1}"


def _gaps_sheet_all_gaps(wb, gaps):
    ws = wb.create_sheet("All Gaps")
    _set_col_widths(ws, [22, 20, 16, 12, 10, 35])
    cols = ["Gap Type", "Resource", "Tenant", "Severity", "Status", "Resolution"]
    _write_table_header(ws, 1, cols)
    ws.freeze_panes = "A2"
    for i, g in enumerate(gaps, 1):
        status = "✓ Resolved" if g.get("resolved") else "Open"
        _write_row(ws, i + 1, [
            (g.get("gap_type") or "").replace("_", " ").title(),
            g.get("resource_name", ""),
            g.get("tenant_name") or "—",
            g.get("severity", "info"),
            status,
            g.get("resolution", ""),
        ], alt=(i % 2 == 0))
    ws.auto_filter.ref = f"A1:F{len(gaps) + 1}"


# ── PDF ──────────────────────────────────────────────────────────────────────

def generate_gaps_pdf_report(
    gaps: List[Dict[str, Any]],
    project_name: str,
    readiness_score: float | None = None,
    quota_summary: Dict | None = None,
    sizing: Dict | None = None,
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.8*cm)

    styles = getSampleStyleSheet()
    _HDR  = colors.HexColor("#1D4ED8")
    _GREY_P = colors.HexColor("#F3F4F6")
    _RED_P  = colors.HexColor("#FEE2E2")
    _YEL_P  = colors.HexColor("#FEF9C3")

    def _h1(text):
        return Paragraph(f"<font color='#1D4ED8'><b>{text}</b></font>",
                         ParagraphStyle("h1", parent=styles["Heading1"], fontSize=14,
                                        spaceAfter=6, textColor=colors.HexColor(_HDR.hexval() if hasattr(_HDR, 'hexval') else "#1D4ED8")))

    def _h2(text):
        return Paragraph(f"<b>{text}</b>",
                         ParagraphStyle("h2", parent=styles["Heading2"], fontSize=11, spaceAfter=4))

    def _body(text):
        return Paragraph(text, ParagraphStyle("body", parent=styles["Normal"], fontSize=8, leading=11))

    story = []

    # Title
    story.append(Paragraph(
        f"<font size='16' color='#1D4ED8'><b>PCD Readiness Action Report</b></font>",
        ParagraphStyle("title", parent=styles["Title"], alignment=TA_CENTER, spaceAfter=4)
    ))
    story.append(Paragraph(
        f"Project: <b>{project_name}</b>  ·  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        ParagraphStyle("sub", parent=styles["Normal"], alignment=TA_CENTER, fontSize=9,
                       textColor=colors.HexColor("#6B7280"), spaceAfter=12)
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E5E7EB")))
    story.append(Spacer(1, 0.3*cm))

    # Score summary row
    score_val = float(readiness_score or 0)
    score_color_hex = "#16a34a" if score_val >= 80 else ("#d97706" if score_val >= 50 else "#dc2626")
    criticals  = sum(1 for g in gaps if g.get("severity") == "critical"  and not g.get("resolved"))
    warnings   = sum(1 for g in gaps if g.get("severity") == "warning"   and not g.get("resolved"))
    resolved_n = sum(1 for g in gaps if g.get("resolved"))

    summary_data = [
        ["Readiness Score", "Critical Gaps", "Warning Gaps", "Resolved"],
        [
            f"{score_val:.1f} / 100" if readiness_score is not None else "N/A",
            str(criticals), str(warnings), str(resolved_n),
        ]
    ]
    t_sum = Table(summary_data, colWidths=[7*cm, 5*cm, 5*cm, 5*cm])
    t_sum.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _HDR),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTSIZE",   (0, 0), (-1, -1), 10),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("BACKGROUND", (1, 1), (1, 1), colors.HexColor("#FEE2E2") if criticals else colors.HexColor("#DCFCE7")),
        ("BACKGROUND", (2, 1), (2, 1), colors.HexColor("#FEF9C3") if warnings else colors.HexColor("#DCFCE7")),
        ("BACKGROUND", (3, 1), (3, 1), colors.HexColor("#DCFCE7")),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t_sum)
    story.append(Spacer(1, 0.5*cm))

    # Action Items table (unresolved only)
    story.append(_h2("Action Items — Unresolved Gaps"))
    story.append(Spacer(1, 0.15*cm))

    unresolved = [g for g in gaps if not g.get("resolved")]
    unresolved.sort(key=lambda g: (0 if g.get("severity") == "critical" else
                                    1 if g.get("severity") == "warning" else 2))

    if not unresolved:
        story.append(_body("✅ No unresolved gaps. All gaps have been marked resolved."))
    else:
        ai_header = [["#", "Gap Type", "Resource / Tenant", "Sev.", "Action Steps"]]
        ai_rows = []
        for idx, g in enumerate(unresolved, 1):
            sev = g.get("severity", "info")
            ai_rows.append([
                str(idx),
                Paragraph((g.get("gap_type") or "").replace("_", " ").title(),
                           ParagraphStyle("s", fontSize=7, leading=9)),
                Paragraph(f"{g.get('resource_name', '')}\\n"
                           f"<font color='#6B7280'>{g.get('tenant_name') or ''}</font>",
                           ParagraphStyle("s", fontSize=7, leading=9)),
                Paragraph(f"<font color='{'#dc2626' if sev=='critical' else '#d97706' if sev=='warning' else '#374151'}'><b>{sev}</b></font>",
                           ParagraphStyle("s", fontSize=7, leading=9)),
                Paragraph(_gap_action_steps(g).replace("\n", "<br/>"),
                           ParagraphStyle("s", fontSize=7, leading=9)),
            ])
        ai_data = ai_header + ai_rows
        cw = [0.8*cm, 4*cm, 5*cm, 1.8*cm, 16*cm]
        t_ai = Table(ai_data, colWidths=cw, repeatRows=1)
        ai_style = [
            ("BACKGROUND", (0, 0), (-1, 0), _HDR),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, 0), 8),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
            ("VALIGN",     (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for i, g in enumerate(unresolved, 1):
            sev = g.get("severity", "info")
            if sev == "critical":
                ai_style.append(("BACKGROUND", (0, i), (3, i), colors.HexColor("#FEE2E2")))
            elif sev == "warning":
                ai_style.append(("BACKGROUND", (0, i), (3, i), colors.HexColor("#FEF9C3")))
            elif i % 2 == 0:
                ai_style.append(("ROWBACKGROUNDS", (0, i), (-1, i), [_GREY_P]))
        t_ai.setStyle(TableStyle(ai_style))
        story.append(t_ai)

    story.append(Spacer(1, 0.5*cm))

    # All Gaps summary table
    story.append(_h2(f"All Gaps ({len(gaps)} total, {resolved_n} resolved)"))
    story.append(Spacer(1, 0.15*cm))

    if not gaps:
        story.append(_body("No gaps recorded. Run 'Run Gap Analysis' in the PCD Readiness tab."))
    else:
        all_header = [["Gap Type", "Resource", "Tenant", "Severity", "Status", "Resolution"]]
        all_rows = [[
            Paragraph((g.get("gap_type") or "").replace("_", " ").title(),
                       ParagraphStyle("s", fontSize=7)),
            Paragraph(g.get("resource_name", ""), ParagraphStyle("s", fontSize=7)),
            Paragraph(g.get("tenant_name") or "—", ParagraphStyle("s", fontSize=7)),
            Paragraph(g.get("severity", ""), ParagraphStyle("s", fontSize=7)),
            Paragraph("✓ Resolved" if g.get("resolved") else "Open", ParagraphStyle("s", fontSize=7)),
            Paragraph(g.get("resolution", ""), ParagraphStyle("s", fontSize=7, leading=9)),
        ] for g in gaps]
        all_data = all_header + all_rows
        cw2 = [4*cm, 4.5*cm, 4.5*cm, 2.3*cm, 2.3*cm, 9.4*cm]
        t_all = Table(all_data, colWidths=cw2, repeatRows=1)
        t_all.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _HDR),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, 0), 8),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _GREY_P]),
            ("VALIGN",     (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE",   (0, 1), (-1, -1), 7),
            ("TOPPADDING",    (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(t_all)

    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.grey)
        txt = (f"Platform9 Migration Planner  ·  PCD Readiness  ·  {project_name}"
               f"  ·  Page {doc.page}")
        canvas.drawCentredString(landscape(A4)[0] / 2, 0.8*cm, txt)
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# Tenant Handoff Sheet PDF
# ══════════════════════════════════════════════════════════════════════════════

def generate_handoff_pdf(
    project_name: str,
    auth_url: str,
    tenants: List[Dict[str, Any]],
    support_text: str = "",
) -> bytes:
    """
    Generate a per-tenant migration handoff PDF (sealed delivery document).

    Each tenant section shows: PCD domain / project, auth endpoint, networks
    (with CIDR / VLAN), and user accounts with plaintext temporary passwords.

    ``tenants`` is a list of dicts; each dict contains:
        tenant_name, target_domain_name, target_project_name, target_display_name,
        pcd_project_id, cohort_name,
        networks: [{source_network, pcd_network_id, vlan_id, cidr, gateway_ip}]
        users:    [{username, email, role, temp_password, is_existing_user, user_type}]
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2.2 * cm,
    )
    styles = getSampleStyleSheet()
    _HDR = colors.HexColor("#1D4ED8")
    _GREY_LT = colors.HexColor("#F9FAFB")
    _BLUE_LT2 = colors.HexColor("#EFF6FF")

    def _h2(text: str):
        return Paragraph(
            f"<b>{text}</b>",
            ParagraphStyle("h2hd", parent=styles["Heading2"], fontSize=9,
                           spaceAfter=3, spaceBefore=6, textColor=colors.HexColor("#1E40AF")),
        )

    def _body(text: str, small: bool = False):
        return Paragraph(
            text,
            ParagraphStyle("bodyh", parent=styles["Normal"],
                           fontSize=7 if small else 8, leading=11),
        )

    story: list = []

    # ── Cover title ───────────────────────────────────────────────────────────
    story.append(Paragraph(
        "<font size='18' color='#1D4ED8'><b>Migration Handoff Sheet</b></font>",
        ParagraphStyle("title", parent=styles["Title"], alignment=TA_CENTER, spaceAfter=4),
    ))
    story.append(Paragraph(
        f"Project: <b>{project_name}</b>  ·  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        ParagraphStyle("sub", parent=styles["Normal"], alignment=TA_CENTER, fontSize=9,
                       textColor=colors.HexColor("#6B7280"), spaceAfter=4),
    ))
    story.append(Paragraph(
        f"PCD Endpoint: <font face='Courier' size='8'>{auth_url or 'Not configured'}</font>",
        ParagraphStyle("sub2", parent=styles["Normal"], alignment=TA_CENTER, fontSize=8,
                       textColor=colors.HexColor("#374151"), spaceAfter=6),
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E5E7EB")))
    story.append(Spacer(1, 0.2 * cm))

    # Confidentiality notice
    notice_data = [[Paragraph(
        "⚠️  <b>CONFIDENTIAL — CONTAINS PLAINTEXT PASSWORDS.</b>  "
        "This document is a sealed handoff packet. Distribute securely and only to the named "
        "tenant owner. Service account passwords are for the migration engine only — do not share.",
        ParagraphStyle("notice", parent=styles["Normal"], fontSize=8, leading=11),
    )]]
    notice_t = Table(notice_data, colWidths=[17 * cm])
    notice_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FEF9C3")),
        ("BOX",        (0, 0), (-1, -1), 0.75, colors.HexColor("#D97706")),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
    ]))
    story.append(notice_t)
    story.append(Spacer(1, 0.3 * cm))

    if support_text:
        story.append(_body(support_text))
        story.append(Spacer(1, 0.25 * cm))

    # ── Per-tenant sections ───────────────────────────────────────────────────
    for tenant in tenants:
        story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#1D4ED8")))
        story.append(Spacer(1, 0.1 * cm))

        tenant_name   = tenant.get("tenant_name", "—")
        domain        = tenant.get("target_domain_name", "—")
        project       = tenant.get("target_project_name", "—")
        display       = tenant.get("target_display_name") or project
        pcd_id        = tenant.get("pcd_project_id") or "⚠️ Not provisioned"
        cohort        = tenant.get("cohort_name", "")

        header_text = f"<b><font color='#1D4ED8' size='11'>{tenant_name}</font></b>"
        if cohort:
            header_text += f"  <font color='#6B7280' size='8'>  {cohort}</font>"
        story.append(Paragraph(header_text,
                               ParagraphStyle("th2", parent=styles["Normal"], spaceAfter=4)))

        # Project identity table
        proj_data = [
            ["Source Tenant", tenant_name],
            ["PCD Domain",    domain],
            ["PCD Project",   project],
            ["Display Name",  display],
            ["PCD Project ID", pcd_id],
            ["Auth URL",      auth_url or "—"],
        ]
        pt = Table(proj_data, colWidths=[3.8 * cm, 13.2 * cm])
        pt.setStyle(TableStyle([
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("FONTNAME",      (0, 0), (0, -1), "Helvetica-Bold"),
            ("TEXTCOLOR",     (0, 0), (0, -1), colors.HexColor("#374151")),
            ("TEXTCOLOR",     (1, 0), (1, -1), colors.HexColor("#111827")),
            ("BACKGROUND",    (0, 0), (-1, -1), _GREY_LT),
            ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
        ]))
        story.append(pt)
        story.append(Spacer(1, 0.15 * cm))

        # Networks
        networks = tenant.get("networks", [])
        if networks:
            story.append(_h2("Networks"))
            net_hdr = [["Source Network", "PCD Network ID", "VLAN", "CIDR", "Gateway"]]
            net_rows = [[
                Paragraph(n.get("source_network_name", "—"),
                          ParagraphStyle("nc", fontSize=7)),
                Paragraph(n.get("target_network_id") or "⚠️ not provisioned",
                          ParagraphStyle("nc", fontSize=7, fontName="Courier")),
                str(n.get("vlan_id", "—")) if n.get("vlan_id") else "—",
                n.get("cidr", "—") or "—",
                n.get("gateway_ip", "—") or "—",
            ] for n in networks]
            nt = Table(net_hdr + net_rows,
                       colWidths=[3.8 * cm, 4.5 * cm, 1.4 * cm, 3.0 * cm, 2.8 * cm])
            nt.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, 0), _HDR),
                ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
                ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",      (0, 0), (-1, -1), 8),
                ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white, colors.HexColor("#F3F4F6")]),
                ("TOPPADDING",    (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ]))
            story.append(nt)
            story.append(Spacer(1, 0.15 * cm))

        # Users / Credentials
        users = tenant.get("users", [])
        if users:
            story.append(_h2("User Accounts"))
            u_hdr = [["Username", "Email", "Role", "Type", "Temp Password", "Notes"]]
            u_rows: list = []
            u_svc_rows: List[int] = []
            for idx, u in enumerate(users, 1):
                is_svc = u.get("user_type") == "service_account"
                if is_svc:
                    u_svc_rows.append(idx)
                    notes = "Migration engine only — do not share with tenant"
                else:
                    notes = "Tenant owner — change password on first login"
                pw = (u.get("temp_password") or
                      ("(existing Keystone user)" if u.get("is_existing_user") else "—"))
                u_rows.append([
                    Paragraph(u.get("username", "—"),
                              ParagraphStyle("uc", fontSize=7, fontName="Courier")),
                    Paragraph(u.get("email") or "—", ParagraphStyle("uc", fontSize=7)),
                    u.get("role", "admin"),
                    "svc-acct" if is_svc else "owner",
                    Paragraph(pw, ParagraphStyle("uc", fontSize=7, fontName="Courier")),
                    Paragraph(notes, ParagraphStyle("uc", fontSize=7)),
                ])
            ut = Table(u_hdr + u_rows,
                       colWidths=[3.5 * cm, 3.3 * cm, 1.4 * cm, 1.6 * cm, 3.4 * cm, 3.8 * cm])
            u_style: list = [
                ("BACKGROUND",    (0, 0), (-1, 0), _HDR),
                ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
                ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",      (0, 0), (-1, -1), 8),
                ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white, colors.HexColor("#F3F4F6")]),
                ("TOPPADDING",    (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ]
            for row_i in u_svc_rows:
                u_style.append(("BACKGROUND", (0, row_i), (-1, row_i), _BLUE_LT2))
            ut.setStyle(TableStyle(u_style))
            story.append(ut)

        story.append(Spacer(1, 0.35 * cm))

    # ── Footer ────────────────────────────────────────────────────────────────
    def _footer_h(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.grey)
        txt = (f"CONFIDENTIAL  ·  Platform9 Migration Handoff  ·  {project_name}"
               f"  ·  Page {doc.page}")
        canvas.drawCentredString(A4[0] / 2, 0.7 * cm, txt)
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer_h, onLaterPages=_footer_h)
    return buf.getvalue()
