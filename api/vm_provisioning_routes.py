"""
VM Provisioning Runbook (Runbook 2) — v1.39.0
==============================================
Multi-step workflow to provision VMs into tenant projects:
  Form or Excel → Validate → Quota-check → Approval → Dry-run → Execute

All VMs boot from a Cinder volume (disk==0 flavors only).
cloud-init / cloudbase-init payload is auto-generated from OS username + password.
"""

import base64
import hashlib
import io
import ipaddress
import json
import logging
import os
import re
import secrets
import string
import threading
import time
import traceback
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

import psycopg2
from psycopg2.extras import RealDictCursor, Json
from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, field_validator

from auth import require_permission, get_current_user, ldap_auth, get_effective_region_filter
from pf9_control import get_client
from cluster_registry import get_registry
from smtp_helper import send_email as smtp_send_email
from db_pool import get_connection, get_pool

logger = logging.getLogger("vm_provisioning")

router = APIRouter(prefix="/api/vm-provisioning", tags=["vm-provisioning"])

# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Encryption constants — Fernet key for os_password at rest
# ---------------------------------------------------------------------------
_VM_PROVISION_KEY     = "vm_provision_key"
_VM_PROVISION_KEY_ENV = "VM_PROVISION_KEY"


def _decrypt_vm_password(stored: str) -> str:
    """Decrypt a stored VM OS password.

    New rows (v1.83.25+) carry a 'fernet:' prefix and are Fernet-encrypted.
    Legacy plaintext rows (created before v1.83.25) are returned unchanged so
    old batches that have not yet been executed continue to work.
    """
    if not stored:
        return ""
    if stored.startswith("fernet:"):
        from crypto_helper import fernet_decrypt as _fd
        return _fd(stored, secret_name=_VM_PROVISION_KEY,
                   env_var=_VM_PROVISION_KEY_ENV,
                   context="vm_provisioning_vms.os_password")
    return stored  # backward-compat: pre-encryption plaintext row


def _encrypt_vm_password(plaintext: str) -> str:
    """Encrypt a plaintext VM OS password for storage.

    Wraps fernet_encrypt and raises HTTPException(500) with a clear message
    if VM_PROVISION_KEY is not configured, rather than propagating a bare
    RuntimeError to the caller.
    """
    from crypto_helper import fernet_encrypt as _fe
    return _fe(plaintext, secret_name=_VM_PROVISION_KEY, env_var=_VM_PROVISION_KEY_ENV)


def _log_activity(conn, resource_type: str, resource_id: str, action: str,
                   details: str, user: str = "system"):
    try:
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO activity_log (resource_type, resource_id, action, details, actor)
                   VALUES (%s, %s, %s, %s, %s)""",
                (resource_type, resource_id, action, Json({"message": details}), user),
            )
        conn.commit()
    except Exception:
        try:
            conn.rollback()  # critical: reset failed-transaction state so connection remains usable
        except Exception:
            pass

# ---------------------------------------------------------------------------
# Table migration (inlined for container compatibility)
# ---------------------------------------------------------------------------
_TABLES_SQL = """
CREATE TABLE IF NOT EXISTS vm_provisioning_batches (
    id              SERIAL PRIMARY KEY,
    name            TEXT NOT NULL,
    status          TEXT NOT NULL DEFAULT 'validated',
    approval_status TEXT NOT NULL DEFAULT 'pending_approval',
    require_approval BOOLEAN NOT NULL DEFAULT TRUE,
    created_by      TEXT,
    domain_name     TEXT NOT NULL,
    project_name    TEXT NOT NULL,
    dry_run_results JSONB,
    created_at      TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    updated_at      TIMESTAMPTZ NOT NULL DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS vm_provisioning_vms (
    id              SERIAL PRIMARY KEY,
    batch_id        INTEGER NOT NULL REFERENCES vm_provisioning_batches(id) ON DELETE CASCADE,
    vm_name_suffix  TEXT NOT NULL,
    count           INTEGER NOT NULL DEFAULT 1,
    image_name      TEXT,
    image_id        TEXT,
    flavor_name     TEXT,
    flavor_id       TEXT,
    volume_gb       INTEGER NOT NULL DEFAULT 20,
    network_name    TEXT,
    network_id      TEXT,
    security_groups JSONB DEFAULT '[]',
    fixed_ip        TEXT,
    hostname        TEXT,
    os_username     TEXT NOT NULL,
    os_password     TEXT NOT NULL,
    extra_cloudinit TEXT,
    os_type         TEXT NOT NULL DEFAULT 'linux',
    delete_on_termination BOOLEAN NOT NULL DEFAULT TRUE,
    pcd_server_ids  JSONB DEFAULT '[]',
    assigned_ips    JSONB DEFAULT '[]',
    status          TEXT NOT NULL DEFAULT 'pending',
    error_msg       TEXT,
    console_log     TEXT,
    created_at      TIMESTAMPTZ NOT NULL DEFAULT NOW()
);
"""

_tables_initialized = False

def _ensure_tables():
    global _tables_initialized
    if _tables_initialized:
        return
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(_TABLES_SQL)
                # Migration: add delete_on_termination if it doesn't exist yet
                cur.execute(
                    """ALTER TABLE vm_provisioning_vms
                       ADD COLUMN IF NOT EXISTS delete_on_termination BOOLEAN NOT NULL DEFAULT TRUE"""
                )
                # Migration: add region_id for multi-cluster support
                cur.execute(
                    """ALTER TABLE vm_provisioning_batches
                       ADD COLUMN IF NOT EXISTS region_id TEXT"""
                )
        _tables_initialized = True
    except Exception as e:
        logger.error("vm_provisioning _ensure_tables failed: %s", e)

# ---------------------------------------------------------------------------
# Naming helpers
# ---------------------------------------------------------------------------
def _tenant_slug(domain_name: str) -> str:
    slug = domain_name.lower()
    slug = re.sub(r"[^a-z0-9\-]", "-", slug)
    slug = re.sub(r"-{2,}", "-", slug).strip("-")
    return slug or "tenant"

def _vm_nova_name(domain_name: str, suffix: str, index: Optional[int] = None, count: int = 1) -> str:
    slug = _tenant_slug(domain_name)
    clean_suffix = re.sub(r"[^a-z0-9\-]", "-", suffix.lower()).strip("-")
    base = f"{slug}_vm_{clean_suffix}"
    if count > 1 and index is not None:
        pad = 3 if count > 99 else 2
        base = f"{base}-{str(index).zfill(pad)}"
    return base

def _vm_hostname(nova_name: str) -> str:
    """Convert Nova name to RFC-1123 hostname (underscores → hyphens)."""
    return nova_name.replace("_", "-")

# ---------------------------------------------------------------------------
# cloud-init payload builder
# ---------------------------------------------------------------------------
def _build_cloudinit_linux(os_username: str, os_password: str, hostname: str,
                             extra: Optional[str] = None) -> str:
    # Hash password for cloud-config using passlib (SHA-512)
    from passlib.hash import sha512_crypt
    salted = sha512_crypt.hash(os_password)
    yaml = f"""#cloud-config
hostname: {hostname}
manage_etc_hosts: true
users:
  - name: {os_username}
    sudo: ALL=(ALL) NOPASSWD:ALL
    shell: /bin/bash
    lock_passwd: false
    passwd: "{salted}"
ssh_pwauth: true
chpasswd:
  expire: false
  list: |
    {os_username}:{os_password}
packages:
  - qemu-guest-agent
runcmd:
  - systemctl enable qemu-guest-agent
  - systemctl start qemu-guest-agent
"""
    if extra and extra.strip():
        yaml += f"\n# --- extra ---\n{extra.strip()}\n"
    return yaml


def _build_cloudinit_windows(os_username: str, os_password: str,
                               extra: Optional[str] = None,
                               static_ip_config: Optional[Dict[str, Any]] = None) -> str:
    """Build a cloudbase-init compatible #ps1_sysnative script.

    Requires cloudbase-init to be installed in the Windows image.
    Handles two cases:
      - os_username == 'administrator': enable + set password on built-in account
      - any other username: enable Administrator + create new local user

    NOTE: If the image was NOT sysprep'd via `cloudbase-init.exe /sysprep`, the
    Windows OOBE wizard will still appear on first boot (before this script runs).
    The correct fix is image-level: prepare the template VM with
    `cloudbase-init.exe /sysprep` so SkipMachineOOBE is set in the sysprep
    unattend.xml and cloudbase-init fires via SetupComplete before any login screen.
    """
    is_builtin_admin = os_username.strip().lower() == "administrator"

    # Unattend snippet written to suppress OOBE on subsequent boots / reboots
    # (helps if the image was not sysprep'd via cloudbase-init's own /sysprep command)
    oobe_suppression = r"""
# Write an unattend.xml to suppress OOBE on subsequent boots
$unattendDir = "C:\Windows\Panther"
if (-not (Test-Path $unattendDir)) { New-Item -ItemType Directory -Path $unattendDir | Out-Null }
$unattendXml = @'
<?xml version="1.0" encoding="utf-8"?>
<unattend xmlns="urn:schemas-microsoft-com:unattend">
  <settings pass="oobeSystem">
    <component name="Microsoft-Windows-Shell-Setup" processorArchitecture="amd64"
               publicKeyToken="31bf3856ad364e35" language="neutral" versionScope="nonSxS"
               xmlns:wcm="http://schemas.microsoft.com/WMIConfig/2002/State">
      <OOBE>
        <HideEULAPage>true</HideEULAPage>
        <HideOEMRegistrationScreen>true</HideOEMRegistrationScreen>
        <HideOnlineAccountScreens>true</HideOnlineAccountScreens>
        <HideWirelessSetupInOOBE>true</HideWirelessSetupInOOBE>
        <SkipMachineOOBE>true</SkipMachineOOBE>
        <SkipUserOOBE>true</SkipUserOOBE>
        <ProtectYourPC>3</ProtectYourPC>
      </OOBE>
    </component>
  </settings>
</unattend>
'@
Set-Content -Path "$unattendDir\unattend.xml" -Value $unattendXml -Encoding UTF8
"""

    if is_builtin_admin:
        # Built-in account already exists — cannot use /add, just set password and activate
        script = f"""#ps1_sysnative
# cloudbase-init: enable and configure built-in Administrator account
net user Administrator "{os_password}"
net user Administrator /active:yes
wmic useraccount where \"name='Administrator'\" set PasswordExpires=False
{oobe_suppression}"""
    else:
        script = f"""#ps1_sysnative
# cloudbase-init: enable Administrator and create user {os_username}
net user Administrator /active:yes
net user "{os_username}" "{os_password}" /add /y
net localgroup Administrators "{os_username}" /add
wmic useraccount where \"name='{os_username}'\" set PasswordExpires=False
{oobe_suppression}"""
    if static_ip_config and static_ip_config.get("fixed_ip"):
        fixed_ip = static_ip_config["fixed_ip"]
        prefix   = int(static_ip_config.get("prefix_length", 24))
        gateway  = static_ip_config.get("gateway_ip") or ""
        dns_list = static_ip_config.get("dns_servers") or ["8.8.8.8", "8.8.4.4"]
        dns_ps   = ", ".join(f'"{d}"' for d in dns_list[:4])
        gw_line  = f"-DefaultGateway '{gateway}'" if gateway else ""
        script += f"""
# Configure static IP assigned at provision time
$_retries = 0
$_adapter = $null
while (-not $_adapter -and $_retries -lt 20) {{
    $_adapter = Get-NetAdapter | Where-Object {{ $_.HardwareInterface -eq $true -and $_.Status -ne 'NotPresent' }} | Select-Object -First 1
    if (-not $_adapter) {{ Start-Sleep -Seconds 3; $_retries++ }}
}}
if ($_adapter) {{
    try {{
        Get-NetIPAddress  -InterfaceIndex $_adapter.InterfaceIndex -ErrorAction SilentlyContinue | Remove-NetIPAddress -Confirm:$false -ErrorAction SilentlyContinue
        Get-NetRoute      -InterfaceIndex $_adapter.InterfaceIndex -ErrorAction SilentlyContinue | Remove-NetRoute     -Confirm:$false -ErrorAction SilentlyContinue
        New-NetIPAddress  -InterfaceIndex $_adapter.InterfaceIndex -IPAddress '{fixed_ip}' -PrefixLength {prefix} {gw_line}
        Set-DnsClientServerAddress -InterfaceIndex $_adapter.InterfaceIndex -ServerAddresses @({dns_ps})
        Write-Host 'Static IP configured: {fixed_ip}/{prefix} via {gateway}'
    }} catch {{
        Write-Host "Warning: static IP config failed — $_"
    }}
}} else {{
    Write-Host 'Warning: no network adapter found for static IP configuration'
}}
"""
    if extra and extra.strip():
        script += f"\n{extra.strip()}\n"
    return script


def _encode_userdata(payload: str) -> str:
    encoded = base64.b64encode(payload.encode("utf-8")).decode("ascii")
    if len(base64.b64decode(encoded)) > 65535:
        raise ValueError("cloud-init payload exceeds Nova 65535-byte limit")
    return encoded

# ---------------------------------------------------------------------------
# Excel template
# ---------------------------------------------------------------------------
def _build_excel_template() -> io.BytesIO:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()

    # ── vms sheet ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "vms"
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(color="FFFFFF", bold=True)
    headers = [
        "domain_name", "project_name", "vm_name_suffix", "count",
        "image_name", "flavor_name", "volume_gb",
        "network_name", "security_groups", "fixed_ip", "hostname",
        "os_username", "os_password", "extra_cloudinit", "delete_on_termination",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 24
    # sample row
    ws.append([
        "mycustomer", "myproject", "web", 1,
        "ubuntu-22.04", "m1.small.bfv", 50,
        "tenant-net", "default", "", "",
        "ubuntu", "REDACTED_PASSWORD", "", True,
    ])

    # ── README sheet ──────────────────────────────────────────────────
    readme = wb.create_sheet("README")
    readme["A1"] = "VM Provisioning — Column Reference"
    readme["A1"].font = Font(bold=True, size=13)
    notes = [
        ("domain_name", "Required. Keystone domain (customer) name."),
        ("project_name", "Required. Project within the domain."),
        ("vm_name_suffix", "Required. Short identifier. Nova name = {domain}_vm_{suffix}."),
        ("count", "Number of identical VMs to create (1–20). Names get -01 suffix."),
        ("image_name", "Glance image name. Must match exactly."),
        ("flavor_name", "Nova flavor name. Must have disk=0 (boot-from-volume)."),
        ("volume_gb", "Boot volume size in GB. Min = image.min_disk or 10."),
        ("network_name", "Neutron network name accessible to the project."),
        ("security_groups", "Comma-separated SG names. Default: 'default'."),
        ("fixed_ip", "Optional. Static IP from subnet allocation pool."),
        ("hostname", "Optional. Auto-derived from vm_name_suffix if blank."),
        ("os_username", "Required. OS user created via cloud-init."),
        ("os_password", "Required. Initial OS password (min 8 chars)."),
        ("extra_cloudinit", "Optional. Additional cloud-init YAML to merge (advanced)."),
        ("delete_on_termination", "true/false — delete boot volume when VM is deleted. Default: true."),
    ]
    for i, (col, note) in enumerate(notes, 3):
        readme.cell(row=i, column=1, value=col).font = Font(bold=True)
        readme.cell(row=i, column=2, value=note)
    readme.column_dimensions["A"].width = 22
    readme.column_dimensions["B"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------
class VmRowRequest(BaseModel):
    vm_name_suffix: str
    count: int = 1
    image_name: Optional[str] = None
    image_id: Optional[str] = None
    flavor_name: Optional[str] = None
    flavor_id: Optional[str] = None
    volume_gb: int = 20
    network_name: Optional[str] = None
    network_id: Optional[str] = None
    security_groups: List[str] = ["default"]
    fixed_ip: Optional[str] = None
    hostname: Optional[str] = None
    os_username: str
    os_password: str
    extra_cloudinit: Optional[str] = None
    os_type: str = "linux"
    delete_on_termination: bool = True

    @field_validator("vm_name_suffix")
    def validate_suffix(cls, v):
        if not v or not v.strip():
            raise ValueError("vm_name_suffix is required")
        if not re.match(r"^[a-z0-9][a-z0-9\-]*$", v.lower()):
            raise ValueError("vm_name_suffix must be lowercase alphanumeric/hyphens")
        if len(v) > 30:
            raise ValueError("vm_name_suffix max 30 chars")
        return v.lower()

    @field_validator("count")
    def validate_count(cls, v):
        if v < 1 or v > 20:
            raise ValueError("count must be 1–20")
        return v

    @field_validator("volume_gb")
    def validate_vol(cls, v):
        if v < 1 or v > 2048:
            raise ValueError("volume_gb must be 1–2048")
        return v

    @field_validator("os_password")
    def validate_pw(cls, v):
        if len(v) < 8:
            raise ValueError("os_password min 8 characters")
        return v


class CreateBatchRequest(BaseModel):
    name: str
    domain_name: str
    project_name: str
    require_approval: bool = True
    region_id: Optional[str] = None  # target PF9 region (Phase 6)
    vms: List[VmRowRequest]

    @field_validator("name")
    def validate_name(cls, v):
        if not v or not v.strip():
            raise ValueError("batch name required")
        return v.strip()[:128]


class DecisionRequest(BaseModel):
    decision: str  # "approve" | "reject"
    comment: Optional[str] = None

    @field_validator("decision")
    def validate_decision(cls, v):
        if v not in ("approve", "reject"):
            raise ValueError("decision must be 'approve' or 'reject'")
        return v

# ---------------------------------------------------------------------------
# Email
# ---------------------------------------------------------------------------
def _send_email(to_addrs: List[str], subject: str, html_body: str):
    if to_addrs:
        smtp_send_email(to_addrs, subject, html_body)


def _build_completion_email(batch: dict, vm_rows: list, activity_steps: list = None) -> str:
    status_icon = "✅" if batch["status"] == "complete" else ("❌" if batch["status"] == "failed" else "⚠️")
    total = sum(r.get("count", 1) for r in vm_rows)
    created = sum(
        len(r.get("pcd_server_ids") or [])
        for r in vm_rows
        if r.get("status") == "complete"
    )
    rows_html = ""
    for vm in vm_rows:
        sids = vm.get("pcd_server_ids") or []
        ips = vm.get("assigned_ips") or []
        count = vm.get("count", 1)
        pairs = list(zip(sids, ips)) if (sids and ips) else ([(None, None)] * count)
        for i, (sid, ip) in enumerate(pairs):
            name = _vm_nova_name(batch["domain_name"], vm["vm_name_suffix"],
                                  i + 1 if count > 1 else None, count)
            st = vm.get("status", "pending")
            if st == "complete":
                st_color = "#22c55e"
                st_bg = "#052e16"
            elif st == "failed":
                st_color = "#f87171"
                st_bg = "#2d1216"
            else:
                st_color = "#93c5fd"
                st_bg = "#0c1a2e"
            error_cell = f'<td style="padding:6px 8px;color:#f87171;font-size:0.8em">{vm.get("error_msg","") or ""}</td>' \
                         if vm.get("error_msg") else '<td style="padding:6px 8px;color:#475569">—</td>'
            rows_html += f"""<tr style="border-top:1px solid #1e293b">
  <td style="padding:6px 8px;font-family:monospace;font-size:0.85em">{name}</td>
  <td style="padding:6px 8px"><span style="background:{st_bg};color:{st_color};padding:2px 8px;border-radius:4px;font-size:0.8em">{st}</span></td>
  <td style="padding:6px 8px">{ip or '—'}</td>
  <td style="padding:6px 8px;font-size:0.82em;color:#94a3b8" title="{vm.get('image_name','')}">{(vm.get('image_name') or '—')[:30]}</td>
  <td style="padding:6px 8px;font-size:0.82em;color:#94a3b8">{vm.get('flavor_name','—') or '—'}</td>
  <td style="padding:6px 8px;font-size:0.82em;text-transform:uppercase;color:#94a3b8">{vm.get('os_type','linux')}</td>
  <td style="padding:6px 8px;font-size:0.82em;color:#94a3b8">{vm.get('volume_gb','—')}</td>
  {error_cell}
</tr>"""

    # Build activity timeline section
    timeline_html = ""
    if activity_steps:
        steps_html = ""
        for step in activity_steps:
            action = step.get("action", "")
            details = step.get("details") or {}
            if isinstance(details, str):
                msg = details
            else:
                msg = details.get("message", "") or str(details)
            ts_raw = step.get("timestamp") or step.get("created_at", "")
            try:
                if isinstance(ts_raw, str):
                    from dateutil.parser import parse as _dtp
                    ts = _dtp(ts_raw).strftime("%H:%M:%S")
                else:
                    ts = str(ts_raw)
            except Exception:
                ts = str(ts_raw)[:19] if ts_raw else "—"
            if "fail" in action or "error" in action:
                act_color = "#f87171"
            elif any(k in action for k in ("complete", "active", "ready", "created", "started")):
                act_color = "#4ade80"
            else:
                act_color = "#93c5fd"
            steps_html += f"""<div style="margin-bottom:4px">
  <span style="color:#475569;margin-right:8px">{ts}</span>
  <span style="color:{act_color};margin-right:8px;font-weight:600">[{action}]</span>
  <span style="color:#cbd5e1">{msg[:200]}</span>
</div>"""
        timeline_html = f"""
  <div style="padding:0 20px 20px">
    <h3 style="color:#94a3b8;font-size:0.85em;text-transform:uppercase;letter-spacing:1px;margin-bottom:8px">Execution Timeline</h3>
    <div style="background:#0f172a;border-radius:8px;padding:12px;font-family:monospace;font-size:0.8em;max-height:300px;overflow:auto">
      {steps_html}
    </div>
  </div>"""

    return f"""<html><body style="font-family:sans-serif;background:#0f172a;color:#e2e8f0;padding:20px">
<div style="max-width:800px;margin:0 auto;background:#1e293b;border-radius:12px;overflow:hidden">
  <div style="background:linear-gradient(135deg,#1e40af,#7c3aed);padding:24px">
    <h2 style="margin:0;color:#fff">{status_icon} VM Provisioning — {batch['name']}</h2>
    <p style="margin:8px 0 0;color:#bfdbfe">
      <strong>{created}/{total}</strong> VMs active &nbsp;·&nbsp;
      Domain: <strong>{batch['domain_name']}</strong> &nbsp;·&nbsp;
      Project: <strong>{batch['project_name']}</strong>
    </p>
    <p style="margin:4px 0 0;color:#93c5fd;font-size:0.85em">
      Submitted by {batch.get('created_by','—')} &nbsp;·&nbsp;
      Status: <strong>{batch['status'].replace('_',' ').title()}</strong>
    </p>
  </div>
  <div style="padding:20px">
    <table style="width:100%;border-collapse:collapse">
      <thead>
        <tr style="background:#0f172a;color:#64748b;font-size:0.8em;text-transform:uppercase;letter-spacing:0.5px">
          <th style="padding:8px 8px;text-align:left">VM Name</th>
          <th style="padding:8px 8px;text-align:left">Status</th>
          <th style="padding:8px 8px;text-align:left">IP Address</th>
          <th style="padding:8px 8px;text-align:left">Image</th>
          <th style="padding:8px 8px;text-align:left">Flavor</th>
          <th style="padding:8px 8px;text-align:left">OS</th>
          <th style="padding:8px 8px;text-align:left">GB</th>
          <th style="padding:8px 8px;text-align:left">Error</th>
        </tr>
      </thead>
      <tbody style="background:#1e293b">{rows_html}</tbody>
    </table>
  </div>
  {timeline_html}
  <div style="padding:12px 20px;background:#0f172a;font-size:0.8em;color:#64748b">
    Platform9 Management System · {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}
  </div>
</div>
</body></html>"""

# ---------------------------------------------------------------------------
# Background execution
# ---------------------------------------------------------------------------
def _execute_batch_thread(batch_id: int, operator_email: Optional[str]):
    conn = get_pool().getconn()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM vm_provisioning_batches WHERE id=%s", (batch_id,))
            batch = dict(cur.fetchone())
            cur.execute("SELECT * FROM vm_provisioning_vms WHERE batch_id=%s ORDER BY id", (batch_id,))
            vm_rows = [dict(r) for r in cur.fetchall()]
        conn.commit()

        # Decrypt os_password in all VM rows now (encrypted since v1.83.25;
        # backward-compat: _decrypt_vm_password returns plaintext rows unchanged).
        for _vm in vm_rows:
            if _vm.get("os_password"):
                _vm["os_password"] = _decrypt_vm_password(_vm["os_password"])

        admin_client = get_client()
        admin_client.authenticate()

        # Resolve target project UUID
        project_id = admin_client.resolve_project_id(
            project_name=batch["project_name"],
            domain_name=batch["domain_name"],
        )

        # Ensure provisionsrv has _member_ role in the target project BEFORE
        # calling scoped_for_project — this must run even when dry-run was skipped.
        from vm_provisioning_service_user import ensure_provisioner_in_project
        ensure_provisioner_in_project(admin_client, project_id)

        # Authenticate as provisionsrv scoped to the target project.
        # This gives a genuine project-scoped token so all Nova/Neutron/Cinder
        # resource lookups (SGs, networks, volumes) resolve within that project.
        project_client = admin_client.scoped_for_project(
            project_name=batch["project_name"],
            domain_name=batch["domain_name"],
        )
        # project_client.project_id is the resolved target project UUID
        project_id = project_client.project_id

        _log_activity(conn, "vm_provisioning", str(batch_id),
                       "execution_started",
                       f"Execution started — project_id={project_id}, "
                       f"cinder_endpoint={project_client.cinder_endpoint}",
                       batch.get("created_by", "system"))

        # Resolve image IDs and flavor IDs by name if needed
        images_cache: Dict[str, Any] = {}
        all_images = admin_client.list_images()
        for img in all_images:
            images_cache[img.get("name", "")] = img
            images_cache[img.get("id", "")] = img

        flavors_cache: Dict[str, Any] = {}
        all_flavors = admin_client.list_flavors()
        for fl in all_flavors:
            flavors_cache[fl.get("name", "")] = fl
            flavors_cache[fl.get("id", "")] = fl

        networks_cache: Dict[str, Any] = {}
        all_networks = admin_client.list_networks(project_id=project_id)
        for net in all_networks:
            networks_cache[net.get("name", "")] = net
            networks_cache[net.get("id", "")] = net

        # Build SG name→id map for the target project so Nova gets UUIDs (not names)
        sgs_name_to_id: Dict[str, str] = {}
        try:
            for sg in admin_client.list_security_groups(project_id=project_id):
                sgs_name_to_id[sg.get("name", "")] = sg.get("id", "")
                sgs_name_to_id[sg.get("id", "")] = sg.get("id", "")  # pass-through UUIDs
        except Exception:
            pass  # if Neutron call fails, fall back to names

        # Cache subnets per network_id for Windows static IP lookup
        subnet_cache: Dict[str, List[Dict[str, Any]]] = {}

        all_failed = False

        for vm in vm_rows:
            vm_id = vm["id"]
            count = vm.get("count", 1)
            suffix = vm["vm_name_suffix"]

            # Skip VMs that already completed (allows re-execute on partial failures)
            if vm.get("status") == "complete":
                continue
            # Reset any previously-failed status back to pending
            if vm.get("status") == "failed":
                _update_vm_status(conn, vm_id, "pending", None)

            # Mark VM as executing so UI shows progress immediately
            _update_vm_status(conn, vm_id, "executing", None)

            # Resolve image
            img_obj = images_cache.get(vm.get("image_id") or "") or \
                      images_cache.get(vm.get("image_name") or "")
            if not img_obj:
                _update_vm_status(conn, vm_id, "failed",
                                   f"Image '{vm.get('image_name')}' not found")
                continue
            image_id = img_obj["id"]
            # Detect OS from stored type OR image metadata OR image name
            raw_os = (
                img_obj.get("os_type", "") +
                img_obj.get("os_distro", "") +
                img_obj.get("name", "")
            ).lower()
            os_type = vm.get("os_type") or (
                "windows" if "windows" in raw_os or "win" in raw_os else "linux"
            )

            # Ensure Windows images carry the required Glance properties so Nova
            # configures the correct virtual hardware (SCSI bus, virtio-scsi model,
            # BIOS firmware).  We patch the image before creating the boot volume so
            # that the properties are present when Cinder and Nova inspect it.
            if os_type == "windows":
                try:
                    admin_client.update_image_properties(image_id, {
                        "os_type":        "windows",
                        "hw_disk_bus":    "scsi",
                        "hw_scsi_model":  "virtio-scsi",
                        "hw_firmware_type": "bios",
                    })
                except Exception as _img_patch_err:
                    _log_activity(conn, "vm_provisioning_vm", str(vm_id),
                                   "image_patch_warning",
                                   f"{vm.get('image_name')}: could not patch Windows image "
                                   f"properties — {_img_patch_err}",
                                   batch.get("created_by", "system"))

            # Resolve flavor
            fl_obj = flavors_cache.get(vm.get("flavor_id") or "") or \
                     flavors_cache.get(vm.get("flavor_name") or "")
            if not fl_obj:
                _update_vm_status(conn, vm_id, "failed",
                                   f"Flavor '{vm.get('flavor_name')}' not found")
                continue
            flavor_id = fl_obj["id"]

            # Resolve network
            net_obj = networks_cache.get(vm.get("network_id") or "") or \
                      networks_cache.get(vm.get("network_name") or "")
            if not net_obj:
                _update_vm_status(conn, vm_id, "failed",
                                   f"Network '{vm.get('network_name')}' not found")
                continue
            network_id = net_obj["id"]

            # Resolve static IP config for Windows VMs with a fixed_ip (count==1 only)
            static_ip_cfg: Optional[Dict[str, Any]] = None
            if os_type == "windows" and vm.get("fixed_ip") and count == 1:
                try:
                    if network_id not in subnet_cache:
                        subnet_cache[network_id] = admin_client.list_subnets(network_id=network_id)
                    for sn in subnet_cache[network_id]:
                        try:
                            net_cidr = ipaddress.ip_network(sn.get("cidr", ""), strict=False)
                            if ipaddress.ip_address(vm["fixed_ip"]) in net_cidr:
                                dns = sn.get("dns_nameservers") or ["8.8.8.8", "8.8.4.4"]
                                static_ip_cfg = {
                                    "fixed_ip":      vm["fixed_ip"],
                                    "prefix_length": net_cidr.prefixlen,
                                    "gateway_ip":    sn.get("gateway_ip") or "",
                                    "dns_servers":   dns,
                                }
                                break
                        except Exception:
                            pass
                except Exception:
                    pass  # subnet lookup failure is non-fatal; Windows will attempt DHCP

            server_ids: List[str] = []
            assigned_ips: List[str] = []
            vm_failed = False

            for idx in range(1, count + 1):
                instance_name = _vm_nova_name(batch["domain_name"], suffix, idx if count > 1 else None, count)
                instance_hostname = _vm_hostname(instance_name)
                if vm.get("hostname"):
                    instance_hostname = vm["hostname"] if count == 1 else f"{vm['hostname']}-{str(idx).zfill(2)}"

                # Build cloud-init payload
                try:
                    if os_type == "windows":
                        payload = _build_cloudinit_windows(
                            vm["os_username"], vm["os_password"], vm.get("extra_cloudinit"),
                            static_ip_config=static_ip_cfg if count == 1 else None)
                    else:
                        payload = _build_cloudinit_linux(
                            vm["os_username"], vm["os_password"],
                            instance_hostname, vm.get("extra_cloudinit"))
                    user_data_b64 = _encode_userdata(payload)
                except Exception as e:
                    _update_vm_status(conn, vm_id, "failed", f"cloud-init build failed: {e}")
                    vm_failed = True
                    break

                # Create boot volume — ensure size >= image virtual/min disk to avoid 400
                try:
                    img_min_disk = int(img_obj.get("min_disk") or 0)
                    img_virtual_gb = int((img_obj.get("virtual_size") or 0) / 1073741824 + 0.999) if img_obj.get("virtual_size") else 0
                    effective_size_gb = max(int(vm["volume_gb"]), img_min_disk, img_virtual_gb, 10)
                    # Windows images need at least 40 GB — guard against un-set virtual_size in Glance
                    if os_type == "windows" and effective_size_gb < 40:
                        effective_size_gb = 40
                    _log_activity(conn, "vm_provisioning_vm", str(vm_id),
                                   "volume_creating",
                                   f"{instance_name}: creating {effective_size_gb}GB boot volume from image {image_id}",
                                   batch.get("created_by", "system"))
                    # Ensure the image is accessible to the provisionsrv project-scoped
                    # token.  Platform9 Cinder resolves imageRef by calling back to
                    # Glance with the caller's auth token; if the image is private
                    # (owned by the service project) and the token is scoped to ORG1
                    # the Glance lookup fails → 400 Invalid image identifier.
                    # Setting visibility to 'community' makes it readable by any project
                    # token without changing ownership or affecting other tenants.
                    admin_client.ensure_image_accessible(image_id)
                    # Create boot volume using the provisionsrv token already scoped to
                    # the target project.  Platform9 Cinder strictly enforces that the
                    # URL project_id matches the token's project scope; using the
                    # admin_client's service-scoped token with an ORG1 project_id in the
                    # URL produces "400 Malformed request url".  project_client's
                    # cinder_endpoint already embeds the correct project_id.
                    vol = project_client.create_boot_volume(
                        name=f"{instance_name}-vol",
                        image_id=image_id,
                        size_gb=effective_size_gb,
                    )
                    volume_id = vol["id"]
                except Exception as e:
                    _update_vm_status(conn, vm_id, "failed", f"Volume create failed: {e}")
                    vm_failed = True
                    break

                # Wait for volume available + bootable (max 20 min — large Windows images need time)
                vol_ready = False
                for _ in range(240):
                    time.sleep(5)
                    try:
                        v = project_client.get_volume(volume_id)
                        vol_status = v.get("status")
                        if vol_status == "error":
                            break
                        if vol_status == "available":
                            # Also verify bootable flag — Cinder sets this after image copy finishes
                            if v.get("bootable") in ("true", True):
                                vol_ready = True
                                break
                            # Still waiting for image copy to complete
                    except Exception:
                        pass
                if not vol_ready:
                    _update_vm_status(conn, vm_id, "failed", "Volume did not become bootable in 20 min — image copy may have failed or timed out")
                    vm_failed = True
                    break

                _log_activity(conn, "vm_provisioning_vm", str(vm_id),
                               "volume_ready",
                               f"{instance_name}: volume {volume_id} ready, booting VM",
                               batch.get("created_by", "system"))

                # Boot VM
                try:
                    sgs = vm.get("security_groups") or ["default"]
                    if isinstance(sgs, str):
                        sgs = json.loads(sgs)
                    # Resolve SG names to UUIDs — Nova fails to find names when using
                    # an admin token with X-Project-Id header pointing to another project
                    sgs = [sgs_name_to_id.get(sg, sg) for sg in sgs]
                    server = project_client.create_server_bfv(
                        name=instance_name,
                        volume_id=volume_id,
                        flavor_id=flavor_id,
                        network_id=network_id,
                        security_group_names=sgs,
                        user_data_b64=user_data_b64,
                        fixed_ip=vm.get("fixed_ip") if count == 1 else None,
                        hostname=instance_hostname,
                        project_id=project_id,  # X-Project-Id so Nova puts VM in target project
                        # adminPass: seeds password via Nova metadata (picked up by cloudbase-init OpenStackService plugin)
                        admin_pass=vm["os_password"] if os_type == "windows" else None,
                        # Pass image_id so Nova reads hw_firmware_type / hw_machine_type from Glance
                        # This ensures UEFI firmware for Windows Server 2019+ GPT images
                        image_id=image_id,
                        delete_on_termination=vm.get("delete_on_termination", True),
                    )
                    server_id = server.get("id")
                except Exception as e:
                    _update_vm_status(conn, vm_id, "failed", f"Server boot failed: {e}")
                    vm_failed = True
                    break

                # Wait for ACTIVE (max 10 min)
                server_active = False
                assigned_ip = None
                for _ in range(120):
                    time.sleep(5)
                    try:
                        srv = project_client.get_server(server_id)
                        task_state = srv.get("OS-EXT-STS:task_state")
                        vm_state = srv.get("OS-EXT-STS:vm_state", "")
                        power_state = srv.get("status", "")
                        if power_state == "ACTIVE":
                            server_active = True
                            # Extract assigned IP
                            addresses = srv.get("addresses", {})
                            for net_name, addr_list in addresses.items():
                                for addr in addr_list:
                                    if addr.get("version") == 4:
                                        assigned_ip = addr.get("addr")
                                        break
                                if assigned_ip:
                                    break
                            break
                        if power_state == "ERROR":
                            fault = srv.get("fault", {}).get("message", "Unknown error")
                            _update_vm_status(conn, vm_id, "failed", f"Nova ERROR: {fault}")
                            vm_failed = True
                            break
                    except Exception:
                        pass

                if not server_active and not vm_failed:
                    _update_vm_status(conn, vm_id, "failed", "Server did not reach ACTIVE in 10 min")
                    vm_failed = True
                    break

                if vm_failed:
                    break

                server_ids.append(server_id)
                assigned_ips.append(assigned_ip or "")

                # Fetch console log
                try:
                    console = admin_client.get_console_log(server_id, 80)
                except Exception:
                    console = ""

                # Log activity
                _log_activity(conn, "vm_provisioning_vm", str(vm_id),
                               "vm_created",
                               f"VM {instance_name} active at {assigned_ip}",
                               batch.get("created_by", "system"))

            # Update vm row with results
            if not vm_failed:
                with conn.cursor() as cur:
                    cur.execute(
                        """UPDATE vm_provisioning_vms
                           SET status='complete', pcd_server_ids=%s, assigned_ips=%s,
                               console_log=%s, error_msg=NULL, os_password=''
                           WHERE id=%s""",
                        (Json(server_ids), Json(assigned_ips), console if count == 1 else None, vm_id),
                    )
                conn.commit()

        # Determine batch final status
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                "SELECT status FROM vm_provisioning_vms WHERE batch_id=%s", (batch_id,)
            )
            statuses = [r["status"] for r in cur.fetchall()]

        if all(s == "complete" for s in statuses):
            final_status = "complete"
        elif any(s == "complete" for s in statuses):
            final_status = "partially_failed"
        else:
            final_status = "failed"

        with conn.cursor() as cur:
            cur.execute(
                "UPDATE vm_provisioning_batches SET status=%s, updated_at=NOW() WHERE id=%s",
                (final_status, batch_id),
            )
        conn.commit()

        _log_activity(conn, "vm_provisioning", str(batch_id),
                       f"provisioning_{final_status}",
                       f"Batch '{batch['name']}' finished as {final_status}",
                       batch.get("created_by", "system"))

        # Send notification email
        final_vms: list = []
        if operator_email:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT * FROM vm_provisioning_vms WHERE batch_id=%s", (batch_id,))
                final_vms = [dict(r) for r in cur.fetchall()]
                vm_db_ids = [str(v["id"]) for v in final_vms]
                cur.execute(
                    "SELECT action, details, created_at FROM activity_log "
                    "WHERE (resource_type='vm_provisioning' AND resource_id=%s) "
                    "   OR (resource_type='vm_provisioning_vm' AND resource_id=ANY(%s)) "
                    "ORDER BY created_at",
                    (str(batch_id), vm_db_ids),
                )
                activity_steps = [dict(r) for r in cur.fetchall()]
            html = _build_completion_email(batch, final_vms, activity_steps)
            _send_email([operator_email], f"VM Provisioning '{batch['name']}' — {final_status}", html)

        # Fire in-app notification bell for all subscribers of this event type
        try:
            from provisioning_routes import _fire_notification
            total_vms = sum(r.get("count", 1) for r in final_vms) if final_vms else 0
            created_vms = sum(
                len(r.get("pcd_server_ids") or [])
                for r in final_vms
                if r.get("status") == "complete"
            ) if final_vms else 0
            event_type = "vm_provisioning_completed" if final_status == "complete" else "vm_provisioning_failed"
            severity = "info" if final_status == "complete" else ("critical" if final_status == "failed" else "warning")
            _fire_notification(
                event_type=event_type,
                summary=f"VM Provisioning '{batch['name']}' {final_status.replace('_', ' ')} — "
                        f"{created_vms}/{total_vms} VMs active · "
                        f"{batch['domain_name']}/{batch['project_name']}",
                severity=severity,
                resource_id=str(batch_id),
                resource_name=batch["name"],
                domain_name=batch["domain_name"],
                project_name=batch["project_name"],
                actor=batch.get("created_by", "system"),
            )
        except Exception as _ne:
            logger.warning("Could not fire in-app notification for batch %s: %s", batch_id, _ne)

    except Exception as e:
        logger.error(f"Batch {batch_id} execution thread error: {traceback.format_exc()}")
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE vm_provisioning_batches SET status='failed', updated_at=NOW() WHERE id=%s",
                    (batch_id,),
                )
            conn.commit()
        except Exception:
            pass
    finally:
        try:
            get_pool().putconn(conn)
        except Exception:
            try:
                conn.close()
            except Exception:
                pass


def _update_vm_status(conn, vm_id: int, status: str, error_msg: Optional[str] = None):
    try:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE vm_provisioning_vms SET status=%s, error_msg=%s WHERE id=%s",
                (status, error_msg, vm_id),
            )
        conn.commit()
    except Exception:
        pass

# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@router.get("/domains")
async def list_domains_projects(
    region_id: Optional[str] = Query(None, description="Filter by region ID"),
    user=Depends(get_current_user),
):
    """Return all Keystone domains, each with its list of projects.
    Used by the Step 1 picker so users never have to type domain/project names."""
    try:
        uname = user["username"] if isinstance(user, dict) else user.username
        effective_region = get_effective_region_filter(uname, region_id)
        admin = get_registry().get_region(effective_region) if effective_region else get_client()
        domains_raw = admin.list_domains()
        result = []
        for d in sorted(domains_raw, key=lambda x: x.get("name", "").lower()):
            projects = admin.list_projects(domain_id=d["id"])
            result.append({
                "id": d["id"],
                "name": d["name"],
                "projects": sorted(
                    [{"id": p["id"], "name": p["name"]} for p in projects],
                    key=lambda x: x["name"].lower(),
                ),
            })
        return result
    except Exception as e:
        logger.exception("list_domains_projects failed")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/resources")
async def get_resources(
    domain_name: str = Query(...),
    project_name: str = Query(...),
    region_id: Optional[str] = Query(None, description="Filter by region ID"),
    request: Request = None,
    user=Depends(get_current_user),
):
    """Return images (Glance), diskless flavors, networks, SGs scoped to domain+project."""
    _ensure_tables()
    try:
        uname = user["username"] if isinstance(user, dict) else user.username
        effective_region = get_effective_region_filter(uname, region_id)
        admin = get_registry().get_region(effective_region) if effective_region else get_client()
        project_id = admin.resolve_project_id(project_name=project_name, domain_name=domain_name)

        images = admin.list_images()
        # Annotate with detected OS type — check metadata AND image name
        for img in images:
            raw = (
                img.get("os_type", "") +
                img.get("os_distro", "") +
                img.get("name", "")
            ).lower()
            img["_detected_os"] = "windows" if "windows" in raw or "win" in raw else "linux"

        flavors = admin.list_diskless_flavors()
        networks = admin.list_networks(project_id=project_id)
        sgs = admin.list_security_groups(project_id=project_id)

        return {
            "images": images,
            "flavors": flavors,
            "networks": networks,
            "security_groups": sgs,
            "project_id": project_id,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/quota")
async def get_quota(
    domain_name: str = Query(...),
    project_name: str = Query(...),
    region_id: Optional[str] = Query(None, description="Filter by region ID"),
    user=Depends(get_current_user),
):
    """Return compute + storage quota with usage for a project."""
    _ensure_tables()
    try:
        uname = user["username"] if isinstance(user, dict) else user.username
        effective_region = get_effective_region_filter(uname, region_id)
        admin = get_registry().get_region(effective_region) if effective_region else get_client()
        project_id = admin.resolve_project_id(project_name=project_name, domain_name=domain_name)
        usage = admin.get_quota_usage(project_id)
        return usage
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/available-ips")
async def get_available_ips(
    domain_name: str = Query(...),
    project_name: str = Query(...),
    network_id: str = Query(...),
    region_id: Optional[str] = Query(None, description="Filter by region ID"),
    user=Depends(get_current_user),
):
    """Return free IPs from the subnet allocation pool for a network."""
    _ensure_tables()
    try:
        uname = user["username"] if isinstance(user, dict) else user.username
        effective_region = get_effective_region_filter(uname, region_id)
        admin = get_registry().get_region(effective_region) if effective_region else get_client()
        project_id = admin.resolve_project_id(project_name=project_name, domain_name=domain_name)
        ips = admin.list_available_ips(network_id, project_id=project_id)
        return {"available_ips": ips}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/template")
async def download_template(user=Depends(get_current_user)):
    """Download the Excel bulk provisioning template."""
    buf = _build_excel_template()
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="vm_provisioning_template.xlsx"'},
    )


@router.post("/upload")
async def upload_excel(
    file: UploadFile = File(...),
    user=Depends(get_current_user),
):
    """Parse and validate an Excel bulk provisioning workbook."""
    _ensure_tables()
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    raw = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)

    if "vms" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Workbook must contain a 'vms' sheet")

    ws = wb["vms"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    errors = []
    parsed_vms: List[Dict[str, Any]] = []

    def _col(row_dict: dict, key: str, default=None):
        return row_dict.get(key) or default

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        row_dict = {h: v for h, v in zip(headers, row) if h}

        domain = _col(row_dict, "domain_name")
        project = _col(row_dict, "project_name")
        suffix = _col(row_dict, "vm_name_suffix")
        os_user = _col(row_dict, "os_username")
        os_pass = str(_col(row_dict, "os_password", ""))

        if not domain:
            errors.append(f"Row {row_num}: domain_name required")
        if not project:
            errors.append(f"Row {row_num}: project_name required")
        if not suffix:
            errors.append(f"Row {row_num}: vm_name_suffix required")
        if not os_user:
            errors.append(f"Row {row_num}: os_username required")
        if len(os_pass) < 8:
            errors.append(f"Row {row_num}: os_password min 6 chars")

        count_val = _col(row_dict, "count", 1)
        try:
            count_val = int(count_val)
        except Exception:
            count_val = 1
        if count_val < 1 or count_val > 20:
            errors.append(f"Row {row_num}: count must be 1–20")

        vol_val = _col(row_dict, "volume_gb", 20)
        try:
            vol_val = int(vol_val)
        except Exception:
            vol_val = 20

        sgs_raw = _col(row_dict, "security_groups", "default")
        sgs = [s.strip() for s in str(sgs_raw).split(",") if s.strip()] if sgs_raw else ["default"]

        parsed_vms.append({
            "domain_name": domain,
            "project_name": project,
            "vm_name_suffix": str(suffix).lower() if suffix else "",
            "count": count_val,
            "image_name": _col(row_dict, "image_name"),
            "flavor_name": _col(row_dict, "flavor_name"),
            "volume_gb": vol_val,
            "network_name": _col(row_dict, "network_name"),
            "security_groups": sgs,
            "fixed_ip": _col(row_dict, "fixed_ip"),
            "hostname": _col(row_dict, "hostname"),
            "os_username": os_user,
            "os_password": os_pass,
            "extra_cloudinit": _col(row_dict, "extra_cloudinit"),
            "delete_on_termination": str(_col(row_dict, "delete_on_termination", "true")).strip().lower() not in ("false", "0", "no"),
        })

    if errors:
        return {"status": "invalid", "errors": errors, "row_count": len(parsed_vms)}

    return {
        "status": "validated",
        "row_count": len(parsed_vms),
        "vms": parsed_vms,
        "errors": [],
    }


@router.post("/batches")
async def create_batch(
    body: CreateBatchRequest,
    request: Request = None,
    user=Depends(get_current_user),
):
    """Create a provisioning batch from form or Excel-validated data."""
    _ensure_tables()
    try:
        with get_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # When approval is not required, mark the batch as already approved
                # so the Execute button is immediately available after dry-run.
                initial_approval = "approved" if not body.require_approval else "pending_approval"
                cur.execute(
                    """INSERT INTO vm_provisioning_batches
                       (name, domain_name, project_name, require_approval, created_by,
                        status, approval_status, region_id)
                       VALUES (%s,%s,%s,%s,%s,'validated',%s,%s)
                       RETURNING id""",
                    (body.name, body.domain_name, body.project_name,
                     body.require_approval, user.username, initial_approval, body.region_id),
                )
                batch_id = cur.fetchone()["id"]

                for vm in body.vms:
                    sgs = vm.security_groups or ["default"]
                    cur.execute(
                        """INSERT INTO vm_provisioning_vms
                           (batch_id, vm_name_suffix, count, image_name, image_id,
                            flavor_name, flavor_id, volume_gb, network_name, network_id,
                            security_groups, fixed_ip, hostname,
                            os_username, os_password, extra_cloudinit, os_type,
                            delete_on_termination)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (batch_id, vm.vm_name_suffix, vm.count,
                         vm.image_name, vm.image_id,
                         vm.flavor_name, vm.flavor_id, vm.volume_gb,
                         vm.network_name, vm.network_id,
                         Json(sgs), vm.fixed_ip, vm.hostname,
                         vm.os_username,
                         _encrypt_vm_password(vm.os_password),
                         vm.extra_cloudinit, vm.os_type,
                         vm.delete_on_termination),
                    )
            _log_activity(conn, "vm_provisioning", str(batch_id), "batch_created",
                           f"Batch '{body.name}' created", user.username)
            return {"batch_id": batch_id, "status": "validated"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/batches")
async def list_batches(
    region_id: Optional[str] = Query(None, description="Filter by region ID"),
    user=Depends(get_current_user),
):
    _ensure_tables()
    uname = user["username"] if isinstance(user, dict) else user.username
    effective_region = get_effective_region_filter(uname, region_id)
    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            if effective_region:
                cur.execute(
                    "SELECT * FROM vm_provisioning_batches WHERE region_id = %s ORDER BY created_at DESC LIMIT 100",
                    (effective_region,),
                )
            else:
                cur.execute(
                    "SELECT * FROM vm_provisioning_batches ORDER BY created_at DESC LIMIT 100"
                )
            rows = [dict(r) for r in cur.fetchall()]
    return rows


@router.get("/batches/{batch_id}")
async def get_batch(batch_id: int, user=Depends(get_current_user)):
    _ensure_tables()
    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM vm_provisioning_batches WHERE id=%s", (batch_id,))
            batch = cur.fetchone()
            if not batch:
                raise HTTPException(status_code=404, detail="Batch not found")
            cur.execute("SELECT * FROM vm_provisioning_vms WHERE batch_id=%s ORDER BY id", (batch_id,))
            vms = [dict(r) for r in cur.fetchall()]
    result = dict(batch)
    # Strip os_password from API response — plaintext (or encrypted) passwords
    # must not be returned to the browser. The field is only needed server-side
    # during batch execution.
    for vm in vms:
        vm["os_password"] = "****" if vm.get("os_password") else ""
    result["vms"] = vms
    return result


@router.get("/batches/{batch_id}/logs")
async def get_batch_logs(batch_id: int, user=Depends(get_current_user)):
    """Return activity_log entries for this batch and its VMs (most recent first)."""
    _ensure_tables()
    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Collect VM ids for this batch
            cur.execute("SELECT id FROM vm_provisioning_vms WHERE batch_id=%s", (batch_id,))
            vm_ids = [str(r["id"]) for r in cur.fetchall()]

            # Fetch batch-level + VM-level log entries
            vm_ids_clause = ""
            params: list = [str(batch_id)]
            if vm_ids:
                placeholders = ",".join(["%s"] * len(vm_ids))
                vm_ids_clause = (
                    f" OR (resource_type='vm_provisioning_vm' AND resource_id IN ({placeholders}))"
                )
                params.extend(vm_ids)

            cur.execute(
                f"""SELECT id, resource_type, resource_id, action, details,
                           actor as created_by, timestamp as created_at
                    FROM activity_log
                    WHERE (resource_type='vm_provisioning' AND resource_id=%s){vm_ids_clause}
                    ORDER BY timestamp DESC
                    LIMIT 200""",
                params,
            )
            logs = [dict(r) for r in cur.fetchall()]
    return logs


@router.post("/batches/{batch_id}/dry-run")
async def dry_run(batch_id: int, user=Depends(get_current_user)):
    """Pre-flight: validate image/flavor/network/SG/quota/IP/name for each VM row."""
    _ensure_tables()
    try:
        with get_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT * FROM vm_provisioning_batches WHERE id=%s", (batch_id,))
                batch = cur.fetchone()
                if not batch:
                    raise HTTPException(status_code=404, detail="Batch not found")
                cur.execute("SELECT * FROM vm_provisioning_vms WHERE batch_id=%s ORDER BY id", (batch_id,))
                vm_rows = [dict(r) for r in cur.fetchall()]

        admin = get_client()
        project_id = admin.resolve_project_id(
            project_name=batch["project_name"],
            domain_name=batch["domain_name"],
        )

        # Ensure provisionsrv has _member_ role in the target project so execution
        # can authenticate as provisionsrv scoped to this project
        from vm_provisioning_service_user import ensure_provisioner_in_project
        ensure_provisioner_in_project(admin, project_id)

        images = {img.get("id"): img for img in admin.list_images()}
        images_by_name = {img.get("name"): img for img in images.values()}
        flavors = {fl.get("id"): fl for fl in admin.list_flavors()}
        flavors_by_name = {fl.get("name"): fl for fl in flavors.values()}
        networks = {net.get("id"): net for net in admin.list_networks(project_id=project_id)}
        networks_by_name = {net.get("name"): net for net in networks.values()}
        sgs_by_name = {sg.get("name"): sg for sg in admin.list_security_groups(project_id=project_id)}
        existing_servers = {s.get("name") for s in admin.list_servers(project_id=project_id, all_tenants=False)}

        quota_usage = admin.get_quota_usage(project_id)
        compute_q = quota_usage.get("compute", {})
        storage_q = quota_usage.get("storage", {})

        def _free(q: dict, key: str) -> int:
            entry = q.get(key, {})
            if isinstance(entry, dict):
                limit = entry.get("limit", -1)
                used = entry.get("in_use", 0)
            else:
                limit = entry
                used = 0
            if limit == -1:
                return 99999
            return limit - used

        vcpu_free = _free(compute_q, "cores")
        ram_free = _free(compute_q, "ram")
        inst_free = _free(compute_q, "instances")
        vol_free = _free(storage_q, "volumes")
        gb_free = _free(storage_q, "gigabytes")

        vcpu_needed = ram_needed = inst_needed = vol_needed = gb_needed = 0

        per_vm_results = []
        has_errors = False

        for vm in vm_rows:
            checks = []
            suffix = vm["vm_name_suffix"]
            count = vm.get("count", 1)

            # Image
            img_obj = images.get(vm.get("image_id") or "") or \
                      images_by_name.get(vm.get("image_name") or "")
            if img_obj:
                checks.append({"check": "image", "status": "ok", "detail": img_obj.get("name")})
                # Warn if requested volume_gb < image virtual/min disk (execute will auto-expand)
                img_min_disk = int(img_obj.get("min_disk") or 0)
                img_virtual_gb = int((img_obj.get("virtual_size") or 0) / 1073741824 + 0.999) if img_obj.get("virtual_size") else 0
                effective_gb = max(int(vm.get("volume_gb", 20)), img_min_disk, img_virtual_gb, 10)
                if effective_gb > int(vm.get("volume_gb", 20)):
                    req_src = f"min_disk={img_min_disk}GB" if img_min_disk >= img_virtual_gb else f"virtual_size={img_virtual_gb}GB"
                    checks.append({"check": "volume_size", "status": "warning",
                                    "detail": f"volume_gb {vm.get('volume_gb')}GB < image {req_src} — will auto-expand to {effective_gb}GB"})
            else:
                checks.append({"check": "image", "status": "error",
                                "detail": f"Image '{vm.get('image_name')}' not found"})
                has_errors = True
                effective_gb = int(vm.get("volume_gb", 20))

            # Flavor (disk == 0)
            fl_obj = flavors.get(vm.get("flavor_id") or "") or \
                     flavors_by_name.get(vm.get("flavor_name") or "")
            if fl_obj:
                if fl_obj.get("disk", -1) != 0:
                    checks.append({"check": "flavor", "status": "error",
                                    "detail": f"Flavor '{fl_obj.get('name')}' has disk={fl_obj.get('disk')} (must be 0)"})
                    has_errors = True
                else:
                    checks.append({"check": "flavor", "status": "ok",
                                    "detail": f"{fl_obj.get('name')} — {fl_obj.get('vcpus')} vCPU / {fl_obj.get('ram')} MB"})
                    vcpu_needed += fl_obj.get("vcpus", 0) * count
                    ram_needed += fl_obj.get("ram", 0) * count
            else:
                checks.append({"check": "flavor", "status": "error",
                                "detail": f"Flavor '{vm.get('flavor_name')}' not found"})
                has_errors = True

            inst_needed += count
            vol_needed += count
            gb_needed += effective_gb * count

            # Network
            net_obj = networks.get(vm.get("network_id") or "") or \
                      networks_by_name.get(vm.get("network_name") or "")
            if net_obj:
                checks.append({"check": "network", "status": "ok", "detail": net_obj.get("name")})
            else:
                checks.append({"check": "network", "status": "error",
                                "detail": f"Network '{vm.get('network_name')}' not found"})
                has_errors = True

            # Security groups
            bad_sgs = []
            sgs_list = vm.get("security_groups") or ["default"]
            if isinstance(sgs_list, str):
                sgs_list = json.loads(sgs_list)
            for sg in sgs_list:
                if sg not in sgs_by_name:
                    bad_sgs.append(sg)
            if bad_sgs:
                checks.append({"check": "security_groups", "status": "error",
                                "detail": f"SGs not found: {bad_sgs}"})
                has_errors = True
            else:
                checks.append({"check": "security_groups", "status": "ok",
                                "detail": ", ".join(sgs_list)})

            # VM name conflict
            for idx in range(1, count + 1):
                nova_name = _vm_nova_name(batch["domain_name"], suffix, idx if count > 1 else None, count)
                if nova_name in existing_servers:
                    checks.append({"check": "name_conflict", "status": "warning",
                                    "detail": f"'{nova_name}' already exists — will skip"})

            # cloud-init size pre-check + Windows cloudbase-init requirement checks
            try:
                hostname_check = _vm_hostname(_vm_nova_name(batch["domain_name"], suffix))
                os_type_check = vm.get("os_type", "linux")
                glance_os_type = ""
                img_name_lower = (img_obj.get("name", "") if img_obj else "").lower()
                if img_obj:
                    raw = (img_obj.get("os_type", "") + img_obj.get("os_distro", "")).lower()
                    glance_os_type = img_obj.get("os_type", "")
                    if "windows" in raw or "win" in raw or "windows" in img_name_lower or "win" in img_name_lower:
                        os_type_check = "windows"
                if os_type_check == "windows":
                    payload = _build_cloudinit_windows(vm["os_username"], vm["os_password"], vm.get("extra_cloudinit"))
                    # Warn about missing os_type Glance property only if we also can't detect
                    # Windows from the image name — if the name includes "win" we already know.
                    if img_obj and not glance_os_type and "win" not in img_name_lower:
                        checks.append({"check": "windows_glance_property", "status": "warning",
                                        "detail": "Image has no os_type=windows property set in Glance — confirm this is a proper OpenStack Windows image"})
                    # Only warn about cloudbase-init if there is no evidence the image includes it.
                    # Evidence: "cloudbase" in image name, or cloudbase_init Glance property present.
                    has_cloudbase = (
                        "cloudbase" in img_name_lower
                        or bool((img_obj or {}).get("cloudbase_init"))
                        or bool((img_obj or {}).get("cloud_init_tool", "").lower() == "cloudbase")
                    )
                    if not has_cloudbase:
                        checks.append({"check": "windows_cloudinit", "status": "warning",
                                        "detail": "Windows VM requires cloudbase-init installed + sysprep'd in image. "
                                                   "Credentials injected via #ps1_sysnative userdata. "
                                                   "Vagrant/VirtualBox images will NOT work — use an image built with cloudbase-init."})
                    # UEFI boot check — Windows 2019+ images with GPT partition table
                    # require hw_firmware_type=uefi set on the Glance image AND OVMF installed
                    # on all compute nodes (nova.conf [libvirt] nvram). Without both, Nova
                    # will either boot with SeaBIOS (GPT not bootable) or fail scheduling.
                    if img_obj and not img_obj.get("hw_firmware_type"):
                        checks.append({"check": "windows_uefi", "status": "warning",
                                        "detail": "Image is missing hw_firmware_type=uefi in Glance. "
                                                   "Windows Server 2019+ uses a GPT partition table which SeaBIOS cannot boot. "
                                                   "Set hw_firmware_type=uefi + hw_machine_type=q35 on this image in Glance — "
                                                   "but only if your compute nodes have OVMF/TianoCore installed and nvram configured in nova.conf. "
                                                   "If UEFI is not supported on your cluster, use an image built with an MBR/hybrid partition table."
                                        })
                else:
                    payload = _build_cloudinit_linux(vm["os_username"], vm["os_password"], hostname_check, vm.get("extra_cloudinit"))
                _encode_userdata(payload)
                checks.append({"check": "cloudinit_size", "status": "ok",
                                "detail": f"{len(payload)} bytes (limit 65535)"})
            except ValueError as e:
                checks.append({"check": "cloudinit_size", "status": "error", "detail": str(e)})
                has_errors = True

            per_vm_results.append({
                "vm_id": vm["id"],
                "vm_name_suffix": suffix,
                "count": count,
                "checks": checks,
            })

        # Quota headroom check
        quota_checks = []
        for label, needed, free in [
            ("vCPU", vcpu_needed, vcpu_free),
            ("RAM MB", ram_needed, ram_free),
            ("Instances", inst_needed, inst_free),
            ("Volumes", vol_needed, vol_free),
            ("Disk GB", gb_needed, gb_free),
        ]:
            if needed > free:
                quota_checks.append({"resource": label, "needed": needed, "free": free,
                                      "status": "error"})
                has_errors = True
            else:
                quota_checks.append({"resource": label, "needed": needed, "free": free,
                                      "status": "ok"})

        dry_run_status = "dry_run_failed" if has_errors else "dry_run_passed"
        results = {"per_vm": per_vm_results, "quota": quota_checks}

        try:
            with get_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "UPDATE vm_provisioning_batches SET status=%s, dry_run_results=%s, updated_at=NOW() WHERE id=%s",
                        (dry_run_status, Json(results), batch_id),
                    )
                _log_activity(conn, "vm_provisioning", str(batch_id),
                               f"vm_provisioning_dry_run_{'passed' if not has_errors else 'failed'}",
                               f"Dry-run {dry_run_status}", user.username)
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

        return {"status": dry_run_status, "results": results}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/batches/{batch_id}/submit")
async def submit_batch(batch_id: int, user=Depends(get_current_user)):
    """Submit batch for admin approval."""
    _ensure_tables()
    try:
        with get_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT * FROM vm_provisioning_batches WHERE id=%s", (batch_id,))
                batch = cur.fetchone()
                if not batch:
                    raise HTTPException(status_code=404, detail="Batch not found")
                if batch["status"] not in ("validated", "dry_run_passed", "dry_run_failed"):
                    raise HTTPException(status_code=400, detail=f"Cannot submit batch in status '{batch['status']}'")
                cur.execute(
                    "UPDATE vm_provisioning_batches SET approval_status='pending_approval', updated_at=NOW() WHERE id=%s",
                    (batch_id,),
                )
            _log_activity(conn, "vm_provisioning", str(batch_id),
                           "vm_provisioning_submitted", "Submitted for approval",
                           user.username)
        return {"status": "pending_approval"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/batches/{batch_id}/decision")
async def batch_decision(batch_id: int, body: DecisionRequest,
                          user=Depends(get_current_user)):
    """Admin: approve or reject a batch."""
    _ensure_tables()
    try:
        with get_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT * FROM vm_provisioning_batches WHERE id=%s", (batch_id,))
                batch = cur.fetchone()
                if not batch:
                    raise HTTPException(status_code=404, detail="Batch not found")
                new_approval = "approved" if body.decision == "approve" else "rejected"
                cur.execute(
                    "UPDATE vm_provisioning_batches SET approval_status=%s, updated_at=NOW() WHERE id=%s",
                    (new_approval, batch_id),
                )
            _log_activity(conn, "vm_provisioning", str(batch_id),
                           f"vm_provisioning_{new_approval}",
                           body.comment or f"Batch {new_approval}",
                           user.username)
        return {"approval_status": new_approval}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/batches/{batch_id}/re-execute")
async def re_execute_batch(batch_id: int, user=Depends(get_current_user)):
    """Re-execute a failed or partially-failed batch, retrying only the failed VMs."""
    _ensure_tables()
    try:
        with get_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT * FROM vm_provisioning_batches WHERE id=%s", (batch_id,))
                batch = cur.fetchone()
            if not batch:
                raise HTTPException(status_code=404, detail="Batch not found")
            if batch["status"] not in ("failed", "partially_failed", "complete"):
                raise HTTPException(
                    status_code=400,
                    detail=f"Re-execute requires a completed/failed batch (current: {batch['status']})"
                )
            if batch["require_approval"] and batch["approval_status"] != "approved":
                raise HTTPException(status_code=400, detail="Batch not yet approved")

            # Reset failed VM rows; complete rows will be skipped by the thread
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE vm_provisioning_vms SET status='pending', error_msg=NULL "
                    "WHERE batch_id=%s AND status='failed'",
                    (batch_id,),
                )
                cur.execute(
                    "UPDATE vm_provisioning_batches SET status='executing', updated_at=NOW() WHERE id=%s",
                    (batch_id,),
                )

        re_operator_email = None
        try:
            info = ldap_auth.get_user_info(user.username)
            re_operator_email = info.get("email") or None
        except Exception:
            pass
        t = threading.Thread(
            target=_execute_batch_thread,
            args=(batch_id, re_operator_email),
            daemon=True,
        )
        t.start()

        return {"status": "executing", "batch_id": batch_id,
                "message": "Re-execution started — retrying failed VMs only"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/batches/{batch_id}/execute")
async def execute_batch(batch_id: int, user=Depends(get_current_user)):
    """Execute batch: provision all VMs in a background thread."""
    _ensure_tables()
    try:
        with get_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT * FROM vm_provisioning_batches WHERE id=%s", (batch_id,))
                batch = cur.fetchone()
            if not batch:
                raise HTTPException(status_code=404, detail="Batch not found")
            if batch["require_approval"] and batch["approval_status"] != "approved":
                raise HTTPException(status_code=400, detail="Batch not yet approved")
            if batch["status"] == "executing":
                raise HTTPException(status_code=400, detail="Batch already executing")
            if batch["status"] == "complete":
                raise HTTPException(status_code=400, detail="Batch already complete — use Re-run to retry failed VMs")

            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE vm_provisioning_batches SET status='executing', updated_at=NOW() WHERE id=%s",
                    (batch_id,),
                )

        operator_email = None
        try:
            info = ldap_auth.get_user_info(user.username)
            operator_email = info.get("email") or None
        except Exception:
            pass
        t = threading.Thread(
            target=_execute_batch_thread,
            args=(batch_id, operator_email),
            daemon=True,
        )
        t.start()

        return {"status": "executing", "batch_id": batch_id,
                "message": "Execution started — poll GET /api/vm-provisioning/batches/{id} for status"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/batches/{batch_id}/reset")
async def reset_batch(batch_id: int, user=Depends(get_current_user)):
    """Force-reset a stuck/executing batch back to 'failed' so it can be re-run."""
    _ensure_tables()
    try:
        with get_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT * FROM vm_provisioning_batches WHERE id=%s", (batch_id,))
                batch = cur.fetchone()
            if not batch:
                raise HTTPException(status_code=404, detail="Batch not found")
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE vm_provisioning_batches SET status='failed', updated_at=NOW() WHERE id=%s",
                    (batch_id,),
                )
                cur.execute(
                    "UPDATE vm_provisioning_vms SET status='failed', error_msg='Force-reset by user' "
                    "WHERE batch_id=%s AND status IN ('pending','executing')",
                    (batch_id,),
                )
            _log_activity(conn, "vm_provisioning", str(batch_id), "batch_reset",
                           "Batch force-reset to failed by user", user.username)
        return {"status": "failed", "batch_id": batch_id, "message": "Batch reset — use Re-run to retry"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/batches/{batch_id}")
async def delete_batch(batch_id: int, user=Depends(get_current_user)):
    _ensure_tables()
    try:
        with get_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT status FROM vm_provisioning_batches WHERE id=%s", (batch_id,))
                row = cur.fetchone()
                if not row:
                    raise HTTPException(status_code=404, detail="Batch not found")
                if row["status"] == "executing":
                    raise HTTPException(status_code=400, detail="Cannot delete a batch that is actively executing")
                cur.execute("DELETE FROM vm_provisioning_batches WHERE id=%s", (batch_id,))
        return {"deleted": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

